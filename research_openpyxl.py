from openpyxl import load_workbook
import os

def inspect_file(path, sheet_name):
    if not os.path.exists(path):
        print(f"File not found: {path}")
        return
    print(f"\nInspecting {path} [{sheet_name}]")
    wb = load_workbook(path, data_only=True)
    if sheet_name not in wb.sheetnames:
        print(f"Sheet {sheet_name} not found. Sheets: {wb.sheetnames}")
        return
    ws = wb[sheet_name]
    for r in range(1, 5):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, 15)]
        print(f"Row {r}: {row_vals}")

inspect_file(r"Templates Master\$index.xlsx", "Schemas")
inspect_file(r"Templates Legacy\$index.xlsm", "Paths")
