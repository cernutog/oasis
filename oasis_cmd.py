"""
cmd.py — OASIS runner script.

Usage examples:
  python cmd.py convert                      # Convert full Templates Legacy folder
  python cmd.py convert listAlerts           # Convert only files matching a name
  python cmd.py inspect schemas              # Print Schemas sheet of generated $index.xlsx
  python cmd.py inspect schemas listAlerts   # Print Schemas sheet of a specific output xlsx
  python cmd.py check arrays                 # Find all root-level array schemas in output
"""
import sys
import shutil
import tempfile
from pathlib import Path

ROOT = Path(__file__).parent
sys.path.insert(0, str(ROOT / "src"))

from legacy_converter import LegacyConverter
import openpyxl

LEGACY_DIR  = ROOT / "Templates Legacy"
MASTER_DIR  = ROOT / "Templates Master"
OUTPUT_DIR  = ROOT / "tmp" / "cmd_output"


def run_convert(filter_name: str = ""):
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    if filter_name:
        # Copy matching files + index into a temp dir
        in_dir = Path(tempfile.mkdtemp(prefix="oasis_in_"))
        for f in LEGACY_DIR.iterdir():
            if "index" in f.name.lower() or filter_name.lower() in f.name.lower():
                shutil.copy(f, in_dir / f.name)
        input_dir = in_dir
    else:
        input_dir = LEGACY_DIR

    lc = LegacyConverter(
        input_dir=str(input_dir),
        output_dir=str(OUTPUT_DIR),
        master_dir=str(MASTER_DIR),
    )
    lc.convert()
    print(f"\nOutput written to: {OUTPUT_DIR}")


def inspect_schemas(file_filter: str = ""):
    target = OUTPUT_DIR / "$index.xlsx"
    if file_filter:
        matches = list(OUTPUT_DIR.rglob(f"*{file_filter}*.xlsx"))
        if matches:
            target = matches[0]

    if not target.exists():
        print(f"File not found: {target}")
        print("Run 'python cmd.py convert' first.")
        return

    wb = openpyxl.load_workbook(target)
    if "Schemas" not in wb.sheetnames:
        print(f"No Schemas sheet in {target.name}. Sheets: {wb.sheetnames}")
        return

    ws = wb["Schemas"]
    print(f"=== Schemas sheet of {target.name} ===")
    for row in ws.iter_rows(values_only=True):
        if any(v for v in row):
            print(row)


def check_arrays():
    target = OUTPUT_DIR / "$index.xlsx"
    if not target.exists():
        print(f"File not found: {target}")
        return

    wb = openpyxl.load_workbook(target)
    if "Schemas" not in wb.sheetnames:
        print("No Schemas sheet.")
        return

    ws = wb["Schemas"]
    print("Root-level ARRAY schemas:")
    for row in ws.iter_rows(values_only=True):
        vals = list(row)
        name   = vals[0] if len(vals) > 0 else None
        parent = vals[1] if len(vals) > 1 else None
        typ    = vals[3] if len(vals) > 3 else None
        items  = vals[4] if len(vals) > 4 else None
        if typ == "array" and not parent and name and name != "Name":
            print(f"  {name:40s}  items={items}")

    # Also check for double-array pattern: property with type=array pointing to an array schema
    print("\nDouble-array candidates (property type=array, schema_name != ''):")
    arr_schemas = set()
    for row in ws.iter_rows(values_only=True):
        vals = list(row)
        parent = vals[1] if len(vals) > 1 else None
        typ    = vals[3] if len(vals) > 3 else None
        items  = vals[4] if len(vals) > 4 else None
        name   = vals[0] if len(vals) > 0 else None
        if typ == "array" and not parent and name and name != "Name":
            arr_schemas.add(name)

    found = False
    for row in ws.iter_rows(values_only=True):
        vals = list(row)
        parent = vals[1] if len(vals) > 1 else None
        typ    = vals[3] if len(vals) > 3 else None
        items  = vals[4] if len(vals) > 4 else None
        schema_name = vals[5] if len(vals) > 5 else None
        prop_name = vals[0] if len(vals) > 0 else None
        if typ == "array" and parent and schema_name and schema_name in arr_schemas:
            print(f"  DOUBLE-ARRAY: property '{prop_name}' (parent={parent}) "
                  f"type=array items_schema={schema_name} which is also array")
            found = True
    if not found:
        print("  (none found — no double-array anti-pattern detected)")


if __name__ == "__main__":
    args = sys.argv[1:]
    if not args:
        print(__doc__)
        sys.exit(0)

    cmd = args[0].lower()

    if cmd == "convert":
        run_convert(args[1] if len(args) > 1 else "")
    elif cmd == "inspect" and len(args) >= 2 and args[1].lower() == "schemas":
        inspect_schemas(args[2] if len(args) > 2 else "")
    elif cmd == "check" and len(args) >= 2 and args[1].lower() == "arrays":
        check_arrays()
    else:
        print(__doc__)
