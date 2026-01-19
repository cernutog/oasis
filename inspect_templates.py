"""Inspect master template files."""
from openpyxl import load_workbook
import os

templates_dir = 'Templates Master'

for filename in os.listdir(templates_dir):
    if filename.endswith('.xlsx'):
        filepath = os.path.join(templates_dir, filename)
        print(f'=== {filename} ===')
        try:
            wb = load_workbook(filepath)
            print(f'Sheets: {wb.sheetnames}')
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                print(f'  {sheet}: max_row={ws.max_row}, max_col={ws.max_column}')
                # Print first 2 rows
                for row in range(1, min(3, ws.max_row + 1)):
                    vals = [ws.cell(row, c).value for c in range(1, min(10, ws.max_column + 1))]
                    print(f'    Row {row}: {vals}')
            wb.close()
        except Exception as e:
            print(f'  Error: {e}')
        print()
