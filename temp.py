import sys
import shutil
from pathlib import Path
import re
import yaml

from src.legacy_converter import LegacyConverter


def verify_legacy_tracing_table_header() -> int:
    legacy_root = Path(__file__).parent / "Templates Legacy"
    if not legacy_root.exists():
        print("SKIP: Templates Legacy not found")
        return 0

    out_dir = Path(__file__).parent / "tmp" / "_verify_tracing_merge"
    if out_dir.exists():
        shutil.rmtree(out_dir)
    out_dir.parent.mkdir(parents=True, exist_ok=True)

    logs = []

    def _log(msg: str):
        logs.append(str(msg))

    converter = LegacyConverter(
        str(legacy_root),
        str(out_dir),
        log_callback=_log,
        include_descriptions_in_collision=False,
        include_examples_in_collision=False,
    )
    ok = converter.convert(tracing_enabled=True)
    if not ok:
        print("FAIL: conversion failed")
        return 2

    text = "\n".join(logs)
    if "| SCHEMA NAME" not in text:
        print("FAIL: tracing table header missing")
        return 1

    print("PASS")
    return 0


def _find_numbering_gaps(names: list[str]) -> list[tuple[str, list[int]]]:
    """Return [(stem, missing_numbers)] for any stem that has a base and numeric variants,
    but missing intermediate suffixes."""
    stems: dict[str, dict[str, object]] = {}
    for n in names:
        s = str(n)
        m = re.search(r"^(.*?)(\d+)$", s)
        if not m:
            stems.setdefault(s, {"base": True, "nums": set()})
            continue
        stem = m.group(1)
        num = int(m.group(2))
        entry = stems.setdefault(stem, {"base": False, "nums": set()})
        entry["nums"].add(num)
    gaps: list[tuple[str, list[int]]] = []
    for stem, info in stems.items():
        if not info.get("nums"):
            continue
        if not info.get("base"):
            continue
        nums = sorted(info["nums"])  # type: ignore[arg-type]
        exp = list(range(1, max(nums) + 1))
        missing = [x for x in exp if x not in nums]
        if missing:
            gaps.append((stem, missing))
    return sorted(gaps, key=lambda t: t[0].lower())


def verify_collision_numbering_contiguous_in_oas() -> int:
    """Checks that in the generated OAS there are no gaps like Foo + Foo2 without Foo1."""
    oas_path = Path(r"C:\\EBA Clearing\\APIs\\Generated OAS\\RT1 API Participants\\generated_oas_3.1.yaml")
    if not oas_path.exists():
        print(f"SKIP: OAS not found: {oas_path}")
        return 0
    spec = yaml.safe_load(oas_path.read_text(encoding="utf-8", errors="ignore"))
    schemas = (spec or {}).get("components", {}).get("schemas", {})
    names = list((schemas or {}).keys())
    gaps = _find_numbering_gaps(names)
    bad = [g for g in gaps if g[0].lower().startswith("errorcodedescription")]
    if bad:
        print("FAIL: numbering gaps in OAS:")
        for stem, missing in bad[:50]:
            print(f"  {stem}: missing {missing}")
        return 1
    print("PASS: OAS collision numbering contiguous (ErrorCodeDescription)")
    return 0


def _run_conversion_to_tmp() -> Path | None:
    # Prefer the real RT1 legacy folder if present; otherwise fall back to local fixtures.
    legacy_root = Path(r"C:\\EBA Clearing\\APIs\\Templates\\RT1 API Participants\\Legacy")
    if not legacy_root.exists():
        legacy_root = Path(__file__).parent / "Templates Legacy"
    if not legacy_root.exists():
        print("SKIP: legacy input not found")
        return None

    out_dir = Path(__file__).parent / "tmp" / "_verify_collision_numbering"
    if out_dir.exists():
        shutil.rmtree(out_dir)
    out_dir.parent.mkdir(parents=True, exist_ok=True)

    logs: list[str] = []

    def _log(msg: str):
        logs.append(str(msg))

    converter = LegacyConverter(
        str(legacy_root),
        str(out_dir),
        log_callback=_log,
        include_descriptions_in_collision=False,
        include_examples_in_collision=False,
    )
    ok = converter.convert(tracing_enabled=False)
    if not ok:
        print("FAIL: conversion failed")
        return None

    idx = out_dir / "$index.xlsx"
    if not idx.exists():
        print("FAIL: conversion produced no $index.xlsx")
        return None

    return out_dir


def verify_collision_numbering_contiguous_in_converted_index() -> int:
    """Checks that in the converted $index.xlsx Schemas sheet there are no gaps."""
    conv_dir = _run_conversion_to_tmp()
    if conv_dir is None:
        return 0
    idx = conv_dir / "$index.xlsx"

    # Lightweight parse via openpyxl to avoid pandas ExcelFile weirdness on windows paths
    import openpyxl
    wb = openpyxl.load_workbook(idx)
    if "Schemas" not in wb.sheetnames:
        print("FAIL: Schemas sheet missing")
        return 2
    ws = wb["Schemas"]

    # Find header row
    header_row = None
    for r in range(1, min(ws.max_row, 40) + 1):
        row_vals = [str(ws.cell(row=r, column=c).value or "").strip().lower() for c in range(1, min(ws.max_column, 30) + 1)]
        if any(v == "name" for v in row_vals) and any(v == "type" for v in row_vals):
            header_row = r
            headers = row_vals
            break
    if not header_row:
        print("FAIL: header row not found")
        return 2
    try:
        name_col = headers.index("name") + 1
    except ValueError:
        print("FAIL: name column not found")
        return 2

    names: list[str] = []
    for r in range(header_row + 1, ws.max_row + 1):
        v = ws.cell(row=r, column=name_col).value
        s = "" if v is None else str(v).strip()
        if not s or s.lower() == "nan":
            continue
        names.append(s)

    gaps = _find_numbering_gaps(names)
    bad = [g for g in gaps if g[0].lower().startswith("errorcodedescription")]
    if bad:
        print("FAIL: numbering gaps in converted index:")
        for stem, missing in bad[:50]:
            print(f"  {stem}: missing {missing}")
        return 1
    print("PASS: converted index collision numbering contiguous (ErrorCodeDescription)")
    return 0


def verify_tracer_no_spurious_bic_differences() -> int:
    """Ensure Bic11/Bic111 don't show Max Val/Regex as differences when identical."""
    conv_dir = _run_conversion_to_tmp()
    if conv_dir is None:
        return 0

    # Run standalone tracer on the freshly converted folder
    logs: list[str] = []

    def _log(msg: str):
        logs.append(str(msg))

    conv = LegacyConverter(str(conv_dir), str(conv_dir), log_callback=_log)
    ok = conv.run_standalone_check(str(conv_dir))
    if not ok:
        print("FAIL: standalone check failed")
        return 2

    # Extract the table row blocks for Bic11 / Bic111
    def collect_block(schema_name: str) -> str:
        start = None
        for i, l in enumerate(logs):
            if l.startswith("|") and (f"| {schema_name}" in l):
                start = i
                break
        if start is None:
            return ""
        out_lines: list[str] = []
        for j in range(start, min(start + 30, len(logs))):
            out_lines.append(logs[j])
            if logs[j].startswith("-") and set(logs[j]) == {"-"}:
                # separator line
                break
        return "\n".join(out_lines)

    b11 = collect_block("Bic11")
    b111 = collect_block("Bic111")
    if not b11 and not b111:
        print("SKIP: Bic11/Bic111 not found in tracer output")
        return 0

    bad_lines = []
    for nm, blk in [("Bic11", b11), ("Bic111", b111)]:
        if not blk:
            continue
        # If either bullet appears in DIFFERENCES, treat as failure
        if re.search(r"-\s*Max\s+Val\s*:", blk, re.IGNORECASE) or re.search(r"-\s*Regex\s*:", blk, re.IGNORECASE):
            bad_lines.append(nm)

    if bad_lines:
        print("FAIL: tracer shows spurious differences for: " + ", ".join(bad_lines))
        return 1

    print("PASS: tracer does not show spurious Bic11/Bic111 differences")
    return 0


if __name__ == "__main__":
    rc = 0
    for fn in [
        verify_legacy_tracing_table_header,
        verify_collision_numbering_contiguous_in_converted_index,
        verify_tracer_no_spurious_bic_differences,
        # OAS check is intentionally skipped here: this runner validates the converter output.
    ]:
        rc = max(rc, int(fn()))
    sys.exit(rc)
