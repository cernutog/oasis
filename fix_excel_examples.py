import pandas as pd
import yaml
import os
import openpyxl
import difflib

# Configuration
INDEX_FILE = "API Templates/$index.xlsm"
REF_YAML = "Expected results/EBACL_FPAD_20251110_OpenApi3.1_FPAD_API_Participant_4.1_v20251212.yaml"
API_TEMPLATES_DIR = "API Templates"

def load_yaml(path):
    with open(path, 'r') as f:
        return yaml.safe_load(f)

def clean_yaml_dump(data):
    # Dumps object to YAML string, ensuring clean block format
    return yaml.dump(data, sort_keys=False, default_flow_style=False).strip()

def find_best_match_file(target_name, directory, files_list):
    """
    Tries to find the actual file in directory that matches target_name.
    Handles exact match, case insensitivity, and minor typos.
    """
    if not target_name:
        return None
    
    # Normalize target
    target = target_name.strip()
    if not target.endswith('.xlsm') and not target.endswith('.xlsx'):
        target += '.xlsm'
        
    # 1. Exact match
    if target in files_list:
        return os.path.join(directory, target)
    
    # 2. Case insensitive
    for f in files_list:
        if f.lower() == target.lower():
            return os.path.join(directory, f)
            
    # 3. Fuzzy match / typo correction
    matches = difflib.get_close_matches(target, files_list, n=1, cutoff=0.6)
    if matches:
        print(f"  Fuzzy matching '{target_name}' to '{matches[0]}'")
        return os.path.join(directory, matches[0])
        
    return None

def main():
    print(f"Loading Reference YAML: {REF_YAML}")
    ref_oas = load_yaml(REF_YAML)
    
    print(f"Loading Index: {INDEX_FILE}")
    # Read Path mapping (File -> Path / Verb)
    paths_df = pd.read_excel(INDEX_FILE, sheet_name="Paths")
    
    # Get all files in directory for lookup
    files_in_dir = os.listdir(API_TEMPLATES_DIR)
    
    # Filter valid rows (where file name exists)
    # print("Columns:", paths_df.columns.tolist())
    # print("Head:\n", paths_df.head().to_string())
    
    paths_df = paths_df[paths_df.iloc[:, 0].notna()]
    print(f"Rows after filtering: {len(paths_df)}")
    
    first_row = True
    for idx, row in paths_df.iterrows():
        # Skip header rows
        filename = str(row.iloc[0]).strip()
        if filename.lower() in ["excel file", "file name", "filename"]: continue

        # Extract values
        filename = row.iloc[0]
        path = row.iloc[1]
        
        # Col 3 is Method/Verb
        verb_raw = row.iloc[3] if len(row) > 3 else None
        verb = str(verb_raw).strip().lower() if pd.notna(verb_raw) else None
        
        # Resolve File Path
        excel_path = find_best_match_file(str(filename), API_TEMPLATES_DIR, files_in_dir)
        if not excel_path:
             print(f"  Error: Could not resolve file for '{filename}'")
             continue

        # 1. Lookup in Ref YAML
        try:
            op = ref_oas["paths"][path][verb]
            content = op.get("requestBody", {}).get("content", {}).get("application/json", {})
            examples = content.get("examples", {})
        except KeyError:
            print(f"Warning: Path/Verb {path} {verb} not found in Reference YAML.")
            # Debug: print closest match?
            # if path in ref_oas["paths"]: print(f"  (Path exists, verbs: {list(ref_oas['paths'][path].keys())})")
            continue
            
        if not examples:
            continue
            
        print(f"\nProcessing {filename} (Path: {path}, Verb: {verb})")
        print(f"Found examples in Ref: {list(examples.keys())}")
        
        # 2. Open Excel
        if not os.path.exists(excel_path):
            print(f"  Error: File {excel_path} not found.")
            continue
            
        try:
            # We use openpyxl to modify without destroying formatting
            # CRITICAL: keep_vba=True for .xlsm files
            wb = openpyxl.load_workbook(excel_path, keep_vba=True)
            if "Body Example" not in wb.sheetnames:
                print("  Skipping: 'Body Example' sheet not found.")
                continue
                
            ws = wb["Body Example"]
            
            # Find columns
            # Header assumed at row 1 (index 1) or we scan
            header_row = None
            col_map = {} # Name -> Col Index (1-based)
            
            # Simple scan for "Example Name" and "Body"
            for r in range(1, 6):
                row_vals = [c.value for c in ws[r]]
                if "Example Name" in row_vals and ("Body" in row_vals or "Value" in row_vals):
                    header_row = r
                    for i, val in enumerate(row_vals):
                        if val: col_map[val.strip()] = i + 1
                    break
            
            if not header_row:
                print("  Error: Could not find header in 'Body Example'.")
                continue
                
            body_col_idx = col_map.get("Body")
            name_col_idx = col_map.get("Example Name")
            
            if not body_col_idx or not name_col_idx:
                 print("  Error: Columns missing.")
                 continue

            # Iterate rows
            updates = 0
            for r in range(header_row + 1, ws.max_row + 1):
                cell_name_obj = ws.cell(row=r, column=name_col_idx)
                ex_name = str(cell_name_obj.value).strip() if cell_name_obj.value else ""
                
                if ex_name in examples:
                    # Get clean YAML
                    ref_ex_obj = examples[ex_name].get("value")
                    clean_yaml = clean_yaml_dump(ref_ex_obj)
                    
                    # Update Cell
                    cell_body = ws.cell(row=r, column=body_col_idx)
                    cell_body.value = clean_yaml
                    updates += 1
            
            if updates > 0:
                print(f"  Updated {updates} examples.")
                wb.save(excel_path)
            else:
                print("  No matching examples found to update.")
                
        except Exception as e:
            print(f"  Error processing Excel: {e}")

if __name__ == "__main__":
    main()
