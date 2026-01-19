"""Check font color extraction."""
import json
from openpyxl import load_workbook

# Check what font color is being used for gray columns
wb = load_workbook('Imported Templates/v5/create-account-assessment.xlsx')
ws = wb['Parameters']

# Check column 4 (Type) font color
c = ws.cell(2, 4)
print(f'Col 4 (Type) font color: {c.font.color.rgb}')
print(f'Expected: FF000000 (black)')

# Check what the extraction produced
config = json.load(open('src/oas_importer/styles_config.json'))
col4_style = config['operation']['Parameters']['column_header_styles'].get('4', {})
font_config = col4_style.get('font', {})
print(f'Extracted col4 font color from config: {font_config.get("color")}')
wb.close()
