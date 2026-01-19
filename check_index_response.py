"""Check $index.xlsx Responses sheet."""
import openpyxl

# filename = 'Imported Templates/$index.xlsx'
filename = 'API Templates/$index.xlsm'
print(f"Checking {filename}...")


try:
    wb = openpyxl.load_workbook(filename)
    if 'Responses' in wb.sheetnames:
        ws = wb['Responses']
        print(f"Sheet 'Responses' found.")
        header = [c.value for c in ws[1]]
        print(f"Header: {header}")
        
        # Dump first 50 rows
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            print(f"Row {i+1}: {row}")
            if i >= 50: break



    else:
        print("Sheet 'Responses' NOT found.")

except Exception as e:
    print(f"Error: {e}")
