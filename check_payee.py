"""Check Custom Extensions for payee-verifications endpoint."""
import openpyxl

wb = openpyxl.load_workbook(r'Imported Templates\$index.xlsx')
ws = wb['Paths']

# Find row with payee-verifications
for r in range(3, 20):
    path = ws.cell(row=r, column=2).value
    if path and 'payee' in str(path):
        print(f"Row {r}: {path}")
        ext = ws.cell(row=r, column=9).value
        if ext:
            print("Custom Extensions:")
            for i, line in enumerate(ext.split('\n')[:30]):
                indent = len(line) - len(line.lstrip())
                print(f"  {i}: [{indent}sp] {line[:70]}")
        else:
            print("  (empty)")
        break
