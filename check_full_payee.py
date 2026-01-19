"""Check FULL Custom Extensions for payee-verifications endpoint."""
import openpyxl

wb = openpyxl.load_workbook(r'Imported Templates\$index.xlsx')
ws = wb['Paths']

for r in range(3, 20):
    path = ws.cell(row=r, column=2).value
    if path and 'payee' in str(path):
        print(f"Row {r}: {path}")
        ext = ws.cell(row=r, column=9).value
        if ext:
            print("Custom Extensions:")
            lines = ext.split('\n')
            for i, line in enumerate(lines):
                # Print all lines to find the 500 block and potential duplicates
                indent = len(line) - len(line.lstrip())
                print(f"{i:03d}: [{indent:02d}sp] {line}")
        else:
            print("  (empty)")
        break
