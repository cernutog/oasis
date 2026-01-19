import openpyxl
import os

filename = 'Imported Templates/$index.xlsx'
print(f"Checking {filename} Responses Sheet...")

wb = openpyxl.load_workbook(filename)
if 'Responses' in wb.sheetnames:
    ws = wb['Responses']
    rows = list(ws.iter_rows(values_only=True))
    for i, row in enumerate(rows[:10]):
        # Print only first 5 cols
        print(f"Row {i+1}: {row[:5]}")
else:
    print("Responses sheet not found")

print(f"Header: {rows[0]}")
# Find 'Schema Name' column index
try:
    schema_idx = rows[0].index("Schema Name\n(for Type or Items Data Type = 'schema'||'header')")
    print(f"Schema Name Index: {schema_idx}")
except ValueError:
    print("Schema Name column not found by exact name.")
    # Try finding approx
    for i, h in enumerate(rows[0]):
        if h and "Schema Name" in h:
            print(f"Found Schema Name at {i}")
            schema_idx = i
            break

for i, row in enumerate(rows[1:10]):
    print(f"Row {i+2}: {row}")
