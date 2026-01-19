import yaml
import openpyxl
import pandas as pd
import os

GOLD_MASTER = "Expected results/EBACL_FPAD_20251110_OpenApi3.1_FPAD_API_Participant_4.1_v20251212.yaml"
INDEX_FILE = "API Templates/$index.xlsm"
TEMPLATES_DIR = "API Templates"


def restore_all():
    print(f"Loading Gold Master: {GOLD_MASTER}")
    with open(GOLD_MASTER, "r") as f:
        gm_data = yaml.safe_load(f)

    print(f"Loading Index: {INDEX_FILE}")
    # Read Paths sheet without header to safely access by integer index
    df_index = pd.read_excel(INDEX_FILE, sheet_name="Paths", header=None)

    # Iterate rows. Assumption: Row 0 is header? Data starts Row 1?
    # We look for rows where Col 0 (File) and Col 1 (Path) and Col 3 (Method) are present.

    for idx, row in df_index.iterrows():
        file_base = row[0]
        ex_path = row[1]
        method = row[3]

        # Skip invalid rows or headers
        if pd.isna(file_base) or pd.isna(ex_path) or pd.isna(method):
            continue

        file_base = str(file_base).strip()
        if file_base.lower() == "file" or file_base.startswith("Unnamed"):
            continue  # Header or junk

        ex_path = str(ex_path).strip()
        method = str(method).strip().lower()

        excel_filename = f"{TEMPLATES_DIR}/{file_base}.xlsm"

        # Check if Excel exists
        if not os.path.exists(excel_filename):
            # Try fuzzy matching?
            # Or just skip. User provided restricted workspace.
            # print(f"Skipping missing file: {excel_filename}")
            continue

        print(f"Processing {excel_filename} (Path: {ex_path}, Method: {method})")

        # Look up in Gold Master
        try:
            path_item = gm_data["paths"][ex_path]
            op = path_item[method]

            if "requestBody" not in op:
                print(f"  No requestBody in Gold Master for this operation. Skipping.")
                continue

            content = op["requestBody"].get("content", {}).get("application/json", {})
            examples = content.get("examples", {})

            if not examples:
                print(f"  No examples found in Gold Master requestBody.")
                continue

            # Prepare restored examples
            restored_data = {}
            for ex_name, ex_val in examples.items():
                if "value" in ex_val:
                    # Dump as block-style YAML
                    clean_yaml = yaml.dump(
                        ex_val["value"], default_flow_style=False, sort_keys=False
                    )
                    restored_data[ex_name] = clean_yaml

            if not restored_data:
                continue

            # Update Excel
            try:
                wb = openpyxl.load_workbook(excel_filename, keep_vba=True)
                if "Body Example" not in wb.sheetnames:
                    print(f"  Sheet 'Body Example' not found. Skipping.")
                    continue

                sheet = wb["Body Example"]
                changes = 0

                # Iterate rows in Body Example sheet
                # Col A = Name, Col B = Body (Data starts row 2)
                for r in sheet.iter_rows(min_row=2, max_col=2):
                    cell_name = r[0]
                    cell_body = r[1]
                    name_str = str(cell_name.value).strip()

                    if name_str in restored_data:
                        print(f"  Updating '{name_str}'")
                        cell_body.value = restored_data[name_str]
                        changes += 1

                if changes > 0:
                    wb.save(excel_filename)
                    print(f"  Saved updates to {excel_filename}")
                else:
                    print(f"  No matching example rows found in sheet.")

            except Exception as e:
                print(f"  Error updating Excel: {e}")

        except KeyError:
            print(f"  Path or Method not found in Gold Master.")
        except Exception as e_gm:
            print(f"  Error reading Gold Master data: {e_gm}")


if __name__ == "__main__":
    restore_all()
