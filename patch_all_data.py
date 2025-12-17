import openpyxl
import os
import glob

base_dir = r"C:\Users\giuse\.gemini\antigravity\scratch\OAS_Generation_Tool\API Templates"

def get_col_map(row):
    headers = {}
    for cell in row:
        if cell.value: headers[str(cell.value).strip()] = cell.column
    return headers

def patch_all():
    files = glob.glob(os.path.join(base_dir, "*.xlsm"))
    for f in files:
        if "$index" in f: continue
        print(f"Processing {os.path.basename(f)}...")
        
        try:
            wb = openpyxl.load_workbook(f, keep_vba=True)
            modified = False
            
            for sheet in wb.worksheets:
                # 1. Parameter Patching (Parameters Sheet or similar)
                if sheet.title == "Parameters" or sheet.title == "Path":
                    # Find Headers in first 5 rows
                    headers = {}
                    header_row_index = 1
                    for r_idx, row in enumerate(sheet.iter_rows(max_row=5), start=1):
                         h = get_col_map(row)
                         if "Name" in h and ("In" in h or "Location" in h):
                              headers = h
                              header_row_index = r_idx
                              break
                    
                    if headers:
                        name_col = headers.get("Name")
                        in_col = headers.get("In") or headers.get("Location")
                        mand_col = headers.get("Mandatory") or headers.get("Required")
                        
                        schema_keys = [
                            "Schema Name", 
                            "Schema Name\n(if Type = schema)",
                            "Schema Name\n(for Type or Items Data Type = 'schema')",
                            "Schema Name\n(for Type or Items Data Type = 'schema'||'header')"
                        ]
                        schema_col = None
                        for k in schema_keys:
                            if k in headers:
                                schema_col = headers[k]
                                break
                        
                        target_params = ["bulkId", "reportId", "fuid"]
                        
                        for row in sheet.iter_rows(min_row=header_row_index+1):
                            nm = row[name_col-1].value if name_col else None
                            if str(nm) in target_params:
                                print(f"  Patching Path Parameter: {nm}")
                                if in_col: row[in_col-1].value = "path"
                                if mand_col: row[mand_col-1].value = "Yes"
                                # Also clear Schema Name if present to avoid invalid refs
                                if schema_col: row[schema_col-1].value = None
                                modified = True

                # 2. General 'fri' Cleanup in ALL sheets
                # Scan for 'fri' rows and clear Schema Name
                # Need to find headers first for every sheet
                headers = {}
                header_row_index = 1
                for r_idx, row in enumerate(sheet.iter_rows(max_row=5), start=1):
                     h = get_col_map(row)
                     if "Name" in h:
                          headers = h
                          header_row_index = r_idx
                          break
                
                if headers:
                    name_col = headers.get("Name")
                    schema_keys = [
                        "Schema Name", 
                        "Schema Name\n(if Type = schema)",
                        "Schema Name\n(for Type or Items Data Type = 'schema')",
                        "Schema Name\n(for Type or Items Data Type = 'schema'||'header')"
                    ]
                    schema_col = None
                    for k in schema_keys:
                        if k in headers:
                            schema_col = headers[k]
                            break
                    
                    if name_col and schema_col:
                        for row in sheet.iter_rows(min_row=header_row_index+1):
                            nm = row[name_col-1].value
                            if str(nm) == "fri":
                                # Check if schema name is FpadResponseIdentifier
                                curr_sch = row[schema_col-1].value
                                if str(curr_sch) == "FpadResponseIdentifier":
                                     print(f"  Clearing Schema Name for 'fri' in {sheet.title}")
                                     row[schema_col-1].value = None
                                     modified = True

            if modified:
                wb.save(f)
                print("  Saved.")
                
        except Exception as e:
            print(f"Error processing {f}: {e}")

if __name__ == "__main__":
    patch_all()
