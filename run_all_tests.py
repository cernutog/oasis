import sys, os, shutil
sys.path.insert(0, os.path.join(os.getcwd(), 'src'))
from legacy_converter import LegacyConverter
from pathlib import Path
import openpyxl

def compare_files(expected_path, actual_path):
    if not actual_path.exists():
        return f"  [FAIL] {actual_path.name} missing"
        
    wb_exp = openpyxl.load_workbook(expected_path, data_only=True)
    wb_act = openpyxl.load_workbook(actual_path, data_only=True)
    
    mismatches = []
    for sn in wb_exp.sheetnames:
        if sn not in wb_act.sheetnames:
            mismatches.append(f"Sheet {sn} missing")
            continue
            
        ws_exp = wb_exp[sn]
        ws_act = wb_act[sn]
        
        sheet_mismatches = 0
        max_r = max(ws_exp.max_row, ws_act.max_row)
        max_c = max(ws_exp.max_column, ws_act.max_column)
        
        for r in range(1, max_r + 1):
            for c in range(1, max_c + 1):
                v_exp = str(ws_exp.cell(r, c).value).strip() if ws_exp.cell(r, c).value is not None else ""
                v_act = str(ws_act.cell(r, c).value).strip() if ws_act.cell(r, c).value is not None else ""
                
                if v_exp != v_act:
                    if sheet_mismatches < 3:
                        mismatches.append(f"    {sn}!R{r}C{c}: exp '{v_exp}', got '{v_act}'")
                    sheet_mismatches += 1
        
        if sheet_mismatches > 0:
            mismatches.append(f"  [FAIL] {sn}: {sheet_mismatches} mismatches")
            
    return mismatches

def run_test(fixture_name):
    print(f"\n>>> Running test: {fixture_name}")
    fixture_dir = Path(f"tests/legacy_converter/fixtures/{fixture_name}")
    input_dir = fixture_dir / "input"
    expected_dir = fixture_dir / "expected"
    output_dir = fixture_dir / "output"
    master_dir = Path("Templates Master")
    
    if output_dir.exists():
        shutil.rmtree(output_dir)
        
    conv = LegacyConverter(str(input_dir), str(output_dir), str(master_dir))
    try:
        conv.convert()
    except Exception as e:
        print(f"  [CRASH] {e}")
        return False

    all_ok = True
    # Compare $index.xlsx if expected exists
    expected_index = expected_dir / "$index.xlsx"
    if expected_index.exists():
        index_mismatches = compare_files(expected_index, output_dir / "$index.xlsx")
        if index_mismatches:
            print(f"  $index.xlsx:")
            for m in index_mismatches:
                print(f"    {m}")
            all_ok = False
        else:
            print(f"  $index.xlsx: [OK]")
    else:
        print(f"  $index.xlsx: [SKIPPED] No expected file")
        
    # Compare endpoints
    for exp_file in expected_dir.glob("*.xlsx"):
        if exp_file.name == "$index.xlsx": continue
        ep_mismatches = compare_files(exp_file, output_dir / exp_file.name)
        if ep_mismatches:
            print(f"  {exp_file.name}:")
            for m in ep_mismatches:
                print(f"    {m}")
            all_ok = False
        else:
            print(f"  {exp_file.name}: [OK]")
            
    return all_ok

fixtures = [
    "R1_basic_structure",
    "R4_array",
    "R5_casing",
    "R8_collision"
]

results = {}
for fix in fixtures:
    results[fix] = run_test(fix)

print("\n" + "="*30)
print("FINAL RESULTS")
print("="*30)
for fix, ok in results.items():
    status = "PASS" if ok else "FAIL"
    print(f"{fix:20}: {status}")
print("="*30)
