from openpyxl import load_workbook
import os

master_path = r"Templates Master\$index.xlsx"
if os.path.exists(master_path):
    wb = load_workbook(master_path, data_only=True)
    print("Sheets in Master:", wb.sheetnames)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n--- {sheet_name} ---")
        for r in range(1, 4):
            row_vals = [ws.cell(row=r, column=c).value for c in range(1, 15)]
            print(f"Row {r}: {row_vals}")
else:
    print("Master not found")
