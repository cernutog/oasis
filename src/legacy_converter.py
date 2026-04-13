"""
Legacy Excel to Modern OAS Converter - Clean Implementation

Converts legacy Excel API specifications to modern OAS format.
Architecture: Simple linear flow with direct sheet transformations.
"""
import os
import shutil
import re
import copy
from pathlib import Path
from collections import OrderedDict
from dataclasses import dataclass
from typing import Dict, List, Tuple, Optional, Any
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
import yaml


@dataclass
class DataType:
    """Data type definition from legacy 'Data Type' sheet."""
    name: str
    type: str
    format: str = ""
    min_val: str = ""
    max_val: str = ""
    description: str = ""
    pattern_eba: str = ""
    regex: str = ""
    allowed_values: str = ""
    example: str = ""
    items_type: str = ""
    source_file: str = "" # Track origin for collision resolution


class LegacyConverter:
    """Converts legacy Excel templates to modern OAS format."""
    
    def __init__(
        self,
        input_dir: str,
        output_dir: str,
        master_dir: str = None,
        log_callback=None,
        detail_log_callback=None,
        include_descriptions_in_collision: bool = False,
        include_examples_in_collision: bool = False,
        capitalize_schema_names: bool = True,
        contact_name: str = "",
        contact_url: str = "",
        release: str = "",
        filename_pattern: str = "",
    ):
        self.input_dir = Path(input_dir)
        self.output_dir = Path(output_dir)
        self.master_dir = Path(master_dir) if master_dir else None
        self.log = log_callback or print
        # detail_log receives verbose technical output (schema table, full trace).
        # Defaults to same as self.log so callers that don't care get current behaviour.
        self.detail_log = detail_log_callback or self.log
        self.tracing_enabled = True # Default

        self.include_descriptions_in_collision = bool(include_descriptions_in_collision)
        self.include_examples_in_collision = bool(include_examples_in_collision)
        self.capitalize_schema_names = bool(capitalize_schema_names)
        self.contact_name = str(contact_name or "").strip()
        self.contact_url = str(contact_url or "").strip()
        self.release = str(release or "").strip()
        self.filename_pattern = str(filename_pattern or "").strip()
        
        # Internal registry
        self.global_schemas: Dict[str, DataType] = {} # out_name -> DataType
        self.used_names = set()
        self.fingerprints = {} # fingerprint -> out_name
        self.output_names = {} # (filename, original_name) -> out_name
        
        # Double-pass state
        self.raw_data_types: Dict[str, Dict[str, DataType]] = {} # file_key -> {norm_name -> DataType}
        self.ordered_filenames: List[str] = [] # list of filenames in index order
        self.schema_usage: Dict[str, List[str]] = {} # out_name -> ["opId.Context.prop[.sub...]", ...]
        
        # Operation IDs mapping
        self.filename_to_opid = {}
        self.filename_to_path: Dict[str, str] = {}  # filename_stem -> API path
        self.emitted_wrappers = set()
        self.emitted_inline_components = set()
        self.inline_component_fingerprints: Dict[Tuple, str] = {}
        self.inline_component_neutral_fingerprints: Dict[Tuple, str] = {}
        self.inline_component_fingerprint_by_name: Dict[str, Tuple] = {}
        self.inline_component_root_desc: Dict[str, str] = {}
        self.inline_component_example_sig: Dict[str, str] = {}
        self.merge_provenance: Dict[str, Dict[str, Dict[str, set]]] = {}  # out_name -> {field -> {value_key -> set(file_keys)}}
        self.inline_component_children: Dict[str, set] = {}  # parent_schema -> {child_schema, ...}
        self.current_ep_name = None
        self.all_tags = set() # Track all tags for the index
        self._current_children_map: Dict[str, frozenset] = {}  # norm_dtype -> frozenset(child_prop_names)
        self.empty_error_responses: Dict[str, str] = {}  # status_code -> description (for empty 4xx/5xx sheets)
        self.error_response_fingerprints: Dict[Tuple, str] = {}  # fingerprint -> ErrorResponse variant name
        self.filename_to_error_response: Dict[str, str] = {}  # ep_filename -> ErrorResponse variant name

    def _resolve_internal_master_dir(self):
        """Finds 'Templates Master' folder as an internal resource."""
        import sys
        # 1. If frozen (PyInstaller EXE)
        if getattr(sys, 'frozen', False):
            base = Path(sys._MEIPASS)
            cand = base / "Templates Master"
            if cand.exists():
                return cand

        # 2. Development mode - Relative to script
        # src/legacy_converter.py -> project_root/Templates Master
        cand = Path(__file__).parent.parent / "Templates Master"
        if cand.exists():
            return cand

        # 3. Fallback to CWD
        cand = Path.cwd() / "Templates Master"
        if cand.exists():
            return cand
            
        return None

    def _clone_schema_variant(self, source_dt: Optional[DataType], target_name: str) -> None:
        """Copy root DataType attributes onto a split variant schema name."""
        if not source_dt or not target_name:
            return
        cloned = copy.deepcopy(source_dt)
        cloned.name = target_name
        self.global_schemas[target_name] = cloned

    def _usage_context_kind(self, usage_ctx: Optional[str]) -> str:
        text = str(usage_ctx or "")
        return "request" if "(Body)" in text else "response"

    def _endpoint_display_name(self, filename_or_stem: str) -> str:
        s = str(filename_or_stem or "").strip()
        if not s:
            return ""
        stem = Path(s).stem
        m = re.match(r"^(.*)\.(\d{6})$", stem)
        return m.group(1) if m else stem

    def convert(self, tracing_enabled=True):
        """Main conversion entry point."""
        if self.master_dir is None:
            self.master_dir = self._resolve_internal_master_dir()
        
        if self.master_dir is None:
            self.log("CRITICAL ERROR: Internal 'Templates Master' folder not found. Please contact support.")
            return False
            
        self.tracing_enabled = tracing_enabled
        if not self.output_dir.exists():
            self.output_dir.mkdir(parents=True)
            
        # 1. Collection Phase
        index_path = self.input_dir / "$index.xlsm"
        if not index_path.exists():
            index_path = self.input_dir / "$index.xlsx"
            
        if index_path.exists():
            self._pre_read_index(index_path)

        # Validate: at least one endpoint file must exist before proceeding.
        ep_candidates = [
            f for f in list(self.input_dir.glob("*.xlsm")) + list(self.input_dir.glob("*.xlsx"))
            if not f.name.startswith(("$", "~"))
        ]
        if not ep_candidates:
            self.log(
                f"ERROR: No endpoint files (*.xlsm / *.xlsx) found in '{self.input_dir}'. "
                "Conversion aborted."
            )
            return False

        self.log("Collecting data types from all endpoints...")
        self._collect_all_data_types()
        
        self.log("Performing naming and usage analysis pass...")
        self._perform_naming_and_usage_pass()
        self.log(f"  Loaded {len(self.global_schemas)} unique global schemas.")

        # 2. Index Phase
        if index_path.exists():
            self._convert_index(index_path)

        # 3. Endpoint Phase
        # Follow the same order as the naming pass for consistency
        for ep_filename in self.ordered_filenames:
            ep_file = self.input_dir / ep_filename
            if ep_file.exists():
                self._convert_endpoint(ep_file)
                
        # Also catch any files not in the index (edge case)
        for ep_file in self.input_dir.glob("*.xlsm"):
            if ep_file.name.startswith(("$", "~")): continue
            if ep_file.name not in self.ordered_filenames:
                self._convert_endpoint(ep_file)
        for ep_file in self.input_dir.glob("*.xlsx"):
            if ep_file.name.startswith(("$", "~")): continue
            if ep_file.name not in self.ordered_filenames:
                self._convert_endpoint(ep_file)

        # 4. Post-processing: write Responses sheet in $index.xlsx
        self._write_responses_sheet()

        # 11. Final Summary
        if self.tracing_enabled:
            self.run_standalone_check(str(self.output_dir))
        self.log(f"Conversion complete. Output: {self.output_dir.as_posix()}")
        return True


    def _collect_all_data_types(self):
        """Collect all data types from all endpoint 'Data Type' sheets."""
        # First from $index (if any global types defined there)
        idx_path = self.input_dir / "$index.xlsm"
        if not idx_path.exists():
            idx_path = self.input_dir / "$index.xlsx"
            
        if idx_path.exists():
            self._collect_data_types_from_file(idx_path, is_global=True)
            
        # Then from all endpoints (unordered first to collect all possible definitions)
        ep_files = list(self.input_dir.glob("*.xlsm")) + list(self.input_dir.glob("*.xlsx"))
        for ep_file in ep_files:
            if not ep_file.name.startswith(("$", "~")):
                self._collect_data_types_from_file(ep_file)

    def _collect_data_types_from_file(self, file_path: Path, is_global: bool = False):
        """Collect data types from a single file's 'Data Type' sheet."""
        xl = None
        try:
            xl = pd.ExcelFile(file_path)
            if "Data Type" not in xl.sheet_names:
                return

            df = pd.read_excel(xl, sheet_name="Data Type", dtype=str, header=None)
            header_keywords = ["name", "type", "description"]
            header_row_idx = self._find_header_row(df, header_keywords)

            if header_row_idx == -1:
                return

            df.columns = [str(c).strip().lower() for c in df.iloc[header_row_idx]]
            df = df.iloc[header_row_idx + 1:].reset_index(drop=True)
            file_key = "$global" if is_global else file_path.name

            def gv(keywords, default=""):
                for col in df.columns:
                    if any(kw in str(col).lower() for kw in keywords):
                        return col
                return None

            name_c = gv(["name"])
            desc_c = gv(["description"])
            type_c = gv(["type"])
            fmt_c = gv(["format"])
            items_c = gv(["items data type", "array only"])
            min_c = gv(["min"])
            max_c = gv(["max"])
            reg_c = gv(["regex"])
            pat_c = gv(["pattern", "eba"])
            all_c = gv(["allowed value", "allowed"])
            ex_c = gv(["example"])

            for _, row in df.iterrows():
                name = self._clean_value(row.get(name_c, ""))
                if not name or str(name).lower() == "nan":
                    continue

                norm_name = self._to_pascal_case(name)
                dt = DataType(
                    name=norm_name,
                    type=self._clean_value(row.get(type_c, "string")),
                    format=self._clean_value(row.get(fmt_c, "")),
                    min_val=self._clean_value(row.get(min_c, "")),
                    max_val=self._clean_value(row.get(max_c, "")),
                    description=self._clean_value(row.get(desc_c, "")),
                    pattern_eba=self._clean_value(row.get(pat_c, "")),
                    regex=self._clean_value(row.get(reg_c, "")),
                    allowed_values=self._clean_value(row.get(all_c, "")),
                    example=self._clean_value(row.get(ex_c, "")),
                    items_type=self._clean_value(row.get(items_c, "")),
                    source_file=file_key
                )

                # Defer registration naming
                if file_key not in self.raw_data_types:
                    self.raw_data_types[file_key] = {}
                self.raw_data_types[file_key][norm_name] = dt
        finally:
            self._close_excel_file(xl)

    def _register_data_type(self, file_key: str, norm_name: str, dt: DataType):
        """Registers a data type, deduplicating by content and handling name collisions."""
        # Fingerprint for deduplication (exclude source_file and pattern_eba).
        # 'name' is INCLUDED so that differently-named DataTypes (e.g. BIC8 vs
        # SenderBIC) are kept separate even when their structural content is identical.
        # Optionally include description/examples depending on preferences.
        excluded = {'source_file', 'pattern_eba'}
        if not self.include_descriptions_in_collision:
            excluded.add('description')
        if not self.include_examples_in_collision:
            excluded.add('example')
        fp_data = {k: v for k, v in dt.__dict__.items() if k not in excluded}
        
        # Normalize lists (ignore separators and extra spaces)
        def normalize_list(v):
            if not isinstance(v, str): return v
            # Replace ; with , and normalize spaces around commas
            v = re.sub(r'\s*;\s*', ',', v)
            v = re.sub(r'\s*,\s*', ',', v)
            return v.strip().strip(',')

        if 'allowed_values' in fp_data:
            fp_data['allowed_values'] = normalize_list(fp_data['allowed_values'])
        if 'example' in fp_data:
            fp_data['example'] = normalize_list(fp_data['example'])
        # Normalize trailing periods in description (e.g. 'BIC searching.' → 'BIC searching')
        if 'description' in fp_data and isinstance(fp_data['description'], str):
            fp_data['description'] = fp_data['description'].rstrip('.')
            
        # Include child property names for object types so that
        # two DataTypes with the same base fields but different children
        # (e.g. Criteria in different endpoints) get distinct fingerprints.
        if dt.type and dt.type.lower() == 'object' and norm_name in self._current_children_map:
            fp_data['_children'] = self._current_children_map[norm_name]
        fingerprint = tuple(sorted(fp_data.items()))
        
        mapping_key = (file_key, norm_name)
        
        # 1. Content-based deduplication
        if fingerprint in self.fingerprints:
            out_name = self.fingerprints[fingerprint]
            self.output_names[mapping_key] = out_name

            # Record merge provenance for neutral fields (description/examples) when they differ.
            try:
                canon = self.global_schemas.get(out_name)
                if canon:
                    if (not self.include_examples_in_collision
                        and not (canon.example or "").strip()
                        and self._is_valid_example_set(dt, dt.example)):
                        # Promote the first valid example set encountered into the canonical
                        # registry entry when examples are excluded from the collision
                        # fingerprint. This keeps the consolidated $index and all downstream
                        # example generation aligned without changing collision grouping.
                        canon.example = (dt.example or "").strip()
                    if not self.include_descriptions_in_collision:
                        desc_in = (dt.description or "").strip()
                        desc_can = (canon.description or "").strip()
                        if desc_in and desc_in != desc_can:
                            self._record_merge_provenance(out_name, "description", desc_in, file_key)
                    if not self.include_examples_in_collision:
                        ex_in = (dt.example or "").strip()
                        ex_can = (canon.example or "").strip()
                        if ex_in and ex_in != ex_can:
                            self._record_merge_provenance(out_name, "example", ex_in, file_key)
            except Exception:
                pass
            return out_name
            
        # 2. Name collision handling for new unique content
        output_name = norm_name
        counter = 1
        while output_name in self.used_names:
            output_name = f"{norm_name}{counter}"
            counter += 1
            
        # 3. Commit to registry
        self.global_schemas[output_name] = dt
        self.used_names.add(output_name)
        self.fingerprints[fingerprint] = output_name
        self.output_names[mapping_key] = output_name
        return output_name

    def _is_valid_example_set(self, dt: Optional[DataType], example_text: Any) -> bool:
        """Return True when every ';'-separated example token is valid for the DataType."""
        if dt is None:
            return False
        raw = str(example_text or "").strip()
        if not raw or raw.lower() == "nan":
            return False
        tokens = [tok.strip() for tok in raw.split(";") if tok and str(tok).strip()]
        if not tokens:
            return False
        return all(self._is_valid_example_token(dt, token) for token in tokens)

    def _is_valid_example_token(self, dt: DataType, token: str) -> bool:
        """Validate a single example token against the DataType constraints."""
        val = str(token or "").strip()
        if not val:
            return False

        allowed_raw = str(dt.allowed_values or "").strip()
        if allowed_raw and allowed_raw.lower() != "nan":
            allowed = [p.strip() for p in re.split(r"[;,]", allowed_raw) if p.strip()]
            if allowed and val not in allowed:
                return False

        regex_raw = str(dt.regex or "").strip()
        if regex_raw and regex_raw.lower() != "nan":
            try:
                if re.fullmatch(regex_raw, val) is None:
                    return False
            except re.error:
                # Ignore malformed regex values here rather than rejecting otherwise
                # useful examples.
                pass

        type_lower = str(dt.type or "string").strip().lower()
        min_raw = str(dt.min_val or "").strip()
        max_raw = str(dt.max_val or "").strip()

        if type_lower in ("integer", "int"):
            try:
                intval = int(val.replace(",", "."))
            except Exception:
                return False
            if min_raw and min_raw.lower() != "nan":
                try:
                    if intval < int(float(min_raw.replace(",", "."))):
                        return False
                except Exception:
                    pass
            if max_raw and max_raw.lower() != "nan":
                try:
                    if intval > int(float(max_raw.replace(",", "."))):
                        return False
                except Exception:
                    pass
            return True

        if type_lower in ("number", "float", "double", "decimal"):
            try:
                numval = float(val.replace(",", "."))
            except Exception:
                return False
            if min_raw and min_raw.lower() != "nan":
                try:
                    if numval < float(min_raw.replace(",", ".")):
                        return False
                except Exception:
                    pass
            if max_raw and max_raw.lower() != "nan":
                try:
                    if numval > float(max_raw.replace(",", ".")):
                        return False
                except Exception:
                    pass
            return True

        if type_lower == "boolean":
            return val.lower() in ("true", "false", "yes", "no", "1", "0")

        # Default string validation: length constraints.
        if min_raw and min_raw.lower() != "nan":
            try:
                if len(val) < int(float(min_raw.replace(",", "."))):
                    return False
            except Exception:
                pass
        if max_raw and max_raw.lower() != "nan":
            try:
                if len(val) > int(float(max_raw.replace(",", "."))):
                    return False
            except Exception:
                pass
        return True

    def _record_merge_provenance(self, out_name: str, field: str, value: str, file_key: str) -> None:
        if not out_name or not field or value is None:
            return
        if out_name not in self.merge_provenance:
            self.merge_provenance[out_name] = {}
        if field not in self.merge_provenance[out_name]:
            self.merge_provenance[out_name][field] = {}
        key = str(value)
        if key not in self.merge_provenance[out_name][field]:
            self.merge_provenance[out_name][field][key] = set()
        if file_key:
            self.merge_provenance[out_name][field][key].add(str(file_key))

    def _resolve_data_type(self, name: str, ep_filename: Optional[str] = None) -> Tuple[Optional[str], Optional[DataType]]:
        """Resolves a named data type to its unique global output name and DataType object."""
        norm_name = self._to_pascal_case(name)
        
        # 1. Try currently registered types (output_names already has the final name)
        # Check specific file context first
        if ep_filename:
            mapping_key = (ep_filename, norm_name)
            if mapping_key in self.output_names:
                out_name = self.output_names[mapping_key]
                return out_name, self.global_schemas.get(out_name)
        
        # Check global context
        mapping_key_global = ("$global", norm_name)
        if mapping_key_global in self.output_names:
            out_name = self.output_names[mapping_key_global]
            return out_name, self.global_schemas.get(out_name)

        # 1.5. Fallback: the type is registered under another file's context but not the
        # current file's or the global context (e.g. a file that references a DataType
        # defined only in peer endpoint files). Return the base registered name directly
        # so the property is not left unresolved (which would emit a raw lowercase dtype).
        if norm_name in self.global_schemas:
            return norm_name, self.global_schemas[norm_name]

        # 2. Not registered yet. This happens during the Naming Pass.
        # Find the best raw definition available
        raw_dt = None
        # Prioritize the current file's raw definition
        if ep_filename and ep_filename in self.raw_data_types and norm_name in self.raw_data_types[ep_filename]:
            raw_dt = self.raw_data_types[ep_filename][norm_name]
            file_key_for_registration = ep_filename
        # Fallback to global raw definition
        elif "$global" in self.raw_data_types and norm_name in self.raw_data_types["$global"]:
            raw_dt = self.raw_data_types["$global"][norm_name]
            file_key_for_registration = "$global"
            
        if raw_dt:
            # During naming pass, we REGISTER it now
            out_name = self._register_data_type(file_key_for_registration, norm_name, raw_dt)
            return out_name, raw_dt
            
        return None, None

    def _perform_naming_and_usage_pass(self):
        """Pass 2: Determine naming priority based on index order and track usage."""
        # 1. Register all Global types from $index (highest priority)
        global_types = self.raw_data_types.get("$global", {})
        for norm_name, dt in global_types.items():
            self._register_data_type("$global", norm_name, dt)
            
        # 2. Register per-endpoint DataTypes in Paths-sheet order so that naming
        # priority (base name vs suffixed variant) matches the schema-building pass,
        # which also follows ordered_filenames. Files not listed in the index are
        # appended in alphabetical order as a deterministic fallback.
        indexed = [f for f in self.ordered_filenames if f in self.raw_data_types]
        extras = sorted(k for k in self.raw_data_types if k != "$global" and k not in set(indexed))
        for file_key in indexed + extras:
            for norm_name, dt in self.raw_data_types[file_key].items():
                self._register_data_type(file_key, norm_name, dt)

    def _track_recursive_usage(self, child_tuple, ep_filename, op_id, sheet):
        """Deprecated: retained for backward compatibility; usage tracking is now done with full property paths."""
        return

    def _log_usage_summary(self):
        """Log the alphabetical list of schemas and where they are used with absolute property values."""
        
        col1_w, col2_w, col3_w = 45, 95, 80
        header = f"| {'SCHEMA NAME':<{col1_w}} | {'USED IN':<{col2_w}} | {'DIFFERENCES':<{col3_w}} |"
        sep = "-" * len(header)
        
        self.log("\n" + "="*len(header))
        self.log(header)
        self.log(sep)
        
        sorted_names = sorted(self.schema_usage.keys(), key=lambda x: x.lower())
        
        # Group schema names by base stem for collision/split analysis.
        # We handle two distinct cases:
        # 1) Split variants where the base itself ends with digits (e.g. Bic11 -> Bic111)
        # 2) Classic collision naming where the base name is absent (e.g. Foo1/Foo2/...)
        def _find_split_base(name: str, all_names: set) -> Optional[Tuple[str, int]]:
            """If `name` is a split/collision variant of an original schema, return (base_name, suffix_int).

            Rule:
            - a digit-ending schema is a valid base when its stem without trailing digits
              does not exist (e.g. Bic8 is a base because Bic does not exist)
            - a longer digit-ending schema is a split/variant only when the candidate base
              exists among known schema names (e.g. Bic81 -> Bic8, Bic111 -> Bic11)
            """
            s = str(name)
            m = re.search(r'(\d+)$', s)
            if not m:
                return None
            digits = m.group(1)
            if len(digits) < 2:
                return None

            # Try trimming 1..len(digits)-1 digits so the remaining prefix exists.
            for k in range(1, len(digits)):
                base = s[:-k]
                suffix = s[-k:]
                if base in all_names:
                    try:
                        return base, int(suffix)
                    except Exception:
                        return None
            return None

        all_names_set = set(sorted_names)
        base_to_variants: Dict[str, List[Tuple[int, str]]] = {}
        variant_to_base: Dict[str, str] = {}

        # 1) Explicit split detection: base may already end with digits (Bic11 -> Bic111)
        for name in sorted_names:
            found = _find_split_base(name, all_names_set)
            if not found:
                continue
            base, suffix_int = found
            base_to_variants.setdefault(base, []).append((suffix_int, name))
            variant_to_base[name] = base

        # Originals are those not classified as a split variant
        base_names: set = set([n for n in sorted_names if n not in variant_to_base])

        # 2) Classic collision variants when base does NOT exist: Foo1/Foo2/... (stem lacks base)
        stem_to_variants: Dict[str, List[Tuple[int, str]]] = {}
        for name in sorted_names:
            if name in variant_to_base:
                continue
            m = re.search(r'^(.*?)(\d+)$', str(name))
            if not m:
                continue
            stem = m.group(1)
            try:
                num = int(m.group(2))
            except Exception:
                continue
            stem_to_variants.setdefault(stem, []).append((num, name))

        final_groups: Dict[str, List[str]] = {}

        # 1. Base names: always show base, plus any numeric variants (even with gaps)
        for base in sorted(base_names, key=lambda x: str(x).lower()):
            variants = []
            variants.extend(base_to_variants.get(base, []))
            variants.extend(stem_to_variants.get(base, []))
            variants = sorted(variants, key=lambda t: t[0])
            members = [base] + [nm for _n, nm in variants if nm != base]
            final_groups[base] = members

        # 2. Stems without base: keep as their own group (variants only)
        # Only group when the smallest numeric suffix is 1: collision naming always starts
        # from 1 (Foo → Foo1 → Foo2 …).  Schemas like Bic8/Bic11/Bic81 have semantic names
        # (not collision-generated) and must NOT be grouped as spurious collision variants.
        for stem, variants in sorted(stem_to_variants.items(), key=lambda kv: str(kv[0]).lower()):
            if stem in base_names:
                continue
            variants = sorted(variants, key=lambda t: t[0])
            if not variants or variants[0][0] != 1:
                continue
            members = [nm for _n, nm in variants]
            if members:
                final_groups[members[0]] = members
            
        # Find which attributes actually differ in each group
        # For DataType-based schemas: compare DataType constraint fields.
        # For promoted inline component schemas: compute property-level diffs.
        mismatches_by_group: Dict[str, List[str]] = {}
        # attrs_display: fields shown in DIFFERENCES column (values only, no arrows)
        attrs = ['type', 'format', 'min_val', 'max_val', 'regex', 'allowed_values', 'items_type']
        # attrs_diff: fields compared when detecting group differences.
        # Keep this aligned with the collision preferences so the tracer explains
        # only the attributes that can actually drive a split.
        attrs_diff = list(attrs)
        if self.include_descriptions_in_collision:
            attrs_diff.append('description')
        if self.include_examples_in_collision:
            attrs_diff.append('example')

        def _short_usage(u: str) -> str:
            if not u:
                return ""
            raw = str(u).strip()

            # Preferred format (stored already clean): 'endpoint (ctx)'
            # Endpoint filenames may contain dots (e.g. Foo.250410), but the formatted
            # string is still authoritative if it contains parentheses.
            if raw.endswith(")") and "(" in raw:
                # Safety-net: if ctx contains extra info like '250410 (Body)' or '260220 (200)',
                # collapse to the inner token.
                m_outer = re.match(r"^(.*?)\((.*)\)\s*$", raw)
                if not m_outer:
                    return raw
                left = m_outer.group(1).strip()
                ctx = m_outer.group(2).strip()
                m_ctx = re.match(r"^\s*\d+\s*\(([^()]*)\)\s*$", ctx)
                if m_ctx:
                    ctx = m_ctx.group(1).strip()
                return f"{left} ({ctx})" if left else f"({ctx})"

            # Legacy dotted format: opId.Context.prop[.sub...]
            parts = raw.split(".")
            if len(parts) < 2:
                return raw
            op_id = parts[0]
            ctx = parts[1]
            ctx_map = {
                "PathParam": "Path",
                "Header": "Header",
            }
            disp_ctx = ctx_map.get(ctx, ctx)
            return f"{op_id} ({disp_ctx})"

        def _inline_fp(name: str):
            return self.inline_component_fingerprint_by_name.get(name)

        def _normalize_constraint_value(attr: str, val: str) -> str:
            s = "" if val is None else str(val)
            s = s.strip()
            if s.lower() == "nan":
                s = ""

            # Normalize numbers (Excel/pandas may coerce ints/floats/strings inconsistently)
            if attr in ("min_val", "max_val"):
                if not s:
                    return ""
                s2 = s.replace(",", ".")
                try:
                    f = float(s2)
                    if f.is_integer():
                        return str(int(f))
                    return str(f)
                except Exception:
                    return s

            # Normalize regex/pattern: ignore whitespace/newlines and undo common Excel escaping
            if attr == "regex":
                if not s:
                    return ""
                # Collapse all whitespace
                s = "".join(s.split())
                # Common double escaping from Excel exports
                s = s.replace("\\\\", "\\")
                return s

            return s

        def _dt_constraints(dt: Optional[DataType]) -> Dict[str, str]:
            if not dt:
                return {}
            out = {}
            for a in attrs_diff:
                v = getattr(dt, a, "")
                out[a] = _normalize_constraint_value(a, v)
            return out

        def _constraint_delta(old_dt: Optional[DataType], new_dt: Optional[DataType]) -> List[str]:
            """Return list of attr names that differ (used for grouping); values rendered per-schema separately."""
            o = _dt_constraints(old_dt)
            n = _dt_constraints(new_dt)
            differing = []
            for k in attrs_diff:
                if o.get(k, "") != n.get(k, ""):
                    differing.append(k)
            return differing

        def _schema_attr_display(dt: Optional[DataType], differing_attrs: List[str]) -> str:
            """Display only the differing attribute values for a given schema (no arrows)."""
            if not dt:
                return ""
            bits = []
            norm = _dt_constraints(dt)
            for k in differing_attrs:
                if k in attrs_diff:
                    label = k.replace('_', ' ').title()
                    disp_raw = norm.get(k, "")
                    disp = disp_raw if disp_raw else "(empty)"
                    bits.append(f"- {label}: {disp}")
            return "\n".join(bits)

        def _fp_to_propmap(fp: Tuple) -> Dict[str, Dict[str, str]]:
            """Convert inline component fingerprint entries to a property map keyed by fully-qualified path."""
            if not fp:
                return {}

            # fp entry structure (see _fp_inline_subtree):
            # (name, parent, desc_part, ex_part, dtype, items_type, mandatory, rules)
            children_by_parent: Dict[str, List[Tuple]] = {}
            node_by_key: Dict[Tuple[str, str], Tuple] = {}
            for e in fp:
                if not isinstance(e, tuple) or len(e) < 8:
                    continue
                nm, parent = e[0], e[1]
                node_by_key[(nm, parent)] = e
                children_by_parent.setdefault(parent or "", []).append(e)

            # Roots are those whose parent is "" (top-level under promoted root)
            out: Dict[str, Dict[str, str]] = {}

            def walk(entry: Tuple, prefix: str = ""):
                nm, parent, desc_part, ex_part, dtype, items_type, mandatory, rules = entry[:8]
                if not nm:
                    return
                path = f"{prefix}.{nm}" if prefix else nm
                out[path] = {
                    "desc": desc_part or "",
                    "dtype": dtype or "",
                    "items": items_type or "",
                    "mandatory": mandatory or "",
                    "rules": rules or "",
                }
                for ch in children_by_parent.get(nm, []):
                    walk(ch, path)

            for r in children_by_parent.get("", []):
                walk(r, "")
            return out

        def _promoted_diffs(base_name: str, other_name: str) -> Dict:
            """Return structured diff dict: {added, removed, changed, mandatory, rules}.
            Each entry carries base_val / other_val so the renderer can display per-schema."""
            base_fp = _inline_fp(base_name)
            other_fp = _inline_fp(other_name)
            if not base_fp or not other_fp:
                return {}
            bmap = _fp_to_propmap(base_fp)
            omap = _fp_to_propmap(other_fp)

            bkeys = set(bmap.keys())
            okeys = set(omap.keys())
            added   = sorted(okeys - bkeys)
            removed = sorted(bkeys - okeys)
            common  = sorted(bkeys & okeys)

            result: Dict = {}
            if added:
                result["added"] = added
            if removed:
                result["removed"] = removed

            changed: Dict[str, Dict] = {}
            mandatory: Dict[str, Dict] = {}
            rules: Dict[str, Dict] = {}
            descriptions: Dict[str, Dict] = {}
            for k in common:
                b = bmap[k]
                o = omap[k]
                b_type  = b.get("dtype", "")
                o_type  = o.get("dtype", "")
                b_items = b.get("items", "")
                o_items = o.get("items", "")
                if (b_type, b_items) != (o_type, o_items):
                    b_dt = self.global_schemas.get(b_type) if b_type in self.global_schemas else None
                    o_dt = self.global_schemas.get(o_type) if o_type in self.global_schemas else None
                    delta_attrs = _constraint_delta(b_dt, o_dt) if (b_dt or o_dt) else []
                    changed[k] = {
                        "base_type":  f"{b_type or '(empty)'}{('[]' if b_items else '')}",
                        "other_type": f"{o_type or '(empty)'}{('[]' if o_items else '')}",
                        "constraint_delta": delta_attrs,
                    }
                b_mand = (b.get("mandatory", "") or "").strip()
                o_mand = (o.get("mandatory", "") or "").strip()
                if b_mand != o_mand:
                    mandatory[k] = {"base_val": b_mand or "(empty)", "other_val": o_mand or "(empty)"}
                b_rules = (b.get("rules", "") or "").strip()
                o_rules = (o.get("rules", "") or "").strip()
                if b_rules != o_rules:
                    rules[k] = {"base_val": b_rules or "(empty)", "other_val": o_rules or "(empty)"}
                if self.include_descriptions_in_collision:
                    b_desc = (b.get("desc", "") or "").strip()
                    o_desc = (o.get("desc", "") or "").strip()
                    if b_desc != o_desc:
                        descriptions[k] = {"base_val": b_desc or "(empty)", "other_val": o_desc or "(empty)"}

            if changed:      result["changed"]      = changed
            if mandatory:    result["mandatory"]    = mandatory
            if rules:        result["rules"]        = rules
            if descriptions: result["descriptions"] = descriptions
            return result

        for root_name, members in final_groups.items():
            if len(members) <= 1:
                continue

            diff_fields = set()
            dts = [self.global_schemas.get(m) for m in members if self.global_schemas.get(m)]
            fps = [(_inline_fp(m), m) for m in members if _inline_fp(m) is not None]

            if dts:
                base = dts[0]
                base_norm = _dt_constraints(base)
                for other in dts[1:]:
                    other_norm = _dt_constraints(other)
                    for attr in attrs_diff:
                        if base_norm.get(attr, "") != other_norm.get(attr, ""):
                            diff_fields.add(attr)

            # Inline fp diffs: run always (not only when dts is absent) to catch
            # mixed groups where base is a DataType but variants are inline-only.
            if fps:
                # Determine the reference name: prefer base without numeric suffix if present.
                base_nm = members[0]
                stem = re.sub(r'(\d+)$', '', base_nm)
                if stem and stem in base_names:
                    base_nm = stem

                # Only variants show diffs vs the base; base itself stays empty.
                for other_nm in members:
                    if other_nm == base_nm:
                        continue
                    # If other_nm has no inline fp but does have a DataType, flag it as
                    # "structure differs" (DataType vs inline promoted).
                    other_fp = _inline_fp(other_nm)
                    base_fp  = _inline_fp(base_nm)
                    if other_fp and not base_fp:
                        if other_nm not in mismatches_by_group:
                            mismatches_by_group[other_nm] = ("inline_fp_diff", base_nm, {"structure": "promoted inline (base is DataType)"})
                    elif base_fp and not other_fp:
                        if other_nm not in mismatches_by_group:
                            mismatches_by_group[other_nm] = ("inline_fp_diff", base_nm, {"structure": "DataType (base is promoted inline)"})
                    elif base_fp and other_fp:
                        promoted_dict = _promoted_diffs(base_nm, other_nm)
                        if promoted_dict:
                            # Store as ("inline_fp_diff", base_nm, promoted_dict) so renderer can access per-schema values
                            mismatches_by_group[other_nm] = ("inline_fp_diff", base_nm, promoted_dict)

            # Flag missing base schema when root is a numbered variant and stem base is absent.
            if root_name and re.search(r'(\d+)$', root_name):
                stem = re.sub(r'(\d+)$', '', root_name)
                if stem and stem not in base_names:
                    diff_fields.add("missing base name")

            # Flag numbering gaps inside a group (e.g. Base + Base2 without Base1)
            if root_name in final_groups and len(final_groups[root_name]) > 1:
                members = final_groups[root_name]
                stem = root_name if not re.search(r'(\d+)$', root_name) else re.sub(r'(\d+)$', '', root_name)
                nums = []
                for nm in members:
                    m = re.search(r'^' + re.escape(str(stem)) + r'(\d+)$', str(nm))
                    if m:
                        nums.append(int(m.group(1)))
                if nums:
                    nums = sorted(set(nums))
                    expected = list(range(1, max(nums) + 1))
                    if nums != expected:
                        diff_fields.add("non-contiguous numbering")

            if diff_fields:
                mismatches_by_group[root_name] = sorted(list(diff_fields))

        def wrap_text(text, width):
            """Robust text wrapping that breaks long words."""
            if not text: return []
            lines = []
            paragraphs = text.split('\n')
            for p in paragraphs:
                words = p.split()
                if not words:
                    lines.append("")
                    continue
                current_line = []
                for word in words:
                    # If word exceeds width, break it
                    if len(word) > width:
                        if current_line:
                            lines.append(" ".join(current_line))
                            current_line = []
                        while len(word) > width:
                            lines.append(word[:width])
                            word = word[width:]
                        current_line = [word]
                    elif sum(len(w) for w in current_line) + len(current_line) + len(word) <= width:
                        current_line.append(word)
                    else:
                        lines.append(" ".join(current_line))
                        current_line = [word]
                if current_line:
                    lines.append(" ".join(current_line))
            return lines

        def _display_schema_name(nm: str) -> str:
            s = "" if nm is None else str(nm)
            m = re.match(r"^(.*)\.(\d{6})(Request|Response)$", s)
            if m:
                return f"{m.group(1)}{m.group(3)}"
            return s

        def _merge_inline_diff_dicts(diff_dicts: List[Dict[str, Any]]) -> Dict[str, Any]:
            merged: Dict[str, Any] = {}
            for diff_dict in diff_dicts:
                if not diff_dict:
                    continue
                if "structure" in diff_dict:
                    merged["structure"] = diff_dict["structure"]
                for list_key in ["added", "removed"]:
                    if diff_dict.get(list_key):
                        merged.setdefault(list_key, [])
                        for item in diff_dict[list_key]:
                            if item not in merged[list_key]:
                                merged[list_key].append(item)
                for dict_key in ["changed", "mandatory", "rules", "descriptions"]:
                    if diff_dict.get(dict_key):
                        merged.setdefault(dict_key, {})
                        merged[dict_key].update(diff_dict[dict_key])
            return merged

        def _render_inline_diff(diff_dict: Dict[str, Any], is_base: bool) -> str:
            top_bits: List[str] = []
            per_prop: Dict[str, List[str]] = {}

            if "structure" in diff_dict:
                top_bits.append(diff_dict["structure"])

            if ("added" in diff_dict) and (not is_base):
                added_list = diff_dict["added"]
                top_bits.append("added:")
                for a in added_list[:10]:
                    top_bits.append(f"- {a}")
                if len(added_list) > 10:
                    top_bits.append(f"- (+{len(added_list) - 10} more)")

            if ("removed" in diff_dict) and (not is_base):
                removed_list = diff_dict["removed"]
                top_bits.append("removed:")
                for r in removed_list[:10]:
                    top_bits.append(f"- {r}")
                if len(removed_list) > 10:
                    top_bits.append(f"- (+{len(removed_list) - 10} more)")

            if "changed" in diff_dict:
                _oas_primitives = {
                    "string", "number", "integer", "boolean",
                    "array", "object", "schema", "(empty)", "",
                }
                for prop, info in list(diff_dict["changed"].items())[:10]:
                    val = info["base_type"] if is_base else info["other_type"]
                    delta = info.get("constraint_delta", [])
                    extra = f" ({', '.join(delta[:4])})" if delta else ""
                    is_array_suffix = val.endswith("[]")
                    base_val = val[:-2] if is_array_suffix else val
                    if base_val.lower() not in _oas_primitives:
                        display_val = base_val
                        for gname in self.global_schemas:
                            if gname.lower() == base_val.lower():
                                display_val = gname
                                break
                        if is_array_suffix:
                            display_val += "[]"
                        per_prop.setdefault(prop, []).append(f"- $ref: {display_val}{extra}")
                    else:
                        per_prop.setdefault(prop, []).append(f"- Type: {val}{extra}")

            if "mandatory" in diff_dict:
                for prop, info in list(diff_dict["mandatory"].items())[:20]:
                    val = info["base_val"] if is_base else info["other_val"]
                    per_prop.setdefault(prop, []).append(f"- Mandatory: {str(val).upper()}")

            if "rules" in diff_dict:
                for prop, info in list(diff_dict["rules"].items())[:20]:
                    raw = info["base_val"] if is_base else info["other_val"]
                    short = (raw[:120] + "...") if len(raw) > 120 else raw
                    per_prop.setdefault(prop, []).append(f"- Validation rules: {short}")

            if "descriptions" in diff_dict:
                for prop, info in list(diff_dict["descriptions"].items())[:20]:
                    raw = info["base_val"] if is_base else info["other_val"]
                    short = (raw[:120] + "...") if len(raw) > 120 else raw
                    per_prop.setdefault(prop, []).append(f"- Description: {short}")

            prop_blocks: List[str] = []
            for prop in sorted(per_prop.keys()):
                prop_blocks.append(f"{prop}:")
                prop_blocks.extend(per_prop[prop])
                prop_blocks.append("")

            return "\n".join(
                [x for x in (top_bits + ([""] if top_bits and prop_blocks else []) + prop_blocks) if x is not None]
            ).rstrip()

        for out_name in sorted_names:
            usages = self.schema_usage[out_name]
            
            # Figure out which root group this name belongs to for mismatch lookup
            # If it's the root itself, use it.
            # If it's a child, we need to know its root. We can inverse map final_groups.
            root_for_mismatch = out_name
            for r_key, members in final_groups.items():
                if out_name in members:
                    root_for_mismatch = r_key
                    break
                    
            relevant_fields = mismatches_by_group.get(out_name) or mismatches_by_group.get(root_for_mismatch, [])

            # If no direct entry found, check if this schema is the BASE of an inline_fp_diff group.
            # In that case, fabricate the tuple with is_base=True so the renderer shows this schema's values.
            group_dt_fields = []
            root_group_fields = mismatches_by_group.get(root_for_mismatch, [])
            if isinstance(root_group_fields, list):
                group_dt_fields = [f for f in root_group_fields if f in attrs_diff]

            direct_inline_entry = None
            if isinstance(mismatches_by_group.get(out_name), tuple):
                entry = mismatches_by_group.get(out_name)
                if len(entry) == 3 and entry[0] == "inline_fp_diff":
                    direct_inline_entry = entry

            base_inline_entries = []
            for _variant, _entry in mismatches_by_group.items():
                if (isinstance(_entry, tuple) and len(_entry) == 3 and
                        _entry[0] == "inline_fp_diff" and _entry[1] == out_name):
                    base_inline_entries.append(_entry)

            diff_parts = []
            dt = self.global_schemas.get(out_name)
            if dt and group_dt_fields:
                body = _schema_attr_display(dt, group_dt_fields)
                if body:
                    diff_parts.append(body)

            if direct_inline_entry:
                _, base_nm_ref, diff_dict = direct_inline_entry
                diff_body = _render_inline_diff(diff_dict, is_base=(out_name == base_nm_ref))
                if diff_body:
                    diff_parts.append(diff_body)
            elif base_inline_entries:
                merged_dict = _merge_inline_diff_dicts([entry[2] for entry in base_inline_entries])
                diff_body = _render_inline_diff(merged_dict, is_base=True)
                if diff_body:
                    diff_parts.append(diff_body)

            diff_str = "\n\n".join([part for part in diff_parts if part]).strip()
            if not diff_str and relevant_fields and not isinstance(relevant_fields, tuple) and not dt:
                diff_str = "; ".join([str(a) for a in relevant_fields])

            # WRAP ALL COLUMNS
            name_lines = wrap_text(_display_schema_name(out_name), col1_w)
            usage_lines = []

            short_usages = sorted(set([_short_usage(x) for x in usages if _short_usage(x)]))
            by_endpoint: Dict[str, List[str]] = {}
            passthrough: List[str] = []

            for u in short_usages:
                m_u = re.match(r"^(.*?)\s*\((.*)\)\s*$", str(u))
                if not m_u:
                    passthrough.append(u)
                    continue
                ep = m_u.group(1).strip()
                ctx = m_u.group(2).strip()
                if not ep or not ctx:
                    passthrough.append(u)
                    continue
                by_endpoint.setdefault(ep, []).append(ctx)

            def _ctx_sort_key(s: str):
                s2 = str(s).strip()
                if re.fullmatch(r"\d+", s2):
                    return (0, int(s2))
                return (1, s2.lower(), s2)

            agg_usages: List[str] = []
            for ep in sorted(by_endpoint.keys(), key=lambda x: (str(x).lower(), str(x))):
                ctxs = sorted(set(by_endpoint.get(ep, [])), key=_ctx_sort_key)
                agg_usages.append(f"{ep} " + " ".join([f"({c})" for c in ctxs]))

            for u in (agg_usages + sorted(set(passthrough))):
                usage_lines.extend(wrap_text(u, col2_w))
            if not usage_lines:
                usage_lines = ["(None)"]
            diff_lines = wrap_text(diff_str, col3_w)
            
            max_lines = max(len(name_lines), len(usage_lines), len(diff_lines), 1)
            for i in range(max_lines):
                n_cell = name_lines[i] if i < len(name_lines) else ""
                u_cell = usage_lines[i] if i < len(usage_lines) else ""
                d_cell = diff_lines[i] if i < len(diff_lines) else ""
                self.log(f"| {n_cell:<{col1_w}} | {u_cell:<{col2_w}} | {d_cell:<{col3_w}} |")
            
            self.log(sep)
            
        self.log("="*len(header) + "\n")

    def run_standalone_check(self, folder_path, inject_refs=False):
        """Analyze an existing converted folder by reading its $index file and all endpoints."""
        p = Path(folder_path)
        index_file = None
        for name in ["$index.xlsx", "$index.xlsm"]:
            if (p / name).exists():
                index_file = p / name
                break
        
        if not index_file:
            self.log(f"Error: Could not find $index file in {folder_path}")
            return False
            
        self.log(f"Analyzing project: {p.name}")
        self.log(f"Reading index: {index_file.name}")
        
        xl_idx = None
        try:
            xl_idx = pd.ExcelFile(index_file)
            
            # 1. Load Schemas for definition details AND build reference graph
            # schema_refs: name -> set of schema names directly referenced (via Schema Name or Items Data Type cols)
            schema_refs: Dict[str, set] = {}
            if "Schemas" in xl_idx.sheet_names:
                df_s = pd.read_excel(xl_idx, sheet_name="Schemas", dtype=str, header=None)
                header_idx = self._find_header_row(df_s, ["name", "type"])
                if header_idx != -1:
                    header_vals = [str(c).strip().lower() for c in df_s.iloc[header_idx]]
                    df_s.columns = header_vals
                    df_s = df_s.iloc[header_idx + 1:].reset_index(drop=True)

                    def get_col_val(r, exact_kws, partial_kws=None):
                        for c in df_s.columns:
                            if str(c) in exact_kws: return self._clean_value(r.get(c))
                        if partial_kws:
                            for c in df_s.columns:
                                if any(kw in str(c) for kw in partial_kws): return self._clean_value(r.get(c))
                        return ""

                    _primitive = {"string", "number", "integer", "boolean", "array", "object"}
                    _ref_cols = [c for c in header_vals
                                 if "schema name" in c or "items data type" in c]

                    # Track schemas by parent status:
                    # - has_parent_schema: at least one row with a parent (sub-fields)
                    # - has_no_parent_schema: at least one row WITHOUT a parent (root-level)
                    # - root_object_schema: root-level schemas whose type is object/schema
                    #   (these need to be inline components to be shown in tracing)
                    has_parent_schema: set = set()
                    has_no_parent_schema: set = set()
                    root_object_type_schema: set = set()

                    _object_types = {"object", "schema"}

                    # Collect property rows per schema to build inline fingerprints for
                    # wrapper schemas (e.g. ErrorResponse variants) so _log_usage_summary
                    # can detect and display structural differences in the DIFFERENCES column.
                    _schema_prop_rows: Dict[str, List[Tuple]] = {}

                    for _, row in df_s.iterrows():
                        name = self._clean_value(row.get('name'))
                        if not name: continue

                        raw_type = (get_col_val(row, ['type']) or 'string').lower()
                        dt = DataType(
                            name=name,
                            type=raw_type,
                            format=get_col_val(row, ['format']),
                            min_val=get_col_val(row, ['minimum'], ['min ']),
                            max_val=get_col_val(row, ['maximum'], ['max ']),
                            description=get_col_val(row, ['description'], ['desc']),
                            regex=get_col_val(row, ['regex', 'pattern']),
                            allowed_values=get_col_val(row, ['enum', 'allowed value']),
                            example=get_col_val(row, ['example'], ['examp']),
                            items_type=get_col_val(row, ['items type'], ['items data type'])
                        )
                        self.global_schemas[name] = dt

                        # Build reference graph:
                        # In $index Schemas sheet, rows with a non-empty 'parent' represent
                        # properties/children of that parent schema. The 'name' in those rows
                        # is a property name (NOT a schema), so we must attach any referenced
                        # schema to the owning schema (the parent).
                        parent = self._clean_value(row.get('parent'))
                        owner_schema = parent if (parent and parent not in _primitive) else name

                        if parent and parent not in _primitive:
                            # Mark that this parent schema has children/properties.
                            has_parent_schema.add(parent)
                            # Collect property row for fingerprinting.
                            # Use the schema reference (Schema Name col) as dtype identifier so
                            # that variants using RequestId vs RequestId1 produce distinct fps.
                            ref_val = ""
                            for rc in _ref_cols:
                                v = self._clean_value(row.get(rc))
                                if v and v not in _primitive:
                                    ref_val = v
                                    break
                            mand_val = (get_col_val(row, ['mandatory']) or "").strip().upper()
                            dtype_fp = (ref_val or raw_type).lower()
                            _schema_prop_rows.setdefault(parent, []).append(
                                (name.lower(), "", "", "", dtype_fp, "", mand_val, "")
                            )
                        else:
                            # Root-level schema definition.
                            has_no_parent_schema.add(name)
                            if raw_type in _object_types:
                                root_object_type_schema.add(name)

                        for rc in _ref_cols:
                            ref_val = self._clean_value(row.get(rc))
                            if ref_val and ref_val not in _primitive:
                                schema_refs.setdefault(owner_schema, set()).add(ref_val)

                    # Build inline fingerprints for root-level object/wrapper schemas so that
                    # _log_usage_summary can compare ErrorResponse, ErrorResponse1, … and show
                    # what differs in the DIFFERENCES column.
                    for _sn, _props in _schema_prop_rows.items():
                        if _props and _sn not in self.inline_component_fingerprint_by_name:
                            self.inline_component_fingerprint_by_name[_sn] = tuple(sorted(_props))


            # 2. Get list of endpoint files from Paths sheet
            # Also collect wrapper name stems (operationId) for later filtering.
            ep_files = []
            op_id_norms: set = set()  # lowercased operationId stems
            if "Paths" in xl_idx.sheet_names:
                df_p = pd.read_excel(xl_idx, sheet_name="Paths", dtype=str, header=None)
                h_idx = self._find_header_row(df_p, ["excel file", "operationid"])
                if h_idx != -1:
                    h_vals = [str(c).strip().lower() for c in df_p.iloc[h_idx]]
                    df_p.columns = h_vals
                    df_p = df_p.iloc[h_idx + 1:].reset_index(drop=True)
                    for _, row in df_p.iterrows():
                        fname = self._clean_value(row.get('excel file'))
                        opid = self._clean_value(row.get('operationid'))
                        api_path = self._clean_value(row.get('path') or row.get('paths') or '')
                        if fname:
                            ep_files.append((fname, opid))
                            if api_path:
                                self.filename_to_path[fname] = api_path
                        if opid:
                            op_id_norms.add(opid.lower())

            # 3. Map usages across all endpoint files
            # bfs_seed_wrappers: root wrapper schemas found in Schema Name col of body/response
            # sheets (e.g. CommandDetailsResponse). They are used as BFS propagation seeds
            # but must NOT appear in the final schema_usage output.
            bfs_seed_wrappers: set = set()
            for fname, op_id in ep_files:
                fpath = p / fname
                if not fpath.exists():
                    self.log(f"Warning: Endpoint file {fname} not found, skipping usage mapping.")
                    continue
                
                xl_ep = None
                try:
                    xl_ep = pd.ExcelFile(fpath)

                    # --- Wrapper usage seeding fallback ---
                    # Some endpoint artifacts do not reliably carry the wrapper name in each sheet.
                    # We seed wrapper usage based on the presence of Body/response sheets and map
                    # to the wrapper schema names consistent with the conversion pass.
                    ep_wrapper_base = self._to_wrapper_case(self._extract_wrapper_base(str(fname)))

                    # Seed Request wrapper if Body sheet exists
                    if "Body" in xl_ep.sheet_names:
                        w_req = f"{ep_wrapper_base}Request"
                        self.schema_usage.setdefault(w_req, [])
                        if f"{self._endpoint_display_name(str(fname))} (Body)" not in self.schema_usage[w_req]:
                            self.schema_usage[w_req].append(f"{self._endpoint_display_name(str(fname))} (Body)")
                        bfs_seed_wrappers.add(w_req)

                    # Seed success Response wrappers based on available response code sheets.
                    # Error responses (>=400) are NOT pre-seeded here because the actual
                    # ErrorResponse variant name (ErrorResponse, ErrorResponse1, …) can only be
                    # determined by reading the Schema Name column from the converted sheet (step 2
                    # below). Pre-seeding with the hardcoded base "ErrorResponse" would cause all
                    # error usages to be attributed to that base variant even when a file actually
                    # uses a different variant (e.g. ErrorResponse1), producing wrong BFS propagation.
                    for sh in xl_ep.sheet_names:
                        if not re.fullmatch(r"\d{3}", str(sh).strip()):
                            continue
                        code = str(sh).strip()
                        if int(code) < 400:
                            w_resp = f"{ep_wrapper_base}Response"
                            self.schema_usage.setdefault(w_resp, [])
                            u = f"{self._endpoint_display_name(str(fname))} ({code})"
                            if u not in self.schema_usage[w_resp]:
                                self.schema_usage[w_resp].append(u)
                            bfs_seed_wrappers.add(w_resp)

                    # Load inline component schemas from this endpoint's Schemas sheet
                    # (register definitions only; usages are resolved transitively below)
                    if "Schemas" in xl_ep.sheet_names:
                        df_ep_s = pd.read_excel(xl_ep, sheet_name="Schemas", dtype=str, header=None)
                        h_ep = self._find_header_row(df_ep_s, ["name", "type"])
                        if h_ep != -1:
                            h_ep_vals = [str(c).strip().lower() for c in df_ep_s.iloc[h_ep]]
                            df_ep_s.columns = h_ep_vals
                            df_ep_s = df_ep_s.iloc[h_ep + 1:].reset_index(drop=True)
                            for _, ep_row in df_ep_s.iterrows():
                                ep_name = self._clean_value(ep_row.get('name'))
                                ep_type = self._clean_value(ep_row.get('type')) or 'object'
                                if ep_name and ep_name not in self.global_schemas:
                                    self.global_schemas[ep_name] = DataType(name=ep_name, type=ep_type)

                    # Read direct schema references from response/body sheets.
                    # In the converted format, each sheet (Body, 200, 400, …) contains
                    # exactly one root-level 'Schema Name' value per data row — that is
                    # the wrapper schema (e.g. CommandDetailsResponse).  We do NOT add
                    # wrappers to schema_usage directly; instead we collect them as BFS
                    # seeds so their children (inline components) inherit the usage.
                    for sheet_name in xl_ep.sheet_names:
                        if sheet_name in ["General Description", "Paths", "Tags", "Schemas"]: continue

                        df = pd.read_excel(xl_ep, sheet_name=sheet_name, dtype=str, header=None)

                        # Find a 'schema name' or 'data type' column.
                        schema_col = -1
                        data_type_col = -1
                        header_row = -1
                        for r_idx in range(min(10, len(df))):
                            try:
                                row_vals = [str(v).lower() for v in df.iloc[r_idx]]
                                for i, v in enumerate(row_vals):
                                    if "schema name" in v and schema_col == -1:
                                        schema_col = i
                                        header_row = r_idx
                                    elif "data type" in v and data_type_col == -1:
                                        data_type_col = i
                                        header_row = r_idx
                                if header_row != -1:
                                    break
                            except: continue

                        ep_display = self._endpoint_display_name(str(fname))
                        disp_ctx = str(sheet_name).strip()
                        # Normalize sheet names like '250410 (Body)' -> 'Body' and '260220 (200)' -> '200'
                        m_sheet = re.match(r"^\s*\d+\s*\(([^()]*)\)\s*$", disp_ctx)
                        if m_sheet:
                            disp_ctx = m_sheet.group(1).strip()
                        usage_str = f"{ep_display} ({disp_ctx})"

                        if schema_col != -1 and header_row != -1:
                            # Converted format: 'Schema Name' col holds the wrapper name.
                            # Use it as a BFS seed only (add to schema_refs propagation).
                            for _, row in df.iloc[header_row+1:].iterrows():
                                wrapper = self._clean_value(row.iloc[schema_col])
                                if wrapper and wrapper not in ["string", "number", "integer",
                                                               "boolean", "array", "object"]:
                                    # Wrapper naming can differ between artifacts:
                                    # - endpoint sheets may use 'FooRequest/FooResponse'
                                    # - $index Schemas may contain 'Foo.YYMMDDRequest/Foo.YYMMDDResponse'
                                    # Map to an existing schema name using the endpoint filename's date suffix.
                                    wrapper_mapped = wrapper
                                    stem = Path(str(fname)).stem
                                    m_date = re.match(r"^(.*)\.(\d{6})$", stem)
                                    if m_date:
                                        ep_base = m_date.group(1)
                                        ep_date = m_date.group(2)
                                        m_wr = re.match(r"^(.*?)(Request|Response)$", str(wrapper))
                                        if m_wr:
                                            wr_base = m_wr.group(1)
                                            wr_suffix = m_wr.group(2)
                                            if wr_base == ep_base:
                                                cand = f"{wr_base}.{ep_date}{wr_suffix}"
                                                if cand in self.global_schemas:
                                                    wrapper_mapped = cand
                                    wrapper = wrapper_mapped
                                    # Seed: give the wrapper a usage entry so BFS can propagate
                                    # to its children.  We'll strip wrappers out later.
                                    if wrapper not in self.schema_usage:
                                        self.schema_usage[wrapper] = []
                                    if usage_str not in self.schema_usage[wrapper]:
                                        self.schema_usage[wrapper].append(usage_str)
                                    bfs_seed_wrappers.add(wrapper)

                        elif data_type_col != -1 and header_row != -1:
                            # Legacy format: 'Data Type' col may hold any schema name directly
                            # (these are the leaf types, equivalent to what _perform_naming_and_usage_pass
                            # collects from legacy templates).
                            for _, row in df.iloc[header_row+1:].iterrows():
                                dtype = self._clean_value(row.iloc[data_type_col])
                                if dtype and dtype not in ["string", "number", "integer",
                                                           "boolean", "array", "object"]:
                                    if dtype not in self.schema_usage:
                                        self.schema_usage[dtype] = []
                                    if usage_str not in self.schema_usage[dtype]:
                                        self.schema_usage[dtype].append(usage_str)
                except Exception as ex:
                    self.log(f"Error reading endpoint {fname}: {ex}")
                finally:
                    self._close_excel_file(xl_ep)

            # 4. Propagate usages transitively through the schema reference graph.
            # For each schema that has direct usages, BFS through schema_refs to
            # propagate those same usages to all transitively referenced children.
            changed = True
            while changed:
                changed = False
                for parent_schema, children in list(schema_refs.items()):
                    parent_usages = self.schema_usage.get(parent_schema, [])
                    if not parent_usages:
                        continue
                    for child in children:
                        if child not in self.schema_usage:
                            self.schema_usage[child] = []
                        for u in parent_usages:
                            if u not in self.schema_usage[child]:
                                self.schema_usage[child].append(u)
                                changed = True

            # 5. Filter schema_usage to match converter tracer output:
            #    - Show ALL root-level schemas present in the converted $index Schemas sheet
            #      (i.e. rows without a parent), including wrapper/root schemas like ErrorResponse.
            #    - Exclude sub-field rows (schemas that ONLY appear with a parent in $index)
            _hnp = locals().get("has_no_parent_schema", set())

            # Ensure every root schema is present even if unused
            for k in _hnp:
                self.schema_usage.setdefault(k, [])

            # Keep only root-level schema names
            self.schema_usage = {k: self.schema_usage.get(k, []) for k in _hnp}

            self._log_usage_summary()

            if inject_refs:
                self._inject_schema_references(p, index_file, ep_files)

            return True
            
        except Exception as e:
            self.log(f"Error during standalone check: {e}")
            return False
        finally:
            self._close_excel_file(xl_idx)
    
    def _inject_schema_references(self, folder: Path, index_file: Path, ep_files: list):
        """Create a 'Schema References' sheet in $index.xlsx with merged cells and hyperlinks."""
        from openpyxl.styles import Font, PatternFill, Border, Side, Color

        # 1. Build display_name → actual filename mapping
        display_to_fname: Dict[str, str] = {}
        for fname, _opid in ep_files:
            disp = self._endpoint_display_name(str(fname))
            if disp and disp not in display_to_fname:
                display_to_fname[disp] = str(fname)

        # 2. Parse schema_usage into flat (schema, endpoint, sheet) triples
        triples: List[Tuple[str, str, str]] = []
        for schema_name, usages in self.schema_usage.items():
            for u in usages:
                u = str(u).strip()
                m = re.match(r"^(.*?)\s*\((.+)\)\s*$", u)
                if not m:
                    continue
                ep = m.group(1).strip()
                ctx = m.group(2).strip()
                if not ep or not ctx:
                    continue
                triples.append((schema_name, ep, ctx))

        if not triples:
            self.log("  No schema references to inject.")
            return

        # 3. Sort: schema (case-insensitive), then endpoint, then sheet (numeric first)
        def _sort_key(t):
            s, e, sh = t
            try:
                sh_num = (0, int(sh))
            except ValueError:
                sh_num = (1, sh.lower())
            return (s.lower(), s, e.lower(), e, sh_num)

        triples.sort(key=_sort_key)

        # 4. Open index workbook and create/replace sheet
        wb = None
        try:
            wb = load_workbook(index_file)
        except Exception as ex:
            self.log(f"  ERROR: Cannot open {index_file.name}: {ex}")
            return

        sheet_title = "Schema References"
        if sheet_title in wb.sheetnames:
            del wb[sheet_title]
        ws = wb.create_sheet(sheet_title)
        ws.sheet_properties.tabColor = "8B0000"  # dark red tab

        # 5. Styles — match the Schemas sheet header (red bold Calibri, theme 7 fill)
        header_font = Font(bold=True, color=Color(rgb="FFFF0000"), size=11, name="Calibri")
        header_fill = PatternFill(fill_type="solid",
                                  fgColor=Color(theme=7, tint=0.7999816888943144))
        thin_side = Side(style="thin", color="CCCCCC")
        med_side = Side(style="medium", color="666666")
        thin_border = Border(left=thin_side, right=thin_side,
                             top=thin_side, bottom=thin_side)
        # Separator: medium TOP border on the first row of each new schema group.
        # Using top (not bottom) because merged cells preserve the top-left cell's
        # top border, while bottom borders on the last row of a merge are lost.
        sep_top_border = Border(left=thin_side, right=thin_side,
                                top=med_side, bottom=thin_side)

        headers = ["Schema Name", "Endpoint", "Sheet"]
        for ci, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            cell.border = thin_border

        # 6. Precompute merge ranges from sorted data
        schema_ranges: List[Tuple[int, int]] = []   # (start_row, end_row)
        ep_ranges: List[Tuple[int, int]] = []        # (start_row, end_row)

        prev_schema = None
        prev_ep = None
        schema_start = 2
        ep_start = 2

        for idx, (schema, ep, _sh) in enumerate(triples):
            current_row = idx + 2  # data starts at row 2
            if schema != prev_schema:
                if prev_schema is not None:
                    schema_ranges.append((schema_start, current_row - 1))
                    ep_ranges.append((ep_start, current_row - 1))
                schema_start = current_row
                ep_start = current_row
                prev_schema = schema
                prev_ep = ep
            elif ep != prev_ep:
                ep_ranges.append((ep_start, current_row - 1))
                ep_start = current_row
                prev_ep = ep

        if prev_schema is not None:
            last_data_row = len(triples) + 1
            schema_ranges.append((schema_start, last_data_row))
            ep_ranges.append((ep_start, last_data_row))

        # Build set of rows that START a new schema group (except the very first)
        separator_start_rows = set(start for start, _end in schema_ranges if start > 2)

        # 7. Write data rows
        link_font = Font(color="0563C1", underline="single", size=10, name="Calibri")
        schema_font = Font(bold=True, size=10, name="Calibri")
        normal_font = Font(size=10, name="Calibri")
        center_align = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(vertical="center")

        for ri, (schema, ep, sheet) in enumerate(triples, start=2):
            row_border = sep_top_border if ri in separator_start_rows else thin_border

            # Schema Name (bold)
            c1 = ws.cell(row=ri, column=1, value=schema)
            c1.font = schema_font
            c1.alignment = left_align
            c1.border = row_border

            # Endpoint
            c2 = ws.cell(row=ri, column=2, value=ep)
            c2.font = normal_font
            c2.alignment = left_align
            c2.border = row_border

            # Sheet (with hyperlink)
            c3 = ws.cell(row=ri, column=3, value=sheet)
            c3.font = link_font
            c3.alignment = center_align
            c3.border = row_border
            actual_fname = display_to_fname.get(ep)
            if actual_fname:
                target = f"{actual_fname}#{sheet}!A1"
                c3.hyperlink = target
                c3.style = "Hyperlink"
                c3.font = link_font    # re-apply after style override
                c3.border = row_border  # re-apply after style override

        # 8. Merge cells (data-driven)
        for start, end in schema_ranges:
            if end > start:
                ws.merge_cells(start_row=start, start_column=1, end_row=end, end_column=1)
                ws.cell(row=start, column=1).alignment = Alignment(vertical="center")

        for start, end in ep_ranges:
            if end > start:
                ws.merge_cells(start_row=start, start_column=2, end_row=end, end_column=2)
                ws.cell(row=start, column=2).alignment = Alignment(vertical="center")

        # 9. Column widths
        ws.column_dimensions["A"].width = 35
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 12

        # 10. Save
        try:
            wb.save(index_file)
            self.log(f"\n  Injected '{sheet_title}' sheet into {index_file.name} "
                     f"({len(triples)} references for {len(set(t[0] for t in triples))} schemas).")
        except Exception as ex:
            self.log(f"  ERROR saving {index_file.name}: {ex}")
        finally:
            self._close_workbook(wb)

    def _pre_read_index(self, index_path: Path):
        """Pre-read index to map filenames to operationIds."""
        xl = None
        try:
            xl = pd.ExcelFile(index_path)
            if "Paths" in xl.sheet_names:
                df = pd.read_excel(xl, sheet_name="Paths", dtype=str, header=None)
                header_row_idx = self._find_header_row(df, ["excel file", "operationid"])
                if header_row_idx == -1: return
                
                # Robust headers
                header_vals = [str(c).strip().lower() for c in df.iloc[header_row_idx]]
                df.columns = header_vals
                df = df.iloc[header_row_idx + 1:].reset_index(drop=True)
                
                # Find indices by keyword to avoid series issues if duplicate columns exist
                file_idx = next((i for i, h in enumerate(header_vals) if "excel file" in h or "file" in h), -1)
                op_idx = next((i for i, h in enumerate(header_vals) if "operationid" in h or "operation id" in h), -1)
                
                if file_idx != -1 and op_idx != -1:
                    for _, row in df.iterrows():
                        entry_name = self._clean_value(row.iloc[file_idx])
                        opid = self._clean_value(row.iloc[op_idx])
                        if entry_name and opid:
                            # Find physical file on disk (index often omits extension)
                            actual_fname = None
                            for ext in [".xlsm", ".xlsx", ""]:
                                test_name = entry_name if entry_name.lower().endswith(tuple([".xlsm", ".xlsx"])) else f"{entry_name}{ext}"
                                if (self.input_dir / test_name).exists():
                                    actual_fname = test_name
                                    break
                            
                            if actual_fname:
                                self.filename_to_opid[actual_fname] = opid
                                if actual_fname not in self.ordered_filenames:
                                    self.ordered_filenames.append(actual_fname)
                                # Also map without extension for robustness
                                self.filename_to_opid[actual_fname.replace(".xlsx", "").replace(".xlsm", "")] = opid
                            else:
                                self.log(f"  WARNING: File '{entry_name}' from index not found in {self.input_dir}")
        except Exception as e:
            self.log(f"Error pre-reading index: {e}")
        finally:
            self._close_excel_file(xl)

    def _convert_index(self, legacy_path: Path):
        """Convert $index.xlsm or $index.xlsx to $index.xlsx in output."""
        self.log(f"Converting index: {legacy_path.name}")
        
        # Copy master template
        master_index = self.master_dir / "$index.xlsm"
        if not master_index.exists():
            master_index = self.master_dir / "$index.xlsx"
            
        output_path = self.output_dir / "$index.xlsx"
        shutil.copy(master_index, output_path)

        wb = None
        xl_legacy = None
        try:
            wb = load_workbook(output_path)
            xl_legacy = pd.ExcelFile(legacy_path)

            # 1. General Description
            self._convert_general_description(wb, xl_legacy)

            # 2. Paths
            self._convert_paths(wb, xl_legacy)

            # 3. Tags
            self._convert_tags(wb)

            # 4. Schemas (most complex - builds from Body/Response structures)
            self._convert_schemas(wb)

            wb.save(output_path)
        finally:
            self._close_excel_file(xl_legacy)
            self._close_workbook(wb)

        # Cosmetic Polish
        polish_wb = None
        try:
            polish_wb = load_workbook(output_path)
            for sheet_name in polish_wb.sheetnames:
                self._autofit_columns(polish_wb[sheet_name])
            polish_wb.save(output_path)
        except Exception as e:
            self.log(f"  WARNING: Could not apply cosmetic polish to index: {e}")
        finally:
            self._close_workbook(polish_wb)
        self.log(f"  Saved: {output_path.as_posix()}")

    def _sync_endpoint_schemas_from_index(self, wb) -> None:
        try:
            ws_ep = wb["Schemas"]
        except Exception:
            try:
                ws_ep = wb.create_sheet("Schemas")
            except Exception:
                return

        idx_path = self.output_dir / "$index.xlsx"
        if not idx_path.exists():
            return

        wb_idx = None
        try:
            wb_idx = load_workbook(idx_path)
        except Exception:
            return

        try:
            if "Schemas" not in wb_idx.sheetnames:
                return
            ws_idx = wb_idx["Schemas"]

            rows: List[List[Any]] = []
            for r in range(1, ws_idx.max_row + 1):
                row_vals = []
                for c in range(1, ws_idx.max_column + 1):
                    row_vals.append(ws_idx.cell(row=r, column=c).value)
                if all(v is None or v == "" for v in row_vals):
                    continue
                else:
                    rows.append(row_vals)

            try:
                for r in range(1, ws_ep.max_row + 1):
                    for c in range(1, ws_ep.max_column + 1):
                        ws_ep.cell(row=r, column=c).value = None
            except Exception:
                pass

            self._write_rows(ws_ep, rows, start_row=1)
        finally:
            self._close_workbook(wb_idx)
    
    def _convert_general_description(self, wb, xl_legacy):
        """Convert General Description sheet."""
        ws = wb["General Description"]
        
        # Read legacy key-value format
        df = pd.read_excel(xl_legacy, sheet_name="General Description", dtype=str, header=None)
        
        # Build key-value map
        kv_map = {}
        for _, row in df.iterrows():
            vals = [str(v).strip() for v in row.values if str(v).lower() != "nan"]
            if len(vals) >= 2:
                key = vals[0].lower()
                
                # Handle servers with 4 columns: [servers url, url, servers description, desc]
                # Handle servers with 4 columns: [servers url, url, servers description, desc]
                if key == "servers url":
                    url = vals[1]
                    # FIX: Ignore labels or placeholders
                    if url.lower() in ["servers url", "server url", "url", "servers description", "server description"]:
                        continue
                        
                    desc = vals[3] if len(vals) >= 4 else ""
                    # FIX: Clean placeholder description
                    if desc.lower() in ["servers description", "server description", "description"]:
                        desc = ""
                        
                    if key not in kv_map:
                        kv_map[key] = []
                    kv_map[key].append((url, desc))
                else:
                    kv_map[key] = vals[1]
        
        # Map to modern fixed positions
        mappings = {
            "info description": (2, 2),
            "info version": (3, 2),
            "info title": (4, 2),
            "info contact name": (5, 2),
            "info contact url": (6, 2),
            "release": (9, 2),
            "filename pattern": (10, 2)
        }
        
        for key, (row, col) in mappings.items():
            if key in kv_map:
                ws.cell(row=row, column=col).value = kv_map[key]

        # Legacy templates often lack these keys entirely; inject from preferences
        # so the converted $index.xlsx can satisfy downstream info.contact checks.
        if self.contact_name:
            ws.cell(row=5, column=2).value = self.contact_name
        if self.contact_url:
            ws.cell(row=6, column=2).value = self.contact_url
        if self.release:
            ws.cell(row=9, column=2).value = self.release
        if self.filename_pattern:
            ws.cell(row=10, column=2).value = self.filename_pattern

        # Servers
        if "servers url" in kv_map:
            server_rows = [7, 8]
            for i, (url, desc) in enumerate(kv_map["servers url"]):
                if i >= len(server_rows):
                    break
                r = server_rows[i]
                ws.cell(row=r, column=2).value = url
                ws.cell(row=r, column=4).value = desc
    
    def _convert_paths(self, wb, xl_legacy):
        """Convert Paths sheet."""
        ws = wb["Paths"]
        
        df = pd.read_excel(xl_legacy, sheet_name="Paths", dtype=str)
        header_row_idx = self._find_header_row(df, ["excel file", "path"])
        
        if header_row_idx == -1:
            return
        
        if header_row_idx != -2:
            df.columns = df.iloc[header_row_idx]
            df = df.iloc[header_row_idx + 1:].reset_index(drop=True)
        
        # Collect tags
        tag_col = self._find_column(df, ["tag", "tags"])
        if tag_col:
            tags = df[tag_col].dropna().unique()
            self.all_tags.update([str(t).strip() for t in tags if str(t).strip() and str(t).lower() != "nan"])
        
        # Ensure .xlsx extension on filenames
        filename_col = self._find_column(df, ["excel file", "file"])
        if filename_col:
            df[filename_col] = df[filename_col].apply(self._ensure_xlsx_extension)

        # Normalize HTTP method values (trim spaces, remove Excel leading quote,
        # and enforce lowercase) to avoid invalid keys like "'post " in output OAS.
        method_col = self._find_column(df, ["method"])
        if method_col:
            df[method_col] = df[method_col].apply(self._normalize_http_method)
        
        # Format Custom Extensions to YAML (no curly braces)
        ext_col = self._find_column(df, ["custom extensions", "extension"])
        if ext_col:
            df[ext_col] = df[ext_col].apply(self._format_extensions)
        
        # Write rows
        rows = df.values.tolist()
        self._write_rows(ws, rows, start_row=3)

    def _normalize_http_method(self, val) -> str:
        """Normalize HTTP method coming from legacy Paths sheet.

        Handles common Excel artifacts like a leading apostrophe used to force text
        (e.g. "'post "), trims whitespace, and lowercases the method.
        """
        s = self._clean_value(val)
        if not s:
            return ""
        # Excel text artifact: leading apostrophe
        if s.startswith("'"):
            s = s[1:]
        s = s.strip().strip('"').strip()
        return s.lower()
    
    def _format_extensions(self, val) -> str:
        """Convert JSON-formatted extensions to clean YAML key: value format.
        
        Only transforms strings that are valid JSON objects (dicts).
        All other strings are passed through unchanged.
        """
        import json
        s = str(val).strip()
        if not s or s.lower() == "nan":
            return ""
        
        # Only convert if the string is a valid JSON object
        if s.startswith("{") and s.endswith("}"):
            try:
                data = json.loads(s)
                if isinstance(data, dict):
                    return "\n".join([f"{k}: {v}" for k, v in data.items()])
            except (json.JSONDecodeError, ValueError):
                pass
        
        # Everything else: return as-is (already YAML or plain text)
        return s
    
    def _convert_tags(self, wb):
        """Convert Tags sheet."""
        ws = wb["Tags"]
        
        # Clear existing
        for row in range(2, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).value = None
        
        tags_data = [[tag, tag] for tag in sorted(self.all_tags)]
        self._write_rows(ws, tags_data, start_row=2)
    
    def _convert_schemas(self, wb):
        """Convert Schemas sheet - builds hierarchy from Body/Response structures and Global Data Types."""
        ws = wb["Schemas"]

        def _collect_refs_from_rows(rows: List[List[Any]]) -> set:
            """Collect schema refs that are actually present in emitted schema rows."""
            refs = set()
            primitives = {"string", "number", "integer", "boolean", "array", "object"}
            for r in rows or []:
                if not r:
                    continue
                # row layout: [Name, Parent, Description, Type, ItemsType, SchemaName, ...]
                t = self._clean_value(r[3] if len(r) > 3 else "").lower()
                items_t = self._clean_value(r[4] if len(r) > 4 else "")
                schema_n = self._clean_value(r[5] if len(r) > 5 else "")
                if t == "schema" and schema_n:
                    refs.add(schema_n)
                elif t == "array" and items_t and items_t.lower() not in primitives:
                    refs.add(items_t)
            return refs
        
        # Clear existing
        for row in range(2, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).value = None
        
        final_rows = []
        referenced_data_types = set()
        param_data_types = set()
        
        # 1. Collect all referenced schemas from endpoints (in Paths sheet order)
        ep_files = []
        seen = set()
        for fname in self.ordered_filenames:
            p = self.input_dir / fname
            if p.exists() and fname not in seen:
                ep_files.append((p, fname))
                seen.add(fname)
        # Fallback: catch any files not in the index
        for p in list(self.input_dir.glob("*.xlsm")) + list(self.input_dir.glob("*.xlsx")):
            if not p.name.startswith(("$", "~")) and p.name not in seen:
                ep_files.append((p, p.name))
                seen.add(p.name)
        for ep_file, ep_fname in ep_files:
            xl = pd.ExcelFile(ep_file)
            
            # 1.1 Params
            for sheet_name in ["Path", "Header"]:
                if sheet_name in xl.sheet_names:
                    df = pd.read_excel(xl, sheet_name=sheet_name, dtype=str, header=None)
                    # Legacy templates use "Type"/"Element" headers; converted use "Name"/"Type"
                    header_row_idx = self._find_header_row(df, ["name", "type", "element", "constraint"])
                    if header_row_idx != -1:
                        df.columns = [str(c).strip() for c in df.iloc[header_row_idx]]
                        df = df.iloc[header_row_idx + 1:]
                        for _, row in df.iterrows():
                            dtype = self._clean_value(row.get("Type", ""))
                            # Case-sensitive: lowercase names are OAS built-in primitives;
                            # PascalCase names like 'Boolean' or 'Number' are named DataTypes.
                            if dtype and dtype not in ["string", "number", "integer", "boolean", "array", "object"]:
                                # Resolve WITH file context: each file's DataType is the source of
                                # truth — a collision variant (e.g. SenderBIC2) is a distinct schema
                                # and must be emitted as such, not remapped to another variant.
                                out_name, _ = self._resolve_data_type(dtype, ep_fname)
                                if out_name:
                                    referenced_data_types.add(out_name)
                                    param_data_types.add(out_name)
            
            # 1.2 Body & Responses (Wrappers)
            # Use FILENAME as wrapper base (matching old tool behaviour),
            # NOT the operationId from the index which may differ.
            wrapper_base = self._to_wrapper_case(self._extract_wrapper_base(ep_fname))
            op_id = wrapper_base
            usage_ep = self._endpoint_display_name(str(ep_fname))
            
            if "Body" in xl.sheet_names:
                children = self._read_legacy_structure(xl, "Body")
                if children:
                    wrapper = f"{op_id}Request"
                    self.emitted_wrappers.add(wrapper)
                    child_rows, refs, extra_blocks = self._build_children_rows(
                        wrapper,
                        children,
                        ep_fname,
                        usage_ctx=f"{usage_ep} (Body)",
                        reserved_names=set(referenced_data_types),
                    )
                    final_rows.append([wrapper, "", "", "object", "", "", "", "", "", "", "", "", "", ""])
                    final_rows.extend(child_rows)
                    final_rows.extend(extra_blocks)
                    referenced_data_types.update(refs)
            
            for code in ["200", "201", "400", "401", "403", "404", "500"]:
                if code in xl.sheet_names:
                    children = self._read_legacy_structure(xl, code)
                    if children:
                        if int(code) >= 400:
                            # ErrorResponse: fingerprint/split mechanism.
                            # Build children with file context, then fingerprint to
                            # detect structural differences (e.g. different DataType
                            # variants for requestId).  Each unique structure gets
                            # its own ErrorResponse variant name.
                            placeholder = "ErrorResponse"
                            candidate_rows, refs, extra_blocks = self._build_children_rows(
                                placeholder,
                                children,
                                ep_fname,
                                usage_ctx=f"{usage_ep} ({code})",
                                reserved_names=set(referenced_data_types),
                            )
                            # Fingerprint: structurally significant fields to detect distinct error body shapes.
                            # Must include r[5] (schema_name / DataType reference) so that error responses
                            # referencing different DataType variants (e.g. RequestId vs RequestId1) are
                            # treated as separate ErrorResponse variants. Without r[5], two structures that
                            # differ only in their DataType variant would share the same ErrorResponse, causing
                            # BFS to attribute all error usages to whichever variant the first-processed file
                            # happened to resolve — producing misleading tracer output.
                            fp_rows = []
                            for r in candidate_rows:
                                par = r[1] if r[1] != placeholder else "__W__"
                                schema_ref = r[5] if len(r) > 5 else ""
                                fp_rows.append((r[0], par, r[3], r[4], schema_ref, r[7]))
                            fp = tuple(fp_rows)

                            if fp in self.error_response_fingerprints:
                                # Existing variant — reuse
                                wrapper = self.error_response_fingerprints[fp]
                                final_rows.extend(extra_blocks)
                                referenced_data_types.update(_collect_refs_from_rows(extra_blocks))

                            else:
                                # New variant — assign unique name
                                wrapper = "ErrorResponse"
                                counter = 1
                                while wrapper in self.emitted_wrappers:
                                    wrapper = f"ErrorResponse{counter}"
                                    counter += 1
                                self.error_response_fingerprints[fp] = wrapper
                                self.emitted_wrappers.add(wrapper)
                                # Fix parent references if variant name differs
                                if wrapper != placeholder:
                                    for r in candidate_rows:
                                        if r[1] == placeholder:
                                            r[1] = wrapper
                                final_rows.append([wrapper, "", "", "object", "", "", "", "", "", "", "", "", "", ""])
                                final_rows.extend(candidate_rows)
                                final_rows.extend(extra_blocks)
                                referenced_data_types.update(refs)

                            self.filename_to_error_response[ep_fname] = wrapper
                        else:
                            wrapper = f"{op_id}Response"
                            if wrapper not in self.emitted_wrappers:
                                self.emitted_wrappers.add(wrapper)
                                child_rows, refs, extra_blocks = self._build_children_rows(
                                    wrapper,
                                    children,
                                    ep_fname,
                                    usage_ctx=f"{usage_ep} ({code})",
                                    reserved_names=set(referenced_data_types),
                                )
                                final_rows.append([wrapper, "", "", "object", "", "", "", "", "", "", "", "", "", ""])
                                final_rows.extend(child_rows)
                                final_rows.extend(extra_blocks)
                                referenced_data_types.update(refs)
                            else:
                                _, _refs, extra_blocks = self._build_children_rows(
                                    wrapper,
                                    children,
                                    ep_fname,
                                    usage_ctx=f"{usage_ep} ({code})",
                                    reserved_names=set(referenced_data_types),
                                )
                                final_rows.extend(extra_blocks)
                                referenced_data_types.update(_collect_refs_from_rows(extra_blocks))

            self._close_excel_file(xl)

        # 2a. Transitively expand referenced_data_types: array DataTypes whose
        #     items_type is itself a named DataType (e.g. AosId → AosIdItem)
        #     are not encountered as direct property types in endpoint sheets,
        #     so they would be missed without this expansion.
        _prim_expand = {"string", "number", "integer", "boolean", "array", "object"}
        _expand_pending = set(referenced_data_types)
        _expand_visited: set = set()
        while _expand_pending:
            _ename = _expand_pending.pop()
            if _ename in _expand_visited:
                continue
            _expand_visited.add(_ename)
            _edt = self.global_schemas.get(_ename)
            if not _edt or _edt.type.lower() != "array" or not _edt.items_type:
                continue
            _eitems = self._recursive_resolve_items(
                _edt.items_type,
                _edt.source_file if _edt.source_file != "$global" else None,
            )
            if _eitems and _eitems.lower() not in _prim_expand and _eitems not in referenced_data_types:
                referenced_data_types.add(_eitems)
                _expand_pending.add(_eitems)

        # 2. Emit actually used Global Schemas
        global_blocks = []
        for out_name in sorted(referenced_data_types):
            if out_name in self.emitted_wrappers: continue
            # Skip schemas already emitted as structured named-array inline components
            # (they were written with proper type=array + children via named_array_key_map)
            if out_name in self.emitted_inline_components: continue

            # Only skip true primitive schema names. Named DataTypes such as
            # 'Boolean' or 'Number' must still be emitted as components when
            # they exist in the legacy templates.
            if out_name in {"string", "number", "integer", "boolean", "array", "object"}:
                continue

            dt = self.global_schemas.get(out_name)
            if not dt: continue
            
            # Resolve Items Type recursively if it's an array
            # Rule: only inline if name is 'object' or 'array'
            items_out = dt.items_type
            if dt.type.lower() == "array" and dt.items_type:
                # Use recursive resolution helper
                items_out = self._recursive_resolve_items(dt.items_type, dt.source_file)
            
            # Only emit pattern_eba when there IS a regex; old tool never emitted
            # EBA-only patterns (e.g. '4!c2!a2!c') as OAS pattern values.
            emit_pat_eba = dt.pattern_eba if dt.regex else ""
            row = [out_name, "", dt.description, dt.type if dt.type else "string",
                   items_out, "", dt.format, "", dt.min_val, dt.max_val,
                   emit_pat_eba, dt.regex, dt.allowed_values, dt.example, ""]
            global_blocks.append((out_name, [row]))

        # 3. Final Write-back with Cosmetics
        all_blocks = []
        # Group final_rows back into blocks for sorting
        curr_b = []
        for r in final_rows:
            if r[1] == "":
                if curr_b: all_blocks.append((curr_b[0][0], curr_b))
                curr_b = [r]
            else: curr_b.append(r)
        if curr_b: all_blocks.append((curr_b[0][0], curr_b))
        
        all_blocks.extend(global_blocks)

        # --- REACHABILITY FILTER: prune schemas not reachable from endpoints ---
        # Entry points: wrappers referenced by endpoint Body/Response sheets,
        # plus DataTypes referenced by endpoint Path/Header parameters.
        entry_points = set(self.emitted_wrappers) | param_data_types
        # Build reference graph: block_name -> set of schema names it references
        _prim = {"string", "number", "integer", "boolean", "array", "object"}
        _ref_graph = {}
        for _bn, _br in all_blocks:
            _refs_set = set()
            for _row in _br:
                _sn = self._clean_value(_row[5] if len(_row) > 5 else "")
                _it = self._clean_value(_row[4] if len(_row) > 4 else "")
                _dt = self._clean_value(_row[3] if len(_row) > 3 else "").lower()
                if _dt == "schema" and _sn:
                    _refs_set.add(_sn)
                elif _dt == "array" and _it and _it.lower() not in _prim:
                    _refs_set.add(_it)
            _ref_graph.setdefault(_bn, set()).update(_refs_set)
        # BFS from entry points
        _reachable = set()
        _queue = list(entry_points)
        while _queue:
            _cur = _queue.pop(0)
            if _cur in _reachable:
                continue
            _reachable.add(_cur)
            for _nxt in _ref_graph.get(_cur, set()):
                if _nxt not in _reachable:
                    _queue.append(_nxt)
        all_blocks = [(n, r) for n, r in all_blocks if n in _reachable]

        # --- COLLISION RENUMBERING: enforce contiguous suffixes after pruning ---
        # Reachability pruning can remove intermediate variants (e.g. Foo1) and leave
        # Foo + Foo2. Business requirement: output names must be contiguous.
        def _collect_refs_from_blocks(_blocks: List[Tuple[str, List[List[Any]]]]) -> set:
            _refs: set = set()
            _prim2 = {"string", "number", "integer", "boolean", "array", "object"}
            for _bn2, _br2 in _blocks:
                for _row2 in _br2:
                    _sn2 = self._clean_value(_row2[5] if len(_row2) > 5 else "")
                    _it2 = self._clean_value(_row2[4] if len(_row2) > 4 else "")
                    _dt2 = self._clean_value(_row2[3] if len(_row2) > 3 else "").lower()
                    if _dt2 == "schema" and _sn2:
                        _refs.add(_sn2)
                    elif _dt2 == "array" and _it2 and _it2.lower() not in _prim2:
                        _refs.add(_it2)
            return _refs

        def _apply_rename_to_blocks(_blocks: List[Tuple[str, List[List[Any]]]], mapping: Dict[str, str]) -> List[Tuple[str, List[List[Any]]]]:
            out_blocks: List[Tuple[str, List[List[Any]]]] = []
            for _bn2, _br2 in _blocks:
                new_bn = mapping.get(_bn2, _bn2)
                new_rows: List[List[Any]] = []
                for _row2 in _br2:
                    rr = list(_row2)
                    # Update Schema Name (col F, index 5)
                    if len(rr) > 5 and isinstance(rr[5], str) and rr[5]:
                        rr[5] = mapping.get(rr[5], rr[5])
                    # Update Items Type (col E, index 4)
                    if len(rr) > 4 and isinstance(rr[4], str) and rr[4]:
                        rr[4] = mapping.get(rr[4], rr[4])
                    # Update root schema name in column A when parent is empty
                    if len(rr) > 1 and (rr[1] == "" or rr[1] is None):
                        if len(rr) > 0 and isinstance(rr[0], str) and rr[0]:
                            rr[0] = new_bn
                    new_rows.append(rr)
                out_blocks.append((new_bn, new_rows))
            return out_blocks

        # Build variant sets by stem
        names_in_blocks = [n for n, _r in all_blocks]
        stem_map: Dict[str, Dict[str, Any]] = {}
        for nm in names_in_blocks:
            m = re.search(r"^(.*?)(\d+)$", str(nm))
            if not m:
                stem_map.setdefault(str(nm), {"base": True, "nums": []})
                continue
            stem = m.group(1)
            num = int(m.group(2))
            entry = stem_map.setdefault(stem, {"base": False, "nums": []})
            entry["nums"].append(num)

        # Decide rename mapping
        rename_map: Dict[str, str] = {}
        used_after: set = set(names_in_blocks)
        # Keep memory of names that existed before reachability pruning / final
        # compaction. This lets us distinguish:
        # - true collision families where an unsuffixed base really existed
        # - semantic numeric names such as BIC8, where "BIC" never existed and
        #   therefore must NOT be synthesized during compaction
        historical_names: set = (
            set(self.global_schemas.keys())
            | set(self.emitted_wrappers)
            | set(self.emitted_inline_components)
            | set(self.inline_component_fingerprint_by_name.keys())
            | set(names_in_blocks)
        )
        for stem, info in stem_map.items():
            if not info.get("nums"):
                continue
            has_base = bool(info.get("base")) and stem in used_after
            had_base_before_pruning = stem in historical_names
            nums = sorted(set(info.get("nums") or []))
            existing_variant_names = [f"{stem}{n}" for n in nums if f"{stem}{n}" in used_after]

            if has_base:
                expected = list(range(1, len(existing_variant_names) + 1))
                if nums[: len(expected)] == expected and len(nums) == len(expected):
                    continue
                rename_pairs = list(
                    zip(
                        existing_variant_names,
                        [f"{stem}{i}" for i in range(1, len(existing_variant_names) + 1)],
                    )
                )
            else:
                if not had_base_before_pruning:
                    # Semantic numeric names (e.g. BIC8 when BIC never existed)
                    # are valid bases and must keep their original name.
                    continue
                # When the unsuffixed base is no longer reachable/emitted, compact the
                # remaining variants so the first live variant becomes the base name,
                # but only for families whose unsuffixed base really existed.
                # Example: Number1, Number2 -> Number, Number1
                target_names = [stem] + [f"{stem}{i}" for i in range(1, len(existing_variant_names))]
                rename_pairs = list(zip(existing_variant_names, target_names))

            for old_nm, new_nm in rename_pairs:
                if old_nm == new_nm:
                    continue
                # Guard against accidental clash outside this stem
                if (new_nm in used_after) and (new_nm not in existing_variant_names):
                    continue
                rename_map[old_nm] = new_nm
                used_after.discard(old_nm)
                used_after.add(new_nm)

        if rename_map:
            all_blocks = _apply_rename_to_blocks(all_blocks, rename_map)

            # Keep registries aligned for tracing / later lookups
            try:
                for old_nm, new_nm in rename_map.items():
                    if old_nm in self.global_schemas:
                        moved_dt = self.global_schemas.pop(old_nm)
                        # Always overwrite the destination slot: a stale pruned schema with
                        # the compacted name may still exist in the registry, but endpoint
                        # mappings must resolve to the renamed live DataType content.
                        moved_dt.name = new_nm
                        self.global_schemas[new_nm] = moved_dt
                self.used_names = {rename_map.get(nm, nm) for nm in self.used_names}
                self.output_names = {
                    key: rename_map.get(out_name, out_name)
                    for key, out_name in self.output_names.items()
                }
                self.fingerprints = {
                    fp: rename_map.get(out_name, out_name)
                    for fp, out_name in self.fingerprints.items()
                }
                self.merge_provenance = {
                    rename_map.get(name, name): prov
                    for name, prov in self.merge_provenance.items()
                }
            except Exception:
                pass

            # Update ErrorResponse variant mapping so endpoint response sheets
            # reference the renumbered name.
            for ep_key, er_name in list(self.filename_to_error_response.items()):
                if er_name in rename_map:
                    self.filename_to_error_response[ep_key] = rename_map[er_name]

        all_blocks.sort(key=lambda x: str(x[0]).lower())
        
        final_final_rows = []
        for i, (_, block) in enumerate(all_blocks):
            final_final_rows.extend(block)
                
        # Compute hierarchy levels for grouping.
        # Level is written to column O (15) and used for Excel outline grouping.
        def _ensure_len_15(r: List[Any]) -> List[Any]:
            rr = list(r)
            while len(rr) < 15:
                rr.append("")
            return rr

        enriched_rows: List[List[Any]] = []
        parent_map: Dict[str, str] = {}
        in_block = False
        for r in final_final_rows:
            if r is None:
                continue
            if all(v is None or v == "" for v in r):
                enriched_rows.append(_ensure_len_15(r))
                parent_map = {}
                in_block = False
                continue

            rr = _ensure_len_15(r)
            name = rr[0] if len(rr) > 0 else ""
            parent = rr[1] if len(rr) > 1 else ""
            if parent == "" or parent is None:
                level = 0
                parent_map = {}
                in_block = True
            else:
                if in_block and isinstance(name, str) and name:
                    parent_map[str(name)] = str(parent) if parent is not None else ""

                level = 1
                cur = str(parent) if parent is not None else ""
                guard = 0
                while cur and guard < 50:
                    p2 = parent_map.get(cur)
                    if not p2:
                        break
                    level += 1
                    cur = p2
                    guard += 1

            rr[14] = level
            enriched_rows.append(rr)

        self._write_rows(ws, enriched_rows, start_row=2)

        try:
            ws.column_dimensions["O"].hidden = True
        except Exception:
            pass

        try:
            ws.sheet_properties.outlinePr.summaryBelow = True
        except Exception:
            pass

        try:
            for i, r in enumerate(enriched_rows, start=2):
                lvl = r[14]
                if isinstance(lvl, int) and lvl >= 0:
                    ws.row_dimensions[i].outlineLevel = min(lvl, 7)
                    ws.row_dimensions[i].hidden = False

                    # Visual indent in column A based on level (does not change cell content)
                    try:
                        cell = ws.cell(row=i, column=1)
                        indent = min(max(lvl, 0) * 2, 15)
                        cell.alignment = Alignment(horizontal="left", wrap_text=False, indent=indent)
                    except Exception:
                        pass
        except Exception:
            pass

        # Add internal hyperlinks: col F ("Schema Name") → root definition in col A
        try:
            from openpyxl.styles import Font
            # 1. Build map: root schema name → row number
            root_row_map: Dict[str, int] = {}
            for ri in range(2, ws.max_row + 1):
                name_val = ws.cell(row=ri, column=1).value
                parent_val = ws.cell(row=ri, column=2).value
                if name_val and (not parent_val or str(parent_val).strip() == ""):
                    sn = str(name_val).strip()
                    if sn not in root_row_map:
                        root_row_map[sn] = ri

            # 2. Add hyperlinks on col F cells that reference a known root schema
            link_font = Font(color="0563C1", underline="single")
            for ri in range(2, ws.max_row + 1):
                schema_ref = ws.cell(row=ri, column=6).value
                if not schema_ref:
                    continue
                ref_str = str(schema_ref).strip()
                target_row = root_row_map.get(ref_str)
                if target_row:
                    cell = ws.cell(row=ri, column=6)
                    cell.hyperlink = f"#Schemas!A{target_row}"
                    cell.font = link_font
        except Exception:
            pass
    
    def _convert_endpoint(self, legacy_path: Path):
        """Convert endpoint *.xlsm to *.xlsx."""
        filename = legacy_path.name.replace(".xlsm", ".xlsx")
        self.log(f"Converting endpoint: {filename}")
        
        master_ep = self.master_dir / "endpoint.xlsm"
        if not master_ep.exists():
            master_ep = self.master_dir / "endpoint.xlsx"
            
        output_path = self.output_dir / filename
        shutil.copy(master_ep, output_path)
        
        self.current_ep_name = legacy_path.name

        wb = None
        xl = None
        try:
            wb = load_workbook(output_path)
            xl = pd.ExcelFile(legacy_path)

            # Use FILENAME as wrapper base (matching old tool behaviour),
            # NOT the operationId from the index which may differ.
            op_id = self._to_wrapper_case(self._extract_wrapper_base(legacy_path.name))

            status_codes = [s for s in xl.sheet_names if s.isdigit()]

            # Remove unused response sheets
            sheets_to_remove = [s for s in wb.sheetnames if s.isdigit() and s not in status_codes]
            for s in sheets_to_remove:
                del wb[s]

            # Create needed response sheets
            if "Response" in wb.sheetnames:
                response_tpl = wb["Response"]
                for code in status_codes:
                    if code not in wb.sheetnames:
                        new_sheet = wb.copy_worksheet(response_tpl)
                        new_sheet.title = code
                del wb["Response"]

            # 1. Parameters
            self._convert_parameters(wb, xl)

            # 2. Body
            if "Body" in xl.sheet_names:
                self._convert_body(wb, xl, op_id, legacy_path.name)
                self._convert_body_example(wb, xl, op_id, legacy_path.name)

            # 3. Response sheets
            if status_codes:
                self._convert_responses(wb, xl, op_id, status_codes, legacy_path.name)

            wb.save(output_path)
        finally:
            self._close_excel_file(xl)
            self._close_workbook(wb)

        # Cosmetic Polish
        polish_wb = None
        try:
            polish_wb = load_workbook(output_path)
            for sheet_name in polish_wb.sheetnames:
                self._autofit_columns(polish_wb[sheet_name])
            polish_wb.save(output_path)
        except Exception as e:
            self.log(f"  WARNING: Could not apply cosmetic polish to {filename}: {e}")
        finally:
            self._close_workbook(polish_wb)
        self.log(f"  Saved: {output_path.as_posix()}")
    
    def _convert_parameters(self, wb, xl):
        """Convert Parameters from Path and Header sheets."""
        ws = wb["Parameters"]
        
        params = []
        
        # Path parameters
        if "Path" in xl.sheet_names:
            df = pd.read_excel(xl, sheet_name="Path", dtype=str)
            params.extend(self._parse_params(df, "path"))
        
        # Header parameters
        if "Header" in xl.sheet_names:
            df = pd.read_excel(xl, sheet_name="Header", dtype=str)
            params.extend(self._parse_params(df, "header"))
            
        ep_filename = wb.properties.title if hasattr(wb, 'properties') and wb.properties.title else None
        # Since openpyxl title might not be set, use a better way to pass the filename
        # Actually, let's just use self.current_ep_filename if we set it
        ep_filename = self.current_ep_name
        
        # Write parameters with data type resolution
        rows = []
        for p in params:
            param_type = p["type"]
            schema_name = ""
            param_format = ""
            
            # Resolve data type to schema reference (R2.6: parameters refer to dynamic schemas).
            # Use file context to find the DataType, then remap to the lowest-numbered registered
            # variant so that file-specific collision variants (e.g. SenderBIC2 from
            # listFundingDefunding) are replaced with the canonical name (SenderBIC1).
            # Force PascalCase since DataType names are always PascalCase.
            out_name, dt = self._resolve_data_type(param_type, ep_filename)
            if out_name:
                param_type = "schema"
                schema_name = out_name
            elif self._is_oas_primitive_name(param_type):
                param_type = str(param_type).strip()
            
            rows.append([
                p["name"], p["description"], p["in"], param_type,
                "", schema_name, param_format, p["mandatory"],
                "", "", "", "", "", ""  # Min, Max, PatternEba, Regex, Allowed, Example (empty in endpoint)
            ])
        
        self._write_rows(ws, rows, start_row=3)
    
    def _convert_body(self, wb, xl_legacy, op_id: str, ep_filename: str):
        """Convert Body sheet - media type row references wrapper schema."""
        ws = wb["Body"]
        
        wrapper_name = f"{op_id}Request"
        if wrapper_name not in self.emitted_wrappers:
            return
            
        # Media Type row with schema reference to wrapper
        # Properties are already defined in the component schema - no need to duplicate inline
        media_row = ["content", "application/json", "", "", "schema", "", wrapper_name, "", "M",
                     "", "", "", "", "", ""]
        self._write_rows(ws, [media_row], start_row=3)
    
    def _convert_responses(self, wb, xl, op_id: str, status_codes: List[str], filename: str):
        """Convert response sheets."""
        for code in status_codes:
            if code not in wb.sheetnames:
                continue
            
            ws = wb[code]
            
            # Determine wrapper
            if int(code) < 400:
                wrapper_name = f"{op_id}Response"
            else:
                # Use the correct ErrorResponse variant for this endpoint
                wrapper_name = self.filename_to_error_response.get(filename, "ErrorResponse")
            
            # Extract description from legacy title
            desc = ""
            try:
                df = pd.read_excel(xl, sheet_name=code, dtype=str, header=None)
                if not df.empty:
                    title_row = df.iloc[0].dropna().tolist()
                    for val in title_row:
                        s = str(val).strip()
                        if s.startswith(f"Response {code}"):
                            if " - " in s:
                                desc = s.split(" - ", 1)[1].strip()
                            break
            except:
                pass
            
            if not desc:
                desc = {"200": "OK", "201": "Created", "400": "Bad Request",
                        "401": "Unauthorized", "403": "Forbidden", "404": "Not Found"}.get(code, "")
            
            # Tab Coloring (R1 colors)
            status_int = int(code) if code.isdigit() else 0
            if 200 <= status_int < 300:
                ws.sheet_properties.tabColor = "FFD3ECB9"  # Light Green
            else:
                ws.sheet_properties.tabColor = "FFFF9999"  # Light Red
            
            # Row 1: Response | code | description
            ws.cell(row=1, column=1).value = "Response"
            ws.cell(row=1, column=2).value = code
            ws.cell(row=1, column=3).value = desc
            
            # Row 3+: media type row with schema reference to wrapper
            children = self._read_legacy_structure(xl, code)
            if children:
                # Media type row with schema reference to wrapper
                # Properties are already defined in the component schema - no need to duplicate inline
                media_row = ["content", "application/json", "", "", "schema", "", wrapper_name, "", "M",
                             "", "", "", "", "", ""]
                self._write_rows(ws, [media_row], start_row=3)

                if str(code) in ("200", "201"):
                    self._add_ok_response_example(
                        ws=ws,
                        xl_legacy=xl,
                        response_code=str(code),
                        response_children=children,
                        ep_filename=filename,
                    )

                if str(code) == "400":
                    self._add_400_response_example(
                        ws=ws,
                        xl_legacy=xl,
                        response_children=children,
                        ep_filename=filename,
                    )
            elif int(code) >= 400 or code == "204":
                # Empty 4xx/5xx/204 sheet: create a ref to a shared ErrorResponse_<code> component
                resp_component = f"ErrorResponse_{code}"
                media_row = ["content", desc, "", desc, "response", "", resp_component, "", "",
                             "", "", "", "", "", ""]
                self._write_rows(ws, [media_row], start_row=3)
                # Register for the global Responses sheet (once per status code)
                if code not in self.empty_error_responses:
                    self.empty_error_responses[code] = desc
    
    def _write_responses_sheet(self):
        """Populate the Responses sheet in $index.xlsx with ErrorResponse_<code> components."""
        if not self.empty_error_responses:
            return

        index_path = self.output_dir / "$index.xlsx"
        if not index_path.exists():
            return

        wb = None
        try:
            wb = load_workbook(index_path)
            if "Responses" not in wb.sheetnames:
                wb.create_sheet("Responses")
            ws = wb["Responses"]

            rows: List[List[Any]] = []
            for code in sorted(self.empty_error_responses.keys()):
                desc = self.empty_error_responses[code]
                comp = f"ErrorResponse_{code}"
                example_value = "" if code == "204" else desc
                rows.extend([
                    ["", comp, "", desc, "", "", "", "", "", "", "", "", "", "", ""],
                    ["", "x-sandbox-request-name", comp, "", "string", "", "", "", "", "", "", "", "", "", ""],
                    ["", "x-sandbox-request-path-params", comp, "", "string", "", "", "", "", "", "", "", "", "", ""],
                    ["", "senderBic", "x-sandbox-request-path-params", "x sandbox request path params", "string", "", "", "", "", "", "", "", "", "", ""],
                    ["content", "text/plain", comp, "", "string", "", "", "", "", "", "", "", "", "", f"TSTBICXX{code}"],
                    ["examples", desc, "text/plain", "Error message", "string", "", "", "", "", "", "", "", "", "", ""],
                    ["examples", "value", desc, "", "string", "", "", "", "", "", "", "", "", "", example_value],
                ])

            self._write_rows(ws, rows, start_row=2)

            # Cosmetic polish
            try:
                self._autofit_columns(ws)
            except Exception:
                pass

            wb.save(index_path)
            self.log(f"  Responses sheet: {len(self.empty_error_responses)} ErrorResponse components written.")
        finally:
            self._close_workbook(wb)

    # === Helper Methods ===
    
    def _read_legacy_structure(self, xl, sheet_name: str) -> List[Tuple]:
        """Read legacy Body/Response structure: (name, parent, description, type, mandatory)."""
        df = pd.read_excel(xl, sheet_name=sheet_name, dtype=str, header=None)
        
        # Representative keywords for header detection
        header_keywords = ["name", "element", "type", "parent", "data type", "description", "mandatory"]
        header_row_idx = self._find_header_row(df, header_keywords)
        
        if header_row_idx == -1:
            return []
        
        # Determine actual header names (case-insensitive)
        headers = [str(c).strip().lower() for c in df.iloc[header_row_idx]]
        df = df.iloc[header_row_idx + 1:].reset_index(drop=True)
        
        def find_idx(candidates):
            # Perfect match first
            for i, h in enumerate(headers):
                if any(str(c).strip().lower() == h for c in candidates):
                    return i
            # Substring match fallback
            for i, h in enumerate(headers):
                if any(str(c).strip().lower() in h for c in candidates):
                    return i
            return -1

        name_idx = find_idx(["element", "name"])
        parent_idx = find_idx(["parent", "parents"])
        desc_idx = find_idx(["description", "description "])
        type_idx = find_idx(["type", "data type", "dtype"])
        items_idx = find_idx(["items data type", "items type", "array only"])
        mand_idx = find_idx(["mandatory", "required"])
        v_rules_idx = find_idx(["validation rules", "validation rule", "validation"])

        children = []
        for _, row in df.iterrows():
            name = self._clean_value(row.iloc[name_idx]) if name_idx != -1 else ""
            if not name or name.lower() == "nan":
                continue
            
            parent = self._clean_value(row.iloc[parent_idx]) if parent_idx != -1 else ""
            desc = self._clean_value(row.iloc[desc_idx]) if desc_idx != -1 else ""
            dtype = self._clean_value(row.iloc[type_idx]) if type_idx != -1 else "string"
            items_type = self._clean_value(row.iloc[items_idx]) if items_idx != -1 else ""
            mandatory = self._clean_value(row.iloc[mand_idx]) if mand_idx != -1 else ""
            v_rules = self._clean_value(row.iloc[v_rules_idx]) if v_rules_idx != -1 else ""
            
            children.append((name, parent, desc, dtype, mandatory, v_rules, items_type))
        
        return children
    
    def _build_children_rows(self, wrapper_name: str, children: List[Tuple], ep_filename: Optional[str] = None, legato_parent: Optional[str] = None, usage_ctx: Optional[str] = None, reserved_names: Optional[set] = None) -> Tuple[List[List], set, List[List]]:
        """Build child rows with proper parent hierarchy and type resolution.
        If legato_parent is provided, use it as Parent for root items.
        Returns (rows, referenced_data_types set, extra_schema_blocks).
        
        When a node resolves to a schema $ref, its children are NOT emitted
        because they already exist in the referenced component schema.
        """
        rows = []
        referenced_data_types = set()
        extra_schema_blocks: List[List] = []

        def _norm(v: Any) -> str:
            if v is None:
                return ""
            s = str(v).strip()
            if not s or s.lower() == "nan":
                return ""
            s = s.lower()
            try:
                s = re.sub(r"\[\s*\]$", "", s)
                s = re.sub(r"\[\d+\]$", "", s)
            except Exception:
                pass
            return s

        children_lower = [(_norm(n), _norm(p)) for (n, p, *_rest) in children]
        has_children = set()
        for n, p_low in children_lower:
            if p_low:
                has_children.add(p_low)

        # Build children map for collision fingerprinting: object DataTypes
        # whose structure (child property names) differs across endpoints
        # (e.g. Criteria) need distinct fingerprints.
        parent_children: Dict[str, set] = {}
        for name_t, parent_t, *_ in children:
            cn = self._clean_value(name_t)
            cp = self._clean_value(parent_t)
            if cp and cn:
                parent_children.setdefault(cp.lower(), set()).add(cn)
        for name_t, parent_t, desc_t, dtype_t, *_ in children:
            cn = self._clean_value(name_t)
            cdtype = self._clean_value(dtype_t)
            if cn and cdtype and cn.lower() in parent_children:
                norm_dt = self._to_pascal_case(cdtype)
                if norm_dt and norm_dt not in self._current_children_map:
                    self._current_children_map[norm_dt] = frozenset(parent_children[cn.lower()])

        def _unique_inline_component_name(base_name: str) -> str:
            candidate = base_name
            i = 1
            reserved = reserved_names or set()
            registered_names = set(self.inline_component_fingerprints.values())
            while (candidate in reserved
                   or candidate in self.emitted_wrappers
                   or candidate in self.emitted_inline_components
                   or candidate in registered_names):
                candidate = f"{base_name}{i}"
                i += 1
            return candidate

        def _row_example(dtype_val: Any, items_val: Any) -> str:
            """Return example text from referenced global DataType (or array items DataType)."""
            def _ex_for_type(type_name: str) -> str:
                if not type_name:
                    return ""
                if self._is_oas_primitive_name(type_name, allow_schema=True):
                    return ""
                out_name, dt = self._resolve_data_type(str(type_name).strip(), ep_filename)
                if dt and getattr(dt, "example", None):
                    return str(dt.example).strip()
                return ""

            ex = _ex_for_type(dtype_val)
            if not ex:
                ex = _ex_for_type(items_val)
            return ex

        def _subtree_example_signature(subtree_children: List[Tuple]) -> str:
            parts = []
            for t_name, t_parent, t_desc, t_dtype, t_mand, t_rules, t_items in subtree_children:
                ex = _row_example(t_dtype, t_items)
                if not ex:
                    continue
                parts.append(f"{str(t_name).strip()}={ex}")
            parts = sorted(set(parts))
            return "; ".join(parts)

        def _fp_inline_subtree(subtree_children: List[Tuple], root_low: str, include_rules: bool = True) -> Tuple:
            entries = []
            for t_name, t_parent, t_desc, t_dtype, t_mand, t_rules, t_items in subtree_children:
                p_low = _norm(t_parent)
                p_low = "" if p_low == root_low else p_low
                desc_part = ""
                if self.include_descriptions_in_collision:
                    desc_part = str(t_desc).strip() if t_desc is not None else ""
                ex_part = ""
                if self.include_examples_in_collision:
                    ex_part = _row_example(t_dtype, t_items)
                entries.append(
                    (
                        _norm(t_name),
                        p_low,
                        desc_part,
                        ex_part,
                        _norm(t_dtype),
                        _norm(t_items),
                        _norm(t_mand),
                        (str(t_rules).strip() if t_rules is not None else "") if include_rules else "",
                    )
                )
            return tuple(sorted(entries))

        def _track_inline_component_usage(schema_name: str) -> None:
            if not usage_ctx:
                return
            if schema_name not in self.schema_usage:
                self.schema_usage[schema_name] = []
            if usage_ctx not in self.schema_usage[schema_name]:
                self.schema_usage[schema_name].append(usage_ctx)

        def _bfs_descendants(root_name_norm: str, root_parent_norm: str) -> List[Tuple]:
            """Collect descendants of a specific (name, parent) row instance using
            document-order greedy assignment.

            Key constraint: each parent expansion collects at most ONE row per
            distinct child-name.  This prevents the first 'dailyThresholds' block
            from absorbing both 'lacAgenda' subtrees when two parallel blocks share
            the same element names in the same flat children list.
            """
            root_idx = next(
                (i for i, (n, p, *_) in enumerate(children)
                 if _norm(n) == root_name_norm and _norm(p) == root_parent_norm),
                -1,
            )
            if root_idx == -1:
                return []

            result: List[Tuple] = []
            claimed: set = {root_idx}
            queue: list = [(root_idx, root_name_norm)]

            while queue:
                parent_idx, parent_name_norm = queue.pop(0)
                # For each distinct child name, claim only the first unclaimed row
                # with matching parent that appears after parent_idx.
                seen_child_names: set = set()
                for i in range(parent_idx + 1, len(children)):
                    if i in claimed:
                        continue
                    t_name, t_parent = children[i][0], children[i][1]
                    if _norm(t_parent) != parent_name_norm:
                        continue
                    t_name_norm = _norm(t_name)
                    if t_name_norm in seen_child_names:
                        # Skip duplicate child names — they belong to a sibling subtree
                        continue
                    seen_child_names.add(t_name_norm)
                    claimed.add(i)
                    result.append(children[i])
                    queue.append((i, t_name_norm))

            return result

        current_ctx_kind = self._usage_context_kind(usage_ctx)

        # Maps keyed by (name_norm, parent_norm) so parallel same-named subtrees each
        # get their own fingerprint, schema name, and emitted block.
        object_key_map: Dict[Tuple[str, str], str] = {}
        object_key_subtrees: Dict[Tuple[str, str], List[Tuple]] = {}
        array_key_map: Dict[Tuple[str, str], str] = {}
        array_key_subtrees: Dict[Tuple[str, str], List[Tuple]] = {}
        # Named-DataType arrays: dtype is a named schema whose resolved type is 'array'.
        # The property should be a $ref to the schema (which IS the array), not an inline
        # array wrapping the schema. The schema is emitted as type=array with children as
        # items.properties — avoiding the double-array anti-pattern.
        named_array_key_map: Dict[Tuple[str, str], str] = {}
        named_array_key_subtrees: Dict[Tuple[str, str], List[Tuple]] = {}
        # Legacy name-only maps kept for has-children checks in schema_ref_nodes pass.
        object_ref_map: Dict[str, str] = {}
        array_items_ref_map: Dict[str, str] = {}
        named_array_items_ref_map: Dict[str, str] = {}
        for name, parent, desc, dtype, mandatory, v_rules, items_type_row in children:
            if _norm(dtype) != "object":
                continue
            name_norm = _norm(name)
            if name_norm not in has_children:
                continue
            parent_norm = _norm(parent)
            key = (name_norm, parent_norm)

            descendants = _bfs_descendants(name_norm, parent_norm)

            if descendants:
                object_key_subtrees[key] = descendants
                ex_sig = _subtree_example_signature(descendants)
                fp = _fp_inline_subtree(descendants, name_norm, include_rules=(current_ctx_kind == "request"))
                neutral_fp = _fp_inline_subtree(descendants, name_norm, include_rules=False)
                existing = None
                if current_ctx_kind == "request":
                    existing = self.inline_component_fingerprints.get(fp)
                else:
                    existing = self.inline_component_neutral_fingerprints.get(neutral_fp) or self.inline_component_fingerprints.get(fp)
                if existing:
                    object_ref_map[name_norm] = existing
                    object_key_map[key] = existing
                    # If descriptions are neutral, record merge provenance when root descriptions differ across endpoints
                    if not self.include_descriptions_in_collision:
                        root_desc = (str(desc).strip() if desc is not None else "")
                        canon_desc = (self.inline_component_root_desc.get(existing) or "").strip()
                        if root_desc and root_desc != canon_desc:
                            self._record_merge_provenance(existing, f"description:{name_norm}", root_desc, ep_filename or "")
                    if not self.include_examples_in_collision:
                        canon_ex = (self.inline_component_example_sig.get(existing) or "").strip()
                        if ex_sig and ex_sig != canon_ex:
                            self._record_merge_provenance(existing, f"example:{name_norm}", ex_sig, ep_filename or "")
                else:
                    comp_name = _unique_inline_component_name(self._to_pascal_case(name))
                    object_ref_map[name_norm] = comp_name
                    object_key_map[key] = comp_name
                    self.inline_component_fingerprints[fp] = comp_name
                    self.inline_component_neutral_fingerprints.setdefault(neutral_fp, comp_name)
                    self.inline_component_fingerprint_by_name[comp_name] = fp
                    self.inline_component_root_desc[comp_name] = (str(desc).strip() if desc is not None else "")
                    self.inline_component_example_sig[comp_name] = ex_sig
            else:
                comp_name = _unique_inline_component_name(self._to_pascal_case(name))
                object_ref_map[name_norm] = comp_name
                object_key_map[key] = comp_name

            _track_inline_component_usage(object_key_map[key])

        for name, parent, desc, dtype, mandatory, v_rules, items_type_row in children:
            low_dtype = _norm(dtype)
            out_name, dt = self._resolve_data_type(str(dtype).strip() if dtype is not None else "", ep_filename)
            is_literal_array = low_dtype == "array"
            # Named-DataType array: dtype is a named schema whose resolved type is 'array'
            # (e.g. dtype="alerts" → DataType Alerts with type=array).  These properties
            # must emit a plain $ref to the schema (not type=array wrapping the schema),
            # otherwise the generator produces a double-array.
            is_named_array_dt = (
                not is_literal_array
                and out_name is not None
                and dt is not None
                and dt.type
                and dt.type.lower() == "array"
            )
            is_array = is_literal_array or is_named_array_dt
            if not is_array:
                continue
            name_norm = _norm(name)
            if name_norm not in has_children:
                continue
            parent_norm = _norm(parent)
            key = (name_norm, parent_norm)

            descendants = _bfs_descendants(name_norm, parent_norm)

            if is_named_array_dt:
                # Named-DataType array with children: register in named_array_key_map.
                # Use out_name as the preferred base name (preserves the DataType identity).
                # Full fingerprint + collision logic applies just like for inline components.
                if descendants:
                    named_array_key_subtrees[key] = descendants
                    ex_sig = _subtree_example_signature(descendants)
                    fp = _fp_inline_subtree(descendants, name_norm, include_rules=(current_ctx_kind == "request"))
                    neutral_fp = _fp_inline_subtree(descendants, name_norm, include_rules=False)
                    existing = None
                    if current_ctx_kind == "request":
                        existing = self.inline_component_fingerprints.get(fp)
                    else:
                        existing = self.inline_component_neutral_fingerprints.get(neutral_fp) or self.inline_component_fingerprints.get(fp)
                    if existing:
                        named_array_items_ref_map[name_norm] = existing
                        named_array_key_map[key] = existing
                        if not self.include_descriptions_in_collision:
                            root_desc = (str(desc).strip() if desc is not None else "")
                            canon_desc = (self.inline_component_root_desc.get(existing) or "").strip()
                            if root_desc and root_desc != canon_desc:
                                self._record_merge_provenance(existing, "description", root_desc, ep_filename or "")
                        if not self.include_examples_in_collision:
                            canon_ex = (self.inline_component_example_sig.get(existing) or "").strip()
                            if ex_sig and ex_sig != canon_ex:
                                self._record_merge_provenance(existing, "example", ex_sig, ep_filename or "")
                    else:
                        # Prefer out_name as schema name; _unique_inline_component_name handles
                        # conflicts if another component already uses the same name.
                        comp_name = _unique_inline_component_name(out_name)
                        named_array_items_ref_map[name_norm] = comp_name
                        named_array_key_map[key] = comp_name
                        self.inline_component_fingerprints[fp] = comp_name
                        self.inline_component_neutral_fingerprints.setdefault(neutral_fp, comp_name)
                        self.inline_component_fingerprint_by_name[comp_name] = fp
                        self.inline_component_root_desc[comp_name] = (str(desc).strip() if desc is not None else "")
                        self.inline_component_example_sig[comp_name] = ex_sig
                        self._clone_schema_variant(dt, comp_name)
                # No-children named-array: not added to named_array_key_map.
                # The property will fall through to primitive-inline behaviour (type=array,
                # items=primitive) in the second pass, which is the desired behaviour for
                # arrays whose item structure is not explicitly defined in the sheet.
                if key in named_array_key_map:
                    _track_inline_component_usage(named_array_key_map[key])
            else:
                # Literal 'array' dtype: existing inline-object-component behaviour.
                if descendants:
                    array_key_subtrees[key] = descendants
                    ex_sig = _subtree_example_signature(descendants)
                    fp = _fp_inline_subtree(descendants, name_norm, include_rules=(current_ctx_kind == "request"))
                    neutral_fp = _fp_inline_subtree(descendants, name_norm, include_rules=False)
                    existing = None
                    if current_ctx_kind == "request":
                        existing = self.inline_component_fingerprints.get(fp)
                    else:
                        existing = self.inline_component_neutral_fingerprints.get(neutral_fp) or self.inline_component_fingerprints.get(fp)
                    if existing:
                        array_items_ref_map[name_norm] = existing
                        array_key_map[key] = existing
                        if not self.include_descriptions_in_collision:
                            root_desc = (str(desc).strip() if desc is not None else "")
                            canon_desc = (self.inline_component_root_desc.get(existing) or "").strip()
                            if root_desc and root_desc != canon_desc:
                                self._record_merge_provenance(existing, "description", root_desc, ep_filename or "")
                        if not self.include_examples_in_collision:
                            canon_ex = (self.inline_component_example_sig.get(existing) or "").strip()
                            if ex_sig and ex_sig != canon_ex:
                                self._record_merge_provenance(existing, "example", ex_sig, ep_filename or "")
                    else:
                        comp_name = _unique_inline_component_name(self._to_pascal_case(name))
                        array_items_ref_map[name_norm] = comp_name
                        array_key_map[key] = comp_name
                        self.inline_component_fingerprints[fp] = comp_name
                        self.inline_component_neutral_fingerprints.setdefault(neutral_fp, comp_name)
                        self.inline_component_fingerprint_by_name[comp_name] = fp
                        self.inline_component_root_desc[comp_name] = (str(desc).strip() if desc is not None else "")
                        self.inline_component_example_sig[comp_name] = ex_sig
                        self._clone_schema_variant(dt, comp_name)
                else:
                    comp_name = _unique_inline_component_name(self._to_pascal_case(name))
                    array_items_ref_map[name_norm] = comp_name
                    array_key_map[key] = comp_name
                    self._clone_schema_variant(dt, comp_name)

                _track_inline_component_usage(array_key_map[key])
        
        # First pass: determine which nodes resolve to a schema reference.
        # Children of such nodes should be suppressed (they live in the component schema).
        schema_ref_nodes = set()
        
        for name, parent, desc, dtype, mandatory, v_rules, items_type_row in children:
            name_norm = _norm(name)
            if name_norm in object_ref_map:
                schema_ref_nodes.add(name_norm)
                continue
            if name_norm in array_items_ref_map:
                schema_ref_nodes.add(name_norm)
                continue
            if name_norm in named_array_items_ref_map:
                schema_ref_nodes.add(name_norm)
                continue
            low_dtype = _norm(dtype)
            out_name, dt = self._resolve_data_type(str(dtype).strip() if dtype is not None else "", ep_filename)
            
            # If it resolves to a named schema (not literal 'object' or 'array'), mark it
            if out_name and low_dtype not in ["object", "array"] and dt and dt.type and dt.type.lower() not in ["object", "array"]:
                schema_ref_nodes.add(name_norm)
                referenced_data_types.add(out_name)
        
        # Build parent lookup (case-insensitive)
        parent_map = {}
        for name, parent, desc, dtype, mandatory, v_rules, items_type_row in children:
            if parent:
                parent_map[_norm(name)] = _norm(parent)
        
        # Second pass: build rows, skipping children of schema-ref nodes
        for name, parent, desc, dtype, mandatory, v_rules, items_type_row in children:
            # Skip this row if its parent is a node that resolves to a schema $ref:
            # the children are already defined in the referenced component schema.
            parent_norm = _norm(parent)
            if parent_norm and parent_norm in schema_ref_nodes:
                continue
            
            # Also skip if any ancestor resolves to a schema ref (deeply nested case)
            skip = False
            ancestor = parent_norm if parent_norm else None
            while ancestor:
                if ancestor in schema_ref_nodes:
                    skip = True
                    break
                ancestor = parent_map.get(ancestor)
            if skip:
                continue
            
            # Determine actual parent
            if legato_parent:
                actual_parent = str(parent).strip() if parent else legato_parent
            else:
                actual_parent = str(parent).strip() if parent else wrapper_name
            
            # Decorate description with validation rules if present (R3.3)
            final_desc = desc
            if v_rules:
                final_desc = f"{desc}\n\n **Validation Rule(s)** {v_rules}" if desc else f" **Validation Rule(s)** {v_rules}"
            
            # Resolve type with recursive inlining rules
            resolved_type = "string"
            schema_name = ""
            items_type = ""
            
            name_out = str(name).strip() if name is not None else ""
            low_dtype = _norm(dtype)
            
            # Try to resolve by dtype first
            out_name, dt = self._resolve_data_type(str(dtype).strip() if dtype is not None else "", ep_filename)
            
            # If dtype is primitive but element name matches any global DataType, prioritize the DataType
            if not dt and low_dtype in ["string", "number", "integer", "boolean"] and name_out:
                # Search for any global DataType that matches this element name
                for global_name, global_dt in self.global_schemas.items():
                    if global_name.lower() == name_out.lower():
                        alt_name = self.output_names.get(("$global", global_name), global_name)
                        out_name, dt = alt_name, global_dt
                        break
            
            # RULE: Only inline if the literal NAME is 'object' or 'array'.
            name_norm = _norm(name)
            parent_norm_emit = _norm(parent)
            key_emit = (name_norm, parent_norm_emit)
            if low_dtype == "object" and key_emit in object_key_map:
                resolved_type = "schema"
                schema_name = object_key_map[key_emit]
            elif low_dtype == "object" and name_norm in object_ref_map:
                resolved_type = "schema"
                schema_name = object_ref_map[name_norm]
            elif low_dtype == "object":
                resolved_type = "object"
                schema_name = ""
            elif low_dtype == "array":
                resolved_type = "array"
                schema_name = ""
                if key_emit in array_key_map:
                    items_type = array_key_map[key_emit]
                elif name_norm in array_items_ref_map:
                    items_type = array_items_ref_map[name_norm]
                else:
                    # Resolve items recursively using the definition of 'array'
                    it_to_resolve = items_type_row
                    if not it_to_resolve and dt:
                        it_to_resolve = dt.items_type
                    items_type = self._recursive_resolve_items(it_to_resolve, dt.source_file if dt and dt.source_file != "$global" else None)
            elif dt and dt.type and dt.type.lower() in ["object", "array"]:
                # Inline resolved object/array schemas to preserve legacy structure
                resolved_type = dt.type.lower()
                schema_name = ""
                if resolved_type == "array":
                    # Named-DataType array with children: emit as $ref to the schema.
                    # The schema (type=array with items.properties) is emitted separately
                    # via named_array_key_map to avoid the double-array anti-pattern.
                    if key_emit in named_array_key_map:
                        resolved_type = "schema"
                        schema_name = named_array_key_map[key_emit]
                        items_type = ""
                        referenced_data_types.add(schema_name)
                    elif name_norm in named_array_items_ref_map:
                        resolved_type = "schema"
                        schema_name = named_array_items_ref_map[name_norm]
                        items_type = ""
                        referenced_data_types.add(schema_name)
                    elif key_emit in array_key_map:
                        items_type = array_key_map[key_emit]
                    elif name_norm in array_items_ref_map:
                        items_type = array_items_ref_map[name_norm]
                    else:
                        # Named-DataType array with no inline children in this
                        # endpoint sheet: emit as $ref so constraints (minItems/
                        # maxItems) from the DataType sheet are preserved.
                        # The component schema is emitted via global_blocks.
                        if out_name and out_name.lower() not in ["array", "object"]:
                            resolved_type = "schema"
                            schema_name = out_name
                            items_type = ""
                            referenced_data_types.add(out_name)
                        else:
                            it_to_resolve = items_type_row
                            if not it_to_resolve:
                                it_to_resolve = dt.items_type or ""
                            items_type = self._recursive_resolve_items(it_to_resolve, dt.source_file if dt and dt.source_file != "$global" else None)
            elif dt:
                resolved_type = "schema"
                schema_name = out_name
                referenced_data_types.add(out_name)
            else:
                # Fallback: only valid OAS primitive types, otherwise schema
                if low_dtype in ["string", "number", "integer", "boolean"]:
                    resolved_type = low_dtype
                else:
                    # Not a valid OAS type and no DataType resolved - treat as schema reference
                    resolved_type = "schema"
                    schema_name = dtype  # Use the dtype as schema name
            
            # Track non-primitive array items as referenced schemas so they
            # get emitted in the global schemas section of the Schemas sheet.
            if resolved_type == "array" and items_type:
                _it_low = items_type.lower()
                if _it_low not in ("string", "number", "integer", "boolean", "object", "array"):
                    referenced_data_types.add(items_type)

            # Convert mandatory
            mand_value = "M" if mandatory in ["M", "m", "Yes", "yes", "Y", "y"] else \
                         "O" if mandatory in ["O", "o", "No", "no", "N", "n"] else mandatory
            
            rows.append([
                name_out, actual_parent, final_desc, resolved_type,
                items_type, schema_name, "", mand_value,
                "", "", "", "", "", "", "" # Col J-O: Min, Max, PatternEba, Regex, Allowed, Example
            ])

        for (prop_low, prop_parent_low), comp_name in object_key_map.items():
            subtree = object_key_subtrees.get((prop_low, prop_parent_low))
            if not subtree:
                continue
            if comp_name in self.emitted_inline_components:
                continue
            self.emitted_inline_components.add(comp_name)

            # Track parent -> child edge so usage can be propagated transitively
            if wrapper_name:
                self.inline_component_children.setdefault(wrapper_name, set()).add(comp_name)

            transformed: List[Tuple] = []
            for t_name, t_parent, t_desc, t_dtype, t_mand, t_rules, t_items in subtree:
                new_parent = "" if _norm(t_parent) == prop_low else t_parent
                transformed.append((t_name, new_parent, t_desc, t_dtype, t_mand, t_rules, t_items))

            child_rows, _refs, nested_blocks = self._build_children_rows(comp_name, transformed, ep_filename, usage_ctx=usage_ctx)
            referenced_data_types.update(_refs)
            extra_schema_blocks.append([comp_name, "", "", "object", "", "", "", "", "", "", "", "", "", ""])
            extra_schema_blocks.extend(child_rows)
            extra_schema_blocks.extend(nested_blocks)

        for (prop_low, prop_parent_low), comp_name in array_key_map.items():
            subtree = array_key_subtrees.get((prop_low, prop_parent_low))
            if not subtree:
                continue
            if comp_name in self.emitted_inline_components:
                continue
            self.emitted_inline_components.add(comp_name)

            # Track parent -> child edge so usage can be propagated transitively
            if wrapper_name:
                self.inline_component_children.setdefault(wrapper_name, set()).add(comp_name)

            transformed: List[Tuple] = []
            for t_name, t_parent, t_desc, t_dtype, t_mand, t_rules, t_items in subtree:
                new_parent = "" if _norm(t_parent) == prop_low else t_parent
                transformed.append((t_name, new_parent, t_desc, t_dtype, t_mand, t_rules, t_items))

            child_rows, _refs, nested_blocks = self._build_children_rows(comp_name, transformed, ep_filename, usage_ctx=usage_ctx)
            referenced_data_types.update(_refs)
            extra_schema_blocks.append([comp_name, "", "", "object", "", "", "", "", "", "", "", "", "", ""])
            extra_schema_blocks.extend(child_rows)
            extra_schema_blocks.extend(nested_blocks)

        for (prop_low, prop_parent_low), comp_name in named_array_key_map.items():
            subtree = named_array_key_subtrees.get((prop_low, prop_parent_low))
            if not subtree:
                continue
            if comp_name in self.emitted_inline_components:
                continue
            self.emitted_inline_components.add(comp_name)

            # Track parent -> child edge so usage can be propagated transitively
            if wrapper_name:
                self.inline_component_children.setdefault(wrapper_name, set()).add(comp_name)

            transformed: List[Tuple] = []
            for t_name, t_parent, t_desc, t_dtype, t_mand, t_rules, t_items in subtree:
                new_parent = "" if _norm(t_parent) == prop_low else t_parent
                transformed.append((t_name, new_parent, t_desc, t_dtype, t_mand, t_rules, t_items))

            child_rows, _refs, nested_blocks = self._build_children_rows(comp_name, transformed, ep_filename, usage_ctx=usage_ctx)
            referenced_data_types.update(_refs)

            # Determine the primitive items type of the named-array DataType so we
            # can emit it correctly as  type=array, items=<primitive>.
            dt_for_schema = self.global_schemas.get(comp_name)
            items_primitive = "object"
            if dt_for_schema and dt_for_schema.items_type:
                items_primitive = self._recursive_resolve_items(
                    dt_for_schema.items_type,
                    dt_for_schema.source_file if dt_for_schema.source_file != "$global" else None,
                )

            # Emit root row as type=array (NOT object) so generator produces
            # the correct  {type: array, items: {properties: ...}}  schema.
            # Carry minItems/maxItems from the original DataType definition.
            _na_min = dt_for_schema.min_val if dt_for_schema else ""
            _na_max = dt_for_schema.max_val if dt_for_schema else ""
            extra_schema_blocks.append([comp_name, "", "", "array", items_primitive, "", "", "", _na_min, _na_max, "", "", "", ""])
            extra_schema_blocks.extend(child_rows)
            extra_schema_blocks.extend(nested_blocks)

        return rows, referenced_data_types, extra_schema_blocks

    
    # =========================================================================
    # Body Example Generation
    # =========================================================================

    def _get_example_values(self, dt: Optional['DataType']) -> List[str]:
        """Return list of example values from a DataType (split by ';'), or [] if none."""
        if dt is None:
            return []
        raw = dt.example or ""
        if not raw or raw.lower() == "nan":
            # Fallback: first allowed value
            av = dt.allowed_values or ""
            if av and av.lower() != "nan":
                first = re.split(r'[;,]', av)[0].strip()
                if first:
                    return [first]
            return []
        parts = [p.strip() for p in raw.split(";") if p.strip()]
        return parts if parts else []

    def _generate_synthetic_value(self, dtype_lower: str) -> Any:
        """Return a type-appropriate synthetic fallback value."""
        if dtype_lower in ("integer", "int"):
            return 0
        if dtype_lower in ("number", "float", "double", "decimal"):
            return 0.0
        if dtype_lower == "boolean":
            return True
        return "string"

    def _resolve_leaf_dt(self, dtype: str, ep_filename: Optional[str]) -> Optional['DataType']:
        """Resolve a field type to its DataType, or None for primitives."""
        if self._is_oas_primitive_name(dtype):
            return None
        _, dt = self._resolve_data_type(dtype, ep_filename)
        return dt

    def _build_body_example_dict(
        self,
        children: List[Tuple],
        ep_filename: Optional[str],
        omit_field: Optional[str] = None,
        override_field: Optional[str] = None,
        override_value: Any = None,
        override_fields: Optional[Dict[str, Any]] = None,
        array_item_index: int = 0,
    ) -> dict:
        """
        Build a Python dict representing one body example from the legacy structure.

        children  : list of (name, parent, desc, dtype, mandatory, v_rules, items_type)
        omit_field: if set, skip this field name (for Bad Request missing-mandatory)
        override_field / override_value: replace value for one field (other Bad Request violations)
        array_item_index: 0 = first array item, 1 = second array item (shifts rotation)
        """
        # Per-type rotation counter (shared across the whole call so siblings rotate)
        type_counters: Dict[str, int] = {}

        def _coerce_numeric_example(raw_value: Any, numeric_type: str) -> Any:
            """Preserve integer-looking number examples without forcing a trailing '.0'."""
            raw_text = str(raw_value).strip()
            if not raw_text:
                return raw_value

            normalized = raw_text.replace(",", ".")

            if numeric_type in ("integer", "int"):
                try:
                    return int(normalized)
                except (ValueError, TypeError):
                    return raw_value

            if numeric_type in ("number", "float", "double", "decimal"):
                # A numeric example that is lexically an integer should remain an integer
                # in generated YAML instead of being normalized to '<n>.0'.
                if re.fullmatch(r"[+-]?\d+", normalized):
                    try:
                        return int(normalized)
                    except (ValueError, TypeError):
                        return raw_value
                try:
                    return float(normalized)
                except (ValueError, TypeError):
                    return raw_value

            return raw_value

        def get_rotated_value(dtype: str, dt: Optional['DataType'], item_offset: int = 0, is_array_item: bool = False) -> Any:
            """Pick example value using per-type rotation + item_offset for array diversity."""
            examples = self._get_example_values(dt)
            low = dtype.lower() if dtype else "string"

            if examples:
                # For array items, return the full examples list (or subset) instead of single value
                if is_array_item:
                    count = type_counters.get(low, 0)
                    # Return all examples as array, or single example if only one exists
                    if len(examples) >= 2:
                        # Rotate starting point for diversity across different arrays
                        start_idx = (count + item_offset) % len(examples)
                        type_counters[low] = count + 1
                        return [examples[(start_idx + i) % len(examples)] for i in range(min(2, len(examples)))]
                    return [examples[0]]
                
                count = type_counters.get(low, 0)
                idx = (count + item_offset) % len(examples)
                type_counters[low] = count + 1
                raw = examples[idx]
                # Coerce to numeric if the DataType says so
                if dt:
                    if dt.type.lower() in ("integer", "int", "number", "float", "double", "decimal"):
                        coerced = _coerce_numeric_example(raw, dt.type.lower())
                        if coerced is not raw:
                            return coerced
                    elif dt.type.lower() == "boolean":
                        return raw.lower() in ("true", "yes", "1")
                return raw
            else:
                # Synthetic fallback
                if dt:
                    low_t = dt.type.lower()
                else:
                    low_t = low
                synthetic = self._generate_synthetic_value(low_t)
                return [synthetic] if is_array_item else synthetic

        # Build parent → children tree.
        # IMPORTANT: legacy structures can contain duplicate field names under different parents.
        # Use unique node ids internally to avoid overwriting nodes by name.

        node_parent: Dict[str, str] = {}
        node_name: Dict[str, str] = {}
        node_dtype: Dict[str, str] = {}
        node_mandatory: Dict[str, str] = {}
        node_items: Dict[str, str] = {}
        ordered_uids: List[str] = []

        name_to_last_uid: Dict[str, str] = {}
        for i, (name, parent, desc, dtype, mandatory, v_rules, items_type) in enumerate(children):
            uid = f"{name}__{i}"
            node_name[uid] = name
            node_dtype[uid] = dtype or "string"
            node_mandatory[uid] = mandatory or ""
            node_items[uid] = items_type or ""

            parent_name = parent or ""
            parent_uid = name_to_last_uid.get(parent_name, "") if parent_name else ""
            node_parent[uid] = parent_uid
            ordered_uids.append(uid)

            # Update last seen uid for this name
            name_to_last_uid[name] = uid

        def is_root(uid: str) -> bool:
            return not node_parent.get(uid, "")

        children_of: Dict[str, List[str]] = {}
        for uid in ordered_uids:
            p_uid = node_parent.get(uid, "")
            if p_uid not in children_of:
                children_of[p_uid] = []
            children_of[p_uid].append(uid)

        def build_node(uid: str, item_offset: int = 0) -> Any:
            dtype = node_dtype.get(uid, "string")
            low_dtype = dtype.lower()
            dt = self._resolve_leaf_dt(dtype, ep_filename)
            my_children = children_of.get(uid, [])

            # --- ARRAY ---
            if low_dtype == "array" or (dt and dt.type.lower() == "array"):
                items_type_name = node_items.get(uid, "")
                if not items_type_name and dt:
                    items_type_name = dt.items_type or ""
                items_dt = self._resolve_leaf_dt(items_type_name, ep_filename) if items_type_name else None
                items_low = items_type_name.lower() if items_type_name else "object"

                if my_children:
                    # Array of objects defined inline
                    item1 = {
                        node_name[c]: build_node(c, item_offset=0)
                        for c in my_children
                        if node_name.get(c) != omit_field or override_field is not None
                    }
                    # Check if we have enough examples for a second distinct item
                    can_make_second = _can_diversify(my_children, item_offset=1)
                    if can_make_second:
                        item2 = {
                            node_name[c]: build_node(c, item_offset=1)
                            for c in my_children
                            if node_name.get(c) != omit_field or override_field is not None
                        }
                        return [item1, item2]
                    return [item1]
                elif items_low == "object" or (items_dt and items_dt.type.lower() == "object"):
                    # Array of objects but no inline children — use empty dict
                    return [{}]
                else:
                    # Array of scalars
                    v1 = get_rotated_value(items_type_name or items_low, items_dt, item_offset=0, is_array_item=True)
                    if isinstance(v1, list) and len(v1) >= 2:
                        # Already got multiple values from examples list
                        return v1
                    # Try to get a second distinct value for diversity
                    v2 = get_rotated_value(items_type_name or items_low, items_dt, item_offset=1, is_array_item=True)
                    if isinstance(v2, list):
                        return [v1[0] if isinstance(v1, list) else v1, v2[0]]
                    return [v1[0] if isinstance(v1, list) else v1, v2] if v2 is not None else [v1[0] if isinstance(v1, list) else v1]

            # --- OBJECT ---
            if low_dtype == "object" or (dt and dt.type.lower() == "object" and my_children):
                result = {}
                for c in my_children:
                    cname = node_name.get(c)
                    if cname == omit_field:
                        continue
                    val = build_node(c, item_offset=item_offset)
                    if cname == override_field:
                        val = override_value
                    if override_fields and cname in override_fields:
                        val = override_fields.get(cname)
                    result[cname] = val
                return result

            # --- SCALAR (possibly a named schema that resolves to a scalar type) ---
            val = get_rotated_value(dtype, dt, item_offset=item_offset)
            if node_name.get(uid) == override_field:
                val = override_value
            if override_fields and node_name.get(uid) in override_fields:
                val = override_fields.get(node_name.get(uid))
            return val

        def _can_diversify(child_uids: List[str], item_offset: int) -> bool:
            """Check whether at least one child has >1 example value (enables 2nd array item)."""
            for c in child_uids:
                dtype = node_dtype.get(c, "string")
                dt = self._resolve_leaf_dt(dtype, ep_filename)
                examples = self._get_example_values(dt)
                if len(examples) >= 2:
                    return True
            return False

        # Step 3: build root dict
        result = {}
        for uid in ordered_uids:
            if not is_root(uid):
                continue
            name = node_name.get(uid)
            if name == omit_field:
                continue
            val = build_node(uid, item_offset=array_item_index)
            if name == override_field:
                val = override_value
            if override_fields and name in override_fields:
                val = override_fields.get(name)
            result[name] = val

        return result

    def _pick_bad_request_violation(
        self,
        children: List[Tuple],
        ep_filename: Optional[str],
    ) -> Tuple[str, Optional[str], Any]:
        """
        Scan body children and choose the best Bad Request violation.
        Returns (violation_type, field_name, bad_value).
        violation_type: 'omit' | 'enum' | 'range' | 'regex' | 'type'
        """
        mandatory_fields = []
        enum_fields = []
        range_fields = []
        regex_fields = []
        any_field = None

        for name, parent, desc, dtype, mandatory, v_rules, items_type in children:
            dt = self._resolve_leaf_dt(dtype, ep_filename)
            any_field = (name, dtype, dt)

            mand = mandatory.upper() if mandatory else ""
            if mand == "M":
                mandatory_fields.append(name)

            if dt:
                if dt.allowed_values and dt.allowed_values.lower() != "nan":
                    enum_fields.append((name, dt))
                if (dt.min_val and dt.min_val.lower() != "nan") or \
                   (dt.max_val and dt.max_val.lower() != "nan"):
                    range_fields.append((name, dt))
                if dt.regex and dt.regex.lower() != "nan":
                    regex_fields.append((name, dt))

        # Priority 1: omit mandatory
        if mandatory_fields:
            return ("omit", mandatory_fields[0], None)

        # Priority 2: invalid enum value
        if enum_fields:
            name, dt = enum_fields[0]
            return ("enum", name, "INVALID_VALUE")

        # Priority 3: out-of-range value
        if range_fields:
            name, dt = range_fields[0]
            low_t = dt.type.lower() if dt else "string"
            if dt.max_val and dt.max_val.lower() != "nan":
                try:
                    bad = int(float(dt.max_val)) + 1
                    return ("range", name, bad)
                except (ValueError, TypeError):
                    pass
            if dt.min_val and dt.min_val.lower() != "nan":
                try:
                    bad = int(float(dt.min_val)) - 1
                    return ("range", name, bad)
                except (ValueError, TypeError):
                    pass

        # Priority 4: regex violation
        if regex_fields:
            name, dt = regex_fields[0]
            return ("regex", name, "INVALID")

        # Priority 5: wrong type fallback
        if any_field:
            name, dtype, dt = any_field
            low_t = (dt.type.lower() if dt else dtype.lower()) if dtype else "string"
            if low_t in ("integer", "number", "float", "double", "decimal"):
                return ("type", name, "not_a_number")
            else:
                return ("type", name, 12345)

        return ("type", None, None)

    def _convert_body_example(self, wb, xl_legacy, op_id: str, ep_filename: str):
        """Generate and write the Body Example sheet with OK and Bad Request examples."""
        if "Body Example" not in wb.sheetnames:
            return

        if "Body" not in xl_legacy.sheet_names:
            return

        children = self._read_legacy_structure(xl_legacy, "Body")
        if not children:
            return

        ws = wb["Body Example"]

        # --- Build OK example ---
        ok_dict = self._build_body_example_dict(children, ep_filename)
        ok_yaml = yaml.dump(
            ok_dict,
            default_flow_style=False,
            allow_unicode=True,
            sort_keys=False,
        ).rstrip()

        # --- Build Bad Request example ---
        violation_type, viol_field, bad_val = self._pick_bad_request_violation(children, ep_filename)

        if violation_type == "omit":
            bad_dict = self._build_body_example_dict(
                children, ep_filename, omit_field=viol_field
            )
        else:
            bad_dict = self._build_body_example_dict(
                children, ep_filename,
                override_field=viol_field,
                override_value=bad_val,
            )

        bad_yaml = yaml.dump(
            bad_dict,
            default_flow_style=False,
            allow_unicode=True,
            sort_keys=False,
        ).rstrip()

        # --- Write to sheet (col A = name, col B = body YAML) ---
        examples = [
            ("OK", ok_yaml),
            ("Bad Request", bad_yaml),
        ]
        for row_idx, (ex_name, ex_body) in enumerate(examples, start=2):
            cell_name = ws.cell(row=row_idx, column=1, value=ex_name)
            cell_body = ws.cell(row=row_idx, column=2, value=ex_body)
            cell_name.alignment = Alignment(vertical="top")
            cell_body.alignment = Alignment(vertical="top", wrap_text=True)

    def _recursive_resolve_items(self, items_type_name: str, source_context: Optional[str]) -> str:
        """Recursively resolve items_type. If it resolves to object or array, return primitive."""
        if not items_type_name:
            return "object"
            
        low_it = items_type_name.lower()
        if low_it in ["string", "number", "integer", "boolean", "object", "array"]:
            return low_it
            
        it_out, it_dt = self._resolve_data_type(items_type_name, source_context)
        if it_dt:
            if it_dt.type.lower() == "object":
                return "object"
            elif it_dt.type.lower() == "array":
                return "array"
            else:
                return it_out or items_type_name
        return it_out or items_type_name

    def _parse_params(self, df: pd.DataFrame, param_in: str) -> List[Dict]:
        """Parse parameters from Path/Header sheet."""
        header_row_idx = self._find_header_row(df, ["element", "name", "type"])
        
        if header_row_idx == -1:
            return []
        
        df.columns = df.iloc[header_row_idx]
        df = df.iloc[header_row_idx + 1:].reset_index(drop=True)
        
        params = []
        for _, row in df.iterrows():
            name = self._clean_value(row.get("Element", row.get("Request Header", row.get("Name", ""))))
            if not name or name.lower() in ["nan", "element", "name"]:
                continue
            
            desc = self._clean_value(row.get("Description ", row.get("Description", "")))
            dtype = self._clean_value(row.get("Type", "string"))
            mandatory = self._clean_value(row.get("Mandatory", ""))
            v_rules = self._clean_value(row.get("Validation Rules", ""))
            
            mand_value = "M" if mandatory in ["M", "m", "Yes", "yes"] else "O"
            
            final_desc = desc
            if v_rules:
                final_desc = f"{desc}\n\n **Validation Rule(s)** {v_rules}" if desc else f" **Validation Rule(s)** {v_rules}"
            
            params.append({
                "name": name,
                "description": final_desc,
                "in": param_in,
                "type": dtype,
                "mandatory": mand_value
            })
        
        return params

    def _pick_first_example_value(self, raw: str) -> str:
        if raw is None:
            return ""
        s = str(raw).strip()
        if not s or s.lower() == "nan":
            return ""
        for sep in [";", "|", ",", "\n"]:
            if sep in s:
                parts = [p.strip() for p in s.split(sep) if p.strip()]
                if parts:
                    return parts[0]
        return s

    def _read_legacy_params_with_examples(self, xl_legacy, sheet_name: str) -> List[Dict[str, str]]:
        if sheet_name not in xl_legacy.sheet_names:
            return []

        try:
            df = pd.read_excel(xl_legacy, sheet_name=sheet_name, dtype=str, header=None)
        except Exception:
            return []

        header_row_idx = self._find_header_row(df, ["element", "name", "type"])
        if header_row_idx == -1:
            return []

        headers = [str(c).strip().lower() for c in df.iloc[header_row_idx].values]
        df = df.iloc[header_row_idx + 1:].reset_index(drop=True)
        df.columns = headers

        def _get_col(row, keys: List[str]) -> str:
            for k in keys:
                if k in df.columns:
                    return self._clean_value(row.get(k))
            # Partial match fallback
            for c in df.columns:
                for k in keys:
                    if k in str(c):
                        return self._clean_value(row.get(c))
            return ""

        out: List[Dict[str, str]] = []
        for _, row in df.iterrows():
            name = _get_col(row, ["element", "request header", "name"])
            if not name or name.lower() in ["nan", "element", "name"]:
                continue

            dtype = _get_col(row, ["type", "data type"])
            example_raw = _get_col(row, ["example"])
            out.append(
                {
                    "name": name,
                    "type": dtype or "string",
                    "example": self._pick_first_example_value(example_raw),
                }
            )
        return out

    def _find_last_used_row(self, ws, max_scan_cols: int = 6) -> int:
        try:
            last = ws.max_row
        except Exception:
            last = 1
        for r in range(last, 0, -1):
            for c in range(1, max_scan_cols + 1):
                try:
                    if ws.cell(row=r, column=c).value not in (None, ""):
                        return r
                except Exception:
                    continue
        return 1

    def _add_ok_response_example(
        self,
        ws,
        xl_legacy,
        response_code: str,
        response_children: List[Tuple],
        ep_filename: str,
    ) -> None:
        try:
            for r in range(1, min(ws.max_row, 200) + 1):
                if str(ws.cell(row=r, column=1).value).strip().lower() == "examples" and \
                   str(ws.cell(row=r, column=2).value).strip() == "OK":
                    return
        except Exception:
            pass

        try:
            ok_dict = self._build_body_example_dict(response_children, ep_filename)
            ok_yaml = yaml.dump(
                ok_dict,
                default_flow_style=False,
                allow_unicode=True,
                sort_keys=False,
            ).rstrip()
        except Exception:
            ok_yaml = ""

        path_params = self._read_legacy_params_with_examples(xl_legacy, "Path")
        headers = self._read_legacy_params_with_examples(xl_legacy, "Header")

        def _param_example(name: str, raw_ex: str) -> str:
            if name.strip().lower() == "senderbic":
                return f"TSTBICXX{response_code}"
            return raw_ex or ""

        rows: List[List[Any]] = []

        rows.append(["examples", "OK", "application/json", "", "", "", "", "", "", "", "", "", "", "", ""])

        if path_params:
            rows.append(["examples", "x-sandbox-request-path-params", "OK", "", "object", "", "", "", "", "", "", "", "", "", ""])
            for p in path_params:
                pname = p.get("name", "")
                exv = _param_example(pname, p.get("example", ""))
                rows.append(["examples", pname, "x-sandbox-request-path-params", "", "string", "", "", "", "", "", "", "", "", "", exv])

        if headers:
            rows.append(["examples", "x-sandbox-request-headers", "OK", "", "object", "", "", "", "", "", "", "", "", "", ""])
            for h in headers:
                hname = h.get("name", "")
                exv = h.get("example", "")
                rows.append(["examples", hname, "x-sandbox-request-headers", "", "string", "", "", "", "", "", "", "", "", "", exv])

        rows.append(["examples", "value", "OK", "", "object", "", "", "", "", "", "", "", "", "", ok_yaml])

        start_row = self._find_last_used_row(ws) + 1
        self._write_rows(ws, rows, start_row=start_row)

    def _add_400_response_example(
        self,
        ws,
        xl_legacy,
        response_children: List[Tuple],
        ep_filename: str,
    ) -> None:
        try:
            for r in range(1, min(ws.max_row, 200) + 1):
                if str(ws.cell(row=r, column=1).value).strip().lower() == "examples" and \
                   str(ws.cell(row=r, column=2).value).strip().lower() == "bad request":
                    return
        except Exception:
            pass

        try:
            bad_dict = self._build_body_example_dict(
                response_children,
                ep_filename,
                override_fields={
                    "errorCode": "SC01",
                    "errorCodeDescription": "Failed Parameter Validation",
                },
            )
            bad_yaml = yaml.dump(
                bad_dict,
                default_flow_style=False,
                allow_unicode=True,
                sort_keys=False,
            ).rstrip()
        except Exception:
            bad_yaml = ""

        path_params = self._read_legacy_params_with_examples(xl_legacy, "Path")
        headers = self._read_legacy_params_with_examples(xl_legacy, "Header")

        def _param_example(name: str, raw_ex: str) -> str:
            if name.strip().lower() == "senderbic":
                return "TSTBICXX400"
            return raw_ex or ""

        rows: List[List[Any]] = []
        rows.append(["examples", "Bad Request", "application/json", "", "", "", "", "", "", "", "", "", "", "", ""])

        if path_params:
            rows.append(["examples", "x-sandbox-request-path-params", "Bad Request", "", "object", "", "", "", "", "", "", "", "", "", ""])
            for p in path_params:
                pname = p.get("name", "")
                exv = _param_example(pname, p.get("example", ""))
                rows.append(["examples", pname, "x-sandbox-request-path-params", "", "string", "", "", "", "", "", "", "", "", "", exv])

        if headers:
            rows.append(["examples", "x-sandbox-request-headers", "Bad Request", "", "object", "", "", "", "", "", "", "", "", "", ""])
            for h in headers:
                hname = h.get("name", "")
                exv = h.get("example", "") or ""
                rows.append(["examples", hname, "x-sandbox-request-headers", "", "string", "", "", "", "", "", "", "", "", "", exv])

        rows.append(["examples", "value", "Bad Request", "", "object", "", "", "", "", "", "", "", "", "", bad_yaml])

        start_row = self._find_last_used_row(ws) + 1
        self._write_rows(ws, rows, start_row=start_row)
    
    def _extract_wrapper_base(self, filename: str) -> str:
        """Extract wrapper base from the API path's last non-parameter segment.
        
        E.g. /files/fileDetails/{senderBic} → fileDetails
        Falls back to filename-based extraction if no path mapping exists.
        """
        # Try path-based extraction first (preferred)
        api_path = self.filename_to_path.get(filename)
        if not api_path:
            # Try without extension variants
            stem_key = filename.replace(".xlsm", "").replace(".xlsx", "")
            api_path = self.filename_to_path.get(stem_key)
        if api_path:
            segments = [s for s in api_path.split('/') if s and not s.startswith('{')]
            if segments:
                return segments[-1]
        # Fallback: derive from filename
        base = filename.replace(".xlsm", "").replace(".xlsx", "")
        parts = base.rsplit(".", 1)
        stem = parts[0] if len(parts) > 1 and parts[1].isdigit() else base
        if "_" in stem:
            stem = stem.split("_", 1)[1]
        return stem
    
    def _find_header_row(self, df: pd.DataFrame, keywords: List[str]) -> int:
        """Find row containing header keywords by scoring matches."""
        best_idx = -1
        max_score = 0
        
        # Search first 15 rows for the best header candidate
        for idx in range(min(15, len(df))):
            row_vals = [str(v).lower().strip() for v in df.iloc[idx].values if v is not None]
            
            # Score this row based on how many keywords it matches
            score = 0
            for kw in keywords:
                # Check if keyword is present as a standalone part of any cell value
                if any(kw in val for val in row_vals):
                    score += 1
            
            if score > max_score:
                max_score = score
                best_idx = idx
        
        # Confidence threshold: at least 2 keywords must match
        return best_idx if max_score >= 2 else -1
    
    def _find_column(self, df: pd.DataFrame, keywords: List[str]) -> Optional[str]:
        """Find column by keywords."""
        for col in df.columns:
            col_lower = str(col).lower()
            if any(kw in col_lower for kw in keywords):
                return col
        return None
    
    def _to_pascal_case(self, name: str) -> str:
        """PascalCase conversion — ALWAYS applied to ALL schema names.
        This is the single source of truth for schema name capitalization.
        Wrapper names (Request/Response) use _to_wrapper_case instead."""
        if not name: return name
        name = str(name).strip()
        if not name or name.lower() == "nan": return name
        if "_" in name or "-" in name:
            import re
            parts = re.split(r'[^a-zA-Z0-9]+', name)
            return "".join(p.capitalize() for p in parts if p)
        return name[0].upper() + name[1:]

    def _to_wrapper_case(self, name: str) -> str:
        """PascalCase for wrapper names (Request/Response), controlled by
        the capitalize_schema_names preference. When the flag is off,
        camelCase is enforced (leading uppercase run is lowered)."""
        if not name: return name
        name = str(name).strip()
        if not self.capitalize_schema_names:
            # camelCase: lowercase leading uppercase run, preserving word boundary
            i = 0
            while i < len(name) and name[i].isupper():
                i += 1
            if i == 0:
                return name
            if i == 1:
                return name[0].lower() + name[1:]
            if i >= len(name):
                return name.lower()
            return name[:i-1].lower() + name[i-1:]
        return self._to_pascal_case(name)
    
    def _clean_value(self, val) -> str:
        """Clean cell value."""
        s = str(val).strip()
        return "" if s.lower() == "nan" else s

    def _is_oas_primitive_name(self, type_name: Any, allow_schema: bool = False) -> bool:
        """
        Return True only for actual lowercase OAS primitive literals.

        Named DataTypes such as 'Number', 'Boolean', 'String' must NOT be treated
        as primitives just because their lowercase form matches an OAS primitive.
        """
        if type_name is None:
            return False
        raw = str(type_name).strip()
        if not raw:
            return True
        primitives = {"string", "number", "integer", "boolean", "array", "object"}
        if allow_schema:
            primitives.add("schema")
        return raw in primitives

    def _close_excel_file(self, excel_file) -> None:
        """Best-effort close for pandas/openpyxl-backed Excel readers."""
        if excel_file is None:
            return
        try:
            close = getattr(excel_file, "close", None)
            if callable(close):
                close()
        except Exception:
            pass
        try:
            book = getattr(excel_file, "book", None)
            close_book = getattr(book, "close", None)
            if callable(close_book):
                close_book()
        except Exception:
            pass

    def _close_workbook(self, workbook) -> None:
        """Best-effort close for openpyxl workbooks."""
        if workbook is None:
            return
        try:
            workbook.close()
        except Exception:
            pass
    
    def _ensure_xlsx_extension(self, filename) -> str:
        """Ensure filename has .xlsx extension."""
        fname = str(filename).strip()
        if not fname or fname.lower() == "nan":
            return fname
        if not fname.endswith(".xlsx") and not fname.endswith(".xlsm"):
            return fname + ".xlsx"
        if fname.endswith(".xlsm"):
            return fname.replace(".xlsm", ".xlsx")
        return fname
    
    def _write_rows(self, ws, rows: List[List], start_row: int = 3):
        """Write rows to worksheet."""
        for r_idx, row in enumerate(rows, start=start_row):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.value = value if value != "" else None
                try:
                    if cell.alignment:
                        cell.alignment = cell.alignment.copy(vertical="top")
                    else:
                        cell.alignment = Alignment(vertical="top")
                except Exception:
                    cell.alignment = Alignment(vertical="top")

    def _autofit_columns(self, ws, max_width: int = 60):
        """Adjust column widths based on content with a maximum limit and forced wrap."""
        try:
            for col in ws.columns:
                max_len = 0
                # Robustly get column letter (col[0] could be a MergedCell)
                first_cell = col[0]
                column_letter = get_column_letter(first_cell.column)
                
                for cell in col:
                    try:
                        if cell.value:
                            # Calculate max line length in the cell
                            lines = str(cell.value).split('\n')
                            for line in lines:
                                max_len = max(max_len, len(line))
                    except:
                        continue
                
                # Padding
                target_width = max_len + 2
                
                if target_width > max_width:
                    ws.column_dimensions[column_letter].width = max_width
                    # Apply wrap only if content is long
                    for cell in col:
                        # MergedCell does not support direct style assignment in some contexts
                        if not hasattr(cell, 'alignment'): continue
                        try:
                            if cell.alignment:
                                cell.alignment = cell.alignment.copy(wrapText=True, vertical="top")
                            else:
                                cell.alignment = Alignment(wrapText=True, vertical="top")
                        except Exception:
                            cell.alignment = Alignment(wrapText=True, vertical="top")
                else:
                    ws.column_dimensions[column_letter].width = max(target_width, 10)
                    # Keep vertical alignment
                    for cell in col:
                        if not hasattr(cell, 'alignment'): continue
                        if not cell.alignment or not cell.alignment.wrapText:
                            try:
                                if cell.alignment:
                                    cell.alignment = cell.alignment.copy(vertical="top")
                                else:
                                    cell.alignment = Alignment(vertical="top")
                            except Exception:
                                cell.alignment = Alignment(vertical="top")
        except Exception as e:
            self.log(f"  Note: Auto-fit skipped for sheet '{ws.title}': {e}")
