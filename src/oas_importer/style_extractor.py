"""
Style Extractor Module

Extracts styles (fonts, fills, borders, column widths) from reference Excel templates
and saves them to a JSON file for reuse when generating new templates.

Usage:
    extractor = StyleExtractor()
    extractor.extract_from_templates('API Templates/')
    extractor.save('src/oas_importer/styles_config.json')
"""

import json
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from typing import Dict, List, Any, Optional
from dataclasses import dataclass, asdict
from copy import copy


@dataclass
class CellStyle:
    """Represents a cell's complete style configuration."""
    font_name: Optional[str] = None
    font_size: Optional[int] = None
    font_bold: bool = False
    font_italic: bool = False
    font_color: Optional[str] = None
    
    fill_type: Optional[str] = None
    fill_fg_color: Optional[str] = None
    fill_bg_color: Optional[str] = None
    
    border_left: Optional[str] = None
    border_right: Optional[str] = None
    border_top: Optional[str] = None
    border_bottom: Optional[str] = None
    border_color: Optional[str] = None
    
    alignment_horizontal: Optional[str] = None
    alignment_vertical: Optional[str] = None
    alignment_wrap_text: bool = False
    
    def to_dict(self) -> Dict:
        """Convert to dictionary, excluding None values."""
        return {k: v for k, v in asdict(self).items() if v is not None and v is not False}


@dataclass  
class SheetStyleConfig:
    """Style configuration for a specific sheet type."""
    sheet_name: str
    header_row_style: CellStyle
    data_row_style: CellStyle
    column_widths: Dict[str, float]
    header_labels: List[str]
    freeze_panes: Optional[str] = None


class StyleExtractor:
    """
    Extracts and manages Excel template styles.
    
    Provides functionality to:
    1. Extract styles from reference templates
    2. Save styles to a JSON configuration file
    3. Load styles from configuration for template generation
    """
    
    def __init__(self):
        self.styles: Dict[str, Any] = {
            'index': {},      # Styles for $index.xlsx sheets
            'operation': {},  # Styles for operation files
            'metadata': {
                'extracted_from': [],
                'version': '1.0'
            }
        }
    
    def extract_from_templates(self, templates_dir: str) -> None:
        """
        Extract styles from reference templates in the given directory.
        
        Args:
            templates_dir: Path to directory containing reference .xlsm files
        """
        # Extract from $index.xlsm
        index_path = os.path.join(templates_dir, '$index.xlsm')
        if os.path.exists(index_path):
            self._extract_index_styles(index_path)
            self.styles['metadata']['extracted_from'].append('$index.xlsm')
        
        # Extract from a sample operation file
        for f in os.listdir(templates_dir):
            if f.endswith('.xlsm') and not f.startswith('$') and not f.startswith('~'):
                op_path = os.path.join(templates_dir, f)
                self._extract_operation_styles(op_path)
                self.styles['metadata']['extracted_from'].append(f)
                break  # Only need one sample
    
    def _extract_index_styles(self, filepath: str) -> None:
        """Extract styles from $index.xlsm file."""
        wb = load_workbook(filepath, data_only=False)
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheet_config = self._extract_sheet_styles(ws, sheet_name)
            self.styles['index'][sheet_name] = sheet_config
        
        wb.close()
    
    def _extract_operation_styles(self, filepath: str) -> None:
        """Extract styles from an operation .xlsm file."""
        wb = load_workbook(filepath, data_only=False)
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheet_config = self._extract_sheet_styles(ws, sheet_name)
            
            # Categorize: Parameters, Body, Body Example, or response code
            if sheet_name == 'Parameters':
                self.styles['operation']['Parameters'] = sheet_config
            elif sheet_name == 'Body':
                self.styles['operation']['Body'] = sheet_config
            elif sheet_name == 'Body Example':
                self.styles['operation']['Body Example'] = sheet_config
            elif sheet_name.isdigit():
                # Response sheet (200, 400, etc.) - use generic 'Response' key
                if 'Response' not in self.styles['operation']:
                    self.styles['operation']['Response'] = sheet_config
        
        wb.close()
    
    def _extract_sheet_styles(self, ws, sheet_name: str) -> Dict[str, Any]:
        """Extract complete style configuration from a worksheet."""
        config = {
            'sheet_name': sheet_name,
            'column_widths': {},
            'header_labels': [],
            'header_row_style': None,
            'data_row_style': None,
            'freeze_panes': str(ws.freeze_panes) if ws.freeze_panes else None,
            'row_height_header': None,
            'row_height_data': None
        }
        
        # Extract column widths
        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            if col_letter in ws.column_dimensions:
                width = ws.column_dimensions[col_letter].width
                if width:
                    config['column_widths'][col_letter] = width
        
        # Find header row (first row with content)
        header_row = None
        for row_idx in range(1, min(10, ws.max_row + 1)):
            cell = ws.cell(row=row_idx, column=1)
            if cell.value is not None:
                header_row = row_idx
                break
        
        if header_row:
            # Extract header row style and labels
            config['row_height_header'] = ws.row_dimensions[header_row].height
            
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=header_row, column=col_idx)
                if cell.value:
                    config['header_labels'].append(str(cell.value))
                    if config['header_row_style'] is None:
                        config['header_row_style'] = self._extract_cell_style(cell)
            
            # Extract data row style (first row after header with content)
            for data_row in range(header_row + 1, min(header_row + 10, ws.max_row + 1)):
                cell = ws.cell(row=data_row, column=1)
                if cell.value is not None:
                    config['data_row_style'] = self._extract_cell_style(cell)
                    config['row_height_data'] = ws.row_dimensions[data_row].height
                    break
        
        return config
    
    def _extract_cell_style(self, cell) -> Dict[str, Any]:
        """Extract style properties from a single cell."""
        style = {}
        
        # Font
        if cell.font:
            font = cell.font
            style['font'] = {
                'name': font.name,
                'size': font.size,
                'bold': font.bold or False,
                'italic': font.italic or False,
                'color': font.color.rgb if font.color and font.color.rgb else None
            }
        
        # Fill
        if cell.fill and cell.fill.fill_type:
            fill = cell.fill
            style['fill'] = {
                'type': fill.fill_type,
                'fg_color': fill.fgColor.rgb if fill.fgColor and fill.fgColor.rgb else None,
                'bg_color': fill.bgColor.rgb if fill.bgColor and fill.bgColor.rgb else None
            }
        
        # Border
        if cell.border:
            border = cell.border
            style['border'] = {
                'left': border.left.style if border.left else None,
                'right': border.right.style if border.right else None,
                'top': border.top.style if border.top else None,
                'bottom': border.bottom.style if border.bottom else None
            }
        
        # Alignment
        if cell.alignment:
            align = cell.alignment
            style['alignment'] = {
                'horizontal': align.horizontal,
                'vertical': align.vertical,
                'wrap_text': align.wrap_text or False
            }
        
        return style
    
    def save(self, filepath: str) -> None:
        """
        Save extracted styles to a JSON file.
        
        Args:
            filepath: Path to save the JSON configuration
        """
        os.makedirs(os.path.dirname(filepath), exist_ok=True)
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(self.styles, f, indent=2, default=str)
    
    @classmethod
    def load(cls, filepath: str) -> 'StyleExtractor':
        """
        Load styles from a JSON configuration file.
        
        Args:
            filepath: Path to the JSON configuration file
            
        Returns:
            StyleExtractor instance with loaded styles
        """
        extractor = cls()
        with open(filepath, 'r', encoding='utf-8') as f:
            extractor.styles = json.load(f)
        return extractor
    
    def get_sheet_config(self, file_type: str, sheet_name: str) -> Optional[Dict]:
        """
        Get style configuration for a specific sheet.
        
        Args:
            file_type: 'index' or 'operation'
            sheet_name: Name of the sheet
            
        Returns:
            Sheet style configuration dict or None
        """
        configs = self.styles.get(file_type, {})
        
        # For operation files, response sheets use 'Response' template
        if file_type == 'operation' and sheet_name.isdigit():
            return configs.get('Response')
        
        return configs.get(sheet_name)


def main():
    """Extract styles from reference templates and save to config file."""
    import sys
    
    templates_dir = sys.argv[1] if len(sys.argv) > 1 else 'API Templates'
    output_file = sys.argv[2] if len(sys.argv) > 2 else 'src/oas_importer/styles_config.json'
    
    print(f"Extracting styles from: {templates_dir}")
    
    extractor = StyleExtractor()
    extractor.extract_from_templates(templates_dir)
    extractor.save(output_file)
    
    print(f"Styles saved to: {output_file}")
    print(f"Extracted from: {extractor.styles['metadata']['extracted_from']}")
    
    # Summary
    print("\nIndex sheets:", list(extractor.styles['index'].keys()))
    print("Operation sheets:", list(extractor.styles['operation'].keys()))


if __name__ == '__main__':
    main()
