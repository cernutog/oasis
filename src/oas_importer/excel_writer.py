"""
Excel Writer Module

Base class for writing Excel files with styled cells.
Uses styles extracted from reference templates via StyleExtractor.

Usage:
    writer = ExcelWriter(styles_config)
    writer.create_workbook()
    writer.add_sheet('Parameters', header_labels, data_rows)
    writer.save('output.xlsx')
"""

import json
import os
from typing import Dict, List, Any, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter


class ExcelWriter:
    """
    Base class for writing styled Excel files.
    
    Applies styles from a pre-extracted configuration file
    to ensure generated files match reference templates.
    """
    
    # Default style config path
    DEFAULT_STYLES_PATH = os.path.join(
        os.path.dirname(__file__), 
        'styles_config.json'
    )
    
    def __init__(self, styles_config_path: Optional[str] = None, 
                 file_type: str = 'operation'):
        """
        Initialize the writer.
        
        Args:
            styles_config_path: Path to styles_config.json
            file_type: 'index' or 'operation' - determines which styles to use
        """
        self.styles_config_path = styles_config_path or self.DEFAULT_STYLES_PATH
        self.file_type = file_type
        self.styles: Dict[str, Any] = {}
        self.workbook: Optional[Workbook] = None
        
        self._load_styles()
    
    def _load_styles(self) -> None:
        """Load styles from configuration file."""
        if os.path.exists(self.styles_config_path):
            with open(self.styles_config_path, 'r', encoding='utf-8') as f:
                all_styles = json.load(f)
                self.styles = all_styles.get(self.file_type, {})
        else:
            print(f"Warning: Styles config not found at {self.styles_config_path}")
            self.styles = {}
    
    def create_workbook(self) -> Workbook:
        """Create a new workbook."""
        self.workbook = Workbook()
        # Remove default sheet
        if 'Sheet' in self.workbook.sheetnames:
            del self.workbook['Sheet']
        return self.workbook
    
    def add_sheet(self, sheet_name: str, 
                  data_rows: List[Dict[str, Any]],
                  header_labels: Optional[List[str]] = None) -> None:
        """
        Add a sheet with header and data rows.
        
        Args:
            sheet_name: Name of the sheet
            data_rows: List of dicts where keys are column headers
            header_labels: Optional custom header labels (overrides style config)
        """
        if self.workbook is None:
            self.create_workbook()
        
        ws = self.workbook.create_sheet(sheet_name)
        
        # Get style config for this sheet
        style_config = self._get_sheet_style(sheet_name)
        
        # Determine headers - from config, custom, or from data keys
        if header_labels:
            headers = header_labels
        elif style_config and 'header_labels' in style_config:
            headers = style_config['header_labels']
        elif data_rows:
            headers = list(data_rows[0].keys())
        else:
            headers = []
        
        # Write header row
        header_row_idx = 1
        header_style = self._create_cell_style(
            style_config.get('header_row_style') if style_config else None
        )
        
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=header_row_idx, column=col_idx, value=header)
            self._apply_style(cell, header_style)
        
        # Apply header row height
        if style_config and style_config.get('row_height_header'):
            ws.row_dimensions[header_row_idx].height = style_config['row_height_header']
        
        # Write data rows
        data_style = self._create_cell_style(
            style_config.get('data_row_style') if style_config else None
        )
        
        for row_idx, row_data in enumerate(data_rows, start=header_row_idx + 1):
            for col_idx, header in enumerate(headers, start=1):
                # Get value by header name
                value = row_data.get(header)
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                self._apply_style(cell, data_style)
            
            # Apply data row height
            if style_config and style_config.get('row_height_data'):
                ws.row_dimensions[row_idx].height = style_config['row_height_data']
        
        # Apply column widths
        if style_config and 'column_widths' in style_config:
            for col_letter, width in style_config['column_widths'].items():
                ws.column_dimensions[col_letter].width = width
        else:
            # Auto-adjust column widths
            self._auto_adjust_columns(ws, headers)
        
        # Apply freeze panes
        if style_config and style_config.get('freeze_panes'):
            ws.freeze_panes = style_config['freeze_panes']
    
    def add_key_value_sheet(self, sheet_name: str,
                            data: Dict[str, Any]) -> None:
        """
        Add a sheet with key-value pairs (like General Description).
        
        Args:
            sheet_name: Name of the sheet
            data: Dict of key-value pairs
        """
        if self.workbook is None:
            self.create_workbook()
        
        ws = self.workbook.create_sheet(sheet_name)
        style_config = self._get_sheet_style(sheet_name)
        
        # Get styles
        header_style = self._create_cell_style(
            style_config.get('header_row_style') if style_config else None
        )
        data_style = self._create_cell_style(
            style_config.get('data_row_style') if style_config else None
        )
        
        # Write title row
        cell = ws.cell(row=1, column=1, value=sheet_name)
        self._apply_style(cell, header_style)
        
        # Write key-value pairs
        row_idx = 2
        for key, value in data.items():
            key_cell = ws.cell(row=row_idx, column=1, value=key)
            val_cell = ws.cell(row=row_idx, column=2, value=value)
            self._apply_style(key_cell, data_style)
            self._apply_style(val_cell, data_style)
            row_idx += 1
        
        # Apply column widths
        if style_config and 'column_widths' in style_config:
            for col_letter, width in style_config['column_widths'].items():
                ws.column_dimensions[col_letter].width = width
    
    def add_response_sheet(self, sheet_name: str,
                           description: str,
                           rows: List[Dict[str, Any]]) -> None:
        """
        Add a response sheet with section headers.
        
        Args:
            sheet_name: HTTP status code (e.g., '200', '400')
            description: Response description
            rows: Flattened response rows with 'Section' column
        """
        if self.workbook is None:
            self.create_workbook()
        
        ws = self.workbook.create_sheet(sheet_name)
        style_config = self._get_sheet_style(sheet_name)
        
        # Response sheets use 'Response' template for styling
        if not style_config:
            style_config = self.styles.get('Response', {})
        
        # Get styles
        header_style = self._create_cell_style(
            style_config.get('header_row_style') if style_config else None
        )
        data_style = self._create_cell_style(
            style_config.get('data_row_style') if style_config else None
        )
        
        # Row 1: Response | Code | Description
        ws.cell(row=1, column=1, value='Response')
        ws.cell(row=1, column=2, value=sheet_name)
        ws.cell(row=1, column=3, value=description)
        for col in range(1, 4):
            self._apply_style(ws.cell(row=1, column=col), header_style)
        
        # Row 2: Column headers
        headers = style_config.get('header_labels', [
            'Section', 'Name', 'Parent', 'Description', 'Type',
            'Items Data Type', 'Schema Name', 'Format', 'Mandatory',
            'Min', 'Max', 'PatternEba', 'Regex', 'Allowed value', 'Example'
        ])
        
        # Skip first 3 headers (Response, code, desc) if present
        if headers and headers[0] in ('Response', 'Section'):
            data_headers = ['Section', 'Name', 'Parent', 'Description', 'Type',
                           'Items Data Type', 'Schema Name', 'Format', 'Mandatory',
                           'Min', 'Max', 'PatternEba', 'Regex', 'Allowed value', 'Example']
        else:
            data_headers = headers
        
        for col_idx, header in enumerate(data_headers, start=1):
            cell = ws.cell(row=2, column=col_idx, value=header)
            self._apply_style(cell, header_style)
        
        # Data rows
        for row_idx, row_data in enumerate(rows, start=3):
            for col_idx, header in enumerate(data_headers, start=1):
                value = row_data.get(header)
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                self._apply_style(cell, data_style)
        
        # Apply column widths
        if style_config and 'column_widths' in style_config:
            for col_letter, width in style_config['column_widths'].items():
                ws.column_dimensions[col_letter].width = width
    
    def _get_sheet_style(self, sheet_name: str) -> Optional[Dict[str, Any]]:
        """Get style config for a sheet, handling response sheet fallback."""
        if sheet_name in self.styles:
            return self.styles[sheet_name]
        
        # For numeric sheet names (response codes), use 'Response' template
        if sheet_name.isdigit():
            return self.styles.get('Response')
        
        return None
    
    def _create_cell_style(self, style_dict: Optional[Dict[str, Any]]) -> Dict[str, Any]:
        """Create openpyxl style objects from config dict."""
        if not style_dict:
            return {}
        
        result = {}
        
        # Font
        if 'font' in style_dict:
            f = style_dict['font']
            color = f.get('color')
            # Handle invalid color values
            if color and isinstance(color, str) and 'class' not in color:
                result['font'] = Font(
                    name=f.get('name', 'Calibri'),
                    size=f.get('size', 11),
                    bold=f.get('bold', False),
                    italic=f.get('italic', False),
                    color=color
                )
            else:
                result['font'] = Font(
                    name=f.get('name', 'Calibri'),
                    size=f.get('size', 11),
                    bold=f.get('bold', False),
                    italic=f.get('italic', False)
                )
        
        # Fill
        if 'fill' in style_dict:
            f = style_dict['fill']
            fg_color = f.get('fg_color')
            # Handle invalid color values
            if fg_color and isinstance(fg_color, str) and 'class' not in fg_color:
                result['fill'] = PatternFill(
                    fill_type=f.get('type', 'solid'),
                    fgColor=fg_color
                )
        
        # Border
        if 'border' in style_dict:
            b = style_dict['border']
            sides = {}
            for side_name in ['left', 'right', 'top', 'bottom']:
                if b.get(side_name):
                    sides[side_name] = Side(style=b[side_name])
                else:
                    sides[side_name] = Side()
            result['border'] = Border(**sides)
        
        # Alignment
        if 'alignment' in style_dict:
            a = style_dict['alignment']
            result['alignment'] = Alignment(
                horizontal=a.get('horizontal'),
                vertical=a.get('vertical'),
                wrap_text=a.get('wrap_text', False)
            )
        
        return result
    
    def _apply_style(self, cell, style: Dict[str, Any]) -> None:
        """Apply a style dict to a cell."""
        if 'font' in style:
            cell.font = style['font']
        if 'fill' in style:
            cell.fill = style['fill']
        if 'border' in style:
            cell.border = style['border']
        if 'alignment' in style:
            cell.alignment = style['alignment']
    
    def _auto_adjust_columns(self, ws, headers: List[str]) -> None:
        """Auto-adjust column widths based on content."""
        for col_idx, header in enumerate(headers, start=1):
            col_letter = get_column_letter(col_idx)
            # Start with header width
            max_width = len(str(header)) + 2
            
            # Check data in first 20 rows
            for row in range(2, min(22, ws.max_row + 1)):
                cell_value = ws.cell(row=row, column=col_idx).value
                if cell_value:
                    max_width = max(max_width, min(len(str(cell_value)) + 2, 50))
            
            ws.column_dimensions[col_letter].width = max_width
    
    def save(self, filepath: str) -> None:
        """Save the workbook to a file."""
        if self.workbook is None:
            raise ValueError("No workbook created. Call create_workbook() first.")
        
        # Ensure directory exists
        os.makedirs(os.path.dirname(filepath) or '.', exist_ok=True)
        
        self.workbook.save(filepath)


def main():
    """Test the Excel writer."""
    writer = ExcelWriter(file_type='operation')
    writer.create_workbook()
    
    # Test Parameters sheet
    params_data = [
        {'Name': 'X-Request-ID', 'Description': 'Unique request identifier', 
         'In': 'header', 'Type': 'string', 'Mandatory': 'M'},
        {'Name': 'accountId', 'Description': 'Account identifier',
         'In': 'path', 'Type': 'string', 'Mandatory': 'M'}
    ]
    
    writer.add_sheet('Parameters', params_data, 
                     ['Name', 'Description', 'In', 'Type', 'Mandatory'])
    
    # Test Body sheet
    body_data = [
        {'Name': 'amount', 'Parent': None, 'Type': 'number', 'Mandatory': 'M'},
        {'Name': 'currency', 'Parent': None, 'Type': 'string', 'Mandatory': 'M'}
    ]
    
    writer.add_sheet('Body', body_data,
                     ['Name', 'Parent', 'Type', 'Mandatory'])
    
    output_path = 'test_excel_writer.xlsx'
    writer.save(output_path)
    print(f"Test file saved to: {output_path}")


if __name__ == '__main__':
    main()
