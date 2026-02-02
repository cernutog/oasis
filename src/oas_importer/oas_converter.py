"""
OAS to Excel Converter

High-level orchestrator that converts an OpenAPI Specification file
into Excel template files using pre-styled master templates.

Usage:
    converter = OASToExcelConverter('path/to/oas.yaml')
    converter.generate_endpoint_files('output_folder')
    converter.generate_index_file('output_folder/$index.xlsx')
"""

import os
import datetime
from typing import Dict, List, Any, Optional
import pandas as pd
from openpyxl import Workbook
from src.generator_pkg.yaml_output import RawNumericValue, OASDumper, SafeLoaderRawNumbers
from .oas_parser import OASParser, OperationInfo
from .schema_flattener import SchemaFlattener, FlatRow
from .template_writer import TemplateExcelWriter



def _to_string(value: Any) -> str:
    """Safely convert value to string, preserving ISO 8601 for datetimes."""
    if isinstance(value, (datetime.datetime, datetime.date)):
        return value.isoformat()
    return str(value)


def _set_cell_as_text(ws, row: int, col: int, value: str) -> None:
    """Write value to cell and force TEXT format to prevent Excel number conversion."""
    cell = ws.cell(row=row, column=col, value=value)
    cell.number_format = '@'  # Text format


class OASToExcelConverter:
    """
    Converts OAS files to OASIS Excel template format.
    
    Uses OASParser for extraction, SchemaFlattener for schema conversion,
    and TemplateExcelWriter for output generation.
    """
    
    def __init__(self, oas_filepath: str, log_callback=None):
        """
        Initialize converter with an OAS file.
        
        Args:
            oas_filepath: Path to the OAS YAML file
            log_callback: Optional function to log messages
        """
        self.oas_filepath = oas_filepath
        self.log_callback = log_callback
        self.parser = OASParser(oas_filepath)
        self.flattener = SchemaFlattener(self.parser.oas)

    def log(self, message: str):
        if self.log_callback:
            self.log_callback(message)
        
    def generate_endpoint_file(self, operation: OperationInfo, 
                                output_path: str) -> str:
        """
        Generate an endpoint Excel file for a single operation.
        
        Args:
            operation: The OperationInfo to export
            output_path: Output file path
            
        Returns:
            Path to the generated file
        """
        writer = TemplateExcelWriter('endpoint')
        writer.load_template()
        
        # Fill Parameters sheet
        param_rows = self._convert_parameters(operation.parameters)
        writer.fill_parameters_sheet(param_rows)
        
        # Fill Body sheet (if request body exists)
        if operation.request_body:
            body_description = self._get_body_description(operation.request_body)
            body_required = 'M' if operation.request_body.get('required', False) else 'O'
            body_rows = self._flatten_request_body(operation.request_body)
            writer.fill_body_sheet(body_rows, description=body_description, 
                                   mandatory=body_required)
            
            # Fill Body Example if available
            examples = self._extract_body_examples(operation.request_body)
            if examples:
                writer.fill_body_example_sheet(examples)
        
        # Add Response sheets
        for status_code, response in operation.responses.items():
            response_rows = self._flatten_response(response)
            writer.add_response_sheet(status_code, response.description or '', 
                                      response_rows)
        
        # Save
        writer.save(output_path)
        return output_path
    
    def generate_all_endpoint_files(self, output_dir: str) -> List[str]:
        """
        Generate endpoint files for all operations in the OAS.
        
        Args:
            output_dir: Directory to save files
            
        Returns:
            List of generated file paths
        """
        os.makedirs(output_dir, exist_ok=True)
        self.log(f"Generating endpoint files in: {output_dir}")
        
        generated_files = []
        operations = self.parser.get_operations()
        total_ops = len(operations)
        
        for i, operation in enumerate(operations):
            # Generate filename from operation ID or path
            filename = self._operation_to_filename(operation)
            filepath = os.path.join(output_dir, filename)
            
            # self.log(f"Processing ({i+1}/{total_ops}): {filename}")
            self.generate_endpoint_file(operation, filepath)
            generated_files.append(filepath)
            
        self.log(f"Generated {len(generated_files)} endpoint files.")
        return generated_files
    
    def generate_index_file(self, output_path: str) -> str:
        """
        Generate the $index.xlsx file with API metadata.
        
        Populates:
        - General Description: API info (title, version, description)
        - Paths: List of all operations with Excel file references
        - Tags: API tags
        - Parameters: Shared component parameters
        - Headers: Shared component headers  
        - Schemas: All component schemas flattened
        - Responses: Shared component responses
        
        Args:
            output_path: Output file path
            
        Returns:
            Path to generated file
        """
        writer = TemplateExcelWriter('index')
        writer.load_template()
        
        # 1. General Description sheet
        self._fill_general_description(writer)
        
        # 2. Paths sheet
        self._fill_paths_sheet(writer)
        
        # 3. Tags sheet
        self._fill_tags_sheet(writer)
        
        # 4. Parameters sheet (component parameters)
        self._fill_component_parameters(writer)
        
        # 5. Headers sheet (component headers)
        self._fill_component_headers(writer)
        
        # 6. Schemas sheet
        self._fill_schemas_sheet(writer)
        
        # 7. Responses sheet (component responses)
        self._fill_component_responses(writer)
        
        writer.save(output_path)
        self.log(f"Generated Master Index: {output_path}")
        return output_path
    
    def _fill_general_description(self, writer: TemplateExcelWriter) -> None:
        """
        Fill General Description sheet with API info.
        
        Template structure (fixed row positions):
        Row 2: info description | value (B)
        Row 3: info version | value (B)
        Row 4: info title | value (B)
        Row 5: info contact name | value (B)
        Row 6: info contact url | value (B)
        Row 7: servers url | value (B) | servers description | value (D)
        Row 8: servers url | value (B) | servers description | value (D)
        Row 9: release | value (B)
        Row 10: filename pattern | value (B)
        """
        info = self.parser.get_info()
        ws = writer.workbook['General Description']
        
        from .template_writer import DEFAULT_ALIGNMENT
        
        def set_cell(row: int, col: int, value):
            """Set cell value and alignment."""
            if value is not None:
                cell = ws.cell(row=row, column=col, value=str(value) if value else '')
                cell.alignment = DEFAULT_ALIGNMENT
        
        # Row 2: info description
        set_cell(2, 2, info.get('description', ''))
        
        # Row 3: info version  
        set_cell(3, 2, info.get('version', ''))
        
        # Row 4: info title
        set_cell(4, 2, info.get('title', ''))
        
        # Row 5: info contact name (using flat key from get_info)
        set_cell(5, 2, info.get('contact_name', ''))
        
        # Row 6: info contact url 
        set_cell(6, 2, info.get('contact_url', ''))
        
        # Rows 7-8: servers (max 2, url in col B, description in col D)
        servers = self.parser.get_servers()[:2]  # Max 2 servers
        for idx, server in enumerate(servers):
            row = 7 + idx  # Row 7 for first server, row 8 for second
            set_cell(row, 2, server.get('url', ''))  # Col B: url
            set_cell(row, 4, server.get('description', ''))  # Col D: description
        
        # Row 9: release
        set_cell(9, 2, info.get('release', ''))
        

        

    
    def _fill_paths_sheet(self, writer: TemplateExcelWriter) -> None:
        """
        Fill Paths sheet with operation list.
        
        Columns: Excel file, Path, Name, Method, Description, Tag, Summary, OperationId, Custom Extensions
        """
        ws = writer.workbook['Paths']
        
        from .template_writer import DEFAULT_ALIGNMENT
        
        operations = self.parser.get_operations()
        for row_idx, op in enumerate(operations, start=3):  # Data starts row 3
            filename = self._operation_to_filename(op)
            ws.cell(row=row_idx, column=1, value=filename)  # Excel file
            ws.cell(row=row_idx, column=2, value=op.path)
            ws.cell(row=row_idx, column=3, value=op.summary or op.operation_id or '')  # Name (prefer summary)
            ws.cell(row=row_idx, column=4, value=op.method.lower())
            ws.cell(row=row_idx, column=5, value=op.description or '')
            ws.cell(row=row_idx, column=6, value=', '.join(op.tags) if op.tags else '')
            ws.cell(row=row_idx, column=7, value=op.summary or '')
            ws.cell(row=row_idx, column=8, value=op.operation_id or '')
            
            # Column 9: Custom Extensions as raw YAML text (preserves original formatting)
            if op.extensions:
                ext_text = self.parser.get_raw_extensions(op.path, op.method)
                # Add 6-space indent to ALL lines to match Excel template style
                # (original templates have top-level x-sandbox keys at 6-space indent)
                if ext_text:
                    lines = ext_text.split('\n')
                    indented_lines = [('      ' + line) if line.strip() else line for line in lines]
                    ext_text = '\n'.join(indented_lines)
                ws.cell(row=row_idx, column=9, value=ext_text)
            
            # Apply alignment to all columns
            for col in range(1, 10):
                ws.cell(row=row_idx, column=col).alignment = DEFAULT_ALIGNMENT
    
    def _fill_tags_sheet(self, writer: TemplateExcelWriter) -> None:
        """Fill Tags sheet."""
        ws = writer.workbook['Tags']
        
        tags = self.parser.get_tags()
        for row_idx, tag in enumerate(tags, start=2):  # Row 1 is header
            ws.cell(row=row_idx, column=1, value=tag.get('name', ''))
            ws.cell(row=row_idx, column=2, value=tag.get('description', ''))
            
            from .template_writer import DEFAULT_ALIGNMENT
            ws.cell(row=row_idx, column=1).alignment = DEFAULT_ALIGNMENT
            ws.cell(row=row_idx, column=2).alignment = DEFAULT_ALIGNMENT
    
    def _fill_component_parameters(self, writer: TemplateExcelWriter) -> None:
        """Fill Parameters sheet with component parameters."""
        components = self.parser.get_components()
        parameters = components.get('parameters', {})
        
        if not parameters:
            return
        
        ws = writer.workbook['Parameters']
        row_idx = 2  # Row 1 is header
        
        from .template_writer import DEFAULT_ALIGNMENT
        
        for name, param_def in parameters.items():
            schema = param_def.get('schema', {})
            
            ws.cell(row=row_idx, column=1, value=name)
            ws.cell(row=row_idx, column=2, value=param_def.get('description', ''))
            ws.cell(row=row_idx, column=3, value=param_def.get('in', ''))
            # Type: show 'schema' if using $ref
            type_val = schema.get('type', '')
            if not type_val and '$ref' in schema:
                type_val = 'schema'
            ws.cell(row=row_idx, column=4, value=type_val)
            ws.cell(row=row_idx, column=5, value=schema.get('$ref', '').split('/')[-1] if '$ref' in schema else '')
            # Items data type (for arrays)
            items = schema.get('items', {})
            ws.cell(row=row_idx, column=6, value=items.get('type', ''))
            ws.cell(row=row_idx, column=7, value=schema.get('format', ''))
            ws.cell(row=row_idx, column=8, value='M' if param_def.get('required', False) else '')
            _set_cell_as_text(ws, row_idx, 9, str(schema.get('minLength', schema.get('minimum', ''))) or '')
            _set_cell_as_text(ws, row_idx, 10, str(schema.get('maxLength', schema.get('maximum', ''))) or '')
            ws.cell(row=row_idx, column=11, value='')  # PatternEba - not standard OAS
            ws.cell(row=row_idx, column=12, value=schema.get('pattern', ''))  # Regex
            # Allowed values (enum) - col 13
            enum = schema.get('enum', [])
            ws.cell(row=row_idx, column=13, value=', '.join(str(e) for e in enum) if enum else '')
            # Example: try param_def.example first (parameter level), then schema.example, then schema.examples[0], then resolve $ref
            example = param_def.get('example')
            if example is None:
                example = schema.get('example')
            if example is None:
                examples_list = schema.get('examples')
                if examples_list and isinstance(examples_list, list) and len(examples_list) > 0:
                    example = examples_list[0]
            if example is None and '$ref' in schema:
                ref_name = schema['$ref'].split('/')[-1]
                ref_schema = self.parser.oas.get('components', {}).get('schemas', {}).get(ref_name, {})
                example = ref_schema.get('example')
                if example is None:
                    examples_list = ref_schema.get('examples')
                    if examples_list and isinstance(examples_list, list) and len(examples_list) > 0:
                        example = examples_list[0]
            ws.cell(row=row_idx, column=14, value=str(example) if example is not None else '')
            
            # Apply alignment
            for col in range(1, 15):
                ws.cell(row=row_idx, column=col).alignment = DEFAULT_ALIGNMENT
            
            row_idx += 1
    
    def _fill_component_headers(self, writer: TemplateExcelWriter) -> None:
        """Fill Headers sheet with component headers."""
        components = self.parser.get_components()
        headers = components.get('headers', {})
        
        if not headers:
            return
        
        ws = writer.workbook['Headers']
        row_idx = 2  # Row 1 is header
        
        from .template_writer import DEFAULT_ALIGNMENT
        
        for name, header_def in headers.items():
            schema = header_def.get('schema', {})
            
            ws.cell(row=row_idx, column=1, value=name)
            ws.cell(row=row_idx, column=2, value=header_def.get('description', ''))
            ws.cell(row=row_idx, column=3, value=schema.get('type', ''))
            ws.cell(row=row_idx, column=4, value=schema.get('$ref', '').split('/')[-1] if '$ref' in schema else '')
            # Items data type (for arrays)
            items = schema.get('items', {})
            ws.cell(row=row_idx, column=5, value=items.get('type', ''))
            ws.cell(row=row_idx, column=6, value=schema.get('format', ''))
            ws.cell(row=row_idx, column=7, value='M' if header_def.get('required', False) else '')
            _set_cell_as_text(ws, row_idx, 8, str(schema.get('minLength', schema.get('minimum', ''))) or '')
            _set_cell_as_text(ws, row_idx, 9, str(schema.get('maxLength', schema.get('maximum', ''))) or '')
            ws.cell(row=row_idx, column=10, value='')  # PatternEba - not standard OAS
            ws.cell(row=row_idx, column=11, value=schema.get('pattern', ''))  # Regex
            enum = schema.get('enum', [])
            ws.cell(row=row_idx, column=12, value=', '.join(str(e) for e in enum) if enum else '')  # Allowed value
            # Example: try header_def.example first, then schema.example, then schema.examples[0], then resolve $ref
            example = header_def.get('example')
            if example is None:
                example = schema.get('example')
            if example is None:
                examples_list = schema.get('examples')
                if examples_list and isinstance(examples_list, list) and len(examples_list) > 0:
                    example = examples_list[0]
            if example is None and '$ref' in schema:
                ref_name = schema['$ref'].split('/')[-1]
                ref_schema = self.parser.oas.get('components', {}).get('schemas', {}).get(ref_name, {})
                example = ref_schema.get('example')
                if example is None:
                    examples_list = ref_schema.get('examples')
                    if examples_list and isinstance(examples_list, list) and len(examples_list) > 0:
                        example = examples_list[0]
            ws.cell(row=row_idx, column=13, value=_to_string(example) if example is not None else '')  # Example
            
            for col in range(1, 14):
                ws.cell(row=row_idx, column=col).alignment = DEFAULT_ALIGNMENT
            
            row_idx += 1
    
    def _fill_schemas_sheet(self, writer: TemplateExcelWriter) -> None:
        """Fill Schemas sheet with all component schemas flattened."""
        schemas = self.parser.get_schemas()
        
        if not schemas:
            return
        
        ws = writer.workbook['Schemas']
        row_idx = 2  # Row 1 is header
        
        from .template_writer import DEFAULT_ALIGNMENT
        
        for schema_name, schema_def in schemas.items():
            # Flatten each schema
            rows = self.flattener.flatten_schema(schema_name)
            
            for flat_row in rows:
                ws.cell(row=row_idx, column=1, value=flat_row.name)
                ws.cell(row=row_idx, column=2, value=flat_row.parent or '')
                ws.cell(row=row_idx, column=3, value=flat_row.description or '')
                ws.cell(row=row_idx, column=4, value=flat_row.type or '')
                ws.cell(row=row_idx, column=5, value=flat_row.items_type or '')
                ws.cell(row=row_idx, column=6, value=flat_row.schema_name or '')
                ws.cell(row=row_idx, column=7, value=flat_row.format or '')
                # Mandatory: keep 'M', map 'O' to empty string to match reference format
                mand = flat_row.mandatory
                if mand == 'O':
                    mand = ''
                ws.cell(row=row_idx, column=8, value=mand or '')
                _set_cell_as_text(ws, row_idx, 9, flat_row.min_value or '')
                _set_cell_as_text(ws, row_idx, 10, flat_row.max_value or '')
                ws.cell(row=row_idx, column=11, value=flat_row.pattern or '')
                ws.cell(row=row_idx, column=12, value=flat_row.regex or '')
                ws.cell(row=row_idx, column=13, value=flat_row.allowed_values or '')
                ws.cell(row=row_idx, column=14, value=flat_row.example or '')
                
                # Apply alignment
                for col in range(1, 15):
                    ws.cell(row=row_idx, column=col).alignment = DEFAULT_ALIGNMENT
                
                row_idx += 1
    
    def _fill_component_responses(self, writer: TemplateExcelWriter) -> None:
        """Fill Responses sheet with component responses."""
        components = self.parser.get_components()
        responses = components.get('responses', {})
        
        if not responses:
            return
        
        ws = writer.workbook['Responses']
        row_idx = 2  # Row 1 is header
        
        from .template_writer import DEFAULT_ALIGNMENT
        
        for resp_name, resp_def in responses.items():
            # Flatten each response with named root row
            flat_rows = self.flattener.flatten_component_response(resp_name, resp_def)
            
            for flat_row in flat_rows:
                ws.cell(row=row_idx, column=1, value=flat_row.section or '')
                ws.cell(row=row_idx, column=2, value=flat_row.name)
                ws.cell(row=row_idx, column=3, value=flat_row.parent or '')
                ws.cell(row=row_idx, column=4, value=flat_row.description or '')
                ws.cell(row=row_idx, column=5, value=flat_row.type or '')
                ws.cell(row=row_idx, column=6, value=flat_row.items_type or '')
                ws.cell(row=row_idx, column=7, value=flat_row.schema_name or '')
                ws.cell(row=row_idx, column=8, value=flat_row.format or '')
                # Mandatory: keep 'M', map 'O' to empty string
                mand = flat_row.mandatory
                if mand == 'O':
                    mand = ''
                ws.cell(row=row_idx, column=9, value=mand or '')
                _set_cell_as_text(ws, row_idx, 10, flat_row.min_value or '')
                _set_cell_as_text(ws, row_idx, 11, flat_row.max_value or '')
                ws.cell(row=row_idx, column=12, value=flat_row.pattern or '')
                ws.cell(row=row_idx, column=13, value=flat_row.regex or '')
                ws.cell(row=row_idx, column=14, value=flat_row.allowed_values or '')
                ws.cell(row=row_idx, column=15, value=flat_row.example or '')
                
                for col in range(1, 16):
                    ws.cell(row=row_idx, column=col).alignment = DEFAULT_ALIGNMENT
                
                row_idx += 1
    
    def _extensions_to_yaml(self, extensions: Dict[str, Any]) -> str:
        """
        Convert extensions dict to YAML with literal block style for multiline strings.
        
        This preserves the format expected by the Excel templates, where multiline
        content like scripts use the YAML literal block style (|).
        """
        import yaml
        
        # Custom string class to force literal block style
        class LiteralStr(str):
            """String subclass that forces literal block style in YAML."""
            pass
        
        def literal_representer(dumper, data):
            """Always use literal block style for LiteralStr."""
            return dumper.represent_scalar('tag:yaml.org,2002:str', data, style='|')
        
        def str_representer(dumper, data):
            """Use literal style for multiline strings, plain for others."""
            if '\n' in data:
                return dumper.represent_scalar('tag:yaml.org,2002:str', data, style='|')
            return dumper.represent_scalar('tag:yaml.org,2002:str', data)
        
        # Convert multiline strings to LiteralStr to force literal style
        def convert_to_literal(obj):
            if isinstance(obj, str):
                if '\n' in obj:
                    return LiteralStr(obj)
                return obj
            elif isinstance(obj, dict):
                return {k: convert_to_literal(v) for k, v in obj.items()}
            elif isinstance(obj, list):
                return [convert_to_literal(item) for item in obj]
            return obj
        
        # Create a custom dumper inheriting from OASDumper to preserve RawNumericValue precision
        class LiteralDumper(OASDumper):
            pass
        
        # Register representers - LiteralStr always uses literal, str checks for newlines
        LiteralDumper.add_representer(LiteralStr, literal_representer)
        LiteralDumper.add_representer(str, str_representer)
        
        converted = convert_to_literal(extensions)
        result = yaml.dump(converted, Dumper=LiteralDumper, 
                          default_flow_style=False, allow_unicode=True, 
                          sort_keys=False, width=10000)
        
        return result.rstrip()

    def _convert_parameters(self, parameters: List) -> List[Dict[str, Any]]:
        """Convert OAS parameters to Parameters sheet format."""
        rows = []
        for param in parameters:
            # check for top-level ref
            is_ref = getattr(param, 'ref', None) is not None
            
            if is_ref:
                # Ref with possible description override (valid in OAS 3.1)
                ref_name = param.ref.split('/')[-1]
                # Use description if present (override), else empty
                param_desc = param.description or ''
                rows.append({
                    'Name': ref_name,
                    'Description': param_desc,  # Include if present (enables x-comment inline resolution)
                    'In': '',  # Empty for refs
                    'Type': 'parameter',  # Singular form
                    'Items Data Type (Array only)': '',
                    "Schema Name\n(for Type or Items Data Type = 'schema')": ref_name,
                    'Format': '',
                    'Mandatory': '',  # Empty - not specified at this level
                    'Min  \nValue/Length/Item': '',
                    'Max  \nValue/Length/Item': '',
                    'PatternEba': '',
                    'Regex': '',
                    'Allowed values': '',
                    'Example': ''  # Empty - not specified at this level
                })
            else:
                # Inline parameter: include all data
                schema_name = param.schema_ref or ''
                type_val = param.schema_type or param.schema_ref or ''
                param_desc = param.description or ''
                
                rows.append({
                    'Name': param.name,
                    'Description': param_desc,
                    'In': param.location,
                    'Type': type_val,
                    'Items Data Type (Array only)': param.items_type or '',
                    "Schema Name\n(for Type or Items Data Type = 'schema')": schema_name,
                    'Format': param.schema_format or '',
                    'Mandatory': 'M' if param.required is True else ('O' if param.required is False else ''),
                    'Min  \nValue/Length/Item': self._format_min(param),
                    'Max  \nValue/Length/Item': self._format_max(param),
                    'PatternEba': '',
                    'Regex': param.pattern or '',
                    'Allowed values': ', '.join(param.enum) if param.enum else '',
                    'Example': _to_string(param.example) if param.example else ''
                })
        return rows
    
    def _flatten_request_body(self, request_body: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Flatten request body schema to Body sheet format."""
        flat_rows = self.flattener.flatten_request_body(request_body)
        return [self._map_row_to_body_format(row) for row in flat_rows]
    
    def _flatten_response(self, response) -> List[Dict[str, Any]]:
        """Flatten response to Response sheet format."""
        # Check for top-level ref
        if getattr(response, 'ref', None):
             ref_name = response.ref.split('/')[-1]
             # Create a single row for the Reference
             # Matches Reference Excel format: Section='content', Name=Description, Type='response', SchemaName=RefName
             return [{
                 'Section': 'content',
                 'Name': response.description or 'Response',
                 'Parent': '',
                 'Description': response.description or '',
                 'Type': 'response',
                 'Items Data Type \n(Array only)': '',
                 "Schema Name\n(for Type or Items Data Type = 'schema')": ref_name,
                 'Format': '',
                 'Mandatory': '',
                 'Min  \nValue/Length/Item': '',
                 'Max  \nValue/Length/Item': '',
                 'PatternEba': '',
                 'Regex': '',
                 'Allowed value': '',
                 'Example': ''
             }]

        # Convert ResponseInfo to dict format expected by flattener
        response_dict = {
            'description': response.description,
            'headers': {h.get('name', ''): h for h in response.headers} if response.headers else {},
            'content': response.content,
            'extensions': response.extensions
        }
        flat_rows = self.flattener.flatten_response(response_dict)
        return [self._map_row_to_response_format(row) for row in flat_rows]
    
    def _map_row_to_body_format(self, row: FlatRow) -> Dict[str, Any]:
        """Map FlatRow to Body sheet header names."""
        return {
            'Section': row.section or '',
            'Name': row.name,
            'Parent': row.parent or '',
            'Description': row.description or '',
            'Type': row.type or '',
            'Items Data Type \n(Array only)': row.items_type or '',
            "Schema Name\n(for Type or Items Data Type = 'schema')": row.schema_name or '',
            'Format': row.format or '',
            'Mandatory': row.mandatory or '',
            'Min  \nValue/Length/Item': row.min_value or '',
            'Max  \nValue/Length/Item': row.max_value or '',
            'PatternEba': row.pattern or '',
            'Regex': row.regex or '',
            'Allowed value': row.allowed_values or '',
            'Example': row.example or ''
        }
    
    def _map_row_to_response_format(self, row: FlatRow) -> Dict[str, Any]:
        """Map FlatRow to Response sheet header names."""
        # Response uses same format as Body
        return self._map_row_to_body_format(row)
    
    def _format_min(self, param) -> str:
        """Format minimum value from various sources."""
        if param.minimum is not None:
            return str(param.minimum)
        if param.min_length is not None:
            return str(param.min_length)
        if param.min_items is not None:
            return str(param.min_items)
        return ''
    
    def _format_max(self, param) -> str:
        """Format maximum value from various sources."""
        if param.maximum is not None:
            return str(param.maximum)
        if param.max_length is not None:
            return str(param.max_length)
        if param.max_items is not None:
            return str(param.max_items)
        return ''
    
    def _get_body_description(self, request_body: Dict[str, Any]) -> str:
        """Extract description from request body."""
        return request_body.get('description', '')
    
    def _coerce_example_types(self, value: Any, schema: Dict[str, Any]) -> Any:
        """
        Recursively coerce example value types to match schema expectations.
        
        If schema type is 'string' but value is a number/date, converts to string.
        This ensures proper YAML serialization with quotes preserved.
        
        Args:
            value: The example value to check
            schema: The schema definition for this value
            
        Returns:
            Coerced value with proper types
        """
        if schema is None:
            return value
            
        # Resolve $ref if present
        if '$ref' in schema:
            ref_path = schema['$ref']
            if ref_path.startswith('#/components/schemas/'):
                ref_name = ref_path.split('/')[-1]
                schema = self.parser.oas.get('components', {}).get('schemas', {}).get(ref_name, {})
        
        schema_type = schema.get('type')
        
        # Handle object: recurse into properties
        if isinstance(value, dict):
            properties = schema.get('properties', {}).copy()
            # Also check combinators for properties
            for combinator in ['oneOf', 'anyOf', 'allOf']:
                if combinator in schema and isinstance(schema[combinator], list):
                    for sub_schema in schema[combinator]:
                        if isinstance(sub_schema, dict):
                            if '$ref' in sub_schema:
                                ref_name = sub_schema['$ref'].split('/')[-1]
                                sub_schema = self.parser.oas.get('components', {}).get('schemas', {}).get(ref_name, {})
                            properties.update(sub_schema.get('properties', {}))

            result = {}
            for k, v in value.items():
                prop_schema = properties.get(k, {})
                result[k] = self._coerce_example_types(v, prop_schema)
            return result
        
        # Handle array: recurse into items
        if isinstance(value, list):
            items_schema = schema.get('items', {})
            # Resolve $ref for items
            if items_schema and isinstance(items_schema, dict) and '$ref' in items_schema:
                ref_name = items_schema['$ref'].split('/')[-1]
                items_schema = self.parser.oas.get('components', {}).get('schemas', {}).get(ref_name, {})
            return [self._coerce_example_types(item, items_schema) for item in value]
        
        # Handle scalar: 
        if isinstance(value, str):
            # PROACTIVE STRIPPING: Always strip redundant quotes from string inputs.
            # This handles cases where schema_type is unknown or resolution failed.
            if isinstance(value, str):
                s_val = value.strip()
                if len(s_val) >= 2 and (
                    (s_val.startswith("'") and s_val.endswith("'")) or 
                    (s_val.startswith('"') and s_val.endswith('"'))
                ):
                    value = s_val[1:-1]
                    
            # Check if it looks like a number
            try:
                float(value)
                
                # PRECISE TYPE DETECTION: If not direct, look into combinators
                if not schema_type and schema and isinstance(schema, dict):
                    for combinator in ['oneOf', 'anyOf', 'allOf']:
                        if combinator in schema and isinstance(schema[combinator], list):
                            for sub_schema in schema[combinator]:
                                if isinstance(sub_schema, dict):
                                    if '$ref' in sub_schema:
                                        ref_name = sub_schema['$ref'].split('/')[-1]
                                        sub_schema = self.parser.oas.get('components', {}).get('schemas', {}).get(ref_name, {})
                                    if sub_schema.get('type'):
                                        schema_type = sub_schema.get('type')
                                        break
                        if schema_type: break

                if schema_type == 'string':
                    # DO NOT WRAP IN QUOTES. 
                    return value
                
                if schema_type in ['number', 'integer']:
                    # STRING PRESERVATION: Return as RawNumericValue string
                    return RawNumericValue(value)
                
                # Default for numbers if type unknown
                return RawNumericValue(value)
            except (ValueError, TypeError):
                pass
            
        # Handle datetime/date specifically to ensure ISO 8601
        if isinstance(value, (datetime.datetime, datetime.date)):
            return value.isoformat()
        
        return value

    def _extract_body_examples(self, request_body: Dict[str, Any]) -> List[Dict[str, str]]:
        """Extract examples from request body in YAML format, preserving string types."""
        import yaml
        
        examples = []
        content = request_body.get('content', {})
        for media_type, media_def in content.items():
            # Get schema for type coercion
            schema = media_def.get('schema', {})
            
            if 'examples' in media_def:
                for name, example in media_def['examples'].items():
                    value = example.get('value', example)
                    # Coerce types based on schema
                    value = self._coerce_example_types(value, schema)
                    if isinstance(value, (dict, list)):
                        # Use OASDumper for clean, precise output in the Body/Example cell
                        value = yaml.dump(value, Dumper=OASDumper, default_flow_style=False, allow_unicode=True, sort_keys=False)
                    examples.append({'name': name, 'body': str(value)})
            elif 'example' in media_def:
                value = media_def['example']
                # Coerce types based on schema
                value = self._coerce_example_types(value, schema)
                if isinstance(value, (dict, list)):
                    value = yaml.dump(value, Dumper=OASDumper, default_flow_style=False, allow_unicode=True, sort_keys=False)
                examples.append({'name': 'default', 'body': str(value)})

        return examples
    
    def _operation_to_filename(self, operation: OperationInfo) -> str:
        """Generate a filename for an operation."""
        # Prefer operationId, fallback to method-path
        if operation.operation_id:
            name = operation.operation_id
        else:
            # Convert /v1/accounts/{id} to v1-accounts-id
            path = operation.path.strip('/').replace('/', '-').replace('{', '').replace('}', '')
            name = f"{operation.method}-{path}"
        
        # Sanitize filename
        name = "".join(c for c in name if c.isalnum() or c in '-_')
        
        # Append current date (YYMMDD) as per user requirement
        from datetime import datetime
        date_str = datetime.now().strftime("%y%m%d")
        
        return f"{name}.{date_str}.xlsx"


def main():
    """Test the converter."""
    import sys
    
    if len(sys.argv) < 2:
        print("Usage: python oas_converter.py <oas_file> [output_dir]")
        return
    
    oas_file = sys.argv[1]
    output_dir = sys.argv[2] if len(sys.argv) > 2 else 'output'
    
    converter = OASToExcelConverter(oas_file)
    files = converter.generate_all_endpoint_files(output_dir)
    
    print(f"Generated {len(files)} files:")
    for f in files:
        print(f"  - {f}")


if __name__ == '__main__':
    main()
