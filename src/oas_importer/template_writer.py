"""
Template-Based Excel Writer Module

Fills pre-styled master templates instead of generating styles from code.
Uses master template files:
- Templates Master/$index.xlsx - for index files
- Templates Master/endpoint.xlsx - for operation files

Usage:
    writer = TemplateExcelWriter('endpoint')
    writer.load_template()
    writer.fill_parameters_sheet(param_rows)
    writer.fill_body_sheet(body_rows, description)
    writer.add_response_sheet('201', 'Created', response_rows)
    writer.add_response_sheet('400', 'Bad Request', error_rows)
    writer.save('output/my_endpoint.xlsx')
"""

import os
import copy
from typing import Dict, List, Any, Optional
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# Global alignment for all data cells
DEFAULT_ALIGNMENT = Alignment(horizontal='left', vertical='top', wrap_text=True)


class TemplateExcelWriter:
    """
    Excel writer that fills pre-styled master templates.
    
    Templates have pre-defined styles - this class only fills in data,
    preserving the original formatting from the template.
    """
    
    # Template paths
    TEMPLATES_DIR = os.path.join(os.path.dirname(__file__), '..', '..', 'Templates Master')
    INDEX_TEMPLATE = '$index.xlsx'
    ENDPOINT_TEMPLATE = 'endpoint.xlsx'
    
    # Response status colors
    SUCCESS_FILL = PatternFill(start_color='FF92D050', end_color='FF92D050', fill_type='solid')  # Green
    ERROR_FILL = PatternFill(start_color='FFFF6B6B', end_color='FFFF6B6B', fill_type='solid')    # Red
    
    def __init__(self, template_type: str = 'endpoint'):
        """
        Initialize the template writer.
        
        Args:
            template_type: 'index' or 'endpoint'
        """
        self.template_type = template_type
        self.workbook: Optional[Workbook] = None
        self._response_template_ws: Optional[Worksheet] = None
        
    def load_template(self) -> Workbook:
        """Load the appropriate master template."""
        if self.template_type == 'index':
            template_path = os.path.join(self.TEMPLATES_DIR, self.INDEX_TEMPLATE)
        else:
            template_path = os.path.join(self.TEMPLATES_DIR, self.ENDPOINT_TEMPLATE)
        
        if not os.path.exists(template_path):
            raise FileNotFoundError(f"Template not found: {template_path}")
        
        self.workbook = load_workbook(template_path)
        
        # Store reference to Response template for cloning
        if 'Response' in self.workbook.sheetnames:
            self._response_template_ws = self.workbook['Response']
        
        return self.workbook
    
    def fill_parameters_sheet(self, data_rows: List[Dict[str, Any]]) -> None:
        """
        Fill the Parameters sheet with data.
        
        Args:
            data_rows: List of dicts where keys match column headers
        """
        if self.workbook is None:
            self.load_template()
        
        ws = self.workbook['Parameters']
        self._fill_sheet_data(ws, data_rows, start_row=3)
    
    def fill_body_sheet(self, data_rows: List[Dict[str, Any]], 
                        description: str = '',
                        mandatory: str = 'M') -> None:
        """
        Fill the Body sheet with data.
        
        Args:
            data_rows: List of dicts where keys match column headers
            description: Request body description (for cell B1)
            mandatory: 'M' or 'O' (for cell C1)
        """
        if self.workbook is None:
            self.load_template()
        
        ws = self.workbook['Body']
        
        # Update title row info
        ws.cell(row=1, column=2, value=description)
        ws.cell(row=1, column=3, value=mandatory)
        
        self._fill_sheet_data(ws, data_rows, start_row=3)
    
    def fill_body_example_sheet(self, examples: List[Dict[str, str]]) -> None:
        """
        Fill the Body Example sheet.
        
        Args:
            examples: List of {'name': str, 'body': str}
        """
        if self.workbook is None:
            self.load_template()
        
        ws = self.workbook['Body Example']
        
        for row_idx, example in enumerate(examples, start=2):
            cell_name = ws.cell(row=row_idx, column=1, value=example.get('name', ''))
            cell_body = ws.cell(row=row_idx, column=2, value=example.get('body', ''))
            
            # Apply global alignment
            cell_name.alignment = DEFAULT_ALIGNMENT
            cell_body.alignment = DEFAULT_ALIGNMENT
    
    def add_response_sheet(self, status_code: str, description: str,
                           data_rows: List[Dict[str, Any]]) -> Worksheet:
        """
        Add a response sheet by cloning the Response template.
        
        Args:
            status_code: HTTP status code (e.g., '201', '400')
            description: Response description
            data_rows: Response body structure rows
            
        Returns:
            The newly created worksheet
        """
        if self.workbook is None:
            self.load_template()
        
        # Clone the Response template sheet
        if self._response_template_ws is None:
            raise ValueError("Response template sheet not found in workbook")
        
        # Create new sheet with status code name
        new_ws = self.workbook.copy_worksheet(self._response_template_ws)
        new_ws.title = status_code
        
        # Update title row
        # Cell A1: "Response" (keep), B1: status code, C1: description
        new_ws.cell(row=1, column=2, value=status_code)
        new_ws.cell(row=1, column=3, value=description)
        
        # Apply TAB COLOR based on status code class
        # (Colors the sheet tab at the bottom, not the cells)
        status_int = int(status_code) if status_code.isdigit() else 0
        if 200 <= status_int < 300:
            tab_color = '92D050'  # Green
        else:
            tab_color = 'FF6B6B'  # Red
        
        new_ws.sheet_properties.tabColor = tab_color
        
        # Fill data rows
        self._fill_sheet_data(new_ws, data_rows, start_row=3)
        
        return new_ws
    
    def finalize(self) -> None:
        """
        Finalize the workbook - remove the template Response sheet.
        Call this before saving.
        """
        if self.workbook and 'Response' in self.workbook.sheetnames:
            # Only remove if we have at least one actual response sheet
            response_sheets = [s for s in self.workbook.sheetnames 
                             if s.isdigit()]
            if response_sheets:
                del self.workbook['Response']
    
    def save(self, filepath: str) -> None:
        """
        Save the workbook to file.
        
        Args:
            filepath: Output file path
        """
        if self.workbook is None:
            raise ValueError("No workbook loaded")
        
        # Ensure directory exists
        os.makedirs(os.path.dirname(filepath), exist_ok=True)
        
        # Finalize before saving
        self.finalize()
        
        self.workbook.save(filepath)
    
    def _fill_sheet_data(self, ws: Worksheet, data_rows: List[Dict[str, Any]], 
                         start_row: int) -> None:
        """
        Fill a sheet with data rows, mapping by header column names.
        
        Preserves the style from the template row.
        
        Args:
            ws: Worksheet to fill
            data_rows: Data to insert
            start_row: First data row (after headers)
        """
        if not data_rows:
            return
        
        # Get header row (row before start_row)
        header_row = start_row - 1
        
        # Build column index by header name
        col_map = {}
        for col in range(1, ws.max_column + 1):
            header_val = ws.cell(row=header_row, column=col).value
            if header_val:
                col_map[str(header_val)] = col
        
        # Get template row style (first data row style from template)
        template_styles = {}
        for col in range(1, ws.max_column + 1):
            template_cell = ws.cell(row=start_row, column=col)
            template_styles[col] = {
                'font': copy.copy(template_cell.font),
                'fill': copy.copy(template_cell.fill),
                'border': copy.copy(template_cell.border),
                'alignment': copy.copy(template_cell.alignment),
                'number_format': template_cell.number_format
            }
        
        # Fill data
        for row_idx, row_data in enumerate(data_rows, start=start_row):
            for header, value in row_data.items():
                if header in col_map:
                    col = col_map[header]
                    cell = ws.cell(row=row_idx, column=col, value=value)
                    
                    # Apply template style (font, fill, border, number_format)
                    if col in template_styles:
                        style = template_styles[col]
                        cell.font = style['font']
                        cell.fill = style['fill']
                        cell.border = style['border']
                        cell.number_format = style['number_format']
                    
                    # Apply global alignment (left, top, word wrap) to ALL cells
                    cell.alignment = DEFAULT_ALIGNMENT


# Backward compatibility wrapper
class ExcelWriter(TemplateExcelWriter):
    """Backward-compatible alias for TemplateExcelWriter."""
    
    def __init__(self, styles_config_path: Optional[str] = None, 
                 file_type: str = 'operation'):
        # Map file_type to template_type
        template_type = 'endpoint' if file_type == 'operation' else 'index'
        super().__init__(template_type=template_type)
    
    def create_workbook(self) -> Workbook:
        """Load template instead of creating empty workbook."""
        return self.load_template()
    
    def add_sheet(self, sheet_name: str, data_rows: List[Dict[str, Any]],
                  column_headers: Optional[List[str]] = None,
                  title: Optional[str] = None) -> None:
        """
        Add data to a sheet (backward compatible method).
        
        For response sheets (numeric names), clones the Response template.
        For other sheets, fills existing template sheet.
        """
        if self.workbook is None:
            self.load_template()
        
        if sheet_name.isdigit():
            # Response sheet - clone template
            desc = title or f"Response {sheet_name}"
            self.add_response_sheet(sheet_name, desc, data_rows)
        elif sheet_name == 'Parameters':
            self.fill_parameters_sheet(data_rows)
        elif sheet_name == 'Body':
            self.fill_body_sheet(data_rows, description=title or '')
        elif sheet_name in self.workbook.sheetnames:
            ws = self.workbook[sheet_name]
            # Determine start row (after headers)
            start_row = 2
            for row in range(1, 5):
                if ws.cell(row=row, column=1).value:
                    start_row = row + 1
                    break
            self._fill_sheet_data(ws, data_rows, start_row=start_row)
        else:
            print(f"Warning: Sheet '{sheet_name}' not in template")
