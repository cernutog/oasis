
import os
import shutil
import openpyxl
from typing import List, Optional
from .v3_models import (
    SchemaGraph,
    EnrichedNode,
    EnrichedOperation,
    NodeClass
)

class V3Writer:
    def __init__(self, graph: SchemaGraph, master_dir: str):
        self.graph = graph
        self.master_dir = master_dir

    def write_all(self, output_dir: str):
        if os.path.exists(output_dir):
            shutil.rmtree(output_dir)
        os.makedirs(output_dir)

        # 1. Generate $index.xlsx
        self._write_index(output_dir)

        # 2. Generate Operation Files
        for op_id, op in self.graph.operations.items():
            self._write_operation_file(output_dir, op)

    def _write_index(self, output_dir: str):
        master_index = os.path.join(self.master_dir, "$index.xlsx")
        target_index = os.path.join(output_dir, "$index.xlsx")
        shutil.copy(master_index, target_index)

        wb = openpyxl.load_workbook(target_index)
        
        # A. Fill Global Schemas (RECURSIVE)
        if "Schemas" in wb.sheetnames:
            ws = wb["Schemas"]
            # Restore clearing to ensure spacers work and old data is removed
            for row in ws.iter_rows(min_row=2):
                for cell in row: cell.value = None
            
            self._write_schemas(ws)

        # B. Fill Paths
        if "Paths" in wb.sheetnames:
            ws = wb["Paths"]
            for idx, (op_id, op) in enumerate(self.graph.operations.items(), start=2):
                ws.cell(row=idx, column=1, value=f"{op_id}.xlsx")
                ws.cell(row=idx, column=8, value=op_id)
        
        wb.save(target_index)
    def _write_schemas(self, ws):
        # Build map for lookups
        all_nodes = {n.name: n for n in self.graph.global_schemas}
        sorted_names = sorted(all_nodes.keys(), key=lambda x: x.lower())
        
        written_schemas = set()
        self.current_row = 2
        
        def write_tree(node_name):
            if node_name in written_schemas or node_name not in all_nodes:
                return
            
            node = all_nodes[node_name]
            # Write the node (and its inline properties via existing recursive logic)
            self.current_row = self._write_node_recursive(ws, self.current_row, [node], is_index=True)
            written_schemas.add(node_name)
            
            # Identify references to other Global Schemas (Children)
            # We want to pull them under this parent
            children_refs = []
            if node.properties:
                for p in node.properties:
                    # Check direct type ref
                    if p.type_literal in all_nodes:
                        children_refs.append(p.type_literal)
                    # Check items type ref (for arrays)
                    elif p.items_type_literal in all_nodes:
                        children_refs.append(p.items_type_literal)
            
            # Sort children to ensure deterministic sub-order (Legacy template order preferred, but alphabetical is safe fallback if mapping not explicit)
            # Actually, let's keep them in property definition order to match "encountered" order.
            
            for child in children_refs:
                if child not in written_schemas:
                    write_tree(child)

        for name in sorted_names:
            if name not in written_schemas:
                # Spacer between Root trees (but not at start)
                if self.current_row > 2:
                    self.current_row += 1
                
                write_tree(name)

    def _write_operation_file(self, output_dir: str, op: EnrichedOperation):
        master_op = os.path.join(self.master_dir, "endpoint.xlsx")
        target_op = os.path.join(output_dir, f"{op.operation_id}.xlsx")
        shutil.copy(master_op, target_op)

        wb = openpyxl.load_workbook(target_op)
        
        if op.request:
            if "Body" in wb.sheetnames:
                ws = wb["Body"]
                for row in ws.iter_rows(min_row=3):
                    for cell in row: cell.value = None
                # Request Root as Reference
                self._write_node_row(ws, 3, op.request, is_index=False, section="Body")

        for status, resp in op.responses.items():
            if status not in wb.sheetnames:
                source = wb["Body"]
                ws = wb.copy_worksheet(source)
                ws.title = status
            else:
                ws = wb[status]
            
            ws.cell(row=1, column=1, value="Response")
            ws.cell(row=1, column=2, value=status)
            ws.cell(row=1, column=3, value=resp.description or f"Response {status}")
            
            for row in ws.iter_rows(min_row=3):
                for cell in row: cell.value = None
            # Response Root as Reference
            self._write_node_row(ws, 3, resp, is_index=False, section="Content")

        wb.save(target_op)

    def _write_node_recursive(self, ws, start_row: int, nodes: List[EnrichedNode], is_index: bool, section: Optional[str] = None, parent_name: Optional[str] = None) -> int:
        current_row = start_row
        for node in nodes:
            # Atomic Row Write
            self._write_node_row(ws, current_row, node, is_index=is_index, section=section, parent_name=parent_name)
            current_row += 1
            
            # Recurse children if complex (object/array)
            # Only recurse if we are defining the structure (not just referencing it)
            # In the Schemas sheet, we ALWAYS define the structure.
            # In Operation sheets, we are now using roots as refs, so properties are empty anyway.
            if node.properties:
                 current_row = self._write_node_recursive(ws, current_row, node.properties, is_index, section, node.name)
                 
        return current_row

    def _write_node_row(self, ws, row: int, node: EnrichedNode, is_index: bool, section: Optional[str] = None, parent_name: Optional[str] = None):
        if is_index:
            m = {
                "name": 1, "parent": 2, "description": 3, "type": 4, 
                "items": 5, "ref": 6, "fmt": 7, "mand": 8,
                "min": 9, "max": 10, "peba": 11, "reg": 12, "allw": 13, "ex": 14
            }
        else:
            m = {
                "sect": 1, "name": 2, "parent": 3, "description": 4, "type": 5,
                "items": 6, "ref": 7, "fmt": 8, "mand": 9,
                "min": 10, "max": 11, "peba": 12, "reg": 13, "allw": 14, "ex": 15
            }

        ws.cell(row=row, column=m["name"], value=node.name)
        ws.cell(row=row, column=m["parent"], value=parent_name or "")
        ws.cell(row=row, column=m["description"], value=node.description)
        
        type_val = node.type_literal
        schema_ref = ""
        primitives = ["object", "array", "string", "number", "integer", "boolean"]
        
        # LOGIC:
        # 1. If it's a non-primitive type AND NOT a root definition in the index...
        # 2. Or if it's intentionally a reference to a registered domain concept...
        # We use 'schema' type + 'ref' value.
        # BUT: For a Domain Concept root row in the index, we want its base identity.
        
        if type_val and type_val.lower() not in primitives:
            # Check if this is a leaf reference or a promoted structural reference
            # If we are in the index AND it's a root (parent_name is empty), 
            # the 'type_literal' might be the schema name itself.
            # We must use its canonical underlying structure.
            
            # For simplicity: if name == type_literal (case-insensitive), it's likely a self-reference, force object.
            if node.name.lower() == type_val.lower() and not parent_name:
                type_val = "object" if node.properties else "string"
            else:
                schema_ref = type_val
                type_val = "schema"
            
        ws.cell(row=row, column=m["type"], value=type_val)
        ws.cell(row=row, column=m["ref"], value=schema_ref)
        ws.cell(row=row, column=m["items"], value=node.items_type_literal)
        ws.cell(row=row, column=m["fmt"], value=node.format)
        ws.cell(row=row, column=m["mand"], value="M" if node.mandatory else "O")
        ws.cell(row=row, column=m["min"], value=node.min_val)
        ws.cell(row=row, column=m["max"], value=node.max_val)
        ws.cell(row=row, column=m["peba"], value=node.pattern_eba)
        ws.cell(row=row, column=m["reg"], value=node.regex)
        ws.cell(row=row, column=m["allw"], value=node.allowed_values)
        ws.cell(row=row, column=m["ex"], value=node.example)
        
        if not is_index and section:
            ws.cell(row=row, column=m["sect"], value=section)
