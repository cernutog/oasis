import openpyxl

FILE = "API Templates/v1_transactions_investigations_reports.251111.xlsm"

def fix():
    print(f"Opening {FILE}...")
    wb = openpyxl.load_workbook(FILE, keep_vba=True)
    if "400" not in wb.sheetnames:
        print("Sheet 400 not found!")
        return
        
    sheet = wb["400"]
    # Target Row 3 (1-based index)
    # Check current values
    val_name = sheet.cell(row=3, column=2).value # Col B
    val_type = sheet.cell(row=3, column=5).value # Col E
    
    print(f"Current Row 3 Col B (Name): {val_name}")
    print(f"Current Row 3 Col E (Type): {val_type}")
    
    if val_name == "Bad Request" and val_type == "response":
        print("Applying fix...")
        sheet.cell(row=3, column=2).value = "fri"
        sheet.cell(row=3, column=5).value = "header"
        wb.save(FILE)
        print("File saved.")
    else:
        print("Values do not match expected incorrect state. verifying 'headers' section...")
        # Maybe row is different?
        # Let's iterate rows 1 to 5 to find 'headers' section
        for r in range(1, 10):
            sec = sheet.cell(row=r, column=1).value
            if sec == "headers":
                print(f"Found headers at row {r}")
                v_name = sheet.cell(row=r, column=2).value
                v_type = sheet.cell(row=r, column=5).value
                print(f"  Name: {v_name}, Type: {v_type}")
                
                if v_name == "Bad Request" and v_type == "response":
                     sheet.cell(row=r, column=2).value = "fri"
                     sheet.cell(row=r, column=5).value = "header"
                     wb.save(FILE)
                     print("File saved.")
                     return

if __name__ == "__main__":
    fix()
