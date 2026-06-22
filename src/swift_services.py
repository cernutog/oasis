from __future__ import annotations

from copy import copy, deepcopy
from pathlib import Path
from typing import Any, Callable

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, range_boundaries


SWIFT_SERVICES_PREFERENCE_KEY = "swift_services"
SWIFT_SERVER_URL_KEY = "swift servers url"
SWIFT_SERVER_DESCRIPTION_KEY = "servers description"
MAX_TEMPLATE_SWIFT_SERVERS = 2


def normalize_swift_services(value: Any) -> dict[str, dict[str, list[dict[str, str]]]]:
    """Normalize the configurable SWIFT service catalog."""
    if not isinstance(value, dict):
        return {}

    services: dict[str, dict[str, list[dict[str, str]]]] = {}
    for raw_name, raw_service in value.items():
        name = str(raw_name or "").strip()
        if not name or not isinstance(raw_service, dict):
            continue

        servers = []
        raw_servers = raw_service.get("servers", [])
        if isinstance(raw_servers, list):
            for raw_server in raw_servers:
                if not isinstance(raw_server, dict):
                    continue
                servers.append(
                    {
                        "url": str(raw_server.get("url", "") or "").strip(),
                        "description": str(raw_server.get("description", "") or "").strip(),
                    }
                )
        services[name] = {"servers": servers}

    return services


def get_swift_service_servers(
    services: Any,
    service_name: str,
    log_callback: Callable[[str], None] | None = None,
) -> list[dict[str, str]]:
    catalog = normalize_swift_services(services)
    service = catalog.get(str(service_name or "").strip())
    if not service:
        return []

    servers = deepcopy(service.get("servers", []))
    if len(servers) > MAX_TEMPLATE_SWIFT_SERVERS and log_callback:
        log_callback(
            f"WARNING: Service '{service_name}' has {len(servers)} configured servers; "
            f"only the first {MAX_TEMPLATE_SWIFT_SERVERS} were written to the template."
        )
    return servers[:MAX_TEMPLATE_SWIFT_SERVERS]


def has_swift_server_rows(index_path: str | Path) -> bool:
    workbook = load_workbook(index_path, read_only=True, data_only=True)
    try:
        if "General Description" not in workbook.sheetnames:
            return False
        ws = workbook["General Description"]
        for row in ws.iter_rows(min_col=1, max_col=1):
            value = row[0].value
            if str(value or "").strip().lower() == SWIFT_SERVER_URL_KEY:
                return True
        return False
    finally:
        workbook.close()


def read_swift_servers_from_general_description(df_info) -> list[dict[str, str]]:
    servers: list[dict[str, str]] = []
    if df_info is None:
        return servers

    for _, row in df_info.iterrows():
        key = str(row.iloc[0]).strip().lower()
        if key != SWIFT_SERVER_URL_KEY:
            continue

        url = row.iloc[1] if len(row) > 1 else ""
        if url is None or str(url).strip().lower() == "nan" or not str(url).strip():
            continue

        server = {"url": str(url).strip()}
        if len(row) > 3:
            desc = row.iloc[3]
            if desc is not None and str(desc).strip().lower() != "nan" and str(desc).strip():
                server["description"] = str(desc).strip()
        servers.append(server)

    return servers


def ensure_swift_server_rows_in_workbook(workbook, servers: list[dict[str, str]] | None = None) -> None:
    if "General Description" not in workbook.sheetnames:
        return

    ws = workbook["General Description"]
    style_source_rows = _find_server_style_rows(ws)
    swift_rows = [
        row_idx
        for row_idx in range(1, ws.max_row + 1)
        if str(ws.cell(row=row_idx, column=1).value or "").strip().lower() == SWIFT_SERVER_URL_KEY
    ]

    if not swift_rows:
        insert_at = _find_swift_insert_row(ws)
        _insert_rows_preserving_merged_ranges(ws, insert_at, MAX_TEMPLATE_SWIFT_SERVERS)
        for offset in range(MAX_TEMPLATE_SWIFT_SERVERS):
            row_idx = insert_at + offset
            _label_swift_row(ws, row_idx)
            swift_rows.append(row_idx)
    elif len(swift_rows) < MAX_TEMPLATE_SWIFT_SERVERS:
        insert_at = swift_rows[-1] + 1
        _insert_rows_preserving_merged_ranges(ws, insert_at, MAX_TEMPLATE_SWIFT_SERVERS - len(swift_rows))
        swift_rows = swift_rows + [
            insert_at + offset
            for offset in range(MAX_TEMPLATE_SWIFT_SERVERS - len(swift_rows))
        ]
        for row_idx in swift_rows:
            _label_swift_row(ws, row_idx)
    else:
        swift_rows = swift_rows[:MAX_TEMPLATE_SWIFT_SERVERS]

    _unmerge_ranges_intersecting_rows(ws, swift_rows)

    for idx, row_idx in enumerate(swift_rows):
        _copy_server_row_style(ws, style_source_rows, idx, row_idx)
        _label_swift_row(ws, row_idx)

    values = list(servers)[:MAX_TEMPLATE_SWIFT_SERVERS] if servers is not None else None
    for idx, row_idx in enumerate(swift_rows):
        if values is None:
            continue
        server = values[idx] if idx < len(values) else {}
        ws.cell(row=row_idx, column=2).value = server.get("url", "")
        ws.cell(row=row_idx, column=4).value = server.get("description", "")


def update_index_swift_servers(index_path: str | Path, servers: list[dict[str, str]]) -> None:
    workbook = load_workbook(index_path)
    try:
        ensure_swift_server_rows_in_workbook(workbook, servers)
        workbook.save(index_path)
    finally:
        workbook.close()


def _find_swift_insert_row(ws) -> int:
    for row_idx in range(1, ws.max_row + 1):
        key = str(ws.cell(row=row_idx, column=1).value or "").strip().lower()
        if key == "release":
            return row_idx
    return min(ws.max_row + 1, 9)


def _label_swift_row(ws, row_idx: int) -> None:
    ws.cell(row=row_idx, column=1).value = SWIFT_SERVER_URL_KEY
    ws.cell(row=row_idx, column=3).value = SWIFT_SERVER_DESCRIPTION_KEY


def _insert_rows_preserving_merged_ranges(ws, insert_at: int, amount: int) -> None:
    shifted_ranges = []
    for merged_range in list(ws.merged_cells.ranges):
        min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
        if min_row >= insert_at:
            ws.unmerge_cells(str(merged_range))
            shifted_ranges.append(
                f"{get_column_letter(min_col)}{min_row + amount}:"
                f"{get_column_letter(max_col)}{max_row + amount}"
            )

    ws.insert_rows(insert_at, amount)

    for shifted_range in shifted_ranges:
        ws.merge_cells(shifted_range)


def _unmerge_ranges_intersecting_rows(ws, rows: list[int]) -> None:
    target_rows = set(rows)
    for merged_range in list(ws.merged_cells.ranges):
        _min_col, min_row, _max_col, max_row = range_boundaries(str(merged_range))
        if any(min_row <= row <= max_row for row in target_rows):
            ws.unmerge_cells(str(merged_range))


def _find_server_style_rows(ws) -> list[int]:
    rows = [
        row_idx
        for row_idx in range(1, ws.max_row + 1)
        if str(ws.cell(row=row_idx, column=1).value or "").strip().lower() == "servers url"
    ]
    return rows[-MAX_TEMPLATE_SWIFT_SERVERS:]


def _copy_server_row_style(ws, source_rows: list[int], source_index: int, target_row: int) -> None:
    if not source_rows:
        return

    source_row = source_rows[min(source_index, len(source_rows) - 1)]
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
    for col_idx in range(1, max(ws.max_column, 4) + 1):
        source_cell = ws.cell(row=source_row, column=col_idx)
        target_cell = ws.cell(row=target_row, column=col_idx)
        target_cell._style = copy(source_cell._style)
