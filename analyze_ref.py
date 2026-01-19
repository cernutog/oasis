"""Analyze reference template structure for row merging and per-column styles."""

from openpyxl import load_workbook

wb = load_workbook('API Templates/account_assessment.251111.xlsm')
ws = wb['Parameters']

print('=== Row 1 Merge Info ===')
for merge in ws.merged_cells.ranges:
    print(f'Merged: {merge}')

print()
print('=== Row 2 Headers and Styles ===')
for col in range(1, 15):
    cell = ws.cell(row=2, column=col)
    try:
        fill_rgb = cell.fill.fgColor.rgb if cell.fill.fgColor else None
        if 'class' in str(fill_rgb):
            fill = f'theme={cell.fill.fgColor.theme}+tint={cell.fill.fgColor.tint:.2f}'
        else:
            fill = fill_rgb
    except:
        fill = f'theme={cell.fill.fgColor.theme}+tint={cell.fill.fgColor.tint:.2f}'
    print(f'Col {col}: "{cell.value}" | fill={fill}')

print()
print('=== Row 1 Style ===')
cell = ws.cell(row=1, column=1)
try:
    fill = f'theme={cell.fill.fgColor.theme}+tint={cell.fill.fgColor.tint:.2f}'
except:
    fill = cell.fill.fgColor.rgb if cell.fill.fgColor else None
print(f'Title: "{cell.value}" | fill={fill}')

wb.close()
