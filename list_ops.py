"""Check operations in imported index."""
import openpyxl

wb = openpyxl.load_workbook(r'Imported Templates\$index.xlsx')
ws = wb['Paths']

print('Operations in imported $index.xlsx:')
for r in range(3, 20):
    path = ws.cell(row=r, column=2).value
    if path:
        method = ws.cell(row=r, column=4).value
        print(f'  Row {r}: {method} {path}')
