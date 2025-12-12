import openpyxl
import os

base_dir = r"C:\Users\giuse\.gemini\antigravity\scratch\OAS_Generation_Tool\API Templates"
file_path = os.path.join(base_dir, "$index.xlsm")

print(f"Loading {file_path}...")
try:
    wb = openpyxl.load_workbook(file_path, keep_vba=True)
    if "Parameters" in wb.sheetnames:
        ws = wb["Parameters"]
        
        # Find headers
        headers = {}
        for cell in ws[1]:
            headers[cell.value] = cell.column
            
        print(f"Headers found: {headers.keys()}")
        
        name_col = headers.get("Name")
        in_col = headers.get("In")
        mand_col = headers.get("Mandatory")
        
        if name_col:
            for row in ws.iter_rows(min_row=2):
                name_val = row[name_col - 1].value
                if name_val == "fuid":
                    print(f"Found 'fuid' at row {row[0].row}")
                    
                    if in_col:
                        old_in = row[in_col - 1].value
                        row[in_col - 1].value = "path"
                        print(f"Updated 'In': '{old_in}' -> 'path'")
                    
                    if mand_col:
                        old_mand = row[mand_col - 1].value
                        row[mand_col - 1].value = "Yes"
                        print(f"Updated 'Mandatory': '{old_mand}' -> 'Yes'")
                    
                    break
            
            wb.save(file_path)
            print("File saved successfully.")
        else:
            print("Could not find 'Name' column.")
    else:
        print("Sheet 'Parameters' not found.")
        
except Exception as e:
    print(f"Error: {e}")
