from __future__ import annotations

from difflib import SequenceMatcher
from pathlib import Path

import openpyxl


def find_header(ws) -> tuple[int, dict[str, int]] | None:
    for row in range(1, min(ws.max_row, 20) + 1):
        headers: dict[str, int] = {}
        for col in range(1, ws.max_column + 1):
            value = str(ws.cell(row=row, column=col).value or "").strip().lower()
            if value:
                headers[value] = col
        if "name" in headers and "example" in headers:
            return row, headers
    return None


def split_values(raw: str) -> list[str]:
    parts = [part.strip() for part in raw.replace(",", ";").split(";") if part.strip()]
    return parts


def shape_signature(value: str) -> str:
    mapped = []
    for ch in value:
        if ch.isupper():
            mapped.append("A")
        elif ch.islower():
            mapped.append("a")
        elif ch.isdigit():
            mapped.append("9")
        elif ch.isspace():
            mapped.append(" ")
        else:
            mapped.append(ch)
    return "".join(mapped)


def values_are_similar(values: list[str]) -> bool:
    if len(values) < 2:
        return False

    shapes = {shape_signature(v) for v in values}
    if len(shapes) == 1:
        return True

    ratios = []
    for i in range(len(values)):
        for j in range(i + 1, len(values)):
            ratios.append(SequenceMatcher(None, values[i], values[j]).ratio())
    return bool(ratios) and min(ratios) >= 0.60


def allowed_values_confirm(values: list[str], allowed: str) -> bool:
    if len(values) < 2 or len(values) > 3 or not allowed:
        return False
    allowed_parts = split_values(allowed)
    return len(allowed_parts) >= len(values) and values == allowed_parts[: len(values)]


def should_normalize(example: str, allowed: str) -> bool:
    if "," not in example:
        return False
    values = split_values(example)
    if len(values) < 2 or len(values) > 3:
        return False
    if allowed:
        allowed_parts = split_values(allowed)
        if len(allowed_parts) >= len(values) and values == allowed_parts[: len(values)]:
            return True
    return allowed_values_confirm(values, allowed) or values_are_similar(values)


def normalize_example(example: str) -> str:
    return "; ".join(split_values(example))


def process_file(path: Path) -> int:
    wb = openpyxl.load_workbook(path)
    updates = 0
    try:
        if "Data Type" not in wb.sheetnames:
            return 0

        ws = wb["Data Type"]
        hdr = find_header(ws)
        if not hdr:
            return 0

        header_row, headers = hdr
        name_col = headers["name"]
        example_col = headers["example"]
        allowed_col = headers.get("allowed value")

        for row in range(header_row + 1, ws.max_row + 1):
            name = str(ws.cell(row=row, column=name_col).value or "").strip()
            if not name:
                continue

            example = str(ws.cell(row=row, column=example_col).value or "").strip()
            if not example:
                continue

            allowed = ""
            if allowed_col:
                allowed = str(ws.cell(row=row, column=allowed_col).value or "").strip()

            if not should_normalize(example, allowed):
                if "," in example:
                    print(f"SKIPPED|{path.name}|{name}|{example}")
                continue

            normalized = normalize_example(example)
            if normalized != example:
                ws.cell(row=row, column=example_col).value = normalized
                updates += 1

        if updates:
            wb.save(path)
        return updates
    finally:
        wb.close()


def main(root: str) -> None:
    base = Path(root)
    if not base.exists():
        raise SystemExit(f"Directory not found: {base}")

    file_updates = 0
    cell_updates = 0
    for path in sorted(base.glob("*.xlsm")):
        if path.name.startswith(("~", "$")):
            continue
        updates = process_file(path)
        if updates:
            file_updates += 1
            cell_updates += updates
            print(f"UPDATED|{path.name}|{updates}")

    print(f"UPDATED_FILES={file_updates}")
    print(f"UPDATED_CELLS={cell_updates}")


if __name__ == "__main__":
    import sys

    if len(sys.argv) != 2:
        raise SystemExit("Usage: normalize_legacy_example_separators.py <legacy_dir>")
    main(sys.argv[1])
