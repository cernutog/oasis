
import os
import sys
from collections import OrderedDict

# Add src to path
sys.path.append(os.path.abspath("src"))

from legacy_converter import LegacyConverter

legacy_dir = r"c:\Users\giuse\.gemini\antigravity\scratch\OASIS\Templates Legacy"
output_dir = r"c:\Users\giuse\.gemini\antigravity\scratch\OASIS\Templates Converted Test"

conv = LegacyConverter(legacy_dir, output_dir)
conv.convert()

print("\n--- DEBUG: SCHEMA NAMING ---")
if "Amount" in conv.global_registry:
    print(f"Amount variants in registry: {len(conv.global_registry['Amount'])}")
    used = [(v) for v in conv.global_registry["Amount"] if ("Amount", v) in conv.used_global_schemas]
    print(f"Used Amount variants: {len(used)}")
    
    mapping = conv.schema_collision_map.get("Amount", {})
    for v_tuple, final_name in mapping.items():
        v_data = conv.global_registry["Amount"][v_tuple]
        print(f"  Variant -> {final_name}: Max={v_data.get('max')}, Min={v_data.get('min')}")

# Check global_schemas_data
print(f"\nTotal in global_schemas_data: {len(conv.global_schemas_data)}")
amount_keys = [k for k in conv.global_schemas_data if "Amount" in k]
print(f"Amount-related keys in global_schemas_data: {amount_keys}")

# Verify file existence
index_path = os.path.join(output_dir, "$index.xlsx")
if os.path.exists(index_path):
    from openpyxl import load_workbook
    wb = load_workbook(index_path, data_only=True)
    ws = wb["Schemas"]
    print("\n--- Schemas in $index.xlsx ---")
    for row in ws.iter_rows(values_only=True):
        if row[0] and "Amount" in str(row[0]):
            print(f"  {row[0]} | Max: {row[9]} | Min: {row[8]}")
