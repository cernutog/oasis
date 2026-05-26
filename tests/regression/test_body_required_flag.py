import os
import sys
from tempfile import NamedTemporaryFile

import pandas as pd
from openpyxl import Workbook


sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "../..")))


from src.excel_parser import load_excel_sheet
from src.generator import OASGenerator
from src.legacy_converter import LegacyConverter


def test_converter_marks_body_c1_as_mandatory_when_request_body_exists():
    wb = Workbook()
    ws = wb.active
    ws.title = "Body"

    converter = LegacyConverter(input_dir=".", output_dir=".")
    converter.emitted_wrappers.add("CommandDetailsRequest")

    converter._convert_body(wb, None, "CommandDetails", "commandDetails.xlsx")

    assert ws["B1"].value in (None, "")
    assert ws["C1"].value == "M"


def test_excel_parser_reads_body_description_from_b1_and_required_from_c1():
    wb = Workbook()
    ws = wb.active
    ws.title = "Body"
    ws["B1"] = "Command details request body."
    ws["C1"] = "M"
    ws.append([""])
    ws.append(
        [
            "Name",
            "Parent",
            "Description",
            "Type",
            "Items Data Type \n(Array only)",
            "Schema Name\n(if Type = schema)",
            "Format",
            "Mandatory",
        ]
    )
    ws.append(
        [
            "content",
            "application/json",
            "",
            "schema",
            "",
            "CommandDetailsRequest",
            "",
            "M",
        ]
    )

    with NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp_path = tmp.name

    try:
        wb.save(tmp_path)
        df = load_excel_sheet(tmp_path, "Body")
        assert df is not None
        assert df.attrs.get("body_description") == "Command details request body."
        assert df.attrs.get("body_required") == "M"
    finally:
        os.remove(tmp_path)


def test_generator_sets_request_body_description_and_required_from_body_metadata():
    body_df = pd.DataFrame(
        [
            {
                "Name": "content",
                "Parent": "application/json",
                "Description": "",
                "Type": "schema",
                "Items Data Type \n(Array only)": "",
                "Schema Name\n(if Type = schema)": "CommandDetailsRequest",
                "Format": "",
                "Mandatory": "M",
            }
        ]
    )
    body_df.attrs["sheet_name"] = "Body"
    body_df.attrs["body_description"] = "Command details request body."
    body_df.attrs["body_required"] = "M"

    generator = OASGenerator(version="3.1.0")
    generator.build_paths(
        [
            {
                "path": "/commandDetails/{senderBic}",
                "method": "post",
                "file": "commandDetails.xlsx",
                "operationId": "commandDetails",
            }
        ],
        {
            "commandDetails.xlsx": {
                "body": body_df,
            }
        },
    )

    request_body = generator.oas["paths"]["/commandDetails/{senderBic}"]["post"]["requestBody"]
    assert request_body["description"] == "Command details request body."
    assert request_body["required"] is True

