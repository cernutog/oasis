import os
import sys

import yaml
from openpyxl import load_workbook


sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "../..")))


from src.oas_importer.oas_converter import OASToExcelConverter


def test_oas_import_index_uses_metadata_overrides_for_release_and_filename_pattern(tmp_path):
    source_oas = tmp_path / "source.yaml"
    source_oas.write_text(
        yaml.safe_dump(
            {
                "openapi": "3.1.0",
                "info": {
                    "title": "Payments API",
                    "version": "4.1.0",
                    "description": "Imported API",
                },
                "paths": {},
            },
            sort_keys=False,
        ),
        encoding="utf-8",
    )

    index_path = tmp_path / "$index.xlsx"
    converter = OASToExcelConverter(str(source_oas))
    converter.generate_index_file(
        str(index_path),
        info_overrides={
            "release": "v2026Q4",
            "filename_pattern": "EBACL_<current_date>_OpenApi<oas_version>_NEXI_FPAD_<api_version>_<release>.yaml",
        },
    )

    wb = load_workbook(index_path, data_only=True)
    try:
        ws = wb["General Description"]
        assert ws.cell(row=9, column=2).value == "v2026Q4"
        assert (
            ws.cell(row=10, column=2).value
            == "EBACL_<current_date>_OpenApi<oas_version>_NEXI_FPAD_<api_version>_<release>.yaml"
        )
    finally:
        wb.close()
