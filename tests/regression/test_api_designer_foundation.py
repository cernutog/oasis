from pathlib import Path
import datetime
import os
import re
import shutil
import warnings

import yaml
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from src.api_designer.models import Change, ChangeStep, create_empty_workspace
from src.api_designer.persistence import FileSystemDesignerRepository
from src.excel_parser import parse_info, parse_paths_index
from src.generator import OASGenerator
from src.legacy_converter import DataType, LegacyConverter
import src.legacy_converter_dialog as legacy_converter_dialog_module
from src.legacy_converter_dialog import LegacyConverterDialog
from src.gui import ImportDialog, OASGenApp
from src.oas_importer.schema_flattener import SchemaFlattener
from src.conversion_metadata_preferences import (
    load_metadata_preferences,
    save_metadata_preferences,
)
from src import main as main_script
from src.oas_output_archive import (
    archive_existing_oas_files,
    build_expected_oas_filenames,
    delete_existing_oas_files,
    describe_archive_destinations,
    filter_previous_oas_files,
    list_existing_oas_files,
)
from src.preferences import (
    DEFAULT_SWIFT_SERVICES,
    DEFAULT_X_INFO_OPTIONS,
    GENERATION_MODE_API_PORTAL_READY,
    GENERATION_MODE_MINIMAL,
    GENERATION_MODE_STANDARD,
    PreferencesManager,
    migrate_saved_preferences,
)
from src.swift_services import (
    ensure_swift_server_rows_in_workbook,
    get_swift_service_servers,
)
from src.update_checker import parse_update_manifest
from src.version import FULL_VERSION


TEST_TEMP_ROOT = Path(__file__).resolve().parents[1] / "_tmp_api_designer_foundation"


def _snapshot_files(root: Path) -> dict[str, str]:
    return {
        str(path.relative_to(root)).replace("\\", "/"): path.read_text(encoding="utf-8")
        for path in sorted(root.rglob("*.yaml"))
    }


def _clean_test_temp() -> Path:
    if TEST_TEMP_ROOT.exists():
        shutil.rmtree(TEST_TEMP_ROOT)
    TEST_TEMP_ROOT.mkdir(parents=True)
    return TEST_TEMP_ROOT


def _create_workbook(path: Path, sheets: dict[str, list[list[str]]] | None = None) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    for sheet_name, rows in (sheets or {}).items():
        sheet = workbook.create_sheet(sheet_name)
        for row in rows:
            sheet.append(row)
    workbook.save(path)


def _create_minimal_converted_index(path: Path, endpoint_file: str = "operation.xlsx") -> None:
    _create_workbook(
        path,
        {
            "General Description": [
                ["Key", "Value"],
                ["Info Title", "Payments API"],
                ["Info Version", "1.0"],
            ],
            "Tags": [["Name", "Description"]],
            "Servers": [["URL", "Description"]],
            "Security": [["Name", "Type", "Scheme", "Format", "Description"]],
            "Paths": [
                ["Excel", "Path", "Description", "Method", "Tag", "Summary", "OperationId"],
                [endpoint_file, "/payments", "Payments", "get", "Payments", "List payments", "listPayments"],
            ],
            "Parameters": [["Name", "Description", "In", "Required", "Schema"]],
            "Headers": [["Name", "Description", "Schema"]],
            "Schemas": [["Name", "Parent", "Description", "Type"]],
            "Responses": [["Name", "Description", "Schema"]],
        },
    )


def _create_minimal_endpoint(path: Path) -> None:
    _create_workbook(
        path,
        {
            "Parameters": [["Name", "Description", "In", "Required", "Schema"]],
            "200": [
                ["Response", "200", "OK"],
                ["Name", "Description", "Type", "Schema Name"],
            ],
        },
    )


def test_oas_generation_rejects_legacy_template_folder_before_writing_outputs():
    temp_root = _clean_test_temp()
    output_dir = temp_root / "out"
    _create_workbook(temp_root / "$index.xlsm", {"General Description": [["Key", "Value"]]})
    logs: list[str] = []

    main_script.generate_oas(
        str(temp_root),
        output_dir=str(output_dir),
        log_callback=logs.append,
    )

    log_text = "\n".join(logs)
    assert "OAS Generation requires valid templates with $index.xlsx" in log_text
    assert "appears to contain legacy templates ($index.xlsm)" in log_text
    assert "Please select a folder containing a complete set of valid templates" in log_text
    assert not list(output_dir.rglob("*.yaml")) if output_dir.exists() else True


def test_converted_template_validation_reports_missing_index_sheets():
    temp_root = _clean_test_temp()
    _create_workbook(temp_root / "$index.xlsx", {"General Description": [["Key", "Value"]]})

    errors = main_script.get_converted_template_validation_errors(str(temp_root))

    assert any("Invalid converted template structure" in line for line in errors)
    assert any("Missing sheet(s):" in line and "Paths" in line and "Schemas" in line for line in errors)
    assert any("complete set of valid templates" in line for line in errors)


def test_converted_template_validation_allows_optional_servers_and_security_sheets():
    temp_root = _clean_test_temp()
    _create_minimal_converted_index(temp_root / "$index.xlsx", endpoint_file="operation.xlsx")
    _create_minimal_endpoint(temp_root / "operation.xlsx")
    workbook = load_workbook(temp_root / "$index.xlsx")
    del workbook["Servers"]
    del workbook["Security"]
    workbook.save(temp_root / "$index.xlsx")
    workbook.close()

    errors = main_script.get_converted_template_validation_errors(str(temp_root))

    assert errors == []


def test_converted_template_validation_reports_endpoint_files_with_missing_sheets():
    temp_root = _clean_test_temp()
    _create_minimal_converted_index(temp_root / "$index.xlsx", endpoint_file="operation.xlsx")
    _create_workbook(temp_root / "operation.xlsx", {"Path": [["Name", "Value"]]})

    errors = main_script.get_converted_template_validation_errors(str(temp_root))

    assert any("Invalid converted template structure" in line for line in errors)
    assert any("operation.xlsx" in line for line in errors)
    assert any("Missing sheet(s):" in line and "Parameters" in line for line in errors)


def test_oas_output_archive_moves_oas_files_and_source_maps_by_modified_date():
    temp_root = _clean_test_temp()
    output_dir = temp_root / "Generated OAS"
    map_dir = output_dir / ".oasis_excel_maps"
    output_dir.mkdir(parents=True)
    map_dir.mkdir()

    first = output_dir / "api.yaml"
    second = output_dir / "api.json"
    ignored = output_dir / "notes.txt"
    first.write_text("openapi: 3.1.0\n", encoding="utf-8")
    second.write_text('{"openapi":"3.0.0"}', encoding="utf-8")
    ignored.write_text("keep me", encoding="utf-8")
    (map_dir / "api.yaml.map.json").write_text('{"source":"api"}', encoding="utf-8")
    (map_dir / "api.json.map.json").write_text('{"source":"json"}', encoding="utf-8")

    first_ts = 1715428800  # 2024-05-11
    second_ts = 1715515200  # 2024-05-12
    for path, ts in [(first, first_ts), (second, second_ts)]:
        path.touch()
        os.utime(path, (ts, ts))

    existing = list_existing_oas_files(output_dir)
    assert [path.name for path in existing] == ["api.json", "api.yaml"]

    moved = archive_existing_oas_files(existing, output_dir)

    assert not first.exists()
    assert not second.exists()
    assert ignored.exists()
    assert (output_dir / "Archive" / "20240511" / "api.yaml").exists()
    assert (output_dir / "Archive" / "20240511" / ".oasis_excel_maps" / "api.yaml.map.json").exists()
    assert (output_dir / "Archive" / "20240512" / "api.json").exists()
    assert (output_dir / "Archive" / "20240512" / ".oasis_excel_maps" / "api.json.map.json").exists()
    assert {dest.name for _, dest in moved} == {"api.yaml", "api.json"}


def test_oas_output_archive_renames_colliding_archived_files_and_can_delete_current_files():
    temp_root = _clean_test_temp()
    output_dir = temp_root / "Generated OAS"
    archive_dir = output_dir / "Archive" / "20240511"
    map_dir = output_dir / ".oasis_excel_maps"
    output_dir.mkdir(parents=True)
    archive_dir.mkdir(parents=True)
    map_dir.mkdir()

    current = output_dir / "api.yaml"
    current.write_text("new", encoding="utf-8")
    (map_dir / "api.yaml.map.json").write_text("map", encoding="utf-8")
    (archive_dir / "api.yaml").write_text("old", encoding="utf-8")

    os.utime(current, (1715428800, 1715428800))

    archive_existing_oas_files([current], output_dir)

    assert (archive_dir / "api.yaml").read_text(encoding="utf-8") == "old"
    assert (archive_dir / "api_1.yaml").read_text(encoding="utf-8") == "new"
    assert (archive_dir / ".oasis_excel_maps" / "api_1.yaml.map.json").exists()

    fresh = output_dir / "fresh.yml"
    fresh.write_text("fresh", encoding="utf-8")
    (map_dir / "fresh.yml.map.json").write_text("fresh-map", encoding="utf-8")
    delete_existing_oas_files([fresh], output_dir)

    assert not fresh.exists()
    assert not (map_dir / "fresh.yml.map.json").exists()


def test_oas_output_archive_describes_actual_archive_folders():
    moved = [
        (Path("api.yaml"), Path(r"C:\out\Archive\20240511\api.yaml")),
        (Path("other.yaml"), Path(r"C:\out\Archive\20240512\other.yaml")),
        (Path("api-31.yaml"), Path(r"C:\out\Archive\20240511\api-31.yaml")),
    ]

    assert describe_archive_destinations(moved) == [
        r"C:\out\Archive\20240511",
        r"C:\out\Archive\20240512",
    ]


def test_oas_output_archive_filters_only_unexpected_or_not_today_files():
    temp_root = _clean_test_temp()
    output_dir = temp_root / "Generated OAS"
    output_dir.mkdir(parents=True)

    current = output_dir / "expected.yaml"
    old_expected = output_dir / "expected-old.yaml"
    unexpected = output_dir / "unexpected.yaml"
    for path in [current, old_expected, unexpected]:
        path.write_text("openapi: 3.1.0\n", encoding="utf-8")

    today = datetime.date(2026, 5, 15)
    today_ts = datetime.datetime(2026, 5, 15, 10, 30).timestamp()
    old_ts = datetime.datetime(2026, 5, 14, 23, 59).timestamp()
    os.utime(current, (today_ts, today_ts))
    os.utime(old_expected, (old_ts, old_ts))
    os.utime(unexpected, (today_ts, today_ts))

    previous = filter_previous_oas_files(
        [current, old_expected, unexpected],
        expected_filenames={"expected.yaml", "expected-old.yaml"},
        today=today,
    )

    assert [path.name for path in previous] == ["expected-old.yaml", "unexpected.yaml"]


def test_oas_output_archive_builds_expected_filenames_from_pattern_and_selected_versions():
    names = build_expected_oas_filenames(
        "EBACL_<current_date>_OpenApi<oas_version>_<customization>OPS_<api_version>_<release>.yaml",
        api_version="4.0",
        release="v20261116",
        gen_30=True,
        gen_31=True,
        gen_swift=True,
        today=datetime.date(2026, 5, 15),
    )

    assert names == {
        "EBACL_20260515_OpenApi3.0_OPS_4.0_v20261116.yaml",
        "EBACL_20260515_OpenApi3.1_OPS_4.0_v20261116.yaml",
        "EBACL_20260515_OpenApi3.0_SWIFT_OPS_4.0_v20261116.yaml",
        "EBACL_20260515_OpenApi3.1_SWIFT_OPS_4.0_v20261116.yaml",
    }


def test_designer_is_hidden_by_default_for_intermediate_releases():
    defaults = PreferencesManager.DEFAULT_PREFERENCES
    assert defaults["enable_api_designer"] is False
    assert defaults["default_tab"] == "OAS Generation"


def test_legacy_example_repair_and_complete_have_separate_defaults():
    defaults = PreferencesManager.DEFAULT_PREFERENCES

    assert defaults["tools_legacy_repair_examples"] is True
    assert defaults["tools_legacy_complete_examples"] is False


def test_legacy_example_migration_keeps_completion_disabled_for_old_combined_flag():
    migrated = migrate_saved_preferences({"tools_legacy_fill_fix_examples": True})

    assert migrated["tools_legacy_repair_examples"] is True
    assert migrated["tools_legacy_complete_examples"] is False


def test_api_designer_workspace_roundtrip_is_deterministic():
    temp_root = _clean_test_temp()
    workspace = create_empty_workspace("Payments Workspace")
    api = workspace.apis[0]
    api.id = "api_payments"
    api.name = "Payments API"
    api.display_label = "Payments API"
    api.version = "1.2.0"
    api.info = {"title": "Payments API", "version": "1.2.0"}
    api.extensions["x-portal-owner"] = "payments"
    workspace.changes.append(
        Change(
            id="chg_cr5262",
            kind="CR",
            external_ref="CR5262",
            title="Add payment status endpoint",
            target_api_id=api.id,
            steps=[
                ChangeStep(
                    id="step_add_path",
                    order=1,
                    kind="add_path",
                    target_id=api.id,
                    path="/payments/{paymentId}/status",
                    after={"method": "get"},
                )
            ],
        )
    )

    try:
        repo = FileSystemDesignerRepository(temp_root / "API Models" / "Payments_Workspace")
        repo.save_workspace(workspace)
        first_snapshot = _snapshot_files(repo.root_path)

        loaded = repo.load_workspace()
        repo.save_workspace(loaded)
        second_snapshot = _snapshot_files(repo.root_path)

        assert first_snapshot == second_snapshot
        assert loaded.name == "Payments Workspace"
        assert loaded.apis[0].id == "api_payments"
        assert loaded.apis[0].extensions["x-portal-owner"] == "payments"
        assert loaded.changes[0].steps[0].kind == "add_path"
    finally:
        shutil.rmtree(temp_root, ignore_errors=True)


def test_api_designer_file_package_shape():
    temp_root = _clean_test_temp()
    workspace = create_empty_workspace("Initial Workspace")
    workspace.apis[0].id = "api_initial"

    try:
        repo = FileSystemDesignerRepository(temp_root / "Initial_Workspace")
        repo.save_workspace(workspace)

        assert (repo.root_path / "workspace.yaml").exists()
        assert (repo.root_path / "apis" / "api_initial.yaml").exists()
        assert (repo.root_path / "libraries").is_dir()
        assert (repo.root_path / "metadata" / "catalog.yaml").exists()
        assert (repo.root_path / "revisions").is_dir()

        manifest = yaml.safe_load((repo.root_path / "workspace.yaml").read_text(encoding="utf-8"))
        assert manifest["schema_version"] == "api-designer.workspace/v1"
        assert manifest["apis"] == [
            {
                "id": "api_initial",
                "name": "New API",
                "display_label": "New API",
                "file": "apis/api_initial.yaml",
            }
        ]
        assert "root_surface" not in manifest["apis"][0]
    finally:
        shutil.rmtree(temp_root, ignore_errors=True)


class _DummyEntry:
    def __init__(self, value: str):
        self._value = value

    def get(self) -> str:
        return self._value


class _DummyPrefs:
    def __init__(self, values: dict[str, str]):
        self._values = values
        self.saved = False

    def get(self, key: str, default=None):
        if key == "remember_paths":
            return True
        return self._values.get(key, default)

    def set(self, key: str, value):
        self._values[key] = value

    def save(self):
        self.saved = True


def test_legacy_converter_browse_initial_dir_uses_nearest_existing_path():
    temp_root = _clean_test_temp()
    missing_leaf = temp_root / "Templates" / "CGS API OPS" / "2026Q4"
    expected = missing_leaf.parent
    expected.mkdir(parents=True)

    try:
        dialog = object.__new__(LegacyConverterDialog)
        dialog.prefs_manager = _DummyPrefs({"last_legacy_src": str(temp_root)})

        initial = dialog._get_initial_dir(_DummyEntry(str(missing_leaf)), "last_legacy_src")

        assert Path(initial) == expected
    finally:
        shutil.rmtree(temp_root, ignore_errors=True)


def test_metadata_preferences_roundtrip_includes_swift_service():
    prefs = _DummyPrefs(
        {
            "tools_legacy_contact_name": "EBA CLEARING",
            "tools_legacy_contact_url": "https://www.ebaclearing.eu",
            "tools_legacy_release": "v20260418",
            "tools_legacy_filename_pattern": "template.yaml",
            "tools_legacy_swift_service": "FPAD",
        }
    )

    defaults = load_metadata_preferences(prefs)

    assert defaults["swift_service"] == "FPAD"

    save_metadata_preferences(
        prefs,
        {
            "contact_name": "NEXI",
            "contact_url": "https://www.nexigroup.com/",
            "release": "v2026Q4",
            "filename_pattern": "next.yaml",
            "swift_service": "R2P",
        },
    )

    assert prefs._values["tools_legacy_swift_service"] == "R2P"
    assert prefs.saved is True


def test_metadata_preferences_partial_save_preserves_existing_swift_service():
    prefs = _DummyPrefs(
        {
            "tools_legacy_contact_name": "EBA CLEARING",
            "tools_legacy_contact_url": "https://www.ebaclearing.eu",
            "tools_legacy_release": "v20260418",
            "tools_legacy_filename_pattern": "template.yaml",
            "tools_legacy_swift_service": "FPAD",
        }
    )

    save_metadata_preferences(
        prefs,
        {
            "contact_name": "NEXI",
            "contact_url": "https://www.nexigroup.com/",
            "release": "v2026Q4",
            "filename_pattern": "next.yaml",
        },
    )

    assert prefs._values["tools_legacy_swift_service"] == "FPAD"


def test_metadata_preferences_explicit_blank_clears_swift_service():
    prefs = _DummyPrefs({"tools_legacy_swift_service": "FPAD"})

    save_metadata_preferences(prefs, {"swift_service": ""})

    assert prefs._values["tools_legacy_swift_service"] == ""


def test_legacy_conversion_metadata_defaults_include_saved_swift_service():
    dialog = object.__new__(LegacyConverterDialog)
    dialog.prefs_manager = _DummyPrefs({"tools_legacy_swift_service": "FPAD"})

    defaults = dialog._get_conversion_metadata_defaults()

    assert defaults["swift_service"] == "FPAD"


def test_legacy_conversion_metadata_prompt_saves_preferences_immediately(monkeypatch):
    metadata_result = {
        "contact_name": "EBA CLEARING",
        "contact_url": "https://www.ebaclearing.eu",
        "release": "v20261113",
        "filename_pattern": "EBACL_RT1_<current_date>_OpenApi<oas_version>_<customization>RT1_API_Participant_<api_version>_<release>.yaml",
        "swift_service": "RT1",
        "swift_servers": [{"url": "https://example.test/rt1", "description": "Live Environment"}],
        "save_in_preferences": True,
    }

    class FakeMetadataDialog:
        def __init__(self, *_args, **_kwargs):
            self.result = dict(metadata_result)

    dialog = object.__new__(LegacyConverterDialog)
    dialog.prefs_manager = _DummyPrefs(
        {
            "tools_legacy_contact_name": "OLD",
            "tools_legacy_contact_url": "https://old.example",
            "tools_legacy_release": "vOLD",
            "tools_legacy_filename_pattern": "OLD_FPAD.yaml",
            "tools_legacy_swift_service": "",
            "swift_services": {},
        }
    )
    dialog.wait_window = lambda _dialog: None
    saved_logs = []
    dialog._log = saved_logs.append
    monkeypatch.setattr(
        legacy_converter_dialog_module,
        "LegacyConversionMetadataDialog",
        FakeMetadataDialog,
    )

    result = dialog._prompt_conversion_metadata_overrides()

    assert result == metadata_result
    assert dialog.prefs_manager._values["tools_legacy_contact_name"] == "EBA CLEARING"
    assert dialog.prefs_manager._values["tools_legacy_contact_url"] == "https://www.ebaclearing.eu"
    assert dialog.prefs_manager._values["tools_legacy_release"] == "v20261113"
    assert dialog.prefs_manager._values["tools_legacy_filename_pattern"].startswith("EBACL_RT1_")
    assert dialog.prefs_manager._values["tools_legacy_swift_service"] == "RT1"
    assert dialog.prefs_manager.saved is True
    assert saved_logs == ["Saved conversion metadata to preferences."]


def test_oas_import_metadata_defaults_include_saved_swift_service_when_oas_info_missing():
    dialog = object.__new__(ImportDialog)
    dialog.prefs_manager = _DummyPrefs({"tools_legacy_swift_service": "FPAD"})
    dialog._log = lambda _message: None

    defaults, _needs_filename_pattern = dialog._get_import_metadata_defaults("missing.yaml")

    assert defaults["swift_service"] == "FPAD"


def test_schema_flattener_does_not_print_debug_logs(capsys):
    flattener = SchemaFlattener(
        {
            "components": {
                "schemas": {
                    "AccountCategory": {
                        "description": "Account category.",
                        "type": "string",
                        "enum": ["BUSINESS", "PERSONAL"],
                    }
                }
            }
        }
    )

    rows = flattener.flatten_schema("AccountCategory")

    assert rows
    captured = capsys.readouterr()
    assert "DEBUG:" not in captured.out


def test_apply_preferences_skips_destroyed_import_dialog_log_area():
    class DummyVar:
        def set(self, _value):
            pass

    class DestroyedLogArea:
        def configure(self, **_kwargs):
            raise AssertionError("destroyed import log area was touched")

    class DestroyedImportDialog:
        import_log_area = DestroyedLogArea()

        def winfo_exists(self):
            return False

    app = object.__new__(OASGenApp)
    app.import_dialog = DestroyedImportDialog()
    app.var_30 = DummyVar()
    app.var_31 = DummyVar()
    app.var_swift = DummyVar()
    app.update_file_list = lambda: None

    app._apply_preferences({"import_log_theme": "Dark"})


def test_import_dialog_close_clears_parent_reference():
    class Parent:
        pass

    parent = Parent()
    dialog = object.__new__(ImportDialog)
    dialog.parent = parent
    dialog.prefs_manager = _DummyPrefs({})
    dialog.destroyed = False
    dialog.destroy = lambda: setattr(dialog, "destroyed", True)
    parent.import_dialog = dialog

    dialog._on_close()

    assert parent.import_dialog is None
    assert dialog.destroyed is True


def test_generation_mode_filters_request_body_examples():
    df = pd.DataFrame(
        [
            {
                "Section": "content",
                "Name": "application/json",
                "Parent": "",
                "Type": "schema",
                "Schema": "PaymentRequest",
                "Mandatory": "M",
            }
        ]
    )
    body_examples = {"OK": "amount: 100", "Bad Request": "amount: bad"}

    standard = OASGenerator(generation_mode=GENERATION_MODE_STANDARD)
    standard_body = standard._build_request_body(df, body_examples)
    assert list(standard_body["content"]["application/json"]["examples"].keys()) == ["OK"]

    minimal = OASGenerator(generation_mode=GENERATION_MODE_MINIMAL)
    minimal_body = minimal._build_request_body(df, body_examples)
    assert "examples" not in minimal_body["content"]["application/json"]

    oas_fragment = {
        "x-keep": "visible",
        "content": {"examples": {"OK": {"value": {"x-sandbox-request-name": "hidden"}}}},
    }
    standard._remove_sandbox_extensions(oas_fragment)
    assert oas_fragment == {
        "x-keep": "visible",
        "content": {"examples": {"OK": {"value": {}}}},
    }


def test_swift_generation_filters_bad_request_body_examples_in_api_portal_ready_mode():
    temp_root = _clean_test_temp()
    input_dir = temp_root / "swift_request_examples_input"
    output_dir = temp_root / "swift_request_examples_output"
    input_dir.mkdir()
    endpoint_file = "postPayment.xlsx"

    _create_workbook(
        input_dir / "$index.xlsx",
        {
            "General Description": [
                ["Key", "Value"],
                ["Info Title", "Payments API"],
                ["Info Version", "1.0"],
                [
                    "Filename Pattern",
                    "TEST_<oas_version>_<customization>API_<api_version>_<release>.yaml",
                ],
                ["Release", "v1"],
            ],
            "Tags": [["Name", "Description"]],
            "Servers": [["URL", "Description"]],
            "Security": [["Name", "Type", "Scheme", "Format", "Description"]],
            "Paths": [
                ["Excel", "Path", "Description", "Method", "Tag", "Summary", "OperationId"],
                [
                    endpoint_file,
                    "/payments",
                    "Payments",
                    "post",
                    "Payments",
                    "Create payment",
                    "createPayment",
                ],
            ],
            "Parameters": [["Name", "Description", "In", "Required", "Schema"]],
            "Headers": [["Name", "Description", "Schema"]],
            "Schemas": [
                ["Name", "Parent", "Description", "Type"],
                ["PaymentRequest", "", "Payment request", "object"],
            ],
            "Responses": [["Name", "Description", "Schema"]],
        },
    )
    _create_workbook(
        input_dir / endpoint_file,
        {
            "Parameters": [["Name", "Description", "In", "Required", "Schema"]],
            "Body": [
                ["Body", "Payment request", "M"],
                [
                    "Section",
                    "Name",
                    "Parent",
                    "Description",
                    "Type",
                    "Items Data Type",
                    "Schema Name",
                    "Format",
                    "Mandatory",
                ],
                ["content", "application/json", "", "", "schema", "", "PaymentRequest", "", "M"],
            ],
            "Body Example": [
                ["Name", "Body"],
                ["OK", "amount: 100"],
                ["Bad Request", "amount: invalid"],
            ],
            "200": [
                ["Response", "200", "OK"],
                ["Name", "Description", "Type", "Schema Name"],
            ],
        },
    )

    main_script.generate_oas(
        str(input_dir),
        gen_30=False,
        gen_31=True,
        gen_swift=True,
        generation_mode=GENERATION_MODE_API_PORTAL_READY,
        output_dir=str(output_dir),
        log_callback=lambda _line: None,
    )

    standard_oas = yaml.safe_load((output_dir / "TEST_3.1_API_1.0_v1.yaml").read_text())
    swift_oas = yaml.safe_load((output_dir / "TEST_3.1_SWIFT_API_1.0_v1.yaml").read_text())
    standard_examples = standard_oas["paths"]["/payments"]["post"]["requestBody"][
        "content"
    ]["application/json"]["examples"]
    swift_examples = swift_oas["paths"]["/payments"]["post"]["requestBody"]["content"][
        "application/json"
    ]["examples"]

    assert list(standard_examples.keys()) == ["OK", "Bad Request"]
    assert list(swift_examples.keys()) == ["OK"]


def test_x_info_extensions_are_enabled_by_default():
    defaults = PreferencesManager.DEFAULT_PREFERENCES

    assert defaults["gen_x_info_creation_date"] is True
    assert defaults["gen_x_info_release"] is True
    assert defaults["gen_x_info_customization"] is True
    assert defaults["gen_x_info_oasis_version"] is True
    assert DEFAULT_X_INFO_OPTIONS == {
        "creation_date": True,
        "release": True,
        "customization": True,
        "oasis_version": True,
    }


def test_swift_services_are_preloaded_in_default_preferences():
    defaults = PreferencesManager.DEFAULT_PREFERENCES
    old_swift_hosts = (
        "https://api.common.swiftnet.sipn.swift.com",
        "https://api-pilot.common.swiftnet.sipn.swift.com",
    )

    assert defaults["swift_services"] == DEFAULT_SWIFT_SERVICES
    assert sorted(defaults["swift_services"].keys()) == [
        "B2B",
        "CGS",
        "CGS-DKK",
        "COR",
        "DCT",
        "FPAD",
        "R2P",
        "RT1",
        "SCT",
    ]
    for service_name, service_config in defaults["swift_services"].items():
        servers = service_config["servers"]
        assert len(servers) == 2, service_name
        assert servers[0]["url"].startswith("https://ebaclapi.swiftnet.sipn.swift.com/")
        assert servers[1]["url"].startswith("https://ebaclapi-pilot.swiftnet.sipn.swift.com/")
        assert "/ebacl-" not in servers[0]["url"]
        assert "/ebacl-" not in servers[1]["url"]
        assert servers[0]["url"].replace(
            "https://ebaclapi.swiftnet.sipn.swift.com",
            "",
        ) == servers[1]["url"].replace(
            "https://ebaclapi-pilot.swiftnet.sipn.swift.com",
            "",
        )
        for old_host in old_swift_hosts:
            assert old_host not in servers[0]["url"]
            assert old_host not in servers[1]["url"]
    assert defaults["swift_services"]["DCT"]["servers"] == [
        {
            "url": "https://ebaclapi.swiftnet.sipn.swift.com/dct/v1",
            "description": "Live Environment",
        },
        {
            "url": "https://ebaclapi-pilot.swiftnet.sipn.swift.com/dct/v1",
            "description": "Test Environment",
        },
    ]
    assert defaults["swift_services"]["R2P"]["servers"][1] == {
        "url": "https://ebaclapi-pilot.swiftnet.sipn.swift.com/r2p/v1",
        "description": "Test Environment",
    }


def test_update_manifest_does_not_force_swift_services_override():
    manifest_path = Path(__file__).resolve().parents[2] / "release-metadata" / "oasis-version.json"

    manifest = parse_update_manifest(manifest_path.read_text(encoding="utf-8"))

    assert manifest.preferences_override is None


def test_build_info_emits_selected_x_info_extensions():
    generator = OASGenerator()

    generator.build_info(
        {
            "version": "1.0.0",
            "title": "Payments API",
            "release": "v20261116",
            "filename_pattern": "ignored.yaml",
        }
    )

    info = generator.oas["info"]
    assert re.fullmatch(r"\d{4}-\d{2}-\d{2}", info["x-info-creation-date"])
    assert info["x-info-release"] == "v20261116"
    assert info["x-info-oasis-version"] == FULL_VERSION
    assert "release" not in info
    assert "filename_pattern" not in info


def test_build_info_can_disable_x_info_extensions_without_leaking_raw_metadata():
    generator = OASGenerator(
        x_info_options={
            "creation_date": False,
            "release": False,
            "oasis_version": False,
        }
    )

    generator.build_info(
        {
            "version": "1.0.0",
            "title": "Payments API",
            "release": "v20261116",
            "filename_pattern": "ignored.yaml",
        }
    )

    info = generator.oas["info"]
    assert "x-info-creation-date" not in info
    assert "x-info-release" not in info
    assert "x-info-oasis-version" not in info
    assert "release" not in info
    assert "filename_pattern" not in info


def test_swift_customization_x_info_can_be_disabled():
    generator = OASGenerator(x_info_options={"customization": False})
    generator.oas = {
        "openapi": "3.0.0",
        "info": {"title": "Payments API", "version": "1.0.0"},
        "paths": {},
        "components": {"schemas": {}, "responses": {}},
    }

    generator.apply_swift_customization()

    assert "x-info-customization" not in generator.oas["info"]


def test_swift_customization_uses_servers_from_template_without_fpad_fallback():
    generator = OASGenerator()
    generator.oas = {
        "openapi": "3.0.0",
        "info": {"title": "Payments API", "version": "1.0.0"},
        "servers": [{"url": "/standard"}],
        "paths": {},
        "components": {"schemas": {}, "responses": {}},
    }
    swift_servers = [{"url": "https://swift.example.test/service", "description": "Custom"}]

    generator.apply_swift_customization(swift_servers=swift_servers)

    assert generator.oas["servers"] == swift_servers


def test_parse_info_keeps_swift_servers_separate_from_standard_servers():
    df_info = pd.DataFrame(
        [
            ["servers url", "/standard", "servers description", "Standard"],
            ["swift servers url", "/swift", "servers description", "SWIFT"],
        ]
    )

    info, servers = parse_info(df_info)

    assert servers == [{"url": "/standard", "description": "Standard"}]
    assert info["swift_servers"] == [{"url": "/swift", "description": "SWIFT"}]


def test_swift_service_catalog_truncates_to_template_capacity_and_warns():
    warnings = []
    services = {
        "FPAD": {
            "servers": [
                {"url": "/one", "description": "One"},
                {"url": "/two", "description": "Two"},
                {"url": "/three", "description": "Three"},
            ]
        }
    }

    servers = get_swift_service_servers(services, "FPAD", warnings.append)

    assert servers == [
        {"url": "/one", "description": "One"},
        {"url": "/two", "description": "Two"},
    ]
    assert warnings


def test_ensure_swift_server_rows_inserts_rows_before_release():
    workbook = Workbook()
    ws = workbook.active
    ws.title = "General Description"
    ws.append(["General Description"])
    ws.append(["info description", "API"])
    ws.append(["info version", "1.0"])
    ws.append(["info title", "Title"])
    ws.append(["info contact name", "Contact"])
    ws.append(["info contact url", "https://example.test"])
    ws.append(["servers url", "/standard", "servers description", "Standard"])
    ws.append(["servers url", "", "servers description", ""])
    ws.append(["release", "v1"])
    ws.append(["filename pattern", "api.yaml"])
    ws.merge_cells("B9:D9")
    ws.merge_cells("B10:D10")
    side = Side(style="thin", color="808080")
    server_label_fill = PatternFill("solid", fgColor="D9D9D9")
    server_value_fill = PatternFill("solid", fgColor="FFFFFF")
    for row in (7, 8):
        ws.row_dimensions[row].height = 18
        for col in range(1, 5):
            cell = ws.cell(row=row, column=col)
            cell.font = Font(name="Calibri", size=11, bold=col in (1, 3), color="1F4E79")
            cell.fill = server_label_fill if col in (1, 3) else server_value_fill
            cell.border = Border(left=side, right=side, top=side, bottom=side)
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)

    ensure_swift_server_rows_in_workbook(
        workbook,
        [{"url": "/swift", "description": "SWIFT"}],
    )

    assert ws.cell(row=9, column=1).value == "swift servers url"
    assert ws.cell(row=9, column=2).value == "/swift"
    assert ws.cell(row=9, column=3).value == "servers description"
    assert ws.cell(row=9, column=4).value == "SWIFT"
    assert ws.cell(row=10, column=1).value == "swift servers url"
    assert ws.cell(row=10, column=3).value == "servers description"
    assert ws.cell(row=11, column=1).value == "release"
    assert ws.cell(row=12, column=1).value == "filename pattern"
    assert "B11:D11" in [str(merged_range) for merged_range in ws.merged_cells.ranges]
    assert "B12:D12" in [str(merged_range) for merged_range in ws.merged_cells.ranges]
    for source_row, swift_row in ((7, 9), (8, 10)):
        assert ws.row_dimensions[swift_row].height == ws.row_dimensions[source_row].height
        for col in range(1, 5):
            assert ws.cell(row=swift_row, column=col)._style == ws.cell(row=source_row, column=col)._style


def test_swift_production_policy_removes_400_examples_in_api_portal_ready_mode():
    generator = OASGenerator(generation_mode=GENERATION_MODE_API_PORTAL_READY)
    generator.oas = {
        "openapi": "3.1.0",
        "info": {"title": "Payments API", "version": "1.0.0"},
        "paths": {
            "/payments": {
                "post": {
                    "responses": {
                        "400": {"$ref": "#/components/responses/BadRequestResponse"}
                    }
                }
            }
        },
        "components": {
            "schemas": {
                "ErrorResponse": {"type": "object"},
                "PaymentError": {"type": "object"},
            },
            "responses": {
                "BadRequestResponse": {
                    "description": "Bad Request",
                    "content": {
                        "application/json": {
                            "schema": {"$ref": "#/components/schemas/PaymentError"},
                            "examples": {
                                "OK": {"value": {"code": "OK"}},
                                "Bad Request": {"value": {"code": "INVALID"}},
                            },
                            "x-sandbox-note": "debug only",
                        }
                    },
                }
            },
        },
    }

    generator.apply_swift_customization()
    swift_oas = yaml.safe_load(generator.get_yaml())
    path_media = swift_oas["paths"]["/payments"]["post"]["responses"]["400"]["content"][
        "application/json"
    ]
    component_media = swift_oas["components"]["responses"]["BadRequestResponse"]["content"][
        "application/json"
    ]

    assert "example" not in path_media
    assert "examples" not in path_media
    assert "example" not in component_media
    assert "examples" not in component_media
    assert "x-sandbox-note" not in path_media
    assert "x-sandbox-note" not in component_media


def test_generation_mode_filters_response_examples_and_optional_placeholders():
    content = {
        "application/json": {
            "schema": {
                "type": "object",
                "required": ["settlementBIC"],
                "properties": {
                    "settlementBIC": {"type": "string"},
                    "preferredAgentBIC": {
                        "type": "string",
                        "pattern": "[A-Z0-9]{4,4}[A-Z]{2,2}[A-Z0-9]{2,2}",
                    },
                },
            },
            "examples": {
                "OK": {
                    "value": {
                        "settlementBIC": "IPSDID21XXX",
                        "preferredAgentBIC": "string",
                    }
                },
                "Bad Request": {"value": {"settlementBIC": "INVALID"}},
            },
        }
    }

    standard = OASGenerator(generation_mode=GENERATION_MODE_STANDARD)
    standard._coerce_content_examples(content, {})

    assert list(content["application/json"]["examples"].keys()) == ["OK"]
    assert content["application/json"]["examples"]["OK"]["value"] == {
        "settlementBIC": "IPSDID21XXX"
    }

    minimal_content = {
        "application/json": {
            "schema": {"type": "object"},
            "examples": {"OK": {"value": {"a": "b"}}},
        }
    }
    minimal = OASGenerator(generation_mode=GENERATION_MODE_MINIMAL)
    minimal._coerce_content_examples(minimal_content, {})

    assert "examples" not in minimal_content["application/json"]


def test_minimal_generation_removes_schema_examples_from_yaml_output():
    generator = OASGenerator(generation_mode=GENERATION_MODE_MINIMAL)
    generator.oas["components"]["schemas"]["Payment"] = {
        "type": "object",
        "example": {"amount": 100},
        "examples": [{"amount": 200}],
        "properties": {
            "amount": {"type": "number", "example": 100},
            "examples": {"type": "string", "example": "business-field"},
        },
    }
    generator.oas["paths"]["/payments"] = {
        "post": {
            "requestBody": {
                "content": {
                    "application/json": {
                        "schema": {"$ref": "#/components/schemas/Payment"},
                        "examples": {"OK": {"value": {"amount": 100}}},
                    }
                }
            },
            "responses": {
                "200": {
                    "description": "OK",
                    "content": {
                        "application/json": {
                            "schema": {"$ref": "#/components/schemas/Payment"},
                            "example": {"amount": 100},
                        }
                    },
                }
            },
        }
    }

    generated = yaml.safe_load(generator.get_yaml())
    payment_schema = generated["components"]["schemas"]["Payment"]
    payment_media = generated["paths"]["/payments"]["post"]["requestBody"]["content"]["application/json"]
    response_media = generated["paths"]["/payments"]["post"]["responses"]["200"]["content"]["application/json"]

    assert "example" not in payment_schema
    assert "examples" not in payment_schema
    assert "example" not in payment_schema["properties"]["amount"]
    assert "examples" in payment_schema["properties"]
    assert "example" not in payment_schema["properties"]["examples"]
    assert "examples" not in payment_media
    assert "example" not in response_media


def test_raw_extension_filter_keeps_only_yaml_extension_blocks():
    generator = OASGenerator(generation_mode=GENERATION_MODE_MINIMAL)

    filtered = generator._filter_raw_extensions(
        "\n".join(
            [
                "[240520] - (SS) Description amended - CR5004",
                "x-visible:",
                "  enabled: true",
                "x-sandbox-rule-type: SCRIPT_JS",
            ]
        )
    )

    assert filtered == "x-visible:\n  enabled: true"


def test_paths_parser_does_not_treat_track_changes_as_custom_extensions():
    df = pd.DataFrame(
        [
            {
                "Excel": "updateSettlementBIC.xlsx",
                "Paths": "/updateSettlementBIC/{senderBic}",
                "Unnamed: 3": "put",
                "Unnamed: 4": "Update description",
                "Unnamed: 5": "Participant Management",
                "Unnamed: 6": "Update Settlement BIC",
                "Unnamed: 7": "updateSettlementBIC",
                "Unnamed: 8": "[240520] - (SS) Description amended - CR5004",
            }
        ]
    )

    operations = parse_paths_index(df)

    assert operations[0]["extensions"] is None


def test_legacy_converter_fill_fix_examples_replaces_invalid_bic_and_creates_config():
    temp_root = _clean_test_temp()
    config_path = temp_root / "prefs" / "legacy_example_seed_values.yaml"
    converter = LegacyConverter(
        str(temp_root),
        str(temp_root / "out"),
        fill_fix_examples=True,
        example_seed_values_path=config_path,
    )
    dtype = DataType(
        name="PreferredAgentBIC",
        type="string",
        pattern_eba="[A-Z0-9]{4,4}[A-Z]{2,2}[A-Z0-9]{2,2}",
        example="string",
    )

    try:
        converter._fill_and_fix_examples_for_data_type(dtype)
        examples = dtype.example.split("; ")

        assert config_path.exists()
        assert len(examples) == 3
        assert "string" not in examples
        assert all(re.fullmatch(r"[A-Z0-9]{4,4}[A-Z]{2,2}[A-Z0-9]{2,2}", value) for value in examples)
    finally:
        shutil.rmtree(temp_root, ignore_errors=True)


def test_legacy_converter_excel_open_suppresses_only_openpyxl_data_validation_warning(monkeypatch):
    def fake_excel_file(path):
        warnings.warn_explicit(
            "Data Validation extension is not supported and will be removed",
            UserWarning,
            filename="openpyxl/worksheet/_reader.py",
            lineno=329,
            module="openpyxl.worksheet._reader",
        )
        warnings.warn_explicit(
            "Important workbook warning",
            UserWarning,
            filename="openpyxl/worksheet/_reader.py",
            lineno=330,
            module="openpyxl.worksheet._reader",
        )
        return f"opened:{path}"

    monkeypatch.setattr("src.legacy_converter.pd.ExcelFile", fake_excel_file)
    converter = LegacyConverter("in", "out")

    with warnings.catch_warnings(record=True) as captured:
        warnings.simplefilter("always")
        result = converter._open_excel_file("legacy.xlsm")

    assert result == "opened:legacy.xlsm"
    assert [str(item.message) for item in captured] == ["Important workbook warning"]


def test_legacy_converter_excel_read_suppresses_lazy_openpyxl_data_validation_warning(monkeypatch):
    expected = pd.DataFrame([{"Name": "field"}])

    def fake_read_excel(*args, **kwargs):
        warnings.warn_explicit(
            "Data Validation extension is not supported and will be removed",
            UserWarning,
            filename="openpyxl/worksheet/_reader.py",
            lineno=329,
            module="openpyxl.worksheet._reader",
        )
        warnings.warn_explicit(
            "Important workbook warning",
            UserWarning,
            filename="openpyxl/worksheet/_reader.py",
            lineno=330,
            module="openpyxl.worksheet._reader",
        )
        return expected

    monkeypatch.setattr("src.legacy_converter.pd.read_excel", fake_read_excel)
    converter = LegacyConverter("in", "out")

    with warnings.catch_warnings(record=True) as captured:
        warnings.simplefilter("always")
        result = converter._read_excel_sheet("legacy.xlsm", sheet_name="Body", dtype=str)

    assert result is expected
    assert [str(item.message) for item in captured] == ["Important workbook warning"]


def test_legacy_converter_fill_fix_examples_respects_allowed_values_limit():
    temp_root = _clean_test_temp()
    converter = LegacyConverter(str(temp_root), str(temp_root / "out"), fill_fix_examples=True)
    dtype = DataType(
        name="Status",
        type="string",
        allowed_values="ACTV; INAC",
        example="string",
    )

    try:
        converter._fill_and_fix_examples_for_data_type(dtype)

        assert dtype.example == "ACTV; INAC"
    finally:
        shutil.rmtree(temp_root, ignore_errors=True)


def test_legacy_converter_repairs_consolidated_schemas_only_and_defers_example_tracer():
    temp_root = _clean_test_temp()
    detail_lines: list[str] = []
    converter = LegacyConverter(
        str(temp_root),
        str(temp_root / "out"),
        fill_fix_examples=True,
        detail_log_callback=detail_lines.append,
    )
    converter.tracing_enabled = True
    raw_dtype = DataType(
        name="PreferredAgentBIC",
        type="string",
        pattern_eba="[A-Z0-9]{4,4}[A-Z]{2,2}[A-Z0-9]{2,2}",
        example="string",
    )
    consolidated_dtype = DataType(
        name="PreferredAgentBIC",
        type="string",
        pattern_eba="[A-Z0-9]{4,4}[A-Z]{2,2}[A-Z0-9]{2,2}",
        example="string",
    )
    converter.raw_data_types = {"endpoint.xlsx": {"PreferredAgentBIC": raw_dtype}}
    converter.global_schemas = {"PreferredAgentBIC": consolidated_dtype}

    try:
        converter._fill_and_fix_consolidated_examples()

        assert raw_dtype.example == "string"
        assert consolidated_dtype.example != "string"
        assert not any("EXAMPLE TRACER" in line for line in detail_lines)

        converter._log_example_repair_report()

        assert any("EXAMPLE TRACER" in line for line in detail_lines)
        assert any("REPAIRED" in line and "PreferredAgentBIC" in line for line in detail_lines)
    finally:
        shutil.rmtree(temp_root, ignore_errors=True)


def test_legacy_converter_uses_plausible_bic8_and_bic11_seed_values():
    temp_root = _clean_test_temp()
    converter = LegacyConverter(str(temp_root), str(temp_root / "out"), fill_fix_examples=True)
    bic8 = DataType(
        name="BIC8",
        type="string",
        pattern_eba="4!c2!a2!c",
        regex="[A-Z0-9]{4,4}[A-Z]{2,2}[A-Z0-9]{2,2}",
        example="AAAAAAAA",
    )
    bic11 = DataType(
        name="BIC11Only",
        type="string",
        pattern_eba="4!c2!a2!c3!c",
        regex="[A-Z0-9]{4,4}[A-Z]{2,2}[A-Z0-9]{2,2}[A-Z0-9]{3,3}",
        example="AAAAAAAAAAA",
    )

    try:
        converter._fill_and_fix_examples_for_data_type(bic8)
        converter._fill_and_fix_examples_for_data_type(bic11)

        bic8_examples = bic8.example.split("; ")
        bic11_examples = bic11.example.split("; ")
        assert bic8_examples == ["IPSDITM1", "DEUTDEFF", "BNPAFRPP"]
        assert bic11_examples == ["IPSDITM1XXX", "DEUTDEFFXXX", "BNPAFRPPXXX"]
    finally:
        shutil.rmtree(temp_root, ignore_errors=True)


def test_legacy_converter_converts_eba_pattern_constraints_for_examples():
    converter = LegacyConverter("in", "out", fill_fix_examples=True)
    generic_bic = DataType(
        name="BIC11",
        type="string",
        pattern_eba="4!c2!a2!c[3!c]",
        example="AAAAAAAA",
    )
    bic11_only = DataType(
        name="Bic11Only",
        type="string",
        min_val="11",
        max_val="11",
        pattern_eba="4!c2!a2!c3!c",
        example="AAAAAAAA[A-Z0-9]{3,3}",
    )

    converter._fill_and_fix_examples_for_data_type(generic_bic)
    converter._fill_and_fix_examples_for_data_type(bic11_only)

    assert converter._classify_example_category(generic_bic) == "bic"
    assert generic_bic.example == "IPSDITM1; DEUTDEFFXXX; BNPAFRPP"
    assert bic11_only.example == "IPSDITM1XXX; DEUTDEFFXXX; BNPAFRPPXXX"


def test_legacy_converter_general_bic_semantics_use_mixed_lengths():
    converter = LegacyConverter("in", "out", fill_fix_examples=True)
    receiver_bic = DataType(
        name="ReceiverBic",
        type="string",
        description="The BIC of the message receiver",
        example="",
    )

    converter._fill_and_fix_examples_for_data_type(receiver_bic)

    assert converter._classify_example_category(receiver_bic) == "bic"
    assert receiver_bic.example == "IPSDITM1; DEUTDEFFXXX; BNPAFRPP"


def test_legacy_converter_schema_compaction_renames_block_root_parent_edges():
    converter = LegacyConverter("in", "out")
    blocks = [
        (
            "ProductList14",
            [
                ["ProductList14", "", "", "array", "object", "", "", "", "", "", "", "", "", ""],
                ["solutionPurposeId", "ProductList14", "", "schema", "", "SolutionPurposeId", "", "O", "", "", "", "", "", ""],
                ["pryRcvgNetAd", "ProductList14", "", "schema", "", "PryRcvgNetAd", "", "O", "", "", "", "", "", ""],
            ],
        ),
    ]

    renamed = converter._apply_schema_rename_to_blocks(blocks, {"ProductList14": "ProductList6"})

    assert renamed[0][0] == "ProductList6"
    assert renamed[0][1][0][0] == "ProductList6"
    assert [row[1] for row in renamed[0][1][1:]] == ["ProductList6", "ProductList6"]
    assert [row[5] for row in renamed[0][1][1:]] == ["SolutionPurposeId", "PryRcvgNetAd"]


def test_legacy_converter_preserves_source_datatype_names_ending_with_digits():
    temp_root = _clean_test_temp()
    input_dir = temp_root / "legacy"
    output_dir = temp_root / "converted"
    input_dir.mkdir()
    output_dir.mkdir()
    endpoint_file = input_dir / "operation.xlsx"
    output_index = output_dir / "$index.xlsx"

    bic_regex = r"[A-Z0-9]{4,4}[A-Z]{2,2}[A-Z0-9]{2,2}"
    bic11_regex = r"[A-Z0-9]{4,4}[A-Z]{2,2}[A-Z0-9]{2,2}([A-Z0-9]{3,3}){0,1}"
    _create_workbook(
        endpoint_file,
        {
            "Data Type": [
                ["Name", "Description", "Type", "Min", "Max", "PatternEba", "Regex", "Example"],
                ["Bic", "BIC8 or BIC11", "string", "8", "11", "4!c2!a2!c[3!c]", bic11_regex, "IPSDITM1"],
                ["Bic8", "BIC8", "string", "8", "8", "4!c2!a2!c", bic_regex, "IPSDITM1"],
                ["Bic11", "BIC11", "string", "8", "11", "4!c2!a2!c[3!c]", bic11_regex, "IPSDITM1XXX"],
            ],
            "Body": [
                ["Type", "Parent", "Element", "Mandatory", "Description"],
                ["object", "", "criteria", "M", "Criteria"],
                ["Bic", "criteria", "bicCode", "M", "BIC code"],
                ["Bic11", "criteria", "debAccoOwnBic", "O", "Debtor BIC"],
            ],
        },
    )
    _create_workbook(
        output_index,
        {
            "Schemas": [["Name", "Parent", "Description", "Type", "Items Data Type", "Schema Name"]],
        },
    )

    converter = LegacyConverter(str(input_dir), str(output_dir))
    converter.ordered_filenames = [endpoint_file.name]
    converter._collect_all_data_types()
    converter._perform_naming_and_usage_pass()

    workbook = load_workbook(output_index)
    converter._convert_schemas(workbook)

    rows = list(workbook["Schemas"].iter_rows(values_only=True))
    deb_row = next(row for row in rows if row and row[0] == "debAccoOwnBic")
    emitted_schema_names = {row[0] for row in rows if row and row[1] in ("", None)}

    assert deb_row[5] == "Bic11"
    assert "Bic11" in emitted_schema_names
    assert "Bic1" not in emitted_schema_names


def test_schema_tracer_does_not_compare_source_names_ending_with_digits_as_collisions():
    logs = []
    converter = LegacyConverter("in", "out", log_callback=logs.append)
    converter.tracing_enabled = True
    converter.source_schema_names.update({"Bic", "Bic8", "Bic11"})
    converter.global_schemas["Bic"] = DataType(
        name="Bic",
        type="string",
        min_val="8",
        max_val="11",
        regex=r"[A-Z0-9]{4,4}[A-Z]{2,2}[A-Z0-9]{2,2}([A-Z0-9]{3,3}){0,1}",
    )
    converter.global_schemas["Bic8"] = DataType(
        name="Bic8",
        type="string",
        min_val="8",
        max_val="8",
        regex=r"[A-Z0-9]{4,4}[A-Z]{2,2}[A-Z0-9]{2,2}",
    )
    converter.global_schemas["Bic11"] = DataType(
        name="Bic11",
        type="string",
        min_val="8",
        max_val="11",
        regex=r"[A-Z0-9]{4,4}[A-Z]{2,2}[A-Z0-9]{2,2}([A-Z0-9]{3,3}){0,1}",
    )
    converter.schema_usage["Bic"] = {"listA (Body)"}
    converter.schema_usage["Bic8"] = {"listB (Body)"}
    converter.schema_usage["Bic11"] = {"listC (Body)"}

    converter._log_usage_summary()

    rendered = "\n".join(logs)
    assert "Bic8" in rendered
    assert "Bic11" in rendered
    assert "- Regex:" not in rendered
    assert "- Max Length:" not in rendered


def test_schema_tracer_orders_case_variants_deterministically():
    logs = []
    converter = LegacyConverter("in", "out", log_callback=logs.append)
    converter.schema_usage["EndToEndId"] = ["messageDetails (200)"]
    converter.schema_usage["EndToEndID"] = ["listPayments (200)"]

    converter._log_usage_summary()

    rendered = "\n".join(logs)
    upper_index = rendered.index("| EndToEndID")
    mixed_index = rendered.index("| EndToEndId")
    assert upper_index < mixed_index


def test_legacy_converter_literal_array_uses_distinct_item_component_when_named_array_exists():
    converter = LegacyConverter("in", "out")
    converter.global_schemas["AosId"] = DataType(
        name="AosId",
        type="array",
        items_type="object",
    )
    converter.output_names[("$global", "AosId")] = "AosId"

    rows, _, extra_blocks = converter._build_children_rows(
        "Wrapper",
        [
            ("aosId", "", "", "array", "O", "", "", "object"),
            ("value", "aosId", "", "string", "M", "", "", ""),
        ],
    )

    array_row = next(row for row in rows if row[0] == "aosId")
    item_component = array_row[4]

    assert array_row[3] == "array"
    assert array_row[5] == ""
    assert item_component != "AosId"
    assert any(row[0] == item_component and row[3] == "object" for row in extra_blocks)
    assert any(row[0] == "value" and row[1] == item_component for row in extra_blocks)


def test_legacy_converter_named_array_datatype_is_property_array_with_item_schema_reference():
    converter = LegacyConverter("in", "out")
    converter.global_schemas["AosId"] = DataType(
        name="AosId",
        type="array",
        items_type="AosIdItem",
        min_val="0",
        max_val="30",
    )
    converter.global_schemas["AosIdItem"] = DataType(
        name="AosIdItem",
        type="string",
        allowed_values="ALIA,LINK,ZZZZ",
    )
    converter.output_names[("$global", "AosId")] = "AosId"
    converter.output_names[("$global", "AosIdItem")] = "AosIdItem"

    rows, referenced, extra_blocks = converter._build_children_rows(
        "Wrapper",
        [
            ("aosId", "", "", "AosId", "O", "", "", ""),
        ],
    )

    array_row = next(row for row in rows if row[0] == "aosId")

    assert array_row[3] == "array"
    assert array_row[4] == "AosId"
    assert array_row[5] == ""
    assert array_row[8] == "0"
    assert array_row[9] == "30"
    assert referenced == {"AosId"}
    assert extra_blocks == []

    item_schema = converter._array_alias_item_schema("AosId", converter.global_schemas["AosId"])
    assert item_schema is not None
    assert item_schema.type == "string"
    assert item_schema.min_val == ""
    assert item_schema.max_val == ""
    assert item_schema.allowed_values == "ALIA,LINK,ZZZZ"


def test_legacy_converter_named_array_with_object_children_is_property_array_with_object_item_component():
    converter = LegacyConverter("in", "out")
    converter.global_schemas["ListReOperationDetails"] = DataType(
        name="ListReOperationDetails",
        type="array",
        items_type="object",
        min_val="1",
        max_val="10",
    )
    converter.output_names[("$global", "ListReOperationDetails")] = "ListReOperationDetails"

    rows, referenced, extra_blocks = converter._build_children_rows(
        "Wrapper",
        [
            ("listReOperationDetails", "", "", "ListReOperationDetails", "O", "", "", ""),
            ("aosId", "listReOperationDetails", "", "AosId", "O", "", "", ""),
        ],
    )

    array_row = next(row for row in rows if row[0] == "listReOperationDetails")

    assert array_row[3] == "array"
    assert array_row[4] == "ListReOperationDetails"
    assert array_row[5] == ""
    assert array_row[8] == "1"
    assert array_row[9] == "10"
    assert referenced == {"ListReOperationDetails"}
    assert any(row[0] == "ListReOperationDetails" and row[3] == "object" for row in extra_blocks)


def test_legacy_converter_repairs_bic_semantic_fields_without_generic_fallback():
    converter = LegacyConverter("in", "out", fill_fix_examples=True)
    cases = [
        (
            DataType(
                name="ReBic",
                type="string",
                description="Reachable Entity BIC (BIC11)",
                pattern_eba="4!c2!a2!c3!c",
                regex=r"[A-Z0-9]{4,4}[A-Z]{2,2}[A-Z0-9]{2,2}([A-Z0-9]{3,3})",
                example="AAAAAAAAAAA",
            ),
            {"bic11"},
        ),
        (
            DataType(
                name="ReceiverBic",
                type="string",
                description="The BIC of the message receiver",
                pattern_eba="4!c2!a2!c",
                regex=r"[A-Z0-9]{4,4}[A-Z]{2,2}[A-Z0-9]{2,2}",
                example="AAAAAAAA",
            ),
            {"bic8"},
        ),
        (
            DataType(
                name="ReachableBic",
                type="string",
                regex=r"[A-Z0-9]{4,4}[A-Z]{2,2}[A-Z0-9]{2,2}([A-Z0-9]{3,3})",
                example="",
            ),
            {"bic11"},
        ),
        (
            DataType(
                name="ReachableEntity",
                type="string",
                description="BIC of the reachable entity (BIC11)",
                pattern_eba="4!c2!a2!c3!c",
                example="AAAAAAAAAAA",
            ),
            {"bic11"},
        ),
        (
            DataType(
                name="FinancialInstitution",
                type="string",
                description="Financial institution BIC",
                pattern_eba="4!c2!a2!c[3!c]",
                example="AAAAAAAA",
            ),
            {"bic", "bic8", "bic11"},
        ),
    ]

    for dt, expected_categories in cases:
        converter._example_trace_rows = []
        converter._fill_and_fix_examples_for_data_type(dt, schema_name=dt.name)

        examples = dt.example.split("; ")
        category = converter._classify_example_category(dt)
        assert category in expected_categories
        assert len(examples) == 3
        assert all(converter._is_valid_example_token(dt, value) for value in examples)
        assert not any(re.fullmatch(r"([A-Z0-9])\1+", value) for value in examples)
        reason = converter._example_trace_rows[-1]["reason"]
        assert "semantic category: bic" in reason or reason == "repaired invalid example: semantic placeholder"


def test_legacy_converter_regex_example_generation_unescapes_literal_dots():
    converter = LegacyConverter("in", "out", fill_fix_examples=True)
    msg_type = DataType(
        name="msgTp",
        type="string",
        regex=r"(pain\.013){1}|(pain\.014){1}|(pacs\.028){1}|(camt\.055){1}|(camt\.029){1}",
        example="string",
    )

    converter._fill_and_fix_examples_for_data_type(msg_type)

    assert msg_type.example == "pain.013; pain.014; pacs.028"
    assert all(re.fullmatch(msg_type.regex, value) for value in msg_type.example.split("; "))


def test_legacy_converter_regex_example_generation_validates_simple_quantifiers_and_groups():
    converter = LegacyConverter("in", "out", fill_fix_examples=True)
    cases = [
        r"MSG[0-9]{2}",
        r"^MSG[0-9]{2}$",
        r"AB+",
        r"(AB)+",
        r"X[A-Z]*",
        r"A{0,3}",
        r"[^0-9]{3}",
        r"(AA|BB)[0-9]{2}",
        r"ABC\+DEF",
        r"AMOUNT\$",
    ]

    for pattern in cases:
        generated = converter._generate_valid_from_regex(pattern)
        assert generated is not None, pattern
        assert re.fullmatch(pattern, generated), (pattern, generated)


def test_legacy_converter_example_generation_respects_numeric_and_length_constraints():
    converter = LegacyConverter("in", "out", fill_fix_examples=True)
    integer = DataType(name="RetryCount", type="integer", min_val="2.2", max_val="4.8", example="string")
    decimal = DataType(name="Amount", type="number", min_val="10.5", max_val="11.5", example="string")
    short_text = DataType(name="ShortCode", type="string", min_val="3", max_val="5", example="string")
    empty_allowed_text = DataType(name="EmptyString", type="string", min_val="0", max_val="0", example="string")

    converter._fill_and_fix_examples_for_data_type(integer)
    converter._fill_and_fix_examples_for_data_type(decimal)
    converter._fill_and_fix_examples_for_data_type(short_text)
    converter._fill_and_fix_examples_for_data_type(empty_allowed_text)

    assert integer.example == "3"
    assert converter._is_valid_example_token(integer, integer.example)
    assert decimal.example == "11.0"
    assert converter._is_valid_example_token(decimal, decimal.example)
    assert converter._is_valid_example_token(short_text, short_text.example)
    assert 3 <= len(short_text.example) <= 5
    assert empty_allowed_text.example == ""


def test_legacy_converter_does_not_write_best_effort_when_constraints_are_impossible():
    converter = LegacyConverter("in", "out", fill_fix_examples=True)
    impossible_integer = DataType(name="ImpossibleInteger", type="integer", min_val="1.2", max_val="1.8", example="")
    impossible_text = DataType(name="ImpossibleText", type="string", min_val="5", max_val="3", example="")

    converter._fill_and_fix_examples_for_data_type(impossible_integer)
    converter._fill_and_fix_examples_for_data_type(impossible_text)

    assert impossible_integer.example == ""
    assert impossible_text.example == ""
    impossible_rows = [row for row in converter._example_trace_rows if row["action"] == "IMPOSSIBLE"]
    assert len(impossible_rows) == 2
    assert {row["severity"] for row in impossible_rows} == {"impossible"}


def test_legacy_converter_marks_too_complex_regex_examples_in_trace():
    converter = LegacyConverter("in", "out", fill_fix_examples=True)
    complex_regex = DataType(name="ComplexPattern", type="string", regex=r"(?=ABC)ABC", example="string")

    converter._fill_and_fix_examples_for_data_type(complex_regex)

    assert complex_regex.example == ""
    assert converter._example_trace_rows[-1]["action"] == "TOO COMPLEX"
    assert converter._example_trace_rows[-1]["severity"] == "complex"


def test_legacy_converter_boolean_examples_respect_allowed_values():
    converter = LegacyConverter("in", "out", fill_fix_examples=True)
    yes_no = DataType(name="ActiveFlag", type="boolean", allowed_values="yes; no", example="maybe")

    converter._fill_and_fix_examples_for_data_type(yes_no)

    assert yes_no.example == "yes; no"
    assert converter._is_valid_example_set(yes_no, yes_no.example)


def test_legacy_converter_example_tracer_does_not_truncate_values():
    temp_root = _clean_test_temp()
    detail_lines: list[str] = []
    converter = LegacyConverter(
        str(temp_root),
        str(temp_root / "out"),
        fill_fix_examples=True,
        detail_log_callback=detail_lines.append,
    )
    long_after = "banking-data.xml; banking-report.csv; banking-document.txt"

    try:
        converter._record_example_trace(
            "FileName",
            DataType(name="FileName", type="string"),
            "COMPLETED",
            "",
            long_after,
            "semantic category: filename",
        )
        converter._log_example_trace_summary()

        joined = "\n".join(detail_lines)
        assert "banking-data.xml" in joined
        assert "banking-report.csv" in joined
        assert "banking-document.txt" in joined
        assert "..." not in joined
    finally:
        shutil.rmtree(temp_root, ignore_errors=True)


def test_legacy_converter_example_tracer_wraps_long_unbroken_values_in_grid():
    detail_lines: list[str] = []
    converter = LegacyConverter("in", "out", fill_fix_examples=True, detail_log_callback=detail_lines.append)
    long_value = (
        "P08M59270578EA544566A8DB7D970D;"
        "P08B59450578EA544526A8DB7D060D;"
        "P08R59805521EA544566A8DB7P998C"
    )
    converter._record_example_trace(
        "MessageId",
        DataType(name="MessageId", type="string"),
        "REPAIRED",
        long_value,
        long_value,
        "semantic category: message_id",
    )

    converter._log_example_trace_summary()

    table_lines = [line for line in detail_lines if line.startswith("| ")]
    assert len(table_lines) > 2
    assert len({len(line) for line in table_lines}) == 1
    assert long_value.replace(";", "") not in "\n".join(table_lines)


def test_legacy_converter_example_tracer_shows_field_usage_context():
    detail_lines: list[str] = []
    converter = LegacyConverter("in", "out", fill_fix_examples=True, detail_log_callback=detail_lines.append)
    converter._record_example_schema_field("BIC8", "InsertSettlementBICRequest.preferredAgentBIC")
    converter._record_example_trace(
        "BIC8",
        DataType(name="BIC8", type="string"),
        "REPAIRED",
        "AAAAAAAA",
        "IPSDITM1; DEUTDEFF; BNPAFRPP",
        "semantic category: bic8",
    )

    converter._log_example_trace_summary()

    joined = "\n".join(detail_lines)
    assert "FIELD/USAGE" in joined
    assert "FIELD USAGE DETAILS" not in joined
    assert "InsertSettlementBICRequest.preferredAgentBIC" in joined


def test_legacy_converter_example_tracer_marks_constraint_severity_lines():
    detail_lines: list[str] = []
    converter = LegacyConverter("in", "out", fill_fix_examples=True, detail_log_callback=detail_lines.append)
    converter._record_example_trace(
        "ImpossibleText",
        DataType(name="ImpossibleText", type="string"),
        "IMPOSSIBLE",
        "",
        "",
        "impossible length range",
        severity="impossible",
    )
    converter._record_example_trace(
        "ComplexPattern",
        DataType(name="ComplexPattern", type="string"),
        "TOO COMPLEX",
        "",
        "",
        "too complex regex",
        severity="complex",
    )

    converter._log_example_trace_summary()

    joined = "\n".join(detail_lines)
    assert "[[OASIS_EXAMPLE_TRACE_IMPOSSIBLE]]" in joined
    assert "[[OASIS_EXAMPLE_TRACE_COMPLEX]]" in joined


def test_legacy_converter_standalone_example_trace_reports_kept_examples():
    temp_root = _clean_test_temp()
    template_path = temp_root / "$index.xlsx"
    detail_lines: list[str] = []
    wb = Workbook()
    ws = wb.active
    ws.title = "Schemas"
    ws.append(["Name", "Parent", "Type", "Pattern", "Example"])
    ws.append(["BIC8", "", "string", "4!c2!a2!c", "IPSDITM1; DEUTDEFF; BNPAFRPP"])
    wb.save(template_path)
    wb.close()

    try:
        converter = LegacyConverter(
            str(temp_root),
            str(temp_root / "out"),
            fill_fix_examples=True,
            log_callback=detail_lines.append,
            detail_log_callback=detail_lines.append,
        )

        assert converter.run_standalone_example_trace(str(temp_root), repair_files=False)

        joined = "\n".join(detail_lines)
        assert "Repair/complete examples: kept 1" in joined
        assert "EXAMPLE TRACER" in joined
        assert "KEPT" in joined
        assert "BIC8" in joined
    finally:
        shutil.rmtree(temp_root, ignore_errors=True)


def test_legacy_converter_semantic_classifier_uses_template_wide_categories():
    converter = LegacyConverter("in", "out", fill_fix_examples=True)

    assert converter._classify_example_category(DataType(name="FileStatus", type="string")) == "status"
    assert converter._classify_example_category(DataType(name="FileReference", type="string")) == "reference"
    assert converter._classify_example_category(DataType(name="NetworkFileName", type="string")) == "filename"
    assert converter._classify_example_category(DataType(name="RequestId", type="string")) == "request_id"
    assert converter._classify_example_category(DataType(name="EndToEndId", type="string")) == "transaction_id"
    assert converter._classify_example_category(DataType(name="ModuleIdentifier", type="string")) == "identifier"
    assert converter._classify_example_category(DataType(name="AOSList", type="string")) == "aos"
    assert converter._classify_example_category(DataType(name="FlowIndicator", type="string")) == "flag"
    assert converter._classify_example_category(DataType(name="LSP", type="string", allowed_values="0,1")) == "flag"
    assert converter._classify_example_category(DataType(name="AdmissionProfile", type="string")) == "profile"
    assert converter._classify_example_category(DataType(name="CmndRfrnc", type="string")) == "reference"
    assert converter._classify_example_category(DataType(name="FileTypeList", type="string")) == "type"
    assert converter._classify_example_category(DataType(name="Requestor", type="string")) == "role"
    assert converter._classify_example_category(DataType(name="LacNumber", type="string")) == "lac"
    assert converter._classify_example_category(DataType(name="BeneficiaryName", type="string")) == "name"
    assert converter._classify_example_category(DataType(name="SentReceived", type="string", allowed_values="SNT,RCV")) == "direction"
    assert converter._classify_example_category(DataType(name="UETR", type="string")) == "uuid"
    assert converter._classify_example_category(DataType(name="TotNumOfMsgRcv", type="string")) == "count"
    assert converter._classify_example_category(DataType(name="ErrorDetails", type="string")) == "validation_detail"
    assert converter._classify_example_category(DataType(name="SecRcvgNetAd", type="string")) == "network_address"
    assert converter._classify_example_category(DataType(name="FreezeUnfreeze", type="string")) == "operation"
    assert converter._classify_example_category(DataType(name="XmlMsgSts", type="string")) == "status"
    assert converter._classify_example_category(DataType(name="XmlStsRsn", type="string")) == "reason"
    assert converter._classify_example_category(DataType(name="PryApiEndpoint", type="string")) == "endpoint"
    assert converter._classify_example_category(DataType(name="MaxFileSize", type="string")) == "file_size"
    assert converter._classify_example_category(DataType(name="Account", type="string")) == "account"
    assert converter._classify_example_category(
        DataType(name="AccptncDtTm", type="string", pattern_eba="YYYY-MM-DDTHH:MM:SS[.M{1,6}]")
    ) == "datetime"
    assert converter._classify_example_category(
        DataType(
            name="dataTime",
            type="string",
            regex=r"[0-9]{4,4}-[0-9]{2,2}-[0-9]{2,2}[T][0-9]{2,2}:[0-9]{2,2}:[0-9]{2,2}",
        )
    ) == "datetime"


def test_legacy_converter_datetime_regex_drives_realistic_examples():
    converter = LegacyConverter("in", "out", fill_fix_examples=True)
    data_time = DataType(
        name="dataTime",
        type="string",
        regex=r"[0-9]{4,4}-[0-9]{2,2}-[0-9]{2,2}[T][0-9]{2,2}:[0-9]{2,2}:[0-9]{2,2}",
        example="not-a-datetime",
    )

    converter._fill_and_fix_examples_for_data_type(data_time)

    examples = data_time.example.split("; ")
    assert examples == [
        "2026-01-31T10:15:30",
        "2026-06-30T12:00:00",
        "2026-12-31T23:59:59",
    ]
    assert all(converter._is_valid_example_token(data_time, value) for value in examples)


def test_legacy_converter_semantic_classifier_uses_configured_overrides_and_rules():
    temp_root = _clean_test_temp()
    rules_path = temp_root / "prefs" / "legacy_example_semantic_rules.yaml"
    rules_path.parent.mkdir(parents=True)
    rules_path.write_text(
        yaml.safe_dump(
            {
                "overrides": {
                    "exact": {"LSP": "service"},
                    "compact_exact": {"Instruction": "instruction_id"},
                },
                "rules": [
                    {"category": "channel", "compact_exact": ["accesspoint"]},
                ],
            },
            sort_keys=False,
        ),
        encoding="utf-8",
    )
    converter = LegacyConverter(
        str(temp_root),
        str(temp_root / "out"),
        fill_fix_examples=True,
        example_semantic_rules_path=rules_path,
    )

    try:
        assert converter._classify_example_category(DataType(name="LSP", type="string", allowed_values="0,1")) == "service"
        assert converter._classify_example_category(DataType(name="Instruction", type="string")) == "instruction_id"
        assert converter._classify_example_category(DataType(name="AccessPoint", type="string")) == "channel"
    finally:
        shutil.rmtree(temp_root, ignore_errors=True)


def test_legacy_converter_semantic_classifier_creates_default_rules_config():
    temp_root = _clean_test_temp()
    rules_path = temp_root / "prefs" / "legacy_example_semantic_rules.yaml"
    converter = LegacyConverter(
        str(temp_root),
        str(temp_root / "out"),
        fill_fix_examples=True,
        example_semantic_rules_path=rules_path,
    )

    try:
        assert converter._classify_example_category(DataType(name="RequestId", type="string")) == "request_id"
        assert rules_path.exists()
        loaded = yaml.safe_load(rules_path.read_text(encoding="utf-8"))
        assert "rules" in loaded
        assert "overrides" in loaded
    finally:
        shutil.rmtree(temp_root, ignore_errors=True)


def test_legacy_converter_standalone_example_trace_repairs_template_schema_examples():
    temp_root = _clean_test_temp()
    template_path = temp_root / "$index.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Schemas"
    ws.append(["Name", "Parent", "Type", "Pattern", "Example"])
    ws.append(["PreferredAgentBIC", "", "string", "[A-Z0-9]{4,4}[A-Z]{2,2}[A-Z0-9]{2,2}", "string"])
    wb.save(template_path)
    wb.close()

    try:
        converter = LegacyConverter(str(temp_root), str(temp_root / "out"), fill_fix_examples=True)
        assert converter.run_standalone_example_trace(str(temp_root), repair_files=True)

        repaired_wb = load_workbook(template_path)
        try:
            repaired = repaired_wb["Schemas"].cell(row=2, column=5).value
        finally:
            repaired_wb.close()

        assert repaired != "string"
        assert all(re.fullmatch(r"[A-Z0-9]{4,4}[A-Z]{2,2}[A-Z0-9]{2,2}", value) for value in repaired.split("; "))
    finally:
        shutil.rmtree(temp_root, ignore_errors=True)


def test_legacy_converter_blocks_conversion_when_index_references_missing_file():
    temp_root = _clean_test_temp()
    index_path = temp_root / "$index.xlsx"
    logs: list[str] = []
    wb = Workbook()
    ws = wb.active
    ws.title = "Paths"
    ws.append(["Path", "Method", "Excel File", "OperationId"])
    ws.append(["/missing", "get", "listLCR.260424", "listLCR"])
    wb.save(index_path)
    wb.close()

    try:
        converter = LegacyConverter(str(temp_root), str(temp_root / "out"), log_callback=logs.append)

        assert converter.convert() is False
        assert converter.missing_index_files == ["listLCR.260424"]
        assert any("ERROR: $index.xlsx references missing endpoint templates" in line for line in logs)
        assert any("ERROR: Conversion aborted because $index.xlsx references missing endpoint templates." in line for line in logs)
    finally:
        shutil.rmtree(temp_root, ignore_errors=True)


def test_legacy_converter_blocks_conversion_when_allowed_values_do_not_match_type():
    temp_root = _clean_test_temp()
    template_path = temp_root / "insertAP.240906.xlsm"
    second_template_path = temp_root / "listThrottlingThreshold.240423.xlsm"
    logs: list[str] = []

    def write_template(path: Path, field_name: str, row_number: int) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "Data Type"
        ws.append(["Data Type"])
        ws.append(["Name", "Description", "Type", "Allowed value", "Example"])
        for _ in range(row_number - 3):
            ws.append(["", "", "", "", ""])
        ws.append([field_name, "", "boolean", "0,1", ""])
        wb.save(path)
        wb.close()

    write_template(template_path, "BeneficiaryBankPosPaymConf", 8)
    write_template(second_template_path, "DefaultId", 12)

    try:
        converter = LegacyConverter(str(temp_root), str(temp_root / "out"), log_callback=logs.append)

        assert converter.convert() is False
        joined = "\n".join(logs)
        assert "=== CONVERSION BLOCKED: TEMPLATE DATA ERRORS ===" in joined
        assert "2 issue(s) found. Fix the Excel templates, then run conversion again." in joined
        assert "+-- ACTION REQUIRED ----------------------------------------+" in joined
        assert "| Fix the Allowed value column in the Data Type sheet.      |" in joined
        assert "| Ensure each Allowed value matches its declared Type.      |" in joined
        assert "| For boolean fields use true,false instead of 0,1.         |" in joined
        assert "| If 0/1 is intended, change Type or use true,false.        |" in joined
        assert "ISSUES" in joined
        assert "[1] insertAP.240906.xlsm" in joined
        assert str(template_path) in joined
        assert "Sheet: Data Type" in joined
        assert "Row: 8" in joined
        assert "Field: BeneficiaryBankPosPaymConf" in joined
        assert "[2] listThrottlingThreshold.240423.xlsm" in joined
        assert str(second_template_path) in joined
        assert "Row: 12" in joined
        assert "Field: DefaultId" in joined
        assert "Type: boolean" in joined
        assert "Allowed value: 0,1" in joined
        assert "Problem: boolean allowed values must be true/false." in joined
        assert "CONVERSION FAILED" in joined
        assert "No files were converted." in joined
    finally:
        shutil.rmtree(temp_root, ignore_errors=True)


def test_legacy_converter_allows_quoted_and_unquoted_string_allowed_values():
    converter = LegacyConverter("in", "out")

    unquoted = DataType(name="DefaultId", type="string", allowed_values="0,1")
    quoted = DataType(name="Status", type="string", allowed_values="'ACTV',\"INAC\"")

    assert converter._allowed_values_type_mismatch_reason(
        unquoted,
        converter._split_allowed_value_tokens(unquoted.allowed_values),
    ) == ""
    assert converter._allowed_values_type_mismatch_reason(
        quoted,
        converter._split_allowed_value_tokens(quoted.allowed_values),
    ) == ""
