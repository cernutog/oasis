"""
Debug the full flow using _convert_body_example as called in the real process
"""
import sys
from pathlib import Path

ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT))

from src.legacy_converter import LegacyConverter
import openpyxl
import pandas as pd

INPUT_DIR = ROOT / "Templates Legacy"
OUTPUT_DIR = ROOT / "tests" / "_full_flow_debug"
MASTER_DIR = ROOT / "Templates Master"

def log(msg):
    print(msg)

# Clean output
if OUTPUT_DIR.exists():
    import shutil
    shutil.rmtree(OUTPUT_DIR)
OUTPUT_DIR.mkdir(parents=True)

converter = LegacyConverter(
    input_dir=str(INPUT_DIR),
    output_dir=str(OUTPUT_DIR),
    master_dir=str(MASTER_DIR),
    log_callback=log,
)

# Load global schemas
converter._collect_all_data_types()
converter._perform_naming_and_usage_pass()

# Convert only retransmitOutFiles using the full flow
ep_filename = "retransmitOutFiles.210702.xlsm"
ep_path = INPUT_DIR / ep_filename

if ep_path.exists():
    print(f"Converting {ep_filename} using full flow...")
    
    # Load the master template
    master_path = MASTER_DIR / "endpoint.xlsx"
    wb = openpyxl.load_workbook(master_path)
    
    # Load the legacy file
    xl = pd.ExcelFile(ep_path)
    op_id = converter.filename_to_opid.get(ep_filename, converter._extract_operation_id(ep_filename))
    
    # Call _convert_body_example as in the real flow
    converter._convert_body_example(wb, xl, op_id, ep_filename)
    
    # Save and read the result
    out_path = OUTPUT_DIR / ep_filename.replace(".xlsm", ".xlsx")
    wb.save(out_path)
    
    # Read the generated Body Example
    wb_result = openpyxl.load_workbook(out_path, read_only=True)
    if "Body Example" in wb_result.sheetnames:
        ws = wb_result["Body Example"]
        print(f"\n=== Body Example for {ep_filename} ===")
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            name = row[0]
            body = row[1]
            if name or body:
                print(f"\n  [{name}]")
                if body:
                    print(f"    Raw body type: {type(body)}")
                    print(f"    Raw body repr: {repr(body)}")
                    for line in str(body).splitlines():
                        print(f"    {line}")
    else:
        print("No Body Example sheet found")
    wb_result.close()
else:
    print(f"Input file not found: {ep_path}")
