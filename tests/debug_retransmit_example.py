"""
Test only retransmitOutFiles Body Example generation
"""
import sys
from pathlib import Path

ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT))

from src.legacy_converter import LegacyConverter
import openpyxl

INPUT_DIR = ROOT / "Templates Legacy"
OUTPUT_DIR = ROOT / "tests" / "_retransmit_debug"

def log(msg):
    print(msg)

# Clean output
if OUTPUT_DIR.exists():
    import shutil
    shutil.rmtree(OUTPUT_DIR)
OUTPUT_DIR.mkdir(parents=True)

converter = LegacyConverter(
    input_dir=str(INPUT_DIR),
    output_dir=str(OUTPUT_DIR),
    master_dir=str(ROOT / "Templates Master"),
    log_callback=log,
)

# Convert only retransmitOutFiles
ep_filename = "retransmitOutFiles.210702.xlsm"
ep_path = INPUT_DIR / ep_filename

if ep_path.exists():
    print(f"Converting {ep_filename}...")
    converter._convert_endpoint(ep_path)
    
    # Read the generated Body Example
    out_path = OUTPUT_DIR / ep_filename.replace(".xlsm", ".xlsx")
    if out_path.exists():
        wb = openpyxl.load_workbook(out_path, read_only=True)
        if "Body Example" in wb.sheetnames:
            ws = wb["Body Example"]
            print(f"\n=== Body Example for {ep_filename} ===")
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                name = row[0]
                body = row[1]
                if name or body:
                    print(f"\n  [{name}]")
                    if body:
                        for line in str(body).splitlines():
                            print(f"    {line}")
        else:
            print("No Body Example sheet found")
        wb.close()
    else:
        print(f"Output file not found: {out_path}")
else:
    print(f"Input file not found: {ep_path}")
