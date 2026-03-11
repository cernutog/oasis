"""
Check which mandatory fields are missing from revokeLcr Body Example
"""
import sys
from pathlib import Path

ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT))

from src.legacy_converter import LegacyConverter
import openpyxl

INPUT_DIR = ROOT / "Templates Legacy"
MASTER_DIR = ROOT / "Templates Master"

def log(msg):
    print(msg)

converter = LegacyConverter(
    input_dir=str(INPUT_DIR),
    output_dir=str(ROOT / "tests" / "_debug_out"),
    master_dir=str(MASTER_DIR),
    log_callback=log,
)

# Load global schemas
converter._collect_all_data_types()
converter._perform_naming_and_usage_pass()

# Get the Body structure for revokeLcr
import pandas as pd
ep_filename = "revokeLcr.230301.xlsm"
ep_path = INPUT_DIR / ep_filename

if ep_path.exists():
    xl = pd.ExcelFile(ep_path)
    children = converter._read_legacy_structure(xl, "Body")
    
    print(f"=== revokeLcr mandatory fields vs generated example ===")
    
    # Get all mandatory fields from Body sheet
    mandatory_fields = []
    for name, parent, desc, dtype, mandatory, v_rules, items_type in children:
        if mandatory and mandatory.strip().upper() == 'M':
            mandatory_fields.append(name)
    
    print(f"\nMandatory fields from Body sheet: {mandatory_fields}")
    
    # Get generated example
    result = converter._build_body_example_dict(children, ep_filename)
    print(f"Generated example keys: {list(result.keys())}")
    
    # Check what's missing
    missing = []
    for field in mandatory_fields:
        if field not in result:
            missing.append(field)
    
    if missing:
        print(f"\n[MISSING] MANDATORY FIELDS: {missing}")
    else:
        print(f"\n[OK] All mandatory fields present")
    
    # Also check the actual generated file
    out_path = ROOT / "Templates Converted Test" / "revokeLcr.230301.xlsx"
    if out_path.exists():
        wb = openpyxl.load_workbook(out_path, read_only=True)
        if "Body Example" in wb.sheetnames:
            ws = wb["Body Example"]
            print(f"\n=== Actual Body Example from file ===")
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                name = row[0]
                body = row[1]
                if name == "OK" and body:
                    print(f"OK example content:")
                    for line in str(body).splitlines():
                        print(f"  {line}")
                    
                    # Parse YAML to check keys
                    import yaml
                    try:
                        parsed = yaml.safe_load(body)
                        print(f"Parsed keys: {list(parsed.keys()) if isinstance(parsed, dict) else 'Not a dict'}")
                        
                        if isinstance(parsed, dict):
                            for key in parsed.keys():
                                if isinstance(parsed[key], dict):
                                    print(f"  {key}: {list(parsed[key].keys())}")
                    except:
                        print("Failed to parse YAML")
        wb.close()
else:
    print(f"File not found: {ep_path}")
