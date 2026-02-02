import openpyxl
import os

master_index = r"c:\Users\giuse\.gemini\antigravity\scratch\OASIS\Templates Master\$index.xlsx"
if os.path.exists(master_index):
    try:
        wb = openpyxl.load_workbook(master_index)
        print(f"Sheet names: {wb.sheetnames}")
        if "General Description" in wb.sheetnames:
            ws = wb["General Description"]
            print("--- General Description Content ---")
            for r in range(1, 15):
                row_vals = [c.value for c in ws[r]]
                print(f"Row {r}: {row_vals}")
        else:
            print("Sheet 'General Description' not found in master index.")
    except Exception as e:
        print(f"Error: {e}")
else:
    print(f"Master index not found at {master_index}")
