"""
Detailed inspection of Body Example sheets - shows all populated files.
"""
import sys
from pathlib import Path

ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT))

import openpyxl

OUTPUT_DIR = ROOT / "tests" / "_body_example_out"

def main():
    for xlsx in sorted(OUTPUT_DIR.glob("*.xlsx")):
        if xlsx.name.startswith("$"):
            continue
        wb = openpyxl.load_workbook(xlsx, read_only=True)
        if "Body Example" not in wb.sheetnames:
            wb.close()
            continue
        ws = wb["Body Example"]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            name = row[0]
            body = row[1]
            if name or body:
                rows.append((name, body))
        wb.close()
        if not rows:
            continue

        print(f"\n{'='*70}")
        print(f"FILE: {xlsx.name}")
        print(f"{'='*70}")
        for ex_name, body in rows:
            print(f"\n  [{ex_name}]")
            if body:
                for line in str(body).splitlines():
                    print(f"    {line}")

if __name__ == "__main__":
    main()
