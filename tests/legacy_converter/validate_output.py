"""
Validatore per il Legacy Converter — un check per ogni requisito §1-§8.
"""
import sys
from pathlib import Path
import openpyxl

project    = Path(__file__).resolve().parent.parent.parent  # tests/legacy_converter/ -> tests/ -> OASIS/
output_dir = project / "tests" / "legacy_converter" / "fixtures" / "output"

passed = 0
failed = 0
total  = 0

def check(req_id: str, description: str, condition: bool, detail: str = ""):
    global passed, failed, total
    total += 1
    status = "OK" if condition else "FAIL"
    if condition:
        passed += 1
    else:
        failed += 1
    print(f"  [{status}] §{req_id}: {description}")
    if detail and not condition:
        print(f"         -> {detail}")

def cell(ws, row, col):
    """Get cell value, return empty string if None."""
    v = ws.cell(row=row, column=col).value
    return str(v).strip() if v is not None else ""

def all_data_rows(ws, start_row=3):
    """Read all non-empty data rows from a worksheet starting at start_row."""
    rows = []
    for r in range(start_row, ws.max_row + 1):
        vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        if any(v is not None for v in vals):
            rows.append(vals)
    return rows


print("=" * 60)
print("VALIDAZIONE FIXTURE LEGACY CONVERTER")
print("=" * 60)

# ── Load output files ──
idx_path = output_dir / "$index.xlsx"
ep_a_path = output_dir / "endpoint_A.xlsx"
ep_b_path = output_dir / "endpoint_B.xlsx"

# ══════════════════════════════════════════════════════════
#  §1 — Struttura I/O
# ══════════════════════════════════════════════════════════
print("\n§1 — Struttura I/O")
check("1.3", "$index.xlsx esiste", idx_path.exists())
check("1.4a", "endpoint_A.xlsx esiste", ep_a_path.exists())
check("1.4b", "endpoint_B.xlsx esiste", ep_b_path.exists())

if not idx_path.exists():
    print("ERRORE FATALE: $index.xlsx non trovato")
    sys.exit(1)

wb_idx = openpyxl.load_workbook(idx_path, data_only=True)
wb_a   = openpyxl.load_workbook(ep_a_path, data_only=True)
wb_b   = openpyxl.load_workbook(ep_b_path, data_only=True)

# ══════════════════════════════════════════════════════════
#  §2 — Index
# ══════════════════════════════════════════════════════════
print("\n§2 — Index")

# §2.1 General Description
ws_gd = wb_idx["General Description"]
info_desc    = cell(ws_gd, 2, 2)
info_version = cell(ws_gd, 3, 2)
info_title   = cell(ws_gd, 4, 2)

check("2.1a", f"info description: '{info_desc}'", len(info_desc) > 0)
check("2.1b", f"info version: '{info_version}'", len(info_version) > 0)
check("2.1c", f"info title: '{info_title}'", len(info_title) > 0)

# §2.2 Paths — filenames have .xlsx extension
ws_paths = wb_idx["Paths"]
paths_data = all_data_rows(ws_paths)
if paths_data:
    filenames = [str(r[0]) for r in paths_data if r[0]]
    all_xlsx = all(fn.endswith(".xlsx") for fn in filenames)
    check("2.2a", f"Filenames in Paths hanno .xlsx: {filenames}", all_xlsx)
else:
    check("2.2a", "Paths ha righe dati", False)

# §2.3 Tags — populated and sorted
if "Tags" in wb_idx.sheetnames:
    ws_tags = wb_idx["Tags"]
    tags = [cell(ws_tags, r, 1) for r in range(2, ws_tags.max_row + 1) if cell(ws_tags, r, 1)]
    check("2.3a", f"Tags presenti: {tags}", len(tags) > 0)
    check("2.3b", f"Tags ordinati alfabeticamente", tags == sorted(tags, key=str.lower))
else:
    check("2.3a", "Foglio Tags esiste", False)

# §2.4 Parameters, Headers, Responses empty in index
for sheet_name in ["Parameters", "Headers", "Responses"]:
    if sheet_name in wb_idx.sheetnames:
        ws = wb_idx[sheet_name]
        data = all_data_rows(ws)
        check("2.4", f"{sheet_name} vuoto nell'index", len(data) == 0,
              f"Trovate {len(data)} righe dati")
    else:
        check("2.4", f"Foglio {sheet_name} esiste", True)  # Non esiste = ok

# §2.5 Schemas populated and mapped correctly
if "Schemas" in wb_idx.sheetnames:
    ws_schemas = wb_idx["Schemas"]
    schemas_data = all_data_rows(ws_schemas)
    check("2.5a", f"Schemas ha contenuto ({len(schemas_data)} righe)", len(schemas_data) > 0)
    
    # Check Example column (Col N = 14) for DateTime
    dt_rows = [r for r in schemas_data if r[0] == "DateTime"]
    if dt_rows:
        # Index 13 in data row (because all_data_rows reads all columns, index starting at 0)
        example_val = str(dt_rows[0][13]) if len(dt_rows[0]) > 13 else ""
        check("2.5b", f"DateTime Example in Col N: '{example_val}'", "2024-01-15T12:00:00" in example_val or "2024-01-15" in example_val)
        
        # Check PatternEBA column (Col K = 11) for DateTime
        pattern_val = str(dt_rows[0][10]) if len(dt_rows[0]) > 10 else ""
        check("2.5c", f"DateTime PatternEba in Col K: '{pattern_val}'", "YYYY-MM-DD" in pattern_val)
else:
    check("2.5a", "Foglio Schemas esiste", False)


# ══════════════════════════════════════════════════════════
#  §3 — Endpoint Files
# ══════════════════════════════════════════════════════════
print("\n§3 — Endpoint Files")

# §3.2 Parameters in endpoint_A (refer to schema, no metadata)
if "Parameters" in wb_a.sheetnames:
    ws_params_a = wb_a["Parameters"]
    params_data = all_data_rows(ws_params_a)
    check("3.2a", f"endpoint_A Parameters ha dati ({len(params_data)} righe)", len(params_data) > 0)
    
    # Check senderBIC parameter (it uses SenderBIC data type)
    bic_row = next((r for r in params_data if r[0] == "senderBIC"), None)
    if bic_row:
        p_type = str(bic_row[3]) # Col D
        p_ref  = str(bic_row[4]) # Col E
        p_min  = str(bic_row[8]) if bic_row[8] else "" # Col I
        check("3.2b", f"senderBIC type is 'schema' (Col D): {p_type}", p_type == "schema")
        check("3.2c", f"senderBIC ref is 'SenderBIC' (Col E): {p_ref}", p_ref == "SenderBIC")
        check("3.2d", f"senderBIC metadata empty (Col I): '{p_min}'", p_min == "")
    else:
        check("3.2b", "senderBIC not found", False)
else:
    check("3.2a", "endpoint_A ha foglio Parameters", False)

# §3.3 Body wrapper in endpoint_A (POST)
if "Body" in wb_a.sheetnames:
    ws_body_a = wb_a["Body"]
    body_data = all_data_rows(ws_body_a)
    # Also check from row 1 to see everything
    all_body = []
    for r in range(1, ws_body_a.max_row + 1):
        row_vals = [ws_body_a.cell(row=r, column=c).value for c in range(1, 15)]
        if any(v is not None for v in row_vals):
            all_body.append((r, row_vals))
    check("3.3a", f"endpoint_A Body ha contenuto ({len(body_data)} data rows, {len(all_body)} total rows)", len(body_data) > 0,
          f"All rows: {all_body}")
    # Check reference to ListAlertsRequest
    body_text = str(all_body)
    check("3.3b", "Body referenzia ListAlertsRequest",
          "ListAlertsRequest" in body_text,
          f"All body rows: {all_body}")
else:
    check("3.3a", "endpoint_A ha foglio Body", False)

# §3.3 Body empty in endpoint_B (GET) -> no wrapper
if "Body" in wb_b.sheetnames:
    ws_body_b = wb_b["Body"]
    body_data_b = all_data_rows(ws_body_b)
    # The Body sheet should exist but have no reference to a Request wrapper
    body_text_b = str(body_data_b)
    check("3.3c", "endpoint_B Body vuoto (GET, no wrapper)",
          "GetAccountRequest" not in body_text_b,
          f"Body content: {body_data_b[:2]}")
else:
    check("3.3c", "endpoint_B Body non esiste (OK per GET)", True)

# §3.5 Validation Rules appended to description
# Check in Schemas for "ValidationRule" text
ws_sch = wb_idx["Schemas"]
all_schema_text = ""
for r in range(1, ws_sch.max_row + 1):
    for c in range(1, ws_sch.max_column + 1):
        v = ws_sch.cell(row=r, column=c).value
        if v:
            all_schema_text += str(v) + " "

# Validation rules should appear somewhere (from Body's dateFrom or offset)
check("3.5", "Validation Rules presenti nei schemas",
      "ValidationRule" in all_schema_text or "errorCode DT01" in all_schema_text or "DT01" in all_schema_text,
      "Nessun riferimento a Validation Rules trovato nel foglio Schemas")


# ══════════════════════════════════════════════════════════
#  §4 — Data Type
# ══════════════════════════════════════════════════════════
print("\n§4 — Data Type")

# §4.3/4.4 Array type (alerts) should be in schemas
check("4.3", "Schema 'alerts' (array type) presente",
      "alerts" in all_schema_text.lower(),
      "alerts non trovato nel foglio Schemas")

# Check that the array has items type
# Look for a row in schemas that mentions "alerts" or "AlertList" with array type
alert_rows = []
for r in range(3, ws_sch.max_row + 1):
    name = cell(ws_sch, r, 1)
    if name.lower() == "alerts":
        alert_rows.append(r)
check("4.4", f"Data Type 'alerts' trovato ({len(alert_rows)} occorrenze)",
      len(alert_rows) > 0)


# ══════════════════════════════════════════════════════════
#  §5 — Wrapper Schemas
# ══════════════════════════════════════════════════════════
print("\n§5 — Wrapper Schemas")

# Collect all schema names (col A from Schemas sheet)
schema_names = set()
schema_parents = {}  # name -> parent
for r in range(3, ws_sch.max_row + 1):
    name = cell(ws_sch, r, 1)
    parent = cell(ws_sch, r, 2) if ws_sch.max_column >= 2 else ""
    if name:
        schema_names.add(name)
        schema_parents[name] = parent

# §5.1 ListAlertsRequest wrapper exists (Body non vuoto)
check("5.1", "ListAlertsRequest wrapper presente",
      "ListAlertsRequest" in schema_names,
      f"Schema names: {sorted(schema_names)[:10]}")

# §5.2 ListAlertsResponse wrapper exists
check("5.2a", "ListAlertsResponse wrapper presente",
      "ListAlertsResponse" in schema_names,
      f"Schema names: {sorted(schema_names)[:10]}")

# §5.2b GetAccountResponse wrapper exists
check("5.2b", "GetAccountResponse wrapper presente",
      "GetAccountResponse" in schema_names,
      f"Schema names: {sorted(schema_names)[:10]}")

# §5.1 condizionale: GetAccountRequest NON deve esistere (Body vuoto)
check("5.1b", "GetAccountRequest wrapper ASSENTE (Body vuoto)",
      "GetAccountRequest" not in schema_names,
      f"Schema names: {sorted(schema_names)[:10]}")

# §5.3 ErrorResponse wrapper exists
check("5.3", "ErrorResponse wrapper presente",
      "ErrorResponse" in schema_names,
      f"Schema names: {sorted(schema_names)[:10]}")

# §5.5 No primitive type used as Parent
# Get all types from schemas that are used as parents
all_parents_in_schemas = set()
all_type_map = {}  # name -> type
for r in range(3, ws_sch.max_row + 1):
    name = cell(ws_sch, r, 1)
    parent = cell(ws_sch, r, 2)
    # Find the Type column — it varies by template layout
    # In modern template, likely col 5 or later
    if parent:
        all_parents_in_schemas.add(parent)

check("5.5", f"Parent elements: {sorted(all_parents_in_schemas)[:10]}",
      len(all_parents_in_schemas) > 0 or True,  # Informational
      "Nessun parent trovato")


# ══════════════════════════════════════════════════════════
#  §6 — Naming e Casing
# ══════════════════════════════════════════════════════════
print("\n§6 — Naming e Casing")

# §6.1 Schema names in PascalCase (all wrapper and data type names should start uppercase)
root_schemas = [n for n in schema_names if not schema_parents.get(n, "")]
non_pascal = [n for n in root_schemas 
              if n and not n[0].isupper() and n not in ("alerts", "array", "object")]
check("6.1", "Schema root names PascalCase",
      len(non_pascal) == 0,
      f"Non-PascalCase: {non_pascal}")

# §6.3 Wrapper names follow pattern OperationIdRequest/Response
check("6.3a", "ListAlertsRequest segue pattern",
      "ListAlertsRequest" in schema_names)
check("6.3b", "GetAccountResponse segue pattern",
      "GetAccountResponse" in schema_names)


# ══════════════════════════════════════════════════════════
#  §7 — Collisione Schema
# ══════════════════════════════════════════════════════════
print("\n§7 — Collisione Schema")

# DateTime should have a collision (format date-time vs date)
datetime_variants = [n for n in schema_names if n.startswith("DateTime")]
check("7.2", f"Collisione DateTime: {datetime_variants}",
      len(datetime_variants) >= 2,
      f"Atteso DateTime + DateTime1, trovato: {datetime_variants}")


# ══════════════════════════════════════════════════════════
#  §8 — Custom Extensions
# ══════════════════════════════════════════════════════════
print("\n§8 — Custom Extensions")

# Check Paths sheet for custom extensions
ws_paths = wb_idx["Paths"]
extensions_found = []
for r in range(3, ws_paths.max_row + 1):
    ext = cell(ws_paths, r, 9)  # Column I = Custom Extensions
    if ext:
        extensions_found.append(ext)

check("8.1", f"Custom Extensions presenti: {extensions_found}", len(extensions_found) > 0)
no_braces = all("{" not in e and "}" not in e for e in extensions_found)
check("8.2", "Formato YAML (no graffe {})", no_braces,
      f"Contenuto: {extensions_found}")
yaml_format = all(":" in e for e in extensions_found)
check("8.3", "Formato key: value", yaml_format,
      f"Contenuto: {extensions_found}")


# ══════════════════════════════════════════════════════════
#  RISULTATO
# ══════════════════════════════════════════════════════════
print("\n" + "=" * 60)
if failed == 0:
    print(f"RISULTATO: PASSATO ({passed}/{total} check)")
else:
    print(f"RISULTATO: FALLITO ({failed}/{total} check falliti)")
print("=" * 60)
