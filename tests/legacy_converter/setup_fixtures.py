"""
Setup fixture di test per il Legacy Converter.
Crea input che copre tutti i requisiti §1-§8.

Struttura:
  input/
    $index.xlsm          — General Description, Paths
    endpoint_A.xlsm      — POST /alerts (listAlerts)
    endpoint_B.xlsm      — GET /accounts/{accountId} (getAccount)
"""
import shutil
from pathlib import Path
from openpyxl import load_workbook


def setup():
    project   = Path(__file__).resolve().parent.parent.parent  # tests/legacy_converter/ -> tests/ -> OASIS/
    input_dir = project / "tests" / "legacy_converter" / "fixtures" / "input"
    master    = project / "Templates Master Legacy"

    # Pulizia robusta: elimina ogni file singolarmente
    if input_dir.exists():
        for f in input_dir.iterdir():
            if f.is_file() and not f.name.startswith("~$"):
                try:
                    f.unlink()
                except PermissionError:
                    print(f"  [WARN] Impossibile eliminare {f.name}, file bloccato")
    input_dir.mkdir(parents=True, exist_ok=True)

    # ══════════════════════════════════════════════════════════
    #  $index.xlsm
    # ══════════════════════════════════════════════════════════
    idx_path = input_dir / "$index.xlsm"
    shutil.copy(master / "$index.xlsm", idx_path)
    wb = load_workbook(idx_path, keep_vba=True)

    # --- General Description (§2.1) ---
    # Layout legacy: R1=title, R2=info description, R3=info version, R4=info title,
    #                R5=servers url + desc, R6=servers url + desc
    ws = wb["General Description"]
    ws.cell(row=2, column=2, value="Regression API for Legacy Converter")
    ws.cell(row=3, column=2, value="1.0.0")
    ws.cell(row=4, column=2, value="Legacy Converter Regression API")

    # --- Paths (§2.2, §8) ---
    # Layout legacy: R1=title, R2=headers, R3+=data
    # Colonne: A=Excel file, B=Path, C=Name, D=Method, E=Description,
    #          F=Tag, G=Summary, H=OperationId, I=Custom Extensions
    ws_p = wb["Paths"]
    paths = [
        # A                    B                       C              D      E                     F            G               H              I
        ["endpoint_A.xlsm",   "/alerts",              "List Alerts", "post", "List all alerts",   "Monitoring", "Get alerts",   "listAlerts",  "x-custom-tag: monitoring"],
        ["endpoint_B.xlsm",   "/accounts/{accountId}", "Get Account", "get", "Get account info",  "Account",    "Account info", "getAccount",  "x-rate-limit: 100"],
    ]
    for r, row in enumerate(paths, start=3):
        for c, val in enumerate(row, start=1):
            ws_p.cell(row=r, column=c, value=val)

    wb.save(idx_path)
    print(f"  Creato: {idx_path.name}")

    # ══════════════════════════════════════════════════════════
    #  endpoint_A.xlsm  (POST /alerts — listAlerts)
    # ══════════════════════════════════════════════════════════
    ep_a = input_dir / "endpoint_A.xlsm"
    shutil.copy(master / "endpoint.xlsm", ep_a)
    wb_a = load_workbook(ep_a, keep_vba=True)

    # --- Data Type (§4) ---
    # Layout: R1=title, R2=headers, R3+=data
    # Colonne: A=Name, B=Description, C=Type, D=Format,
    #          E=Items Data Type, F=Min, G=Max,
    #          H=PatternEba, I=Regex, J=Allowed value, K=Example
    ws_dt = wb_a["Data Type"]
    data_types_a = [
        # Name        Description      Type      Format     Items  Min  Max   PatternEba              Regex                   Allowed        Example
        ["DateTime",  "ISO date-time", "string", "date-time", "",   "",  "",   "YYYY-MM-DDTHH:MM:SS",  "",                     "",            "2024-01-15T12:00:00"],
        ["EventType", "Alert type",    "string", "",          "",   "",  "",   "",                     "",                     "SEC,ERR,WRN", "SEC"],
        ["EventDesc", "Description",   "string", "",          "",   "",  "1000", "",                   "",                     "",            "Alert triggered"],
        ["BoolFlag",  "True/False",    "boolean","",          "",   "",  "",   "",                     "",                     "",            "true"],
        ["alerts",    "Alert list",    "array",  "",          "object","","",  "",                     "",                     "",            ""],
        ["Offset",    "Page token",    "string", "",          "",   "",  "36", "",                     "",                     "",            "abc123"],
        ["SenderBIC", "Sender BIC",    "string", "",          "",   "8", "11", "4!c2!a2!c",            "[A-Z0-9]{4}[A-Z]{2}[A-Z0-9]{2}([A-Z0-9]{3})?", "", "TESTIT88XXX"],
        ["object",    "",              "object", "",          "",   "",  "",   "",                     "",                     "",            ""],
        ["array",     "",              "array",  "",          "object","","",  "",                     "",                     "",            ""],
    ]
    for r, row in enumerate(data_types_a, start=3):
        for c, val in enumerate(row, start=1):
            ws_dt.cell(row=r, column=c, value=val)

    # --- Path (§3.2, §6.4) ---
    # Layout: R1=title, R2=headers, R3+=data
    # Colonne: A=Type, B=Element, C=Mandatory, D=Constraint, E=Description, F=Validation Rules
    ws_path = wb_a["Path"]
    path_params = [
        ["SenderBIC", "senderBIC", "M", "", "The BIC of the sender", "Must match participant BIC (errorCode PY01)"],
    ]
    for r, row in enumerate(path_params, start=3):
        for c, val in enumerate(row, start=1):
            ws_path.cell(row=r, column=c, value=val)

    # --- Header (§3.2) ---
    ws_hdr = wb_a["Header"]
    hdr_params = [
        ["DateTime", "X-Request-Date", "M", "", "Date of the request", ""],
    ]
    for r, row in enumerate(hdr_params, start=3):
        for c, val in enumerate(row, start=1):
            ws_hdr.cell(row=r, column=c, value=val)

    # --- Body (§3.3, §5.1, §5.4, §5.5) ---
    # Layout: R1=title, R2=headers, R3+=data
    # Colonne: A=Type, B=Parent, C=Element, D=Parents, E=Mandatory,
    #          F=Constraint, G=Description, H=Validation Rules
    ws_body = wb_a["Body"]
    body_rows = [
        # Type        Parent            Element          Parents           Mand  Constraint  Description          ValidationRules
        ["DateTime",  "",               "dateTime",      "",               "M",  "",         "Request timestamp", ""],
        ["object",    "",               "searchCriteria","",               "M",  "",         "Search filters",    ""],
        ["DateTime",  "searchCriteria", "dateFrom",      "searchCriteria", "O",  "",         "Start date",        "Must be <= current date (errorCode DT01)"],
        ["DateTime",  "searchCriteria", "dateTo",        "searchCriteria", "O",  "",         "End date",          "Must be >= dateFrom (errorCode DT01)"],
        ["EventType", "searchCriteria", "eventType",     "searchCriteria", "O",  "",         "Filter by type",    ""],
        ["Offset",    "",               "offset",        "",               "O",  "",         "Pagination token",  "Must match previous response value"],
    ]
    for r, row in enumerate(body_rows, start=3):
        for c, val in enumerate(row, start=1):
            ws_body.cell(row=r, column=c, value=val)

    # --- 200 Response (§5.2) ---
    ws_200 = wb_a["200"]
    resp_200 = [
        ["DateTime",  "",        "dateTime",         "",        "M", "", "Response timestamp", ""],
        ["alerts",    "",        "alerts",           "",        "M", "", "Alert results",      ""],
        ["DateTime",  "alerts",  "dateTime",         "alerts",  "M", "", "Alert timestamp",    ""],
        ["EventType", "alerts",  "eventType",        "alerts",  "M", "", "Alert type",         ""],
        ["EventDesc", "alerts",  "eventDescription", "alerts",  "M", "", "Alert text",         ""],
        ["BoolFlag",  "",        "endOfList",        "",        "M", "", "Pagination flag",    ""],
        ["Offset",    "",        "offset",           "",        "O", "", "Next page token",    ""],
    ]
    for r, row in enumerate(resp_200, start=3):
        for c, val in enumerate(row, start=1):
            ws_200.cell(row=r, column=c, value=val)

    # --- 400 Response (§5.3, §6) ---
    ws_400 = wb_a["400"]
    resp_400 = [
        ["DateTime",  "", "dateTime",     "", "M", "", "Response timestamp", ""],
        ["EventType", "", "errorCode",    "", "M", "", "Error code",         ""],
        ["EventDesc", "", "errorMessage", "", "M", "", "Error description",  ""],
    ]
    for r, row in enumerate(resp_400, start=3):
        for c, val in enumerate(row, start=1):
            ws_400.cell(row=r, column=c, value=val)

    wb_a.save(ep_a)
    print(f"  Creato: {ep_a.name}")

    # ══════════════════════════════════════════════════════════
    #  endpoint_B.xlsm  (GET /accounts/{accountId} — getAccount)
    # ══════════════════════════════════════════════════════════
    ep_b = input_dir / "endpoint_B.xlsm"
    shutil.copy(master / "endpoint.xlsm", ep_b)
    wb_b = load_workbook(ep_b, keep_vba=True)

    # --- Data Type (§4, §7 collision: DateTime con format diverso) ---
    ws_dt_b = wb_b["Data Type"]
    data_types_b = [
        # Name           Description      Type      Format  Items  Min  Max              PatternEba  Regex          Allowed          Example
        ["DateTime",     "Account date",  "string", "date", "",    "",  "",              "YYYY-MM-DD", "",          "",              "2024-01-15"],
        ["AccountId",    "Account IBAN",  "string", "",     "",    "",  "34",            "",         "[A-Z]{2}\\d+", "",              "IT60X054281"],
        ["Amount",       "Currency amt",  "number", "double","",   "0", "999999999.99",  "",         "",            "",              "1500.50"],
        ["Currency",     "Currency code", "string", "",     "",    "3", "3",             "",         "[A-Z]{3}",    "EUR,USD,GBP",   "EUR"],
        ["AccountStatus","Account status","string", "",     "",    "",  "",              "",         "",            "active,closed", "active"],
        ["object",       "",              "object", "",     "",    "",  "",              "",         "",            "",              ""],
        ["array",        "",              "array",  "",     "object","","",              "",         "",            "",              ""],
    ]
    for r, row in enumerate(data_types_b, start=3):
        for c, val in enumerate(row, start=1):
            ws_dt_b.cell(row=r, column=c, value=val)

    # --- Path (§3.2) ---
    ws_path_b = wb_b["Path"]
    path_b = [
        ["AccountId", "accountId", "M", "", "The account identifier", ""],
    ]
    for r, row in enumerate(path_b, start=3):
        for c, val in enumerate(row, start=1):
            ws_path_b.cell(row=r, column=c, value=val)

    # --- Header: VUOTO ---
    # --- Body: VUOTO (è un GET → §3.3: no wrapper Request generato) ---

    # --- 200 Response (§5.2) ---
    ws_200_b = wb_b["200"]
    resp_200_b = [
        ["AccountId",     "", "id",       "", "M", "", "Account ID",       ""],
        ["AccountStatus", "", "status",   "", "M", "", "Account status",   ""],
        ["Amount",        "", "balance",  "", "M", "", "Current balance",  ""],
        ["Currency",      "", "currency", "", "M", "", "Account currency", ""],
        ["DateTime",      "", "openDate", "", "O", "", "Date opened",      ""],
    ]
    for r, row in enumerate(resp_200_b, start=3):
        for c, val in enumerate(row, start=1):
            ws_200_b.cell(row=r, column=c, value=val)

    # --- 400 Response (§5.3) ---
    ws_400_b = wb_b["400"]
    resp_400_b = [
        ["DateTime",  "", "dateTime",  "", "M", "", "Response timestamp", ""],
        ["EventType", "", "errorCode", "", "M", "", "Error code",         ""],
    ]
    for r, row in enumerate(resp_400_b, start=3):
        for c, val in enumerate(row, start=1):
            ws_400_b.cell(row=r, column=c, value=val)

    wb_b.save(ep_b)
    print(f"  Creato: {ep_b.name}")

    print(f"\nFixture creati in: {input_dir}")


if __name__ == "__main__":
    setup()
