"""
Quick smoke test for the Body Example generation feature.
Runs the legacy converter on Templates Legacy -> a temp output folder,
then reads back the Body Example sheet of one endpoint and prints the result.
"""
import sys
import os
import shutil
from pathlib import Path

# Make sure src is importable
ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT))

from src.legacy_converter import LegacyConverter

INPUT_DIR  = ROOT / "Templates Legacy"
OUTPUT_DIR = ROOT / "tests" / "_body_example_out"
MASTER_DIR = ROOT / "Templates Master"

def main():
    if OUTPUT_DIR.exists():
        shutil.rmtree(OUTPUT_DIR)
    OUTPUT_DIR.mkdir(parents=True)

    def log(msg):
        print(msg)

    converter = LegacyConverter(
        input_dir=str(INPUT_DIR),
        output_dir=str(OUTPUT_DIR),
        master_dir=str(MASTER_DIR),
        log_callback=log,
    )
    ok = converter.convert(tracing_enabled=False)
    if not ok:
        print("CONVERSION FAILED")
        sys.exit(1)

    # Pick a file that has a Body sheet - e.g. amendChangeSettlementBIC
    import openpyxl
    results = []
    for xlsx in sorted(OUTPUT_DIR.glob("*.xlsx")):
        if xlsx.name.startswith("$"):
            continue
        wb = openpyxl.load_workbook(xlsx, read_only=True)
        if "Body Example" not in wb.sheetnames:
            wb.close()
            continue
        ws = wb["Body Example"]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            name = row[0]
            body = row[1]
            if name or body:
                rows.append((name, body))
        wb.close()
        if rows:
            results.append((xlsx.name, rows))

    if not results:
        print("\n[WARN] No Body Example sheets were populated.")
        sys.exit(0)

    print(f"\n{'='*70}")
    print(f"Body Example populated in {len(results)} file(s)")
    print(f"{'='*70}")
    for fname, rows in results[:3]:  # show first 3
        print(f"\n--- {fname} ---")
        for name, body in rows:
            print(f"  [{name}]")
            if body:
                for line in str(body).splitlines():
                    print(f"    {line}")
    print(f"\n{'='*70}")
    print("DONE")

if __name__ == "__main__":
    main()
