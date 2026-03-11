"""
Debug revokeLcr Body Example generation
"""
import sys
from pathlib import Path

ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT))

from src.legacy_converter import LegacyConverter
import openpyxl

INPUT_DIR = ROOT / "Templates Legacy"
MASTER_DIR = ROOT / "Templates Master"

def log(msg):
    print(msg)

converter = LegacyConverter(
    input_dir=str(INPUT_DIR),
    output_dir=str(ROOT / "tests" / "_debug_out"),
    master_dir=str(MASTER_DIR),
    log_callback=log,
)

# Load global schemas
converter._collect_all_data_types()
converter._perform_naming_and_usage_pass()

# Get the Body structure for revokeLcr
import pandas as pd
ep_filename = "revokeLcr.230301.xlsm"
ep_path = INPUT_DIR / ep_filename

if ep_path.exists():
    xl = pd.ExcelFile(ep_path)
    children = converter._read_legacy_structure(xl, "Body")
    
    print(f"=== Debug revokeLcr Body Example ===")
    
    # Print all children structure
    print(f"\nAll children from _read_legacy_structure:")
    for i, child in enumerate(children, 1):
        name, parent, desc, dtype, mandatory, v_rules, items_type = child
        print(f"  {i}. {name:20s} | parent: {parent or 'ROOT':10s} | dtype: {dtype:15s} | mandatory: {mandatory or 'O':1s}")
    
    # Test _build_body_example_dict
    result = converter._build_body_example_dict(children, ep_filename)
    print(f"\n_build_body_example_dict result: {result}")
    
    # Check the actual generated file
    out_path = ROOT / "Templates Converted Test" / "revokeLcr.230301.xlsx"
    if out_path.exists():
        wb = openpyxl.load_workbook(out_path, read_only=True)
        if "Body Example" in wb.sheetnames:
            ws = wb["Body Example"]
            print(f"\n=== Body Example from generated file ===")
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                name = row[0]
                body = row[1]
                if name or body:
                    print(f"\n  [{name}]")
                    if body:
                        print(f"    Raw body type: {type(body)}")
                        print(f"    Raw body repr: {repr(body)}")
                        for line in str(body).splitlines():
                            print(f"    {line}")
        wb.close()
    else:
        print("Generated file not found")
else:
    print(f"File not found: {ep_path}")
