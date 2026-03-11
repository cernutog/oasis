"""
Regenerate retransmitOutFiles with the latest v2.1.45 fix
"""
import sys
from pathlib import Path

ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT))

from src.legacy_converter import LegacyConverter

INPUT_DIR = ROOT / "Templates Legacy"
OUTPUT_DIR = ROOT / "Templates Converted Test"
MASTER_DIR = ROOT / "Templates Master"

def log(msg):
    print(msg)

# Clean existing retransmitOutFiles
existing_file = OUTPUT_DIR / "retransmitOutFiles.210702.xlsx"
if existing_file.exists():
    import os
    os.remove(existing_file)
    print(f"Removed old file: {existing_file}")

converter = LegacyConverter(
    input_dir=str(INPUT_DIR),
    output_dir=str(OUTPUT_DIR),
    master_dir=str(MASTER_DIR),
    log_callback=log,
)

# Convert only retransmitOutFiles
ep_filename = "retransmitOutFiles.210702.xlsm"
ep_path = INPUT_DIR / ep_filename

if ep_path.exists():
    print(f"Regenerating {ep_filename} with v2.1.45 fix...")
    converter._convert_endpoint(ep_path)
    print(f"Generated: {OUTPUT_DIR / ep_filename.replace('.xlsm', '.xlsx')}")
    
    # Verify the Body Example
    import openpyxl
    out_path = OUTPUT_DIR / ep_filename.replace(".xlsm", ".xlsx")
    wb = openpyxl.load_workbook(out_path, read_only=True)
    if "Body Example" in wb.sheetnames:
        ws = wb["Body Example"]
        print(f"\n=== Body Example (NEW) ===")
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            name = row[0]
            body = row[1]
            if name or body:
                print(f"\n  [{name}]")
                if body:
                    for line in str(body).splitlines():
                        print(f"    {line}")
    wb.close()
else:
    print(f"Input file not found: {ep_path}")
