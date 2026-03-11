"""
Debug using the full LegacyConverter as initialized in the real flow
"""
import sys
from pathlib import Path

ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT))

from src.legacy_converter import LegacyConverter
import openpyxl

INPUT_DIR = ROOT / "Templates Legacy"
OUTPUT_DIR = ROOT / "Templates Converted Test"
MASTER_DIR = ROOT / "Templates Master"

def log(msg):
    print(msg)

# Use the full LegacyConverter as in the real application
converter = LegacyConverter(
    input_dir=str(INPUT_DIR),
    output_dir=str(OUTPUT_DIR),
    master_dir=str(MASTER_DIR),
    log_callback=log,
)

# Run the full conversion process as in the real app
print("=== Running full LegacyConverter conversion ===")
ok = converter.convert(tracing_enabled=False)

if ok:
    print("Conversion completed successfully")
    
    # Check the retransmitOutFiles result
    out_path = OUTPUT_DIR / "retransmitOutFiles.210702.xlsx"
    if out_path.exists():
        wb = openpyxl.load_workbook(out_path, read_only=True)
        if "Body Example" in wb.sheetnames:
            ws = wb["Body Example"]
            print(f"\n=== Body Example from full conversion ===")
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                name = row[0]
                body = row[1]
                if name or body:
                    print(f"\n  [{name}]")
                    if body:
                        print(f"    Raw body type: {type(body)}")
                        print(f"    Raw body repr: {repr(body)}")
                        for line in str(body).splitlines():
                            print(f"    {line}")
        wb.close()
    else:
        print("retransmitOutFiles.210702.xlsx not found")
else:
    print("Conversion failed")
