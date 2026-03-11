"""
Check retransmitOutFiles Body Example directly
"""
import sys
from pathlib import Path

ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT))

import openpyxl

# Read the actual file
file_path = ROOT / "Templates Converted Test" / "retransmitOutFiles.210702.xlsx"

if file_path.exists():
    wb = openpyxl.load_workbook(file_path, read_only=True)
    
    print(f"=== Checking {file_path.name} ===")
    
    if "Body Example" in wb.sheetnames:
        ws = wb["Body Example"]
        print(f"\n=== Body Example Sheet ===")
        
        # Read all rows
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if row_idx == 1:
                print(f"Header: {row}")
            else:
                name = row[0] if len(row) > 0 else ""
                body = row[1] if len(row) > 1 else ""
                if name or body:
                    print(f"\nRow {row_idx}: [{name}]")
                    if body:
                        print(f"Body content:")
                        for line in str(body).splitlines():
                            print(f"    {line}")
    else:
        print("No Body Example sheet found")
    
    wb.close()
else:
    print(f"File not found: {file_path}")
