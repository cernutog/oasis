"""Inspect Templates Converted Test/$index.xlsx Schemas entries for RevokeLcrRequest and nested nodes."""
import sys
from pathlib import Path

ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT))

import openpyxl

index_path = ROOT / "Templates Converted Test" / "$index.xlsx"
if not index_path.exists():
    print(f"Not found: {index_path}")
    raise SystemExit(1)

wb = openpyxl.load_workbook(index_path, read_only=True)
if "Schemas" not in wb.sheetnames:
    print("No Schemas sheet")
    raise SystemExit(1)

ws = wb["Schemas"]
rows = list(ws.iter_rows(values_only=True))
header = rows[0]
print("Header:", header)

# find column indices
col = {str(v).strip().lower(): i for i, v in enumerate(header) if v}

def get(row, key, default=""):
    i = col.get(key)
    if i is None or i >= len(row):
        return default
    return row[i]

# common keys in this project
name_key = "name"
parent_key = "parent"
type_key = "type"
mand_key = "mandatory"
schema_name_key = "schema name\n(for type or items data type = 'schema')".lower()

print("\n--- Rows for RevokeLcrRequest block ---")
block = []
in_block = False
block_indent_parent = None

for r in rows[1:]:
    name = get(r, name_key)
    parent = get(r, parent_key)
    if name == "RevokeLcrRequest" and (parent is None or str(parent).strip()==""):
        in_block = True
    if in_block:
        # stop at next root schema
        if name and name != "RevokeLcrRequest" and (parent is None or str(parent).strip()=="") and block:
            break
        block.append(r)

for r in block:
    name = get(r, name_key)
    parent = get(r, parent_key)
    typ = get(r, type_key)
    mand = get(r, mand_key)
    schema_name = None
    # try schema name col by fuzzy match
    for k, idx in col.items():
        if "schema name" in k:
            schema_name = r[idx]
            break
    print(f"name={name!s:25} parent={parent!s:25} type={typ!s:10} mandatory={mand!s:3} schema_name={schema_name!s}")

wb.close()
