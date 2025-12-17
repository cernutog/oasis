import openpyxl
import os
import glob

base_dir = r"C:\Users\giuse\.gemini\antigravity\scratch\OAS_Generation_Tool\API Templates"
target_file_pattern = "account_assessment_vop_bulk_{bulkId}*.xlsm"

def patch_bulk():
    files = glob.glob(os.path.join(base_dir, target_file_pattern))
    for f in files:
        print(f"Processing {os.path.basename(f)}...")
        try:
            wb = openpyxl.load_workbook(f, keep_vba=True)
            if "Parameters" in wb.sheetnames:
                ws = wb["Parameters"]
                headers = {}
                header_row_idx = 1
                for r_idx, row in enumerate(ws.iter_rows(max_row=5), start=1):
                    vals = [c.value for c in row]
                    if "Name" in vals:
                        header_row_idx = r_idx
                        for cell in row:
                           if cell.value: headers[cell.value] = cell.column
                        break
                
                name_col = headers.get("Name")
                in_col = headers.get("In")
                mand_col = headers.get("Mandatory")
                
                # If 'In' or 'Mandatory' columns are missing in header row, we might need to assume column index or fail
                # But typically they exist, just empty for that row.
                # If they don't exist in headers dict, it means the column header is missing/different.
                if not in_col:
                    print("  'In' column not found! Attempting to find empty column or heuristic?")
                    # This happens if column header is missing. 
                    # inspect_debug showed columns "Request Parameters", "Unnamed".
                    # Row 0 was "Request Parameters". Row 1 was "Name".
                    # Maybe "In" is unnamed?
                    pass
                
                if name_col:
                    found = False
                    for row in ws.iter_rows(min_row=header_row_idx+1):
                        val = row[name_col-1].value
                        if val == "bulkId":
                            print("  Patching bulkId...")
                            if in_col: row[in_col-1].value = "path"
                            if mand_col: row[mand_col-1].value = "Yes"
                            found = True
                            break
                    if not found:
                        print("  bulkId not found in rows.")
                
                wb.save(f)
                print("  Saved.")
                
        except Exception as e:
            print(f"Error: {e}")

if __name__ == "__main__":
    patch_bulk()
