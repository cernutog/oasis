
import openpyxl
import os

def inspect_order():
    path = "Templates Converted V3/$index.xlsx"
    if not os.path.exists(path):
        print("File not found.")
        return

    wb = openpyxl.load_workbook(path)
    if "Schemas" not in wb.sheetnames:
        print("Schemas sheet not found.")
        return
    
    ws = wb["Schemas"]
    print(f"Inspecting {path} - Schemas Sheet")
    
    last_root = None
    row_count = 0
    
    # Iterate explicitly
    for i, row in enumerate(ws.iter_rows(min_row=2, max_col=2, values_only=True), start=2):
        name = row[0]
        parent = row[1]
        
        if name is None and parent is None:
            print(f"Row {i}: [EMPTY SPACER]")
            continue
            
        print(f"Row {i}: Name='{name}', Parent='{parent}'")
        row_count += 1
        
        if row_count > 50: # Limit output
            print("... truncating output ...")
            break

if __name__ == "__main__":
    inspect_order()
