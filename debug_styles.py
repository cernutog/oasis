"""
Deep style inspection - compare reference vs generated Excel files.
"""

import os
import sys
sys.path.insert(0, 'src')

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill


def inspect_cell_styles(filepath, sheet_name, max_rows=5):
    """Inspect actual cell styles in a worksheet."""
    print(f"\n{'='*60}")
    print(f"FILE: {filepath}")
    print(f"SHEET: {sheet_name}")
    print('='*60)
    
    wb = load_workbook(filepath, data_only=False)
    
    if sheet_name not in wb.sheetnames:
        print(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
        return
    
    ws = wb[sheet_name]
    
    for row_idx in range(1, min(max_rows + 1, ws.max_row + 1)):
        print(f"\n--- Row {row_idx} ---")
        for col_idx in range(1, min(5, ws.max_column + 1)):
            cell = ws.cell(row=row_idx, column=col_idx)
            
            if cell.value is None:
                continue
                
            print(f"\nCell ({row_idx}, {col_idx}): '{str(cell.value)[:30]}...'")
            
            # Font details
            if cell.font:
                f = cell.font
                print(f"  FONT: name={f.name}, size={f.size}, bold={f.bold}")
                
                # Detailed color inspection
                if f.color:
                    print(f"  FONT COLOR object: {type(f.color)}")
                    print(f"    .type = {getattr(f.color, 'type', 'N/A')}")
                    print(f"    .theme = {getattr(f.color, 'theme', 'N/A')}")
                    print(f"    .indexed = {getattr(f.color, 'indexed', 'N/A')}")
                    print(f"    .tint = {getattr(f.color, 'tint', 'N/A')}")
                    try:
                        print(f"    .rgb = {f.color.rgb}")
                    except Exception as e:
                        print(f"    .rgb = ERROR: {e}")
            
            # Fill details
            if cell.fill:
                fill = cell.fill
                print(f"  FILL: type={fill.fill_type}")
                
                if fill.fgColor:
                    print(f"  FILL FG object: {type(fill.fgColor)}")
                    print(f"    .type = {getattr(fill.fgColor, 'type', 'N/A')}")
                    print(f"    .theme = {getattr(fill.fgColor, 'theme', 'N/A')}")
                    print(f"    .indexed = {getattr(fill.fgColor, 'indexed', 'N/A')}")
                    print(f"    .tint = {getattr(fill.fgColor, 'tint', 'N/A')}")
                    try:
                        print(f"    .rgb = {fill.fgColor.rgb}")
                    except Exception as e:
                        print(f"    .rgb = ERROR: {e}")
                
                if fill.bgColor:
                    print(f"  FILL BG object: {type(fill.bgColor)}")
                    print(f"    .type = {getattr(fill.bgColor, 'type', 'N/A')}")
                    print(f"    .theme = {getattr(fill.bgColor, 'theme', 'N/A')}")
                    try:
                        print(f"    .rgb = {fill.bgColor.rgb}")
                    except Exception as e:
                        print(f"    .rgb = ERROR: {e}")
    
    wb.close()


def main():
    # Inspect reference template
    inspect_cell_styles(
        'API Templates/account_assessment.251111.xlsm',
        'Parameters',
        max_rows=3
    )
    
    # Inspect generated file
    inspect_cell_styles(
        'Imported Templates/create-account-assessment.xlsx',
        'Parameters',
        max_rows=3
    )


if __name__ == '__main__':
    main()
